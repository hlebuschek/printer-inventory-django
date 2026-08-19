# integrations/management/commands/glpi_printer_avg.py
"""
Management command: средняя нагрузка принтеров по данным GLPI.

Читает xlsx со списком устройств (серийные номера), ищет каждое в GLPI по API,
берёт первую и последнюю запись PrinterLog и считает среднюю печать в месяц
по той же методике, что printer_monthly_avg: дельта счётчиков / количество месяцев.

С опцией --poll-stale-days N устройства, чья последняя запись в GLPI старше N дней,
дополнительно опрашиваются по SNMP по IP-адресам из карточки GLPI (перебираются все
адреса портов). Свежий счётчик становится последней точкой расчёта. Серийник из
SNMP-ответа сверяется с ожидаемым — чужое устройство на переехавшем IP отбрасывается.

Использование:
    python manage.py glpi_printer_avg /path/to/devices.xlsx
    python manage.py glpi_printer_avg devices.xlsx --output /tmp/glpi_avg.xlsx
    python manage.py glpi_printer_avg devices.xlsx --format csv
    python manage.py glpi_printer_avg devices.xlsx --serial-column "Серийный номер"
    python manage.py glpi_printer_avg devices.xlsx --poll-stale-days 30 --community public
"""

import csv
import logging
import os
import xml.etree.ElementTree as ET
from datetime import date, datetime, timedelta

import openpyxl
import requests
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from django.core.management.base import BaseCommand, CommandError

from integrations.glpi.client import GLPIAPIError, GLPIClient

logger = logging.getLogger(__name__)

RESULT_HEADERS = [
    "GLPI",
    "GLPI ID",
    "Состояние",
    "Первая запись",
    "Последняя запись",
    "Месяцев с данными",
    "Счётчик (последний)",
    "Среднее ЧБ / мес",
    "Среднее Цвет / мес",
    "Среднее всего / мес",
    "IP (GLPI)",
    "Опрос: IP",
    "Опрос: счётчик",
]

STATUS_LABELS = {
    "FOUND_SINGLE": "найден",
    "FOUND_MULTIPLE": "найден (несколько!)",
    "NOT_FOUND": "не найден",
    "ERROR": "ошибка",
}


class Command(BaseCommand):
    help = "Средняя нагрузка принтеров из xlsx-списка по данным GLPI PrinterLog"

    def add_arguments(self, parser):
        parser.add_argument("xlsx_path", help="Путь к xlsx со списком устройств")
        parser.add_argument(
            "--serial-column",
            default="Серийный номер",
            help="Название столбца с серийными номерами (по умолчанию «Серийный номер»)",
        )
        parser.add_argument(
            "--output",
            help="Путь к файлу вывода (по умолчанию: ./glpi_avg_YYYYMMDD.xlsx или .csv)",
        )
        parser.add_argument(
            "--format",
            choices=["excel", "csv"],
            default="excel",
            dest="fmt",
            help="Формат вывода: excel (по умолчанию) или csv",
        )
        parser.add_argument(
            "--poll-stale-days",
            type=int,
            metavar="N",
            help="Опросить по SNMP устройства, чья последняя запись в GLPI старше N дней (по умолчанию выключено)",
        )
        parser.add_argument(
            "--community",
            default="public",
            help="SNMP community для дополнительного опроса (по умолчанию public)",
        )

    def handle(self, *args, **options):
        xlsx_path = options["xlsx_path"]
        serial_column = options["serial_column"]
        output_path = options["output"]
        fmt = options["fmt"]
        poll_stale_days = options["poll_stale_days"]
        community = options["community"]

        if not os.path.exists(xlsx_path):
            raise CommandError(f"Файл не найден: {xlsx_path}")

        if poll_stale_days is not None:
            from inventory.services import _validate_glpi_installation

            glpi_ok, glpi_msg = _validate_glpi_installation()
            if not glpi_ok:
                raise CommandError(f"SNMP-опрос недоступен: {glpi_msg}")

        source_headers, devices = _read_devices(xlsx_path, serial_column)
        self.stdout.write(f"Устройств в файле: {len(devices)}")

        rows = []
        found = not_found = errors = polled = 0

        try:
            with GLPIClient() as client:
                for idx, (source_row, serial) in enumerate(devices, 1):
                    result = _probe_serial(client, serial)

                    if poll_stale_days is not None and result["glpi_id"] and _is_stale(result, poll_stale_days):
                        _poll_and_apply(client, serial, result, community)
                        if result["poll_ip"]:
                            polled += 1

                    rows.append((source_row, result))

                    status = result["status"]
                    if status in ("FOUND_SINGLE", "FOUND_MULTIPLE"):
                        found += 1
                    elif status == "NOT_FOUND":
                        not_found += 1
                    else:
                        errors += 1

                    label = STATUS_LABELS.get(status, status)
                    avg = result["avg_total"]
                    avg_str = f", среднее {avg}/мес" if avg is not None else ""
                    poll_str = f", опрошен {result['poll_ip']}" if result["poll_ip"] else ""
                    self.stdout.write(f"[{idx}/{len(devices)}] {serial}: {label}{avg_str}{poll_str}")
        except GLPIAPIError as e:
            raise CommandError(f"GLPI API: {e}")

        summary = f"Найдено: {found}, не найдено: {not_found}, ошибок: {errors}"
        if poll_stale_days is not None:
            summary += f", опрошено по SNMP: {polled}"
        self.stdout.write(summary)

        today_str = date.today().strftime("%Y%m%d")
        ext = "xlsx" if fmt == "excel" else "csv"
        if not output_path:
            output_path = os.path.join(os.getcwd(), f"glpi_avg_{today_str}.{ext}")

        headers = source_headers + RESULT_HEADERS
        if fmt == "excel":
            _write_excel(headers, rows, len(source_headers), output_path)
        else:
            _write_csv(headers, rows, output_path)

        self.stdout.write(self.style.SUCCESS(f"Файл сохранён: {output_path}"))


def _read_devices(xlsx_path: str, serial_column: str):
    """Возвращает (заголовки исходного файла, [(строка, серийник), ...])."""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active

    rows_iter = ws.iter_rows(values_only=True)
    headers = None
    serial_idx = None
    for row in rows_iter:
        values = [str(v).strip() if v is not None else "" for v in row]
        if serial_column in values:
            headers = values
            serial_idx = values.index(serial_column)
            break

    if headers is None:
        raise CommandError(f"Не найден столбец «{serial_column}» в файле {xlsx_path}")

    devices = []
    for row in rows_iter:
        values = list(row[: len(headers)])
        serial = values[serial_idx]
        serial = str(serial).strip() if serial is not None else ""
        if not serial:
            continue
        devices.append((values, serial))

    wb.close()
    return headers, devices


def _probe_serial(client: GLPIClient, serial: str) -> dict:
    """Ищет серийник в GLPI и считает среднюю печать в месяц по PrinterLog."""
    result = {
        "status": "ERROR",
        "glpi_id": None,
        "glpi_serial": "",
        "state": "",
        "first_date": None,
        "last_date": None,
        "num_months": None,
        "first_counters": {},
        "last_counter": None,
        "avg_bw": None,
        "avg_color": None,
        "avg_total": None,
        "ips": [],
        "poll_ip": "",
        "poll_counter": None,
    }

    status, items, error = client.search_printer_by_serial(serial)
    result["status"] = status
    if status not in ("FOUND_SINGLE", "FOUND_MULTIPLE") or not items:
        if error:
            logger.warning(f"GLPI поиск {serial}: {error}")
        return result

    glpi_id = items[0].get("2") or items[0].get("id")
    if not glpi_id:
        return result
    glpi_id = int(glpi_id)
    result["glpi_id"] = glpi_id
    result["glpi_serial"] = str(items[0].get("5") or items[0].get("serial") or "").strip()
    result["state"] = items[0].get("31") or ""

    first_logs = _get_printer_log(client, glpi_id, order="ASC")
    last_logs = _get_printer_log(client, glpi_id, order="DESC")
    if not first_logs or not last_logs:
        return result

    first, last = first_logs[0], last_logs[0]
    first_date = _parse_date(first.get("date"))
    last_date = _parse_date(last.get("date"))
    if not first_date or not last_date:
        return result

    result["first_date"] = first_date
    result["last_date"] = last_date
    num_months = (last_date.year - first_date.year) * 12 + (last_date.month - first_date.month) + 1
    result["num_months"] = num_months

    def to_int(record, field):
        try:
            return int(record[field])
        except (KeyError, ValueError, TypeError):
            return None

    result["last_counter"] = to_int(last, "total_pages")
    result["first_counters"] = {
        "total": to_int(first, "total_pages"),
        "bw": to_int(first, "bw_pages"),
        "color": to_int(first, "color_pages"),
    }

    def safe_delta(field):
        v_first, v_last = to_int(first, field), to_int(last, field)
        if v_first is None or v_last is None:
            return None
        return max(0, v_last - v_first)

    delta_total = safe_delta("total_pages")
    delta_bw = safe_delta("bw_pages")
    delta_color = safe_delta("color_pages")

    if delta_bw is not None:
        result["avg_bw"] = round(delta_bw / num_months, 1)
    if delta_color is not None:
        result["avg_color"] = round(delta_color / num_months, 1)
    if delta_total is not None:
        result["avg_total"] = round(delta_total / num_months, 1)
    elif delta_bw is not None or delta_color is not None:
        result["avg_total"] = round(((delta_bw or 0) + (delta_color or 0)) / num_months, 1)

    return result


def _get_printer_log(client: GLPIClient, printer_id: int, order: str):
    """Одна запись PrinterLog (ASC — самая ранняя, DESC — самая свежая).

    Запрос выполняется здесь, а не через client.get_printer_log(), чтобы команда
    оставалась одним файлом и работала со старой версией клиента на сервере.
    """
    try:
        response = requests.get(
            f"{client.url}/Printer/{printer_id}/PrinterLog/",
            headers=client._get_headers(with_session=True),
            params={"sort": "date", "order": order, "range": "0-0"},
            timeout=10,
            verify=client.verify_ssl,
        )
        # 206 Partial Content — штатный ответ GLPI на неполную выборку (range)
        if response.status_code in (200, 206):
            return response.json()
        logger.debug(f"PrinterLog {printer_id} ({order}): {response.status_code}")
        return None
    except requests.RequestException as e:
        logger.error(f"Ошибка получения PrinterLog для {printer_id}: {e}")
        return None


def _is_stale(result: dict, stale_days: int) -> bool:
    """Пора ли опрашивать: записей нет вообще или последняя старше stale_days."""
    if result["last_date"] is None:
        return True
    return result["last_date"] < date.today() - timedelta(days=stale_days)


def _get_printer_ips(client: GLPIClient, printer_id: int) -> list:
    """Все IP принтера из GLPI: NetworkPort → NetworkName → IPAddress (как get_printer_ip, но списком)."""
    try:
        response = requests.get(
            f"{client.url}/Printer/{printer_id}/NetworkPort",
            headers=client._get_headers(with_session=True),
            timeout=10,
            verify=client.verify_ssl,
        )
        if response.status_code not in (200, 206):
            return []
        ports = response.json() or []
    except requests.RequestException as e:
        logger.error(f"Ошибка получения NetworkPort принтера {printer_id}: {e}")
        return []

    ips = []
    for port in ports:
        ip = port.get("ip") or port.get("NetworkName_ip")
        if not ip:
            netname_id = port.get("NetworkName_id") or port.get("networknames_id")
            if not netname_id:
                netname_id = client._get_port_networkname_id(int(port.get("id")))
            if netname_id:
                ip = client._get_networkname_ip(int(netname_id))
        if ip:
            ip = str(ip).strip()
            if ip and ip not in ips:
                ips.append(ip)
    return ips


def _poll_and_apply(client: GLPIClient, serial: str, result: dict, community: str):
    """Опрашивает принтер по SNMP по всем IP из GLPI и подставляет свежий счётчик как последнюю точку."""
    ips = _get_printer_ips(client, result["glpi_id"])
    result["ips"] = ips
    if not ips:
        return

    expected_serials = {s.lower() for s in (serial, result["glpi_serial"]) if s}

    for ip in ips:
        polled = _poll_ip(ip, community)
        if not polled:
            continue

        polled_serial = (polled.get("serial") or "").lower()
        # IP мог перейти другому устройству — чужой серийник отбрасываем.
        # Пустой серийник в ответе допускаем: IP взят из карточки самого принтера.
        if polled_serial and expected_serials and polled_serial not in expected_serials:
            logger.warning(f"SNMP {ip}: серийник {polled['serial']} не совпал с {serial} — пропуск")
            continue

        total = polled.get("total")
        if total is None:
            continue

        result["poll_ip"] = ip
        result["poll_counter"] = total
        result["last_counter"] = total
        result["last_date"] = date.today()

        first = result["first_counters"]
        if result["first_date"] and first.get("total") is not None:
            num_months = (
                (result["last_date"].year - result["first_date"].year) * 12
                + (result["last_date"].month - result["first_date"].month)
                + 1
            )
            result["num_months"] = num_months
            result["avg_total"] = round(max(0, total - first["total"]) / num_months, 1)
            if polled.get("bw") is not None and first.get("bw") is not None:
                result["avg_bw"] = round(max(0, polled["bw"] - first["bw"]) / num_months, 1)
            if polled.get("color") is not None and first.get("color") is not None:
                result["avg_color"] = round(max(0, polled["color"] - first["color"]) / num_months, 1)
        return


def _poll_ip(ip: str, community: str):
    """SNMP-опрос одного IP через GLPI Agent. Возвращает {'serial', 'total', 'bw', 'color'} или None."""
    from inventory.services import _build_glpi_command, _cleanup_xml, _get_glpi_discovery_path, _possible_xml_paths
    from inventory.utils import run_glpi_command

    try:
        _cleanup_xml(ip)
        cmd = _build_glpi_command(_get_glpi_discovery_path(), ip, community)
    except ValueError as e:
        logger.warning(f"SNMP {ip}: {e}")
        return None

    ok, out = run_glpi_command(cmd)
    if not ok:
        logger.warning(f"SNMP {ip}: агент завершился с ошибкой: {out[:200] if out else ''}")
        return None

    for candidate in _possible_xml_paths(ip, prefer="inv"):
        if os.path.exists(candidate):
            return _parse_agent_xml(candidate)

    logger.warning(f"SNMP {ip}: XML после опроса не найден")
    return None


def _parse_agent_xml(xml_path: str):
    """Достаёт SERIAL и PAGECOUNTERS (TOTAL, BW_*, COLOR_*) из XML GLPI Agent."""
    serial = None
    counters = {}
    try:
        in_pagecounters = False
        for event, elem in ET.iterparse(xml_path, events=("start", "end")):
            tag = str(elem.tag).split("}", 1)[-1].upper()
            if tag == "PAGECOUNTERS":
                in_pagecounters = event == "start"
            elif event == "end":
                if tag == "SERIAL" and serial is None:
                    serial = (elem.text or "").strip() or None
                elif in_pagecounters and tag not in counters:
                    val = (elem.text or "").strip()
                    if val.isdigit():
                        counters[tag] = int(val)
    except ET.ParseError as e:
        logger.warning(f"Ошибка парсинга {xml_path}: {e}")
        return None

    def summed(*tags):
        vals = [counters[t] for t in tags if t in counters]
        return sum(vals) if vals else None

    return {
        "serial": serial,
        "total": counters.get("TOTAL"),
        "bw": summed("BW_A4", "BW_A3") if ("BW_A4" in counters or "BW_A3" in counters) else counters.get("BW"),
        "color": (
            summed("COLOR_A4", "COLOR_A3")
            if ("COLOR_A4" in counters or "COLOR_A3" in counters)
            else counters.get("COLOR")
        ),
    }


def _parse_date(date_str):
    if not date_str:
        return None
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            return datetime.strptime(str(date_str), fmt).date()
        except (ValueError, TypeError):
            continue
    return None


def _result_to_list(result: dict):
    return [
        STATUS_LABELS.get(result["status"], result["status"]),
        result["glpi_id"],
        result["state"],
        result["first_date"],
        result["last_date"],
        result["num_months"],
        result["last_counter"],
        result["avg_bw"],
        result["avg_color"],
        result["avg_total"],
        "; ".join(result["ips"]),
        result["poll_ip"],
        result["poll_counter"],
    ]


def _write_excel(headers, rows, n_source_cols, path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Средняя нагрузка (GLPI)"

    bold = Font(bold=True)
    col_widths = [len(h) + 2 for h in headers]

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = bold

    date_cols = {n_source_cols + 4, n_source_cols + 5}
    float_cols = {n_source_cols + 8, n_source_cols + 9, n_source_cols + 10}

    for row_idx, (source_row, result) in enumerate(rows, 2):
        values = list(source_row) + _result_to_list(result)
        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if col_idx in date_cols and value is not None:
                cell.number_format = "dd.mm.yyyy"
            elif col_idx in float_cols and value is not None:
                cell.number_format = "0.0"

            cell_len = len(str(value)) if value is not None else 0
            if col_idx <= len(col_widths) and cell_len > col_widths[col_idx - 1]:
                col_widths[col_idx - 1] = cell_len

    for col_idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = min(width + 2, 50)

    wb.save(path)


def _write_csv(headers, rows, path):
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f, delimiter=";")
        writer.writerow(headers)
        for source_row, result in rows:
            values = list(source_row) + _result_to_list(result)
            formatted = []
            for v in values:
                if isinstance(v, date):
                    formatted.append(v.strftime("%d.%m.%Y"))
                elif v is None:
                    formatted.append("")
                elif isinstance(v, float):
                    formatted.append(str(v).replace(".", ","))
                else:
                    formatted.append(v)
            writer.writerow(formatted)
