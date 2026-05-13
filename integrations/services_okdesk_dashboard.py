"""
Сервисный слой для Service Desk dashboard'а Okdesk.
Бизнес-логика подсчёта статистики, группировки по статусам, формирования Excel.
"""

from datetime import date, datetime, timedelta
from io import BytesIO

from django.db.models import Count, Max, Q
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .models import OkdeskComment, OkdeskIssue

# Статусы, которые считаем активными (заявка не закрыта, не отменена)
ACTIVE_STATUSES = (
    "Открыта",
    "В работе",
    "Заявка собрана",
    "Ожидает запчасть",
    "Требует решения",
    "Запчасть на согласовании",
    "Отправлено в сторонний сервис",
)
CLOSED_STATUS = "Закрыта"


def get_user_okdesk_name(user):
    """Маппинг Django-пользователя на формат имени в Okdesk: «Фамилия Имя»
    (без отчества, как делается при создании заявок). Используется фильтром
    «только мои».

    Возвращает строку или None если фамилия и имя пустые (нечем фильтровать)."""
    if not user or not user.is_authenticated:
        return None
    last_name = (user.last_name or "").strip()
    first_name_raw = (user.first_name or "").strip()
    first_name_only = first_name_raw.split()[0] if first_name_raw else ""
    name = f"{last_name} {first_name_only}".strip()
    return name or None


def _mine_filter(qs, user):
    """Применяет фильтр «только мои» к queryset OkdeskIssue: автор=я ИЛИ создал=я."""
    name = get_user_okdesk_name(user)
    cond = Q(created_by=user)
    if name:
        cond |= Q(author_name=name)
    return qs.filter(cond)


def _apply_search(qs, search):
    """Поиск по теме, компании, серийнику или организации устройства."""
    search = (search or "").strip()
    if not search:
        return qs
    return qs.filter(
        Q(title__icontains=search)
        | Q(company_name__icontains=search)
        | Q(serial_numbers__icontains=search)
        | Q(contract_device__serial_number__icontains=search)
        | Q(contract_device__organization__name__icontains=search)
    )


def _author_q(author, field="author_name"):
    """Строит Q для фильтра по инициатору. `author` — строка или список строк
    (несколько → OR через icontains). Возвращает None, если фильтр пустой."""
    if not author:
        return None
    items = [author] if isinstance(author, str) else list(author)
    items = [(a or "").strip() for a in items if a]
    items = [a for a in items if a]
    if not items:
        return None
    q = Q()
    for a in items:
        q |= Q(**{f"{field}__icontains": a})
    return q


def _apply_author(qs, author):
    """Фильтр по инициатору (автору) заявки. Поддерживает строку или список."""
    q = _author_q(author)
    return qs.filter(q) if q is not None else qs


def get_distinct_authors(search="", limit=50):
    """Уникальные значения author_name из заявок Okdesk — для автодополнения
    в фильтре «Инициатор». При непустом `search` — фильтр icontains."""
    qs = OkdeskIssue.objects.exclude(author_name="").exclude(author_name__isnull=True)
    search = (search or "").strip()
    if search:
        qs = qs.filter(author_name__icontains=search)
    return list(qs.values_list("author_name", flat=True).distinct().order_by("author_name")[:limit])


def _matching_issue_ids(search, author):
    """QuerySet с distinct issue_id, удовлетворяющими фильтрам. Возвращается
    как QS (не list) — Django подставит его в `__in=` как SQL-подзапрос,
    без round-trip и без огромного списка параметров.

    Используется для привязки комментариев к отфильтрованным заявкам:
    комментарии хранятся с issue_id без FK."""
    qs = OkdeskIssue.objects.all()
    qs = _apply_search(qs, search)
    qs = _apply_author(qs, author)
    return qs.values_list("issue_id", flat=True).distinct()


def _parse_date(value):
    """Принимает строку 'YYYY-MM-DD' или date — возвращает date (или today если invalid)."""
    if isinstance(value, date) and not isinstance(value, datetime):
        return value
    if isinstance(value, datetime):
        return value.date()
    if not value:
        return timezone.localdate()
    try:
        return datetime.strptime(str(value), "%Y-%m-%d").date()
    except (ValueError, TypeError):
        return timezone.localdate()


def _parse_optional_date(value):
    """То же, что `_parse_date`, но при пустом/невалидном — None."""
    if not value:
        return None
    if isinstance(value, date) and not isinstance(value, datetime):
        return value
    if isinstance(value, datetime):
        return value.date()
    try:
        return datetime.strptime(str(value), "%Y-%m-%d").date()
    except (ValueError, TypeError):
        return None


def _apply_date_range(qs, date_from, date_to, field):
    """Фильтр по диапазону дат [date_from, date_to] включительно по обе стороны.
    Конец дня для date_to обрабатывается как `< date_to+1`. None — игнор."""
    df = _parse_optional_date(date_from)
    dt = _parse_optional_date(date_to)
    if df:
        start = timezone.make_aware(datetime.combine(df, datetime.min.time()))
        qs = qs.filter(**{f"{field}__gte": start})
    if dt:
        end = timezone.make_aware(datetime.combine(dt, datetime.min.time())) + timedelta(days=1)
        qs = qs.filter(**{f"{field}__lt": end})
    return qs


def get_daily_stats(target_date, user=None, mine=False, search="", author=""):
    """Числовая статистика за день. С `mine=True` — фильтр «только мои заявки»
    (для комментариев — только мои комментарии за день).
    `search` — текст для поиска по теме/компании/серийнику/организации.
    `author` — фильтр по инициатору заявки (icontains)."""
    target_date = _parse_date(target_date)
    day_start = timezone.make_aware(datetime.combine(target_date, datetime.min.time()))
    day_end = day_start + timedelta(days=1)

    issues_qs = OkdeskIssue.objects.all()
    comments_qs = OkdeskComment.objects.filter(created_at__gte=day_start, created_at__lt=day_end)
    if mine and user:
        issues_qs = _mine_filter(issues_qs, user)
        my_name = get_user_okdesk_name(user)
        if my_name:
            comments_qs = comments_qs.filter(author_name=my_name)
        else:
            comments_qs = comments_qs.none()

    issues_qs = _apply_search(issues_qs, search)
    issues_qs = _apply_author(issues_qs, author)

    if search:
        # Комментарии — те, что относятся к заявкам, попавшим в поиск.
        comments_qs = comments_qs.filter(issue_id__in=_matching_issue_ids(search, ""))
    author_q_for_comments = _author_q(author)
    if author_q_for_comments is not None:
        comments_qs = comments_qs.filter(author_q_for_comments)

    created_today = (
        issues_qs.filter(created_at__gte=day_start, created_at__lt=day_end).values("issue_id").distinct().count()
    )
    closed_today = (
        issues_qs.filter(
            status_name=CLOSED_STATUS,
            completed_at__gte=day_start,
            completed_at__lt=day_end,
        )
        .values("issue_id")
        .distinct()
        .count()
    )
    comments_count = comments_qs.count()

    return {
        "date": target_date.isoformat(),
        "created_today": created_today,
        "closed_today": closed_today,
        "comments_count": comments_count,
        "mine": bool(mine),
    }


def get_daily_comments(target_date, page=1, per_page=50, user=None, mine=False, search="", author=""):
    """Постраничный список комментариев за день. С `mine=True` — только мои.
    `search` фильтрует по заявкам (комментарии связаны через issue_id).
    `author` — по автору комментария (icontains)."""
    from django.core.paginator import Paginator

    target_date = _parse_date(target_date)
    day_start = timezone.make_aware(datetime.combine(target_date, datetime.min.time()))
    day_end = day_start + timedelta(days=1)

    qs = OkdeskComment.objects.filter(created_at__gte=day_start, created_at__lt=day_end).order_by("-created_at")

    if mine and user:
        my_name = get_user_okdesk_name(user)
        qs = qs.filter(author_name=my_name) if my_name else qs.none()
    if search:
        qs = qs.filter(issue_id__in=_matching_issue_ids(search, ""))
    author_q = _author_q(author)
    if author_q is not None:
        qs = qs.filter(author_q)
    paginator = Paginator(qs, per_page)
    page_obj = paginator.get_page(page)

    issue_ids = {c.issue_id for c in page_obj.object_list}
    titles_by_id = {}
    if issue_ids:
        for issue_id, title in (
            OkdeskIssue.objects.filter(issue_id__in=issue_ids).values_list("issue_id", "title").distinct()
        ):
            titles_by_id.setdefault(issue_id, title)

    comments_list = [
        {
            "id": c.comment_id,
            "issue_id": c.issue_id,
            "issue_title": titles_by_id.get(c.issue_id, "") or f"Заявка #{c.issue_id}",
            "author": c.author_name,
            "content_preview": (c.content or "").strip().replace("\n", " ")[:200],
            "created_at": c.created_at.isoformat() if c.created_at else None,
            "is_public": c.is_public,
        }
        for c in page_obj.object_list
    ]

    return {
        "date": target_date.isoformat(),
        "page": page_obj.number,
        "per_page": per_page,
        "total": paginator.count,
        "total_pages": paginator.num_pages,
        "comments": comments_list,
    }


def get_active_grouped_by_status(user=None, mine=False, search="", author="", date_from=None, date_to=None):
    """Активные заявки сгруппированы по статусу.

    С `mine=True` — только заявки текущего пользователя (по author_name или created_by).
    `search`/`author` — дополнительные фильтры (см. _apply_search/_apply_author).
    `date_from`/`date_to` — диапазон по дате создания заявки.

    Раньше делалось 1 + N запросов (N=число активных статусов): один для counts,
    потом для каждого статуса .order_by('-created_at')[:30] с select_related
    и dedup в Python. Теперь — 2 запроса: counts + window-функция выбирает
    топ-5 представителей на каждый статус одним SELECT'ом.
    """
    from django.db.models import F, Window
    from django.db.models.functions import RowNumber

    base = OkdeskIssue.objects.filter(status_name__in=ACTIVE_STATUSES)
    if mine and user:
        base = _mine_filter(base, user)
    base = _apply_search(base, search)
    base = _apply_author(base, author)
    base = _apply_date_range(base, date_from, date_to, field="created_at")

    rep_qs = _distinct_issue_qs(base)  # SQL-distinct по issue_id (один представитель)

    counts_rows = rep_qs.values("status_name").annotate(count=Count("id")).order_by("-count")
    counts = [(r["status_name"], r["count"]) for r in counts_rows]

    # Window: row_number partition by status_name. Django 5.1+ умеет фильтровать
    # по аннотации с Window через автоматическое оборачивание в subquery.
    ranked = rep_qs.annotate(
        rn=Window(
            expression=RowNumber(),
            partition_by=[F("status_name")],
            order_by=F("created_at").desc(),
        )
    )
    sample_ids = list(ranked.filter(rn__lte=5).values_list("id", flat=True))

    samples_by_status = {}
    if sample_ids:
        for issue in OkdeskIssue.objects.filter(id__in=sample_ids).select_related(
            "contract_device", "contract_device__organization"
        ):
            samples_by_status.setdefault(issue.status_name, []).append(issue)

    return [
        {
            "status": status,
            "count": count,
            "samples": [
                _serialize_issue(i)
                for i in sorted(samples_by_status.get(status, []), key=lambda x: x.created_at or "", reverse=True)
            ],
        }
        for status, count in counts
    ]


def _distinct_issue_qs(base_qs):
    """Превращает базовый qs OkdeskIssue в queryset distinct-по-issue_id:
    один представитель на issue_id (тот, у кого максимальный id — обычно это
    более свежая запись), плюс select_related для сериализации.

    Раньше дедупликация была в Python через .iterator() + set(), что грузило
    весь queryset в память на каждый запрос. Теперь dedup в SQL — Paginator
    может считать COUNT(*) и применять LIMIT/OFFSET без чтения всего набора.
    """
    rep_ids = base_qs.values("issue_id").annotate(rep=Max("id")).values_list("rep", flat=True)
    return OkdeskIssue.objects.filter(id__in=rep_ids).select_related("contract_device", "contract_device__organization")


def get_issues_by_status(
    status_name,
    page=1,
    per_page=50,
    user=None,
    mine=False,
    search="",
    author="",
    date_from=None,
    date_to=None,
):
    """Список заявок в указанном статусе с пагинацией (distinct по issue_id)."""
    from django.core.paginator import Paginator

    base = OkdeskIssue.objects.filter(status_name=status_name)
    if mine and user:
        base = _mine_filter(base, user)
    base = _apply_search(base, search)
    base = _apply_author(base, author)
    base = _apply_date_range(base, date_from, date_to, field="created_at")

    qs = _distinct_issue_qs(base).order_by("-created_at")
    paginator = Paginator(qs, per_page)
    page_obj = paginator.get_page(page)

    return {
        "page": page_obj.number,
        "total_pages": paginator.num_pages,
        "total": paginator.count,
        "issues": [_serialize_issue(i) for i in page_obj.object_list],
    }


def get_closed_issues(
    page=1,
    per_page=50,
    search="",
    user=None,
    mine=False,
    author="",
    date_from=None,
    date_to=None,
):
    """Закрытые заявки с пагинацией. Поиск — по теме/компании/серийнику/организации.
    Диапазон дат фильтрует по `completed_at` (когда заявка была закрыта)."""
    from django.core.paginator import Paginator

    base = OkdeskIssue.objects.filter(status_name=CLOSED_STATUS)
    base = _apply_search(base, search)
    base = _apply_author(base, author)
    base = _apply_date_range(base, date_from, date_to, field="completed_at")
    if mine and user:
        base = _mine_filter(base, user)

    qs = _distinct_issue_qs(base).order_by("-completed_at", "-created_at")
    paginator = Paginator(qs, per_page)
    page_obj = paginator.get_page(page)

    return {
        "page": page_obj.number,
        "total_pages": paginator.num_pages,
        "total": paginator.count,
        "issues": [_serialize_issue(i) for i in page_obj.object_list],
    }


def get_issue_detail(issue_id):
    """Детальная информация по заявке + все комментарии."""
    rows = (
        OkdeskIssue.objects.filter(issue_id=issue_id)
        .select_related("contract_device", "contract_device__organization")
        .order_by("contract_device_id")
    )
    if not rows.exists():
        return None
    primary = rows.first()
    detail = _serialize_issue(primary)
    detail["devices"] = [
        {
            "contract_device_id": r.contract_device_id,
            "serial_number": r.contract_device.serial_number if r.contract_device else "",
            "organization": (
                r.contract_device.organization.name
                if r.contract_device and r.contract_device.organization
                else r.company_name
            ),
            "address": r.contract_device.address if r.contract_device else "",
            "model": (str(r.contract_device.model) if r.contract_device and r.contract_device.model else ""),
        }
        for r in rows
        if r.contract_device_id is not None
    ]
    detail["serial_numbers"] = primary.serial_numbers
    detail["comments"] = [
        {
            "id": c.comment_id,
            "author": c.author_name,
            "content": c.content,
            "is_public": c.is_public,
            "created_at": c.created_at.isoformat() if c.created_at else None,
        }
        for c in OkdeskComment.objects.filter(issue_id=issue_id).order_by("created_at")
    ]
    return detail


def _serialize_issue(issue):
    return {
        "issue_id": issue.issue_id,
        "title": issue.title,
        "status_name": issue.status_name,
        "priority_name": issue.priority_name,
        "author_name": issue.author_name,
        "assignee_name": issue.assignee_name,
        "company_name": issue.company_name,
        "created_at": issue.created_at.isoformat() if issue.created_at else None,
        "completed_at": issue.completed_at.isoformat() if issue.completed_at else None,
        "deadline_at": issue.deadline_at.isoformat() if issue.deadline_at else None,
        "is_overdue": issue.is_overdue,
        "contract_device_id": issue.contract_device_id,
        "serial_number": (issue.contract_device.serial_number if issue.contract_device else ""),
        "organization": (
            issue.contract_device.organization.name
            if issue.contract_device and issue.contract_device.organization
            else issue.company_name
        ),
    }


# ──────────────────────────────────────────────────────────────────────────────
# Excel-экспорт
# ──────────────────────────────────────────────────────────────────────────────

_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill(start_color="495057", end_color="495057", fill_type="solid")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _style_header_row(ws, row=1):
    for cell in ws[row]:
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN


def _autosize(ws, max_width=60):
    for column_cells in ws.columns:
        col_letter = get_column_letter(column_cells[0].column)
        length = max((len(str(c.value)) for c in column_cells if c.value is not None), default=10)
        ws.column_dimensions[col_letter].width = min(length + 2, max_width)


def _write_issues_sheet(ws, issues_iter, title):
    ws.title = title[:31]
    headers = [
        "ID заявки",
        "Заголовок",
        "Статус",
        "Приоритет",
        "Автор",
        "Исполнитель",
        "Компания",
        "Серийный номер",
        "Создана",
        "Дедлайн",
        "Завершена",
        "Просрочена",
    ]
    ws.append(headers)
    _style_header_row(ws)
    for issue in issues_iter:
        ws.append(
            [
                issue.issue_id,
                issue.title,
                issue.status_name,
                issue.priority_name,
                issue.author_name,
                issue.assignee_name,
                issue.company_name,
                issue.contract_device.serial_number if issue.contract_device else "",
                _fmt_dt(issue.created_at),
                _fmt_dt(issue.deadline_at),
                _fmt_dt(issue.completed_at),
                "Да" if issue.is_overdue else "",
            ]
        )
    _autosize(ws)


def _fmt_dt(dt):
    if not dt:
        return ""
    if hasattr(dt, "astimezone"):
        dt = timezone.localtime(dt) if timezone.is_aware(dt) else dt
    return dt.strftime("%d.%m.%Y %H:%M")


def _distinct_by_issue_id(base_qs):
    """Возвращает queryset с одной строкой на issue_id (SQL-distinct через
    Max(id)) и select_related для записи в Excel. Стримится через iterator()
    в _write_issues_sheet — память не растёт линейно от размера набора."""
    return _distinct_issue_qs(base_qs).order_by("-created_at")


def export_created_excel(target_date):
    target_date = _parse_date(target_date)
    day_start = timezone.make_aware(datetime.combine(target_date, datetime.min.time()))
    day_end = day_start + timedelta(days=1)
    base = OkdeskIssue.objects.filter(created_at__gte=day_start, created_at__lt=day_end)
    wb = Workbook()
    _write_issues_sheet(wb.active, _distinct_by_issue_id(base).iterator(chunk_size=500), f"Создано {target_date}")
    return _wb_bytes(wb), f"okdesk_created_{target_date}.xlsx"


def export_closed_excel(target_date):
    target_date = _parse_date(target_date)
    day_start = timezone.make_aware(datetime.combine(target_date, datetime.min.time()))
    day_end = day_start + timedelta(days=1)
    base = OkdeskIssue.objects.filter(
        status_name=CLOSED_STATUS,
        completed_at__gte=day_start,
        completed_at__lt=day_end,
    )
    wb = Workbook()
    _write_issues_sheet(wb.active, _distinct_by_issue_id(base).iterator(chunk_size=500), f"Закрыто {target_date}")
    return _wb_bytes(wb), f"okdesk_closed_{target_date}.xlsx"


def export_by_status_excel(status_name):
    base = OkdeskIssue.objects.filter(status_name=status_name)
    wb = Workbook()
    safe = status_name.replace("/", "-").replace("\\", "-")
    _write_issues_sheet(wb.active, _distinct_by_issue_id(base).iterator(chunk_size=500), safe)
    return _wb_bytes(wb), f"okdesk_status_{safe}.xlsx"


def export_all_active_excel():
    """Все активные заявки, по листу на статус. Для отчёта подрядчику —
    видно сразу сколько заявок висит и в каком состоянии."""
    wb = Workbook()
    wb.remove(wb.active)
    for status in ACTIVE_STATUSES:
        base = OkdeskIssue.objects.filter(status_name=status)
        qs = _distinct_by_issue_id(base)
        if not qs.exists():
            continue
        ws = wb.create_sheet(status[:31])
        _write_issues_sheet(ws, qs.iterator(chunk_size=500), status[:31])
    if not wb.sheetnames:
        wb.create_sheet("Нет активных заявок")
    today = timezone.localdate().isoformat()
    return _wb_bytes(wb), f"okdesk_active_{today}.xlsx"


def export_active_filtered_excel(user=None, mine=False, search="", author="", date_from=None, date_to=None):
    """Все активные заявки с применением фильтров. Один лист."""
    base = OkdeskIssue.objects.filter(status_name__in=ACTIVE_STATUSES)
    if mine and user:
        base = _mine_filter(base, user)
    base = _apply_search(base, search)
    base = _apply_author(base, author)
    base = _apply_date_range(base, date_from, date_to, field="created_at")

    wb = Workbook()
    _write_issues_sheet(wb.active, _distinct_by_issue_id(base).iterator(chunk_size=500), "Активные (фильтр)")
    today = timezone.localdate().isoformat()
    return _wb_bytes(wb), f"okdesk_active_filtered_{today}.xlsx"


def export_closed_filtered_excel(user=None, mine=False, search="", author="", date_from=None, date_to=None):
    """Закрытые заявки с применением фильтров (поиск/автор/диапазон по completed_at)."""
    base = OkdeskIssue.objects.filter(status_name=CLOSED_STATUS)
    if mine and user:
        base = _mine_filter(base, user)
    base = _apply_search(base, search)
    base = _apply_author(base, author)
    base = _apply_date_range(base, date_from, date_to, field="completed_at")

    wb = Workbook()
    _write_issues_sheet(wb.active, _distinct_by_issue_id(base).iterator(chunk_size=500), "Закрытые (фильтр)")
    today = timezone.localdate().isoformat()
    return _wb_bytes(wb), f"okdesk_closed_filtered_{today}.xlsx"


def _wb_bytes(wb):
    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()
