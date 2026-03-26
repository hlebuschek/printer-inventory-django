"""
Celery задачи для интеграций.
"""

import logging
from uuid import uuid4

from celery import shared_task
from django.conf import settings
from django.contrib.auth import get_user_model

from contracts.models import ContractDevice

from .glpi.monthly_report_export import export_counters_to_glpi
from .glpi.services import check_device_in_glpi, cross_check_with_glpi

logger = logging.getLogger(__name__)
User = get_user_model()


@shared_task(bind=True, queue="high_priority")
def export_monthly_report_to_glpi(self, month=None):
    """
    Выгружает счетчики из monthly_report в GLPI с отслеживанием прогресса.

    Args:
        month: Месяц для выгрузки (ISO format string) или None для последнего закрытого

    Returns:
        dict: Результат выгрузки со статистикой
    """
    from datetime import datetime

    logger.info(f"Starting GLPI export task, month={month}")

    # Конвертируем month из строки в datetime если указан
    month_dt = None
    if month:
        try:
            month_dt = datetime.fromisoformat(month)
        except (ValueError, TypeError) as e:
            logger.error(f"Invalid month format: {month}, error: {e}")
            return {"success": False, "message": f"Неверный формат месяца: {month}"}

    # Callback для обновления прогресса
    def progress_callback(current, total, message):
        """Обновляет состояние задачи с текущим прогрессом"""
        self.update_state(
            state="PROGRESS",
            meta={
                "current": current,
                "total": total,
                "message": message,
                "percent": int((current / total) * 100) if total > 0 else 0,
            },
        )
        logger.debug(f"Progress: {current}/{total} - {message}")

    try:
        # Запускаем выгрузку
        result = export_counters_to_glpi(month=month_dt, progress_callback=progress_callback)

        logger.info(
            f"GLPI export completed: exported={result.get('exported', 0)}, " f"errors={result.get('errors', 0)}"
        )

        return result

    except Exception as exc:
        logger.exception(f"Fatal error in GLPI export task: {exc}")
        return {
            "success": False,
            "message": f"Критическая ошибка: {str(exc)}",
            "total": 0,
            "exported": 0,
            "errors": 0,
            "error_details": [],
        }


@shared_task(bind=True, max_retries=3, queue="high_priority", time_limit=3600)
def check_all_devices_in_glpi(self, update_contract_field=False, skip_check=False):
    """
    Ежедневная задача: проверяет все устройства в GLPI.

    Проходит по всем активным устройствам из ContractDevice,
    проверяет их наличие в GLPI и сохраняет результаты.

    Args:
        update_contract_field: Если True, обновляет поле "Заявлен в договоре" в GLPI
        skip_check: Если True, пропускает проверку наличия и только обновляет договор
                    (использует сохраненные ранее GLPI ID)

    Динамически получает актуальный список устройств при каждом запуске.
    """
    import time

    from integrations.glpi.client import GLPIClient

    start_time = time.time()

    logger.info("=" * 70)
    logger.info("🚀 НАЧАЛО ПРОВЕРКИ УСТРОЙСТВ В GLPI")
    if skip_check and update_contract_field:
        logger.info("   📝 Режим: Только обновление поля договора (без проверки)")
    elif update_contract_field:
        logger.info("   📝 Режим: Проверка + обновление поля договора")
    else:
        logger.info("   📝 Режим: Только проверка наличия")
    logger.info("=" * 70)

    try:
        # Получаем системного пользователя для фоновых задач
        # Или создаем специального пользователя 'glpi_sync'
        try:
            system_user = User.objects.get(username="glpi_sync")
            logger.info("✓ Используется пользователь: glpi_sync")
        except User.DoesNotExist:
            # Используем первого суперпользователя
            system_user = User.objects.filter(is_superuser=True).first()
            if not system_user:
                logger.error("❌ No superuser found for GLPI sync task")
                return {"status": "error", "message": "No user available for sync"}
            logger.info(f"✓ Используется суперпользователь: {system_user.username}")

        # Динамически получаем все устройства с серийными номерами
        devices = (
            ContractDevice.objects.filter(serial_number__isnull=False)
            .exclude(serial_number="")
            .select_related("organization", "model")
        )

        total_devices = devices.count()
        logger.info(f"📊 Найдено устройств для проверки: {total_devices}")
        logger.info("-" * 70)

        # Инициализация GLPI клиента для обновления поля договора
        glpi_client = None
        if update_contract_field:
            try:
                glpi_client = GLPIClient()
                glpi_client.init_session()
                logger.info("✓ GLPI клиент инициализирован для обновления договоров")
            except Exception as e:
                logger.error(f"❌ Ошибка подключения к GLPI: {e}")
                logger.warning("⚠️  Продолжаем без обновления поля договора")
                glpi_client = None

        # Статистика
        stats = {
            "total": total_devices,
            "checked": 0,
            "found_single": 0,
            "found_multiple": 0,
            "not_found": 0,
            "errors": 0,
            "conflicts": [],  # Список ID устройств с конфликтами
            "contract_updated": 0,  # Количество обновленных договоров
            "contract_errors": 0,  # Ошибки при обновлении договоров
            "contract_error_details": [],  # Детали ошибок для вывода
        }

        # Обновляем состояние задачи
        self.update_state(state="PROGRESS", meta={"current": 0, "total": total_devices, "status": "Начало проверки..."})

        # Проверяем каждое устройство
        for idx, device in enumerate(devices, 1):
            try:
                logger.debug(f"Checking device {device.id}: {device.serial_number}")

                if skip_check:
                    # Режим "только обновление" - берем существующую запись из БД
                    from integrations.models import GLPISync

                    try:
                        sync = GLPISync.objects.filter(contract_device=device).latest("checked_at")
                        stats["checked"] += 1
                        logger.debug(f"Используем кэшированную запись от {sync.checked_at}")
                    except GLPISync.DoesNotExist:
                        logger.warning(f"⚠️  Устройство {device.serial_number} не найдено в кэше, пропускаем")
                        stats["errors"] += 1
                        continue
                else:
                    # Обычный режим - проверяем в GLPI
                    sync = check_device_in_glpi(
                        device, user=system_user, force_check=False  # Используем кэш если есть свежие данные
                    )
                    stats["checked"] += 1

                # Обновляем статистику
                if sync.status == "FOUND_SINGLE":
                    stats["found_single"] += 1

                    # Обновляем поле "Заявлен в договоре" если включена опция
                    if glpi_client and sync.glpi_ids:
                        try:
                            glpi_printer_id = sync.glpi_ids[0]
                            success, error = glpi_client.update_contract_field(
                                printer_id=glpi_printer_id, is_in_contract=True
                            )

                            if success:
                                stats["contract_updated"] += 1
                                logger.debug(
                                    f"✓ Договор обновлен для устройства {device.serial_number} (GLPI ID: {glpi_printer_id})"
                                )
                            else:
                                stats["contract_errors"] += 1
                                logger.error(f"❌ Ошибка обновления договора для {device.serial_number}: {error}")
                                # Сохраняем детали первых 5 ошибок для вывода
                                if len(stats["contract_error_details"]) < 5:
                                    stats["contract_error_details"].append(
                                        {"serial": device.serial_number, "glpi_id": glpi_printer_id, "error": error}
                                    )
                        except Exception as e:
                            stats["contract_errors"] += 1
                            logger.error(f"❌ Исключение при обновлении договора для {device.serial_number}: {e}")
                            # Сохраняем детали первых 5 ошибок для вывода
                            if len(stats["contract_error_details"]) < 5:
                                stats["contract_error_details"].append(
                                    {
                                        "serial": device.serial_number,
                                        "glpi_id": glpi_printer_id if "glpi_printer_id" in locals() else "N/A",
                                        "error": str(e),
                                    }
                                )

                elif sync.status == "FOUND_MULTIPLE":
                    stats["found_multiple"] += 1
                    stats["conflicts"].append(
                        {
                            "device_id": device.id,
                            "serial": device.serial_number,
                            "count": sync.glpi_count,
                            "glpi_ids": sync.glpi_ids,
                        }
                    )
                elif sync.status == "NOT_FOUND":
                    stats["not_found"] += 1
                elif sync.status == "ERROR":
                    stats["errors"] += 1

                # Логируем прогресс каждые 10 устройств
                if idx % 10 == 0:
                    progress_percent = int((idx / total_devices) * 100)
                    progress_msg = (
                        f"📈 Прогресс: {idx}/{total_devices} ({progress_percent}%) | "
                        f"Найдено: {stats['found_single']}, Конфликтов: {stats['found_multiple']}, "
                        f"Не найдено: {stats['not_found']}, Ошибок: {stats['errors']}"
                    )
                    if glpi_client:
                        progress_msg += f" | Договоров обновлено: {stats['contract_updated']}"
                    logger.info(progress_msg)

                    # Обновляем состояние задачи
                    self.update_state(
                        state="PROGRESS",
                        meta={
                            "current": idx,
                            "total": total_devices,
                            "percent": progress_percent,
                            "status": f"Проверено {idx} из {total_devices} устройств",
                            "stats": stats,
                        },
                    )

            except Exception as e:
                logger.error(f"❌ Error checking device {device.id}: {e}")
                stats["errors"] += 1

        # Закрываем GLPI сессию
        if glpi_client:
            try:
                glpi_client.kill_session()
                logger.info("✓ GLPI сессия завершена")
            except Exception as e:
                logger.warning(f"⚠️  Ошибка при закрытии GLPI сессии: {e}")

        # Финальный отчет
        elapsed_time = time.time() - start_time
        logger.info("=" * 70)
        logger.info("✅ ПРОВЕРКА ЗАВЕРШЕНА")
        logger.info("=" * 70)
        logger.info(f"⏱️  Время выполнения: {elapsed_time:.1f}с ({elapsed_time / 60:.1f}м)")
        logger.info(f"📊 Проверено устройств: {stats['checked']}/{stats['total']}")
        logger.info(f"✓  Найдено (1 карточка): {stats['found_single']}")
        logger.info(f"⚠️  Конфликты (>1 карточки): {stats['found_multiple']}")
        logger.info(f"❌ Не найдено в GLPI: {stats['not_found']}")
        logger.info(f"❗ Ошибок при проверке: {stats['errors']}")

        # Статистика обновления договоров
        if update_contract_field:
            logger.info("-" * 70)
            logger.info("📝 ОБНОВЛЕНИЕ ПОЛЯ ДОГОВОРА:")
            logger.info(f"✓  Обновлено успешно: {stats['contract_updated']}")
            if stats["contract_errors"] > 0:
                logger.error(f"❌ Ошибок обновления: {stats['contract_errors']}")

                # Выводим детали первых ошибок
                if stats["contract_error_details"]:
                    logger.error("")
                    logger.error(f"Примеры ошибок (первые {len(stats['contract_error_details'])}):")
                    for detail in stats["contract_error_details"]:
                        logger.error(
                            f"  • Serial: {detail['serial']} | GLPI ID: {detail['glpi_id']}\n"
                            f"    Ошибка: {detail['error']}"
                        )

        # Если есть конфликты, логируем их детали
        if stats["conflicts"]:
            logger.warning("-" * 70)
            logger.warning(f"⚠️  ОБНАРУЖЕНО {len(stats['conflicts'])} КОНФЛИКТОВ:")
            for conflict in stats["conflicts"]:
                logger.warning(
                    f"  • Device #{conflict['device_id']} ({conflict['serial']}): "
                    f"{conflict['count']} карточек в GLPI - IDs: {conflict['glpi_ids']}"
                )

        logger.info("=" * 70)

        return stats

    except Exception as exc:
        elapsed_time = time.time() - start_time
        logger.error("=" * 70)
        logger.exception(f"❌ КРИТИЧЕСКАЯ ОШИБКА после {elapsed_time:.1f}с: {exc}")
        logger.error("=" * 70)
        # Retry with exponential backoff: 5min, 15min, 45min
        raise self.retry(exc=exc, countdown=60 * 5 * (2**self.request.retries))


@shared_task
def check_single_device_in_glpi(device_id, user_id=None, force_check=False):
    """
    Проверяет одно устройство в GLPI (асинхронная версия).

    Args:
        device_id: ID устройства
        user_id: ID пользователя, запустившего проверку
        force_check: Принудительная проверка (игнорировать кэш)

    Returns:
        dict: Результат проверки
    """
    try:
        device = ContractDevice.objects.get(id=device_id)

        # Получаем пользователя
        user = None
        if user_id:
            try:
                user = User.objects.get(id=user_id)
            except User.DoesNotExist:
                pass

        # Если пользователь не указан, используем системного
        if not user:
            user = User.objects.filter(is_superuser=True).first()

        sync = check_device_in_glpi(device, user=user, force_check=force_check)

        return {
            "ok": True,
            "device_id": device_id,
            "status": sync.status,
            "glpi_count": sync.glpi_count,
            "glpi_ids": sync.glpi_ids,
        }

    except ContractDevice.DoesNotExist:
        logger.error(f"Device {device_id} not found")
        return {"ok": False, "error": "Device not found"}
    except Exception as e:
        logger.exception(f"Error checking device {device_id}: {e}")
        return {"ok": False, "error": str(e)}


@shared_task(bind=True, max_retries=2, queue="low_priority", time_limit=7200)
def cross_check_glpi_task(self):
    """
    Кросс-проверка: находит офлайн/неопрашиваемые устройства,
    которые при этом активно опрашиваются в GLPI.

    Запускается ежедневно по расписанию или вручную.
    """
    batch_id = str(uuid4())
    freshness_days = getattr(settings, "GLPI_FRESHNESS_DAYS", 7)

    logger.info(f"Запуск кросс-проверки GLPI (batch={batch_id}, freshness={freshness_days} дней)")

    self.update_state(state="PROGRESS", meta={"status": "Начало кросс-проверки...", "batch_id": batch_id})

    try:
        stats = cross_check_with_glpi(batch_id=batch_id, freshness_days=freshness_days)
        logger.info(f"Кросс-проверка GLPI завершена: {stats}")
        return {"ok": True, "batch_id": batch_id, "stats": stats}

    except Exception as exc:
        logger.exception(f"Ошибка кросс-проверки GLPI: {exc}")
        raise self.retry(exc=exc, countdown=60 * 5 * (2**self.request.retries))


# ─── Okdesk ──────────────────────────────────────────────────────────────


@shared_task(bind=True, max_retries=3, queue="low_priority", time_limit=14400)
def sync_okdesk_issues(self, full_sync=False):
    """
    Периодическая синхронизация заявок из Okdesk API с обогащением серийниками.

    По умолчанию (full_sync=False) — быстрая синхронизация:
    пропускает заявки, которые уже закрыты в нашей БД.
    При full_sync=True — обновляет все заявки (на случай переоткрытия).

    Обогащение серийниками (если не найдены в equipment):
    1. Поиск в title по справочнику ContractDevice — без доп. запросов
    2. Если не нашли — запрос описания из API, парсинг HTML-таблицы + поиск по тексту
    3. Если не нашли — поиск в Excel-вложениях заявки
    """
    import io
    import re
    import time

    import requests
    import urllib3
    from django.utils import timezone
    from django.utils.dateparse import parse_datetime
    from lxml import html

    from .models import OkdeskIssue

    urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

    api_token = getattr(settings, "OKDESK_API_TOKEN", None)
    if not api_token:
        logger.warning("OKDESK_API_TOKEN не настроен — синхронизация пропущена")
        return {"ok": False, "error": "OKDESK_API_TOKEN не настроен"}

    api_url = getattr(settings, "OKDESK_API_URL", "https://abikom.okdesk.ru/api/v1")

    sync_mode = "полная" if full_sync else "быстрая (без закрытых)"
    logger.info(f"Начало синхронизации заявок Okdesk ({sync_mode})...")

    # Справочник серийников из ContractDevice для обогащения
    reference_serials = set()
    for sn in ContractDevice.objects.exclude(serial_number="").values_list("serial_number", flat=True):
        if sn:
            reference_serials.add(sn.strip())
    reference_lookup = {re.sub(r"[-_\s]", "", s).upper(): s for s in reference_serials}
    logger.info(f"Справочник серийников: {len(reference_serials)} из ContractDevice")

    # При быстрой синхронизации собираем ID закрытых заявок для пропуска
    closed_issue_ids = set()
    if not full_sync:
        closed_issue_ids = set(OkdeskIssue.objects.filter(status_name="Закрыта").values_list("issue_id", flat=True))
        logger.info(f"Пропускаем {len(closed_issue_ids)} закрытых заявок")

    page = 1
    total_created = 0
    total_updated = 0
    total_fetched = 0
    total_skipped = 0
    enrich_stats = {"from_equipment": 0, "from_title": 0, "from_table": 0, "from_text": 0, "from_excel": 0}

    def _normalize(val):
        return re.sub(r"[-_\s]", "", val).upper()

    def _find_serials_in_text(text):
        """Ищет эталонные серийники в тексте."""
        clean = re.sub(r"<[^>]+>", " ", text)
        clean = re.sub(r"[*_|#\-]", " ", clean).upper()
        return [s for s in reference_serials if s.upper() in clean]

    def _parse_html_table(description):
        """Извлекает серийники из HTML-таблицы в описании."""
        try:
            doc = html.fromstring(description)
        except Exception:
            return []
        serials = []
        for table in doc.xpath("//table"):
            headers = table.xpath(".//thead//th | .//tr[1]//th | .//tr[1]//td")
            serial_col = None
            for i, th in enumerate(headers):
                text = (th.text_content() or "").strip().lower()
                if "серийн" in text:
                    serial_col = i
                    break
            if serial_col is None:
                continue
            rows = table.xpath(".//tbody//tr | .//tr[position()>1]")
            for row in rows:
                cells = row.xpath(".//td")
                if serial_col < len(cells):
                    val = (cells[serial_col].text_content() or "").strip()
                    if val and val not in ("-", "—") and "---" not in val:
                        normalized = _normalize(val)
                        fixed = reference_lookup.get(normalized)
                        serials.append(fixed if fixed else val)
        return serials

    def _search_excel_attachments(issue_id, attachments):
        """Скачивает Excel-вложения и извлекает серийники."""
        import openpyxl

        serials = []
        for att in attachments:
            name = (att.get("attachment_file_name") or "").lower()
            if not name.endswith((".xlsx", ".xls")):
                continue
            try:
                att_resp = requests.get(
                    f"{api_url}/issues/{issue_id}/attachments/{att['id']}",
                    params={"api_token": api_token},
                    verify=False,
                    timeout=15,
                )
                att_resp.raise_for_status()
                url = att_resp.json().get("attachment_url")
                if not url:
                    continue
                file_resp = requests.get(url, verify=False, timeout=30)
                file_resp.raise_for_status()

                wb = openpyxl.load_workbook(io.BytesIO(file_resp.content), read_only=True)
                ws = wb.active
                serial_col = None
                for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
                    for j, cell in enumerate(row):
                        if cell and "серийный номер" in str(cell).lower():
                            serial_col = j
                            break
                if serial_col is not None:
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        val = row[serial_col] if serial_col < len(row) else None
                        if val:
                            val = str(val).strip()
                            if val and val != "-" and val.lower() != "отсутствует":
                                normalized = _normalize(val)
                                fixed = reference_lookup.get(normalized)
                                serials.append(fixed if fixed else val)
                wb.close()
            except Exception as e:
                logger.debug(f"Ошибка Excel-вложения для #{issue_id}: {e}")
        return serials

    def _deduplicate(serials):
        seen = set()
        unique = []
        for s in serials:
            if s not in seen:
                seen.add(s)
                unique.append(s)
        return ", ".join(unique)

    try:
        while True:
            resp = requests.get(
                f"{api_url}/issues/list",
                params={
                    "api_token": api_token,
                    "page[number]": page,
                    "page[size]": 50,
                },
                verify=False,
                timeout=30,
            )
            resp.raise_for_status()
            issues_data = resp.json()

            if not issues_data:
                break

            total_fetched += len(issues_data)

            for item in issues_data:
                issue_id = item.get("id")
                if not issue_id:
                    continue

                # Быстрая синхронизация: пропускаем закрытые заявки
                if not full_sync and issue_id in closed_issue_ids:
                    total_skipped += 1
                    continue

                # Серийники из equipment (есть в list endpoint)
                serial_numbers = ""
                equipments = item.get("equipments") or []
                if equipments:
                    serials = [eq.get("serial_number", "") for eq in equipments if eq.get("serial_number")]
                    serial_numbers = ", ".join(serials)

                # Обогащение: если equipment не дал серийников
                if not serial_numbers and reference_serials:
                    title = item.get("title", "") or ""
                    found = []

                    # Шаг 1: поиск в title (бесплатно, без доп. запросов)
                    found = _find_serials_in_text(title)
                    if found:
                        serial_numbers = _deduplicate(found)
                        enrich_stats["from_title"] += 1

                    # Шаг 2: запрос описания → HTML-таблица + текстовый поиск
                    if not serial_numbers:
                        try:
                            detail_resp = requests.get(
                                f"{api_url}/issues/{issue_id}/",
                                params={"api_token": api_token},
                                verify=False,
                                timeout=15,
                            )
                            detail_resp.raise_for_status()
                            detail = detail_resp.json()
                            description = detail.get("description", "") or ""
                            attachments = detail.get("attachments") or []
                            time.sleep(0.1)

                            # 2a: HTML-таблица
                            if description:
                                table_serials = _parse_html_table(description)
                                if table_serials:
                                    serial_numbers = _deduplicate(table_serials)
                                    enrich_stats["from_table"] += 1

                            # 2b: текстовый поиск по description
                            if not serial_numbers and description:
                                text_serials = _find_serials_in_text(title + " " + description)
                                if text_serials:
                                    serial_numbers = _deduplicate(text_serials)
                                    enrich_stats["from_text"] += 1

                            # Шаг 3: Excel-вложения
                            if not serial_numbers and attachments:
                                excel_serials = _search_excel_attachments(issue_id, attachments)
                                if excel_serials:
                                    serial_numbers = _deduplicate(excel_serials)
                                    enrich_stats["from_excel"] += 1

                        except requests.RequestException as e:
                            logger.debug(f"Ошибка получения описания #{issue_id}: {e}")
                            time.sleep(0.5)
                    else:
                        # serial_numbers уже найден из equipment
                        pass

                if serial_numbers and not found and equipments:
                    enrich_stats["from_equipment"] += 1

                # Имя исполнителя
                assignee = item.get("assignee") or {}
                assignee_name = assignee.get("name", "")

                # Компания
                company = item.get("company") or {}
                company_name = company.get("name", "")

                defaults = {
                    "title": item.get("title", ""),
                    "created_at": parse_datetime(item["created_at"]) if item.get("created_at") else None,
                    "completed_at": parse_datetime(item["completed_at"]) if item.get("completed_at") else None,
                    "status_name": (item.get("status") or {}).get("name", ""),
                    "priority_name": (item.get("priority") or {}).get("name", ""),
                    "assignee_name": assignee_name,
                    "company_name": company_name,
                    "serial_numbers": serial_numbers,
                    "is_overdue": item.get("overdue", False),
                    "synced_at": timezone.now(),
                }

                _, created = OkdeskIssue.objects.update_or_create(
                    issue_id=issue_id,
                    defaults=defaults,
                )

                # Не перезаписываем source у заявок созданных через сайт
                if created:
                    OkdeskIssue.objects.filter(issue_id=issue_id).update(source=OkdeskIssue.SOURCE_SYNC)
                    total_created += 1
                else:
                    total_updated += 1

            # Прогресс каждую страницу
            if page % 10 == 0:
                logger.info(
                    f"Okdesk sync: стр.{page}, получено {total_fetched}, "
                    f"обогащено (title:{enrich_stats['from_title']}, "
                    f"table:{enrich_stats['from_table']}, "
                    f"text:{enrich_stats['from_text']}, "
                    f"excel:{enrich_stats['from_excel']})"
                )

            page += 1
            time.sleep(0.2)  # Rate limiting

        result = {
            "ok": True,
            "full_sync": full_sync,
            "fetched": total_fetched,
            "created": total_created,
            "updated": total_updated,
            "skipped_closed": total_skipped,
            "enrich": enrich_stats,
        }
        logger.info(f"Синхронизация Okdesk завершена: {result}")
        return result

    except requests.RequestException as exc:
        logger.exception(f"Ошибка синхронизации Okdesk (page={page}): {exc}")
        raise self.retry(exc=exc, countdown=60 * 5 * (2**self.request.retries))


@shared_task(bind=True, max_retries=2, queue="low_priority", time_limit=7200)
def enrich_okdesk_serials_task(self):
    """
    Периодическая задача: обогащает заявки Okdesk серийными номерами.
    Берёт заявки с пустым serial_numbers, запрашивает описание из API,
    парсит серийники и сопоставляет с ContractDevice.
    """
    import re
    import time

    import requests
    import urllib3
    from lxml import html

    from contracts.models import ContractDevice

    from .models import OkdeskIssue

    urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

    api_token = getattr(settings, "OKDESK_API_TOKEN", None)
    if not api_token:
        logger.warning("OKDESK_API_TOKEN не настроен — обогащение пропущено")
        return {"ok": False, "error": "OKDESK_API_TOKEN не настроен"}

    api_url = getattr(settings, "OKDESK_API_URL", "https://abikom.okdesk.ru/api/v1")

    # Справочник серийников из ContractDevice
    reference_serials = set()
    for sn in ContractDevice.objects.exclude(serial_number="").values_list("serial_number", flat=True):
        if sn:
            reference_serials.add(sn.strip())
    reference_lookup = {re.sub(r"[-_\s]", "", s).upper(): s for s in reference_serials}

    logger.info(f"Обогащение Okdesk: справочник из {len(reference_serials)} серийников ContractDevice")

    # Заявки без серийника
    issues = OkdeskIssue.objects.filter(serial_numbers="").order_by("issue_id")
    total = issues.count()
    if total == 0:
        logger.info("Обогащение Okdesk: нет заявок без серийников")
        return {"ok": True, "processed": 0, "enriched": 0}

    logger.info(f"Обогащение Okdesk: {total} заявок без серийников")

    stats = {
        "processed": 0,
        "found_in_table": 0,
        "found_in_text": 0,
        "found_in_excel": 0,
        "not_found": 0,
        "errors": 0,
    }

    for issue in issues.iterator():
        stats["processed"] += 1

        try:
            resp = requests.get(
                f"{api_url}/issues/{issue.issue_id}/",
                params={"api_token": api_token},
                verify=False,
                timeout=15,
            )
            resp.raise_for_status()
            data = resp.json()

            description = data.get("description", "") or ""
            title = data.get("title", "") or ""
            attachments = data.get("attachments") or []

            final_serials = []

            # Шаг 1: Парсим HTML-таблицу
            if description:
                try:
                    doc = html.fromstring(description)
                    for table in doc.xpath("//table"):
                        headers = table.xpath(".//thead//th | .//tr[1]//th | .//tr[1]//td")
                        serial_col = None
                        for i, th in enumerate(headers):
                            text = (th.text_content() or "").strip().lower()
                            if "серийн" in text:
                                serial_col = i
                                break
                        if serial_col is None:
                            continue
                        rows = table.xpath(".//tbody//tr | .//tr[position()>1]")
                        for row in rows:
                            cells = row.xpath(".//td")
                            if serial_col < len(cells):
                                val = (cells[serial_col].text_content() or "").strip()
                                if val and val not in ("-", "—") and "---" not in val:
                                    normalized = re.sub(r"[-_\s]", "", val).upper()
                                    fixed = reference_lookup.get(normalized)
                                    final_serials.append(fixed if fixed else val)
                except Exception:
                    pass

            if final_serials:
                stats["found_in_table"] += 1

            # Шаг 2: Поиск по тексту
            if not final_serials:
                clean = re.sub(r"<[^>]+>", " ", title + " " + description)
                clean = re.sub(r"[*_|#\-]", " ", clean).upper()
                for serial in reference_serials:
                    if serial.upper() in clean:
                        final_serials.append(serial)
                if final_serials:
                    stats["found_in_text"] += 1

            # Шаг 3: Поиск в Excel-вложениях
            if not final_serials and attachments:
                try:
                    import io

                    import openpyxl

                    for att in attachments:
                        name = (att.get("attachment_file_name") or "").lower()
                        if not name.endswith((".xlsx", ".xls")):
                            continue
                        att_resp = requests.get(
                            f"{api_url}/issues/{issue.issue_id}/attachments/{att['id']}",
                            params={"api_token": api_token},
                            verify=False,
                            timeout=15,
                        )
                        att_resp.raise_for_status()
                        url = att_resp.json().get("attachment_url")
                        if not url:
                            continue
                        file_resp = requests.get(url, verify=False, timeout=30)
                        file_resp.raise_for_status()

                        wb = openpyxl.load_workbook(io.BytesIO(file_resp.content), read_only=True)
                        ws = wb.active
                        serial_col = None
                        for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
                            for j, cell in enumerate(row):
                                if cell and "серийный номер" in str(cell).lower():
                                    serial_col = j
                                    break
                        if serial_col is not None:
                            for row in ws.iter_rows(min_row=2, values_only=True):
                                val = row[serial_col] if serial_col < len(row) else None
                                if val:
                                    val = str(val).strip()
                                    if val and val != "-" and val.lower() != "отсутствует":
                                        normalized = re.sub(r"[-_\s]", "", val).upper()
                                        fixed = reference_lookup.get(normalized)
                                        final_serials.append(fixed if fixed else val)
                        wb.close()
                    if final_serials:
                        stats["found_in_excel"] += 1
                except Exception as e:
                    logger.debug(f"Ошибка Excel-вложения для #{issue.issue_id}: {e}")

            # Сохраняем
            if final_serials:
                seen = set()
                unique = [s for s in final_serials if s not in seen and not seen.add(s)]
                issue.serial_numbers = ", ".join(unique)
                issue.save(update_fields=["serial_numbers"])
            else:
                stats["not_found"] += 1

            # Прогресс каждые 100
            if stats["processed"] % 100 == 0:
                self.update_state(
                    state="PROGRESS",
                    meta={"current": stats["processed"], "total": total, "stats": stats},
                )
                logger.info(
                    f"Обогащение Okdesk: {stats['processed']}/{total} "
                    f"(таблица: {stats['found_in_table']}, текст: {stats['found_in_text']}, "
                    f"excel: {stats['found_in_excel']})"
                )

            time.sleep(0.15)

        except requests.RequestException as e:
            stats["errors"] += 1
            logger.warning(f"Обогащение Okdesk: ошибка API для #{issue.issue_id}: {e}")
            time.sleep(1)
        except Exception as e:
            stats["errors"] += 1
            logger.error(f"Обогащение Okdesk: ошибка для #{issue.issue_id}: {e}")

    enriched = stats["found_in_table"] + stats["found_in_text"] + stats["found_in_excel"]
    logger.info(
        f"Обогащение Okdesk завершено: обработано {stats['processed']}, "
        f"обогащено {enriched}, не найдено {stats['not_found']}, ошибок {stats['errors']}"
    )
    return {"ok": True, **stats, "enriched": enriched}
