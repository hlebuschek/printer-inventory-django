"""
Общие функции обогащения заявок Okdesk серийными номерами.

Используется в:
- integrations/tasks.py (sync_okdesk_issues)
- integrations/management/commands/enrich_okdesk_serials.py
"""

import io
import logging
import re

import requests
from django.conf import settings
from lxml import html

from contracts.models import ContractDevice

logger = logging.getLogger(__name__)

# Слова-мусор, которые не являются серийниками
JUNK_PATTERNS = re.compile(
    r"^(отсутствует|нет|см\..*|не\s+указан|не\s+определ|н/д|n/a|"
    r"без\s+серийн|серийный\s+номер|serial\s+number|---+|—+)$",
    re.IGNORECASE,
)

# Минимальная длина серийного номера
MIN_SERIAL_LENGTH = 4


def normalize_serial(serial):
    """Убирает лишние символы для нечёткого сравнения."""
    return re.sub(r"[-_\s]", "", serial).upper()


def build_reference_serials():
    """
    Собирает справочник серийников из ContractDevice.

    Returns:
        tuple: (reference_serials: set, reference_lookup: dict)
            reference_serials — множество оригинальных серийников
            reference_lookup — {normalized: original} для нечёткого поиска
    """
    serials = set()
    for sn in ContractDevice.objects.exclude(serial_number="").values_list("serial_number", flat=True):
        if sn:
            serials.add(sn.strip())
    lookup = {normalize_serial(s): s for s in serials}
    return serials, lookup


def _is_valid_serial(val):
    """Проверяет, что значение похоже на серийный номер, а не на мусор."""
    if not val or len(val) < MIN_SERIAL_LENGTH:
        return False
    val = val.strip()
    if JUNK_PATTERNS.match(val):
        return False
    # Слишком много пробелов — скорее всего текст, а не серийник
    if val.count(" ") > 3:
        return False
    # Содержит || — мусор из Okdesk (например "Усть-Илимск||")
    if "||" in val:
        return False
    # Кириллическое слово длиннее 3 символов — скорее всего название города/описание
    if re.search(r"[а-яА-ЯёЁ]{4,}", val):
        return False
    return True


def find_serials_in_text(text, reference_serials):
    """
    Ищет эталонные серийники из ContractDevice в произвольном тексте.
    Возвращает только серийники, которые есть в справочнике.
    """
    if not text or not reference_serials:
        return []
    clean = re.sub(r"<[^>]+>", " ", text)
    clean = re.sub(r"[*_|#\-]", " ", clean).upper()
    return [s for s in reference_serials if s.upper() in clean]


def parse_html_table(description, reference_lookup):
    """
    Извлекает серийники из HTML-таблицы в описании заявки.

    Записывает только значения, которые:
    1. Найдены в reference_lookup (нормализованное совпадение) — записывает эталонный вид
    2. ИЛИ проходят валидацию _is_valid_serial — записывает как есть

    Мусорные значения ("отсутствует", "см. вложение" и т.д.) отбрасываются.
    """
    if not description:
        return []
    try:
        doc = html.fromstring(description)
    except Exception:
        return []

    serials = []
    for table in doc.xpath("//table"):
        headers = table.xpath(".//thead//th | .//tr[1]//th | .//tr[1]//td")
        serial_col = None
        for i, th in enumerate(headers):
            text = (th.text_content() or "").strip().lower()
            if "серийн" in text:
                serial_col = i
                break
        if serial_col is None:
            continue
        rows = table.xpath(".//tbody//tr | .//tr[position()>1]")
        for row in rows:
            cells = row.xpath(".//td")
            if serial_col < len(cells):
                val = (cells[serial_col].text_content() or "").strip()
                if not val or val in ("-", "—") or "---" in val:
                    continue
                # Пробуем найти в справочнике
                normalized = normalize_serial(val)
                fixed = reference_lookup.get(normalized)
                if fixed:
                    serials.append(fixed)
                elif _is_valid_serial(val):
                    # Нет в справочнике, но похоже на серийник
                    serials.append(val)
                # Иначе — мусор, пропускаем
    return serials


def search_excel_attachments(issue_id, attachments, api_token, reference_lookup):
    """
    Скачивает Excel-вложения заявки и извлекает серийные номера.

    Записывает только значения из справочника или прошедшие валидацию.
    """
    import openpyxl

    api_url = getattr(settings, "OKDESK_API_URL", "https://abikom.okdesk.ru/api/v1")
    serials = []

    for att in attachments:
        name = (att.get("attachment_file_name") or "").lower()
        if not name.endswith((".xlsx", ".xls")):
            continue

        try:
            resp = requests.get(
                f"{api_url}/issues/{issue_id}/attachments/{att['id']}",
                params={"api_token": api_token},
                verify=getattr(settings, "OKDESK_VERIFY_SSL", True),
                timeout=15,
            )
            resp.raise_for_status()
            url = resp.json().get("attachment_url")
            if not url:
                continue

            file_resp = requests.get(url, verify=getattr(settings, "OKDESK_VERIFY_SSL", True), timeout=30)
            file_resp.raise_for_status()

            wb = openpyxl.load_workbook(io.BytesIO(file_resp.content), read_only=True)
            ws = wb.active
            serial_col = None
            for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
                for j, cell in enumerate(row):
                    if cell and "серийный номер" in str(cell).lower():
                        serial_col = j
                        break
            if serial_col is not None:
                for row in ws.iter_rows(min_row=2, values_only=True):
                    val = row[serial_col] if serial_col < len(row) else None
                    if val:
                        val = str(val).strip()
                        normalized = normalize_serial(val)
                        fixed = reference_lookup.get(normalized)
                        if fixed:
                            serials.append(fixed)
                        elif _is_valid_serial(val):
                            serials.append(val)
            wb.close()
        except Exception as e:
            logger.debug(f"Ошибка при чтении Excel вложения заявки {issue_id}: {e}")

    return serials


def deduplicate_serials(serials):
    """Дедупликация с сохранением порядка, возвращает строку через запятую."""
    seen = set()
    unique = []
    for s in serials:
        if s not in seen:
            seen.add(s)
            unique.append(s)
    return ", ".join(unique)


def enrich_issue(
    issue_id,
    title,
    equipments,
    reference_serials,
    reference_lookup,
    api_token,
    api_url,
):
    """
    Обогащает одну заявку серийниками. Ленивые доп. запросы.

    Returns:
        tuple: (serial_numbers: str, source: str|None)
            source: "equipment", "title", "table", "text", "excel", None
    """
    # Шаг 0: серийники из equipment (есть в list endpoint)
    if equipments:
        serials = []
        for eq in equipments:
            sn = (eq.get("serial_number") or "").strip()
            if not sn:
                continue
            # Проверяем по справочнику или по валидации
            normalized = normalize_serial(sn)
            fixed = reference_lookup.get(normalized)
            if fixed:
                serials.append(fixed)
            elif _is_valid_serial(sn):
                serials.append(sn)
        if serials:
            return deduplicate_serials(serials), "equipment"

    if not reference_serials:
        return "", None

    # Шаг 1: поиск в title (без доп. запросов)
    found = find_serials_in_text(title or "", reference_serials)
    if found:
        return deduplicate_serials(found), "title"

    # Шаг 2: запрос описания → HTML-таблица + текстовый поиск
    try:
        detail_resp = requests.get(
            f"{api_url}/issues/{issue_id}/",
            params={"api_token": api_token},
            verify=getattr(settings, "OKDESK_VERIFY_SSL", True),
            timeout=15,
        )
        detail_resp.raise_for_status()
        detail = detail_resp.json()
        description = detail.get("description", "") or ""
        attachments = detail.get("attachments") or []

        # 2a: HTML-таблица
        if description:
            table_serials = parse_html_table(description, reference_lookup)
            if table_serials:
                return deduplicate_serials(table_serials), "table"

        # 2b: текстовый поиск по description
        if description:
            text_serials = find_serials_in_text((title or "") + " " + description, reference_serials)
            if text_serials:
                return deduplicate_serials(text_serials), "text"

        # Шаг 3: Excel-вложения
        if attachments:
            excel_serials = search_excel_attachments(issue_id, attachments, api_token, reference_lookup)
            if excel_serials:
                return deduplicate_serials(excel_serials), "excel"

    except requests.RequestException as e:
        logger.debug(f"Ошибка получения описания #{issue_id}: {e}")

    return "", None
