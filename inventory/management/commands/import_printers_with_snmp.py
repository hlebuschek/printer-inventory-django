"""
Умная команда импорта принтеров с автоопросом через SNMP.

Workflow:
1. Читает Excel с IP адресами
2. Для каждого IP запускает SNMP опрос
3. Извлекает модель, производителя, серийник, MAC
4. Fuzzy matching модели в справочнике DeviceModel
5. Проверяет серийник на дубликаты в ContractDevice
6. Обнаруживает прошитые модели (MAC совпадает, серийник другой)
7. Создает принтер с уже проставленной моделью

Формат Excel:
| IP адрес      | MAC адрес (опц) | Организация |
|---------------|-----------------|-------------|
| 192.168.1.100 |                 | Новая орг   |
| 192.168.1.101 | AA:BB:CC:DD:EE:FF | Новая орг   |

Примеры:
# Dry-run (без сохранения)
python manage.py import_printers_with_snmp printers.xlsx --dry-run --show-details

# Реальный импорт
python manage.py import_printers_with_snmp printers.xlsx --auto-create-org

# С низким порогом fuzzy matching (более мягкий)
python manage.py import_printers_with_snmp printers.xlsx --fuzzy-threshold 0.65
"""

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.conf import settings
from inventory.models import Printer, Organization, PollingMethod
from inventory.model_matcher import (
    extract_and_match_model,
    extract_serial_from_xml,
    fuzzy_match_model,
)
from inventory.services import _validate_glpi_installation, _get_glpi_discovery_path, _build_glpi_command, _cleanup_xml, _possible_xml_paths
from inventory.utils import run_glpi_command, extract_mac_address, xml_to_json
from contracts.models import ContractDevice
import openpyxl
import os
import logging

logger = logging.getLogger(__name__)


class Command(BaseCommand):
    help = (
        "Умный импорт принтеров с автоопросом через SNMP и fuzzy matching моделей.\n"
        "Формат: IP адрес | MAC адрес (опц) | Организация"
    )

    def add_arguments(self, parser):
        parser.add_argument(
            "xlsx_file",
            type=str,
            help="Путь к Excel файлу"
        )
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Только показать результаты без сохранения"
        )
        parser.add_argument(
            "--auto-create-org",
            action="store_true",
            help="Автоматически создавать организации"
        )
        parser.add_argument(
            "--show-details",
            action="store_true",
            help="Детальный вывод по каждому принтеру"
        )
        parser.add_argument(
            "--fuzzy-threshold",
            type=float,
            default=0.75,
            help="Порог fuzzy matching (0.0-1.0, по умолчанию 0.75)"
        )
        parser.add_argument(
            "--skip-poll-failures",
            action="store_true",
            help="Пропускать принтеры с ошибками SNMP опроса"
        )

    def handle(self, *args, **opts):
        xlsx_file = opts["xlsx_file"]
        dry_run = opts["dry_run"]
        auto_create_org = opts["auto_create_org"]
        show_details = opts["show_details"]
        fuzzy_threshold = opts["fuzzy_threshold"]
        skip_poll_failures = opts["skip_poll_failures"]

        if not os.path.exists(xlsx_file):
            raise CommandError(f"Файл не найден: {xlsx_file}")

        self.stdout.write(self.style.SUCCESS("=== Умный импорт принтеров с SNMP опросом ==="))
        self.stdout.write(f"Файл: {xlsx_file}")
        self.stdout.write(f"Fuzzy matching threshold: {fuzzy_threshold}")

        # Проверка GLPI
        glpi_ok, glpi_msg = _validate_glpi_installation()
        if not glpi_ok:
            raise CommandError(f"GLPI Agent не доступен: {glpi_msg}")

        # Загрузка Excel
        try:
            wb = openpyxl.load_workbook(xlsx_file, data_only=True)
            ws = wb.active
        except Exception as e:
            raise CommandError(f"Ошибка открытия Excel: {e}")

        self.stdout.write(f"Лист: {ws.title}")

        # Парсинг заголовков
        headers_row = 1
        headers = {}
        for col_idx, cell in enumerate(ws[headers_row], start=1):
            if cell.value:
                header = str(cell.value).strip().lower()
                headers[header] = col_idx

        # Обязательный столбец: IP адрес
        ip_col = None
        for h in ['ip адрес', 'ip_address', 'ip']:
            if h in headers:
                ip_col = headers[h]
                break

        if not ip_col:
            raise CommandError("Не найден столбец 'IP адрес'")

        # Опциональные столбцы
        mac_col = headers.get('mac адрес') or headers.get('mac_address') or headers.get('mac')
        org_col = headers.get('организация') or headers.get('organization') or headers.get('org')
        snmp_col = headers.get('snmp community') or headers.get('snmp')

        self.stdout.write(f"\nНайдены столбцы:")
        self.stdout.write(f"  IP адрес: колонка {ip_col}")
        if mac_col:
            self.stdout.write(f"  MAC адрес: колонка {mac_col}")
        if org_col:
            self.stdout.write(f"  Организация: колонка {org_col}")
        if snmp_col:
            self.stdout.write(f"  SNMP Community: колонка {snmp_col}")

        # Счетчики
        stats = {
            'total': 0,
            'success': 0,
            'poll_failed': 0,
            'model_matched': 0,
            'model_not_matched': 0,
            'serial_collision_contracts': 0,
            'flashed_detected': 0,
            'ip_collision': 0,
            'mac_collision': 0,
            'org_not_found': 0,
            'org_created': 0,
        }

        problems = {
            'poll_failures': [],
            'model_not_matched': [],
            'serial_collisions': [],
            'flashed_models': [],
            'ip_collisions': [],
            'mac_collisions': [],
        }

        # Кэш существующих принтеров
        existing_ips = set(Printer.objects.values_list('ip_address', flat=True))
        existing_macs = {}  # MAC -> Printer
        for p in Printer.objects.exclude(mac_address__isnull=True).exclude(mac_address=''):
            existing_macs[p.mac_address.upper()] = p

        # Кэш серийников в contracts
        contract_serials = set(
            ContractDevice.objects.exclude(serial_number__isnull=True)
            .exclude(serial_number='')
            .values_list('serial_number', flat=True)
        )

        # Кэш организаций
        org_cache = {org.name.strip().lower(): org for org in Organization.objects.all()}

        to_create = []

        # Парсинг строк
        for row_idx in range(headers_row + 1, ws.max_row + 1):
            row = ws[row_idx]

            ip_value = row[ip_col - 1].value
            if not ip_value:
                continue

            ip_address = str(ip_value).strip()
            provided_mac = str(row[mac_col - 1].value).strip().upper() if mac_col and row[mac_col - 1].value else None
            org_name = str(row[org_col - 1].value).strip() if org_col and row[org_col - 1].value else ""
            snmp_community = str(row[snmp_col - 1].value).strip() if snmp_col and row[snmp_col - 1].value else "public"

            stats['total'] += 1

            # Валидация IP
            if not self._validate_ip(ip_address):
                if show_details:
                    self.stdout.write(self.style.WARNING(f"  SKIP (строка {row_idx}): Невалидный IP {ip_address}"))
                continue

            # Проверка коллизий IP
            if ip_address in existing_ips:
                stats['ip_collision'] += 1
                problems['ip_collisions'].append({'row': row_idx, 'ip': ip_address})
                if show_details:
                    self.stdout.write(self.style.ERROR(f"  IP_COLLISION (строка {row_idx}): {ip_address} уже существует"))
                continue

            # ═══════════════════════════════════════════════════════════════
            # SNMP ОПРОС
            # ═══════════════════════════════════════════════════════════════

            self.stdout.write(f"\n📡 Опрос {ip_address} (строка {row_idx})...")

            disc_exe = _get_glpi_discovery_path()
            _cleanup_xml(ip_address)

            cmd = _build_glpi_command(disc_exe, ip_address, snmp_community)

            ok, out = run_glpi_command(cmd)
            if not ok:
                stats['poll_failed'] += 1
                problems['poll_failures'].append({'row': row_idx, 'ip': ip_address, 'error': out})

                if show_details:
                    self.stdout.write(self.style.ERROR(f"  POLL_FAILED: {out}"))

                if not skip_poll_failures:
                    self.stdout.write(self.style.ERROR(
                        f"\nSNMP опрос не удался для {ip_address}. Используйте --skip-poll-failures для пропуска таких принтеров."
                    ))
                    return
                continue

            # Поиск XML файла
            xml_path = None
            for candidate in _possible_xml_paths(ip_address, prefer="disc"):
                if os.path.exists(candidate):
                    xml_path = candidate
                    break

            if not xml_path:
                stats['poll_failed'] += 1
                problems['poll_failures'].append({'row': row_idx, 'ip': ip_address, 'error': 'XML не найден'})

                if show_details:
                    self.stdout.write(self.style.ERROR(f"  POLL_FAILED: XML не найден"))

                if not skip_poll_failures:
                    return
                continue

            # ═══════════════════════════════════════════════════════════════
            # ИЗВЛЕЧЕНИЕ ДАННЫХ ИЗ XML
            # ═══════════════════════════════════════════════════════════════

            data = xml_to_json(xml_path)

            # Извлекаем MAC
            mac_from_snmp = extract_mac_address(data)
            mac_address = (provided_mac or mac_from_snmp or '').upper()

            # Извлекаем серийник
            serial_number = extract_serial_from_xml(xml_path) or ''

            # Извлекаем модель и ищем совпадение
            model_text, manufacturer_text, device_model = extract_and_match_model(xml_path, fuzzy_threshold)

            if show_details:
                self.stdout.write(f"  SNMP данные:")
                self.stdout.write(f"    Модель: {model_text or '—'}")
                self.stdout.write(f"    Производитель: {manufacturer_text or '—'}")
                self.stdout.write(f"    Серийник: {serial_number or '—'}")
                self.stdout.write(f"    MAC: {mac_address or '—'}")

            # ═══════════════════════════════════════════════════════════════
            # FUZZY MATCHING МОДЕЛИ
            # ═══════════════════════════════════════════════════════════════

            if device_model:
                stats['model_matched'] += 1
                if show_details:
                    self.stdout.write(self.style.SUCCESS(f"    ✅ Модель найдена: {device_model}"))
            else:
                stats['model_not_matched'] += 1
                problems['model_not_matched'].append({
                    'row': row_idx,
                    'ip': ip_address,
                    'model': model_text or 'N/A'
                })

                if show_details:
                    self.stdout.write(self.style.WARNING(f"    ⚠️  Модель не найдена в справочнике"))

                    # Показываем top-3 похожих моделей
                    if model_text:
                        matches = fuzzy_match_model(model_text, manufacturer_text, threshold=0.5, top_n=3)
                        if matches:
                            self.stdout.write(f"    Похожие модели:")
                            for dm, score in matches:
                                self.stdout.write(f"      - {dm} (score={score:.2f})")

            # ═══════════════════════════════════════════════════════════════
            # ПРОВЕРКА СЕРИЙНИКА В CONTRACTS
            # ═══════════════════════════════════════════════════════════════

            if serial_number and serial_number in contract_serials:
                stats['serial_collision_contracts'] += 1
                problems['serial_collisions'].append({
                    'row': row_idx,
                    'ip': ip_address,
                    'serial': serial_number
                })

                if show_details:
                    self.stdout.write(self.style.WARNING(
                        f"    ⚠️  Серийник {serial_number} уже есть в ContractDevice"
                    ))

            # ═══════════════════════════════════════════════════════════════
            # ОБНАРУЖЕНИЕ ПРОШИТЫХ МОДЕЛЕЙ
            # ═══════════════════════════════════════════════════════════════

            if mac_address and mac_address in existing_macs:
                existing_printer = existing_macs[mac_address]

                # Если MAC совпадает, но серийник другой → прошитая модель
                if serial_number and existing_printer.serial_number and serial_number != existing_printer.serial_number:
                    stats['flashed_detected'] += 1
                    problems['flashed_models'].append({
                        'row': row_idx,
                        'ip': ip_address,
                        'mac': mac_address,
                        'serial_new': serial_number,
                        'serial_old': existing_printer.serial_number,
                        'existing_ip': existing_printer.ip_address
                    })

                    if show_details:
                        self.stdout.write(self.style.WARNING(
                            f"    🔧 ПРОШИТАЯ МОДЕЛЬ: MAC {mac_address} используется принтером {existing_printer.ip_address} "
                            f"(SN: {existing_printer.serial_number}), но серийник отличается ({serial_number})"
                        ))

                # Если MAC совпадает полностью → коллизия
                else:
                    stats['mac_collision'] += 1
                    problems['mac_collisions'].append({
                        'row': row_idx,
                        'ip': ip_address,
                        'mac': mac_address,
                        'existing_ip': existing_printer.ip_address
                    })

                    if show_details:
                        self.stdout.write(self.style.ERROR(
                            f"    ❌ MAC_COLLISION: {mac_address} уже используется принтером {existing_printer.ip_address}"
                        ))
                    continue

            # ═══════════════════════════════════════════════════════════════
            # ОБРАБОТКА ОРГАНИЗАЦИИ
            # ═══════════════════════════════════════════════════════════════

            organization = None
            if org_name:
                org_key = org_name.lower()
                if org_key in org_cache:
                    organization = org_cache[org_key]
                elif auto_create_org:
                    if not dry_run:
                        organization = Organization.objects.create(name=org_name)
                        org_cache[org_key] = organization
                    stats['org_created'] += 1

                    if show_details:
                        self.stdout.write(self.style.SUCCESS(f"    ✅ Организация создана: {org_name}"))
                else:
                    stats['org_not_found'] += 1
                    if show_details:
                        self.stdout.write(self.style.WARNING(f"    ⚠️  Организация не найдена: {org_name}"))
                    continue

            # ═══════════════════════════════════════════════════════════════
            # СОЗДАНИЕ ПРИНТЕРА
            # ═══════════════════════════════════════════════════════════════

            printer = Printer(
                ip_address=ip_address,
                serial_number=serial_number or '',
                mac_address=mac_address or None,
                model=model_text or '',  # Текстовое поле
                device_model=device_model,  # FK к DeviceModel
                snmp_community=snmp_community,
                organization=organization,
                polling_method=PollingMethod.SNMP
            )

            to_create.append(printer)

            # Добавляем в кэш
            existing_ips.add(ip_address)
            if mac_address:
                existing_macs[mac_address] = printer

            stats['success'] += 1

            if show_details:
                org_str = f"[{organization.name}]" if organization else "[без организации]"
                model_str = str(device_model) if device_model else model_text or "—"
                self.stdout.write(self.style.SUCCESS(
                    f"  ✅ OK: {ip_address} {org_str} | SN:{serial_number or '—'} | Модель:{model_str}"
                ))

        # ═══════════════════════════════════════════════════════════════════
        # СОХРАНЕНИЕ
        # ═══════════════════════════════════════════════════════════════════

        if not dry_run and to_create:
            self.stdout.write(f"\n💾 Сохраняем {len(to_create)} принтеров...")
            try:
                with transaction.atomic():
                    Printer.objects.bulk_create(to_create)
                    self.stdout.write(self.style.SUCCESS("✅ Все принтеры успешно сохранены!"))
            except Exception as e:
                raise CommandError(f"Ошибка при сохранении: {e}")
        elif dry_run:
            self.stdout.write(self.style.WARNING(
                f"\n🔍 DRY-RUN: {len(to_create)} принтеров будет создано при реальном запуске"
            ))

        # ═══════════════════════════════════════════════════════════════════
        # ИТОГОВАЯ СТАТИСТИКА
        # ═══════════════════════════════════════════════════════════════════

        self.stdout.write("\n" + "=" * 70)
        self.stdout.write(self.style.SUCCESS("📊 ИТОГОВАЯ СТАТИСТИКА:"))
        self.stdout.write(f"  Всего строк обработано: {stats['total']}")
        self.stdout.write(self.style.SUCCESS(f"  ✅ Успешно подготовлено: {stats['success']}"))
        self.stdout.write(f"  📡 SNMP опросов выполнено: {stats['total'] - stats['poll_failed']}")

        if stats['model_matched'] > 0:
            self.stdout.write(self.style.SUCCESS(f"  🎯 Моделей найдено в справочнике: {stats['model_matched']}"))
        if stats['model_not_matched'] > 0:
            self.stdout.write(self.style.WARNING(f"  ⚠️  Моделей НЕ найдено: {stats['model_not_matched']}"))
        if stats['poll_failed'] > 0:
            self.stdout.write(self.style.ERROR(f"  ❌ SNMP опросов не удалось: {stats['poll_failed']}"))
        if stats['serial_collision_contracts'] > 0:
            self.stdout.write(self.style.WARNING(f"  ⚠️  Серийников уже в contracts: {stats['serial_collision_contracts']}"))
        if stats['flashed_detected'] > 0:
            self.stdout.write(self.style.WARNING(f"  🔧 Прошитых моделей обнаружено: {stats['flashed_detected']}"))
        if stats['org_created'] > 0:
            self.stdout.write(self.style.SUCCESS(f"  ✅ Организаций создано: {stats['org_created']}"))

        # ═══════════════════════════════════════════════════════════════════
        # ДЕТАЛИ ПРОБЛЕМ
        # ═══════════════════════════════════════════════════════════════════

        if problems['poll_failures']:
            self.stdout.write(f"\n{self.style.ERROR('❌ ОШИБКИ SNMP ОПРОСА:')}")
            for prob in problems['poll_failures'][:10]:
                self.stdout.write(f"  Строка {prob['row']}: {prob['ip']} - {prob['error']}")

        if problems['model_not_matched']:
            self.stdout.write(f"\n{self.style.WARNING('⚠️  МОДЕЛИ НЕ НАЙДЕНЫ В СПРАВОЧНИКЕ:')}")
            for prob in problems['model_not_matched'][:10]:
                self.stdout.write(f"  Строка {prob['row']}: {prob['ip']} - модель '{prob['model']}'")

        if problems['serial_collisions']:
            self.stdout.write(f"\n{self.style.WARNING('⚠️  СЕРИЙНИКИ УЖЕ В CONTRACTS:')}")
            for prob in problems['serial_collisions'][:10]:
                self.stdout.write(f"  Строка {prob['row']}: {prob['ip']} - серийник '{prob['serial']}'")

        if problems['flashed_models']:
            self.stdout.write(f"\n{self.style.WARNING('🔧 ПРОШИТЫЕ МОДЕЛИ ОБНАРУЖЕНЫ:')}")
            for prob in problems['flashed_models'][:10]:
                self.stdout.write(
                    f"  Строка {prob['row']}: {prob['ip']} - MAC {prob['mac']} "
                    f"(существующий принтер: {prob['existing_ip']}, SN:{prob['serial_old']} → {prob['serial_new']})"
                )

        # ═══════════════════════════════════════════════════════════════════
        # РЕКОМЕНДАЦИИ
        # ═══════════════════════════════════════════════════════════════════

        if stats['model_not_matched'] > 0:
            self.stdout.write(f"\n{self.style.WARNING('💡 РЕКОМЕНДАЦИИ:')}")
            self.stdout.write("  Добавьте отсутствующие модели в справочник contracts.DeviceModel")
            self.stdout.write("  Или попробуйте снизить порог fuzzy matching: --fuzzy-threshold 0.65")

        if stats['flashed_detected'] > 0:
            self.stdout.write(f"\n{self.style.WARNING('💡 ПРОШИТЫЕ МОДЕЛИ:')}")
            self.stdout.write("  Обнаружены принтеры с одинаковым MAC, но разными серийниками")
            self.stdout.write("  Это нормально для прошитых моделей - MAC используется как уникальный ID")

        if dry_run:
            self.stdout.write(f"\n{self.style.SUCCESS('✅ Для применения изменений запустите без --dry-run')}")
        else:
            self.stdout.write(f"\n{self.style.SUCCESS('✅ Импорт завершен!')}")

            if stats['success'] > 0:
                self.stdout.write(f"\n{self.style.SUCCESS('📝 СЛЕДУЮЩИЕ ШАГИ:')}")
                self.stdout.write("Принтеры готовы к использованию:")
                self.stdout.write("  - Модели уже проставлены (где найдены)")
                self.stdout.write("  - Серийники и MAC адреса заполнены")
                self.stdout.write("  - Можно сразу запускать повторный опрос для обновления счетчиков")
                self.stdout.write("\nЧерез 3 месяца:")
                self.stdout.write("  python manage.py contracts_import_xlsx договор.xlsx")
                self.stdout.write("  python manage.py link_devices_by_serial")

    def _validate_ip(self, ip):
        """Простая валидация IP адреса"""
        try:
            parts = ip.split('.')
            if len(parts) != 4:
                return False
            for part in parts:
                num = int(part)
                if num < 0 or num > 255:
                    return False
            return True
        except:
            return False
