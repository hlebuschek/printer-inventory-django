# inventory/management/commands/printer_monthly_avg.py
"""
Management command: средняя нагрузка принтеров по месяцам.

Алгоритм: берём первый и последний SUCCESS-снапшот PageCounter,
вычитаем дельты, делим на количество месяцев между ними.

Использование:
    python manage.py printer_monthly_avg
    python manage.py printer_monthly_avg --org-id 3 --format csv
    python manage.py printer_monthly_avg --org-name "Больница" --output /tmp/stats.xlsx
    python manage.py printer_monthly_avg --org-id 3 --include-inactive
"""

import csv
import logging
import os
from datetime import date
from itertools import groupby

import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from django.core.management.base import BaseCommand, CommandError
from django.utils import timezone

from inventory.models import Organization, PageCounter, Printer

logger = logging.getLogger(__name__)

HEADERS = [
    "Организация",
    "IP-адрес",
    "Серийный номер",
    "Модель",
    "Первый опрос",
    "Последний опрос",
    "Месяцев с данными",
    "Среднее ЧБ А4 / мес",
    "Среднее Цвет А4 / мес",
    "Среднее ЧБ А3 / мес",
    "Среднее Цвет А3 / мес",
    "Итого среднее / мес",
]


class Command(BaseCommand):
    help = "Экспорт средней нагрузки принтеров по месяцам в Excel или CSV"

    def add_arguments(self, parser):
        parser.add_argument(
            "--org-id",
            type=int,
            metavar="ID",
            help="Фильтр по ID организации",
        )
        parser.add_argument(
            "--org-name",
            metavar="НАЗВАНИЕ",
            help="Фильтр по имени организации (частичное совпадение)",
        )
        parser.add_argument(
            "--output",
            metavar="ПУТЬ",
            help="Путь к файлу вывода (по умолчанию: ./printer_avg_YYYYMMDD.xlsx или .csv)",
        )
        parser.add_argument(
            "--format",
            choices=["excel", "csv"],
            default="excel",
            dest="fmt",
            help="Формат вывода: excel (по умолчанию) или csv",
        )
        parser.add_argument(
            "--include-inactive",
            action="store_true",
            help="Включить неактивные принтеры (по умолчанию только активные)",
        )

    def handle(self, *args, **options):
        org_id = options["org_id"]
        org_name = options["org_name"]
        output_path = options["output"]
        fmt = options["fmt"]
        include_inactive = options["include_inactive"]

        # ── Выборка принтеров ─────────────────────────────────────────────────
        printers_qs = Printer.objects.select_related(
            "organization", "device_model", "device_model__manufacturer"
        )

        if not include_inactive:
            printers_qs = printers_qs.filter(is_active=True)

        if org_id:
            try:
                org = Organization.objects.get(pk=org_id)
            except Organization.DoesNotExist:
                raise CommandError(f"Организация с ID={org_id} не найдена")
            printers_qs = printers_qs.filter(organization=org)
            self.stdout.write(f"Фильтр: организация «{org.name}»")
        elif org_name:
            printers_qs = printers_qs.filter(organization__name__icontains=org_name)
            self.stdout.write(f"Фильтр: организация содержит «{org_name}»")

        printers_qs = printers_qs.order_by("organization__name", "ip_address")
        printers = list(printers_qs)

        if not printers:
            self.stdout.write(self.style.WARNING("Принтеры не найдены по заданным критериям."))
            return

        self.stdout.write(f"Принтеров найдено: {len(printers)}")

        # ── Загрузка всех PageCounter одним запросом ──────────────────────────
        printer_ids = [p.id for p in printers]
        counters_qs = (
            PageCounter.objects.filter(
                task__printer_id__in=printer_ids,
                task__status="SUCCESS",
            )
            .select_related("task")
            .order_by("task__printer_id", "recorded_at")
            .values("task__printer_id", "recorded_at", "bw_a4", "color_a4", "bw_a3", "color_a3")
        )

        # Группируем по printer_id → {printer_id: [counter_dicts]}
        counters_by_printer: dict[int, list] = {}
        for printer_id, group in groupby(counters_qs, key=lambda c: c["task__printer_id"]):
            counters_by_printer[printer_id] = list(group)

        # ── Расчёт статистики ─────────────────────────────────────────────────
        rows = []
        for printer in printers:
            row = _build_row(printer, counters_by_printer.get(printer.id, []))
            rows.append(row)

        self.stdout.write(f"Строк для экспорта: {len(rows)}")

        # ── Вывод ─────────────────────────────────────────────────────────────
        today_str = date.today().strftime("%Y%m%d")
        ext = "xlsx" if fmt == "excel" else "csv"

        if not output_path:
            output_path = os.path.join(os.getcwd(), f"printer_avg_{today_str}.{ext}")

        if fmt == "excel":
            _write_excel(rows, output_path)
        else:
            _write_csv(rows, output_path)

        self.stdout.write(self.style.SUCCESS(f"Файл сохранён: {output_path}"))


# ── Вспомогательные функции ───────────────────────────────────────────────────

def _build_row(printer: Printer, counters: list) -> dict:
    """Рассчитывает строку статистики для одного принтера."""
    org_name = printer.organization.name if printer.organization_id else "—"

    if not counters:
        return {
            "org": org_name,
            "ip": printer.ip_address,
            "serial": printer.serial_number,
            "model": printer.model_display,
            "first_date": None,
            "last_date": None,
            "num_months": None,
            "avg_bw_a4": None,
            "avg_color_a4": None,
            "avg_bw_a3": None,
            "avg_color_a3": None,
            "avg_total": None,
        }

    first = counters[0]
    last = counters[-1]

    first_dt = first["recorded_at"]
    last_dt = last["recorded_at"]

    # recorded_at может быть datetime или date
    if hasattr(first_dt, "date"):
        first_date = first_dt.date() if hasattr(first_dt, "date") else first_dt
    else:
        first_date = first_dt

    if hasattr(last_dt, "date"):
        last_date = last_dt.date() if hasattr(last_dt, "date") else last_dt
    else:
        last_date = last_dt

    # Количество месяцев: включая оба граничных
    num_months = (last_date.year - first_date.year) * 12 + (last_date.month - first_date.month) + 1

    def safe_delta(field: str) -> float:
        v_first = first.get(field)
        v_last = last.get(field)
        if v_first is None or v_last is None:
            return 0.0
        delta = v_last - v_first
        return max(0.0, delta)  # Отрицательные дельты (сброс) → 0

    delta_bw_a4 = safe_delta("bw_a4")
    delta_color_a4 = safe_delta("color_a4")
    delta_bw_a3 = safe_delta("bw_a3")
    delta_color_a3 = safe_delta("color_a3")

    avg_bw_a4 = round(delta_bw_a4 / num_months, 1)
    avg_color_a4 = round(delta_color_a4 / num_months, 1)
    avg_bw_a3 = round(delta_bw_a3 / num_months, 1)
    avg_color_a3 = round(delta_color_a3 / num_months, 1)
    avg_total = round((delta_bw_a4 + delta_color_a4 + delta_bw_a3 + delta_color_a3) / num_months, 1)

    return {
        "org": org_name,
        "ip": printer.ip_address,
        "serial": printer.serial_number,
        "model": printer.model_display,
        "first_date": first_date,
        "last_date": last_date,
        "num_months": num_months,
        "avg_bw_a4": avg_bw_a4,
        "avg_color_a4": avg_color_a4,
        "avg_bw_a3": avg_bw_a3,
        "avg_color_a3": avg_color_a3,
        "avg_total": avg_total,
    }


def _row_to_list(row: dict):
    """Конвертирует словарь строки в список значений для записи."""
    return [
        row["org"],
        row["ip"],
        row["serial"],
        row["model"],
        row["first_date"],
        row["last_date"],
        row["num_months"],
        row["avg_bw_a4"],
        row["avg_color_a4"],
        row["avg_bw_a3"],
        row["avg_color_a3"],
        row["avg_total"],
    ]


def _write_excel(rows: list, path: str) -> None:
    """Записывает данные в Excel-файл."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Средняя нагрузка"

    bold = Font(bold=True)
    col_widths = [len(h) + 2 for h in HEADERS]

    # Заголовки
    for col_idx, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = bold

    # Форматы
    date_fmt = "dd.mm.yyyy"
    num_fmt = "0.0"
    int_fmt = "0"

    date_cols = {5, 6}   # Первый опрос, Последний опрос
    int_cols = {7}        # Месяцев с данными
    float_cols = {8, 9, 10, 11, 12}  # Средние значения

    for row_idx, row in enumerate(rows, 2):
        values = _row_to_list(row)
        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)

            if col_idx in date_cols and value is not None:
                cell.number_format = date_fmt
            elif col_idx in float_cols and value is not None:
                cell.number_format = num_fmt
            elif col_idx in int_cols and value is not None:
                cell.number_format = int_fmt

            # Обновляем ширину столбца
            cell_len = len(str(value)) if value is not None else 0
            if cell_len > col_widths[col_idx - 1]:
                col_widths[col_idx - 1] = cell_len

    # Применяем ширину столбцов
    for col_idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = min(width + 2, 50)

    wb.save(path)


def _write_csv(rows: list, path: str) -> None:
    """Записывает данные в CSV-файл (разделитель ;, кодировка utf-8-sig)."""
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f, delimiter=";")
        writer.writerow(HEADERS)
        for row in rows:
            values = _row_to_list(row)
            # Даты → строки для CSV
            formatted = []
            for v in values:
                if isinstance(v, date):
                    formatted.append(v.strftime("%d.%m.%Y"))
                elif v is None:
                    formatted.append("")
                else:
                    formatted.append(str(v).replace(".", ",") if isinstance(v, float) else v)
            writer.writerow(formatted)
