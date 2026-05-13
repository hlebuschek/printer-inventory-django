"""
Экспорт принтеров в Excel с дополнительными колонками "В GLPI" и "Подключён к ПК".

Команда идёт в GLPI API по серийному номеру каждого принтера, читает вкладку
"Подключение" (Computer_Item) и собирает список ПК, к которым этот принтер
привязан в GLPI. Итог — xlsx с тем же набором колонок, что у веб-экспорта
inventory.views.export_views.export_excel, плюс:
    - "В GLPI"          — статус поиска (FOUND_SINGLE / FOUND_MULTIPLE / NOT_FOUND / ERROR)
    - "Кол-во ПК"       — сколько компьютеров подключено к принтеру
    - "ПК (имя | IP | serial)" — перечисление через "; "

Примеры запуска:
    python manage.py export_glpi_printer_connections out.xlsx
    python manage.py export_glpi_printer_connections out.xlsx --limit 20
    python manage.py export_glpi_printer_connections out.xlsx --organization "ООО Рога"
    python manage.py export_glpi_printer_connections out.xlsx --only-with-serial
"""

import logging
import time
from pathlib import Path

import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from django.core.management.base import BaseCommand, CommandError
from django.utils import timezone
from django.utils.timezone import localtime

from integrations.glpi.client import GLPIAPIError, GLPIClient
from inventory.models import InventoryTask, Printer
from inventory.services import get_printer_inventory_status

logger = logging.getLogger(__name__)

HEADERS = [
    "Организация",
    "IP-адрес",
    "Серийный №",
    "MAC-адрес",
    "Производитель",
    "Модель",
    "ЧБ A4",
    "Цвет A4",
    "ЧБ A3",
    "Цвет A3",
    "Всего",
    "Тонер K",
    "Тонер C",
    "Тонер M",
    "Тонер Y",
    "Барабан K",
    "Барабан C",
    "Барабан M",
    "Барабан Y",
    "Fuser Kit",
    "Transfer Kit",
    "Waste Toner",
    "Правило",
    "Дата последнего опроса",
    "Статус последнего опроса",
    "Последняя ошибка",
    "В GLPI",
    "Кол-во ПК",
    "ПК (имя | IP | serial)",
]

RULE_LABEL = {
    "SN_MAC": "Серийник+MAC",
    "MAC_ONLY": "Только MAC",
    "SN_ONLY": "Только серийник",
}

GLPI_STATUS_LABEL = {
    "FOUND_SINGLE": "Найден",
    "FOUND_MULTIPLE": "Дубликаты",
    "NOT_FOUND": "Не найден",
    "ERROR": "Ошибка",
    "NO_SERIAL": "Без серийника",
}


def _format_computers(computers):
    """Превращает список dict-ов с ПК в строку 'name | ip | serial; ...'."""
    parts = []
    for c in computers:
        name = c.get("name") or "—"
        ip = c.get("ip") or "—"
        serial = c.get("serial") or "—"
        parts.append(f"{name} | {ip} | {serial}")
    return "; ".join(parts)


class Command(BaseCommand):
    help = "Экспорт принтеров в Excel с колонкой подключений к ПК из GLPI."

    def add_arguments(self, parser):
        parser.add_argument("output", help="Путь к выходному .xlsx файлу")
        parser.add_argument(
            "--limit",
            type=int,
            default=0,
            help="Ограничить число принтеров (0 = без ограничения). Удобно для теста.",
        )
        parser.add_argument(
            "--organization",
            default="",
            help="Фильтр по названию организации (icontains).",
        )
        parser.add_argument(
            "--only-with-serial",
            action="store_true",
            help="Пропустить принтеры без серийного номера (в GLPI их всё равно не найти).",
        )
        parser.add_argument(
            "--rate-limit",
            type=float,
            default=0.1,
            help="Пауза между GLPI-запросами в секундах (по умолчанию 0.1).",
        )

    def handle(self, *args, **opts):
        output_path = Path(opts["output"]).resolve()
        if output_path.suffix.lower() != ".xlsx":
            raise CommandError("Файл должен иметь расширение .xlsx")
        output_path.parent.mkdir(parents=True, exist_ok=True)

        qs = (
            Printer.objects.filter(is_active=True)
            .select_related("organization", "device_model", "device_model__manufacturer")
            .order_by("ip_address")
        )

        if opts["organization"]:
            qs = qs.filter(organization__name__icontains=opts["organization"])
        if opts["only_with_serial"]:
            qs = qs.exclude(serial_number__isnull=True).exclude(serial_number="")
        if opts["limit"] > 0:
            qs = qs[: opts["limit"]]

        total = qs.count()
        if not total:
            self.stdout.write(self.style.WARNING("Принтеров для экспорта не найдено."))
            return

        self.stdout.write(f"Принтеров к обработке: {total}")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Printers + GLPI PC"

        for col_idx, header in enumerate(HEADERS, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)

        col_widths = [len(h) for h in HEADERS]
        date_col_idx = HEADERS.index("Дата последнего опроса") + 1

        stats = {"found_single": 0, "found_multiple": 0, "not_found": 0, "errors": 0, "no_serial": 0, "with_pc": 0}
        rate_limit = opts["rate_limit"]

        try:
            with GLPIClient() as client:
                for idx, printer in enumerate(qs, 1):
                    glpi_status_label, computers = self._query_glpi(client, printer, stats)
                    row_values = self._build_row(printer, glpi_status_label, computers)

                    for col_idx, val in enumerate(row_values, 1):
                        cell = ws.cell(row=idx + 1, column=col_idx, value=val)
                        if col_idx == date_col_idx and val:
                            cell.number_format = "dd.mm.yyyy hh:mm"

                        disp = val if val is not None else ""
                        if hasattr(val, "strftime"):
                            disp = val.strftime("%d.%m.%Y %H:%M")
                        col_widths[col_idx - 1] = max(col_widths[col_idx - 1], len(str(disp)))

                    if computers:
                        stats["with_pc"] += 1

                    if idx % 20 == 0 or idx == total:
                        self.stdout.write(f"  {idx}/{total}")

                    if idx < total and rate_limit > 0:
                        time.sleep(rate_limit)

        except GLPIAPIError as e:
            raise CommandError(f"Ошибка GLPI API: {e}")

        for i, w in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = min(max(w + 2, 10), 60)

        wb.save(output_path)

        self.stdout.write(self.style.SUCCESS(f"Готово: {output_path}"))
        self.stdout.write(
            f"  Найдено в GLPI: {stats['found_single']} (одиночные), " f"{stats['found_multiple']} (дубликаты)"
        )
        self.stdout.write(
            f"  Не найдено: {stats['not_found']}, без серийника: {stats['no_serial']}, " f"ошибок: {stats['errors']}"
        )
        self.stdout.write(f"  Принтеров с подключёнными ПК: {stats['with_pc']}")

    # ──────────────────────────────────────────────────────────────────
    # Внутренние помощники
    # ──────────────────────────────────────────────────────────────────

    def _query_glpi(self, client, printer, stats):
        """
        Ищет принтер в GLPI по серийнику и тянет подключённые компьютеры.

        Returns:
            (label, computers) — human-readable статус и список dict про ПК.
        """
        serial = (printer.serial_number or "").strip()
        if not serial:
            stats["no_serial"] += 1
            return GLPI_STATUS_LABEL["NO_SERIAL"], []

        try:
            status, items, error = client.search_printer_by_serial(serial)
        except Exception as e:
            logger.exception(f"GLPI search error for {serial}: {e}")
            stats["errors"] += 1
            return f"Ошибка: {e}", []

        if status == "ERROR":
            stats["errors"] += 1
            return f"Ошибка: {error or ''}", []
        if status == "NOT_FOUND":
            stats["not_found"] += 1
            return GLPI_STATUS_LABEL["NOT_FOUND"], []

        if status == "FOUND_SINGLE":
            stats["found_single"] += 1
        elif status == "FOUND_MULTIPLE":
            stats["found_multiple"] += 1

        # Собираем ПК со всех найденных GLPI-карточек принтера (на случай дубликатов)
        computers = []
        seen_ids = set()
        for item in items:
            glpi_id = item.get("2") or item.get("id")
            if not glpi_id:
                continue
            try:
                connected = client.get_printer_connected_computers(int(glpi_id))
            except Exception as e:
                logger.exception(f"GLPI get_printer_connected_computers({glpi_id}) error: {e}")
                continue
            for comp in connected:
                cid = comp.get("computer_id")
                if cid in seen_ids:
                    continue
                seen_ids.add(cid)
                computers.append(comp)

        return GLPI_STATUS_LABEL.get(status, status), computers

    def _build_row(self, p, glpi_status_label, computers):
        """Возвращает значения для одной строки xlsx, в порядке HEADERS."""
        inv_status = get_printer_inventory_status(p.id)
        counters = inv_status.get("counters", {})

        dt_value = None
        if inv_status.get("timestamp"):
            try:
                dt = timezone.datetime.fromisoformat(inv_status["timestamp"].replace("Z", "+00:00"))
                dt_value = localtime(dt).replace(tzinfo=None)
            except Exception:
                pass

        last_task = InventoryTask.objects.filter(printer=p).order_by("-task_timestamp").first()
        if last_task:
            try:
                status_map = dict(InventoryTask.STATUS_CHOICES)
                last_status = status_map.get(last_task.status, last_task.status or "—")
            except Exception:
                last_status = last_task.status or "—"
        else:
            last_status = "—"

        last_error = ""
        if last_task and last_task.status in ("FAILED", "VALIDATION_ERROR", "HISTORICAL_INCONSISTENCY"):
            last_error = (last_task.error_message or "")[:100]

        return [
            p.organization.name if p.organization_id else "—",
            p.ip_address,
            p.serial_number,
            p.mac_address or "",
            getattr(p.device_model.manufacturer, "name", "—") if p.device_model_id else "—",
            p.model_display,
            counters.get("bw_a4", ""),
            counters.get("color_a4", ""),
            counters.get("bw_a3", ""),
            counters.get("color_a3", ""),
            counters.get("total_pages", ""),
            counters.get("toner_black", ""),
            counters.get("toner_cyan", ""),
            counters.get("toner_magenta", ""),
            counters.get("toner_yellow", ""),
            counters.get("drum_black", ""),
            counters.get("drum_cyan", ""),
            counters.get("drum_magenta", ""),
            counters.get("drum_yellow", ""),
            counters.get("fuser_kit", ""),
            counters.get("transfer_kit", ""),
            counters.get("waste_toner", ""),
            RULE_LABEL.get(p.last_match_rule, "—"),
            dt_value,
            last_status,
            last_error,
            glpi_status_label,
            len(computers),
            _format_computers(computers),
        ]
