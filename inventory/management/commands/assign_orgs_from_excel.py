# inventory/management/commands/assign_orgs_from_excel.py
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from openpyxl import load_workbook

from inventory.models import Printer, Organization


class Command(BaseCommand):
    help = (
        "Проставить организациям принтеров значения из Excel по точному совпадению серийника.\n"
        "Ожидается колонка 'серийный номер оборудования' и колонка с названием организации."
    )

    def add_arguments(self, parser):
        parser.add_argument("--file", required=True, help="Путь к XLSX файлу.")
        parser.add_argument("--sheet", default=None, help="Имя листа (по умолчанию активный).")
        parser.add_argument(
            "--serial-header",
            default="серийный номер оборудования",
            help="Имя заголовка колонки серийника (по умолчанию указанное).",
        )
        parser.add_argument(
            "--org-header",
            default=None,
            help="Имя заголовка колонки организации. Если не задано — ищем первую, где есть подстрока 'организа'.",
        )
        parser.add_argument(
            "--header-row-limit",
            type=int,
            default=10,
            help="Сколько верхних строк просматривать при поиске строки заголовка.",
        )
        parser.add_argument("--dry-run", action="store_true", help="Только показывать изменения.")
        parser.add_argument(
            "--create-missing-orgs",
            action="store_true",
            help="Создавать отсутствующие организации.",
        )

    def handle(self, *args, **opts):
        xlsx_path = opts["file"]
        sheet_name = opts["sheet"]
        serial_header_name = (opts["serial_header"] or "").strip().lower()
        org_header_name = (opts["org_header"] or "")
        org_header_name_l = org_header_name.strip().lower() if org_header_name else None
        header_row_limit = int(opts["header_row_limit"])
        dry_run = bool(opts["dry_run"])
        create_missing_orgs = bool(opts["create_missing_orgs"])

        try:
            wb = load_workbook(xlsx_path, data_only=True)
        except Exception as e:
            raise CommandError(f"Не удалось открыть файл '{xlsx_path}': {e}")

        try:
            ws = wb[sheet_name] if sheet_name else wb.active
        except KeyError:
            raise CommandError(f"Лист '{sheet_name}' не найден в книге.")

        # 1) Ищем строку заголовка
        header_row = None
        max_row_to_scan = min(header_row_limit, ws.max_row)
        for r in range(1, max_row_to_scan + 1):
            for cell in ws[r]:
                val = str(cell.value).strip().lower() if cell.value is not None else ""
                if val == serial_header_name:
                    header_row = r
                    break
            if header_row:
                break
        if not header_row:
            raise CommandError(
                f"Не нашли заголовок '{serial_header_name}' в первых {header_row_limit} строках."
            )

        # 2) Словарь заголовков
        headers = {}
        for col in range(1, ws.max_column + 1):
            raw = ws.cell(row=header_row, column=col).value
            if raw is None:
                continue
            key = str(raw).strip().lower()
            headers[key] = col

        # 3) Колонка серийника — строго по имени
        serial_col = headers.get(serial_header_name)
        if not serial_col:
            raise CommandError(
                f"Колонка с серийником '{serial_header_name}' не найдена в строке заголовков."
            )

        # 4) Колонка организации
        org_col = None
        if org_header_name_l:
            org_col = headers.get(org_header_name_l)
            if not org_col:
                raise CommandError(
                    f"Колонка организации '{org_header_name}' не найдена в строке заголовков."
                )
        else:
            for key, idx in headers.items():
                if "организа" in key:
                    org_col = idx
                    break
            if not org_col:
                raise CommandError(
                    "Колонка с названием организации не найдена (нет заголовка с подстрокой 'организа')."
                )

        # 5) Собираем mapping serial -> org_name
        mapping = {}
        duplicates = set()
        for r in range(header_row + 1, ws.max_row + 1):
            serial_raw = ws.cell(row=r, column=serial_col).value
            org_raw = ws.cell(row=r, column=org_col).value
            serial = str(serial_raw).strip() if serial_raw is not None else ""
            org_name = str(org_raw).strip() if org_raw is not None else ""
            if not serial or not org_name:
                continue
            if serial in mapping and mapping[serial] != org_name:
                duplicates.add(serial)
            mapping[serial] = org_name  # последняя запись выигрывает

        if not mapping:
            self.stdout.write(self.style.WARNING("Не найдено ни одной пары «серийник–организация»."))
            return

        if duplicates:
            self.stdout.write(
                self.style.WARNING(
                    f"В файле есть серийники с разными организациями (берём последнюю запись): {len(duplicates)} шт."
                )
            )

        # 6) Готовим организации
        org_names = set(mapping.values())
        org_qs = Organization.objects.filter(name__in=org_names)
        org_by_name = {o.name: o for o in org_qs}
        missing_orgs = sorted(org_names - set(org_by_name.keys()))

        if missing_orgs and create_missing_orgs:
            to_create = [Organization(name=name, active=True) for name in missing_orgs]
            Organization.objects.bulk_create(to_create, ignore_conflicts=True)
            org_qs = Organization.objects.filter(name__in=org_names)
            org_by_name = {o.name: o for o in org_qs}
            missing_orgs = sorted(org_names - set(org_by_name.keys()))

        # 7) Подтягиваем принтеры по точному совпадению серийника
        serials = list(mapping.keys())
        printers = list(Printer.objects.filter(serial_number__in=serials))
        p_by_serial = {p.serial_number: p for p in printers}

        to_update = []
        missing_serials = []

        for serial, org_name in mapping.items():
            p = p_by_serial.get(serial)
            if not p:
                missing_serials.append(serial)
                continue
            org = org_by_name.get(org_name)
            if not org:
                # нет такой организации в справочнике — пропускаем
                continue
            if p.organization_id != org.id:
                p.organization_id = org.id
                to_update.append(p)

        # 8) Итого / сохранение
        self.stdout.write(f"Найдено в файле пар: {len(mapping)}")
        self.stdout.write(f"Принтеров, найденных по серийнику: {len(printers)}")
        self.stdout.write(f"К обновлению: {len(to_update)}")
        self.stdout.write(f"Серийники, которых нет в БД: {len(missing_serials)}")
        self.stdout.write(f"Организаций, отсутствующих в справочнике: {len(missing_orgs)}")

        if missing_orgs[:10]:
            self.stdout.write("Примеры отсутствующих организаций: " + ", ".join(missing_orgs[:10]))
        if missing_serials[:10]:
            self.stdout.write("Примеры отсутствующих серийников: " + ", ".join(missing_serials[:10]))

        if dry_run:
            self.stdout.write(self.style.WARNING("DRY-RUN: изменения не сохранены."))
            return

        if to_update:
            with transaction.atomic():
                Printer.objects.bulk_update(to_update, ["organization"])

        self.stdout.write(self.style.SUCCESS(f"Готово. Обновлено принтеров: {len(to_update)}."))
