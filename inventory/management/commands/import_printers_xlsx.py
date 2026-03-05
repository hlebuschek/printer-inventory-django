# inventory/management/commands/import_printers_xlsx.py
"""
Импорт принтеров из Excel в модуль "Опрос устройств" (inventory).

Ожидаемая структура листа (строка заголовков):
№ | Производитель | Модель оборудования | Серийный номер | IP

Примеры запуска:
    python manage.py import_printers_xlsx printers.xlsx
    python manage.py import_printers_xlsx printers.xlsx --sheet "Лист1"
    python manage.py import_printers_xlsx printers.xlsx --dry-run
    python manage.py import_printers_xlsx printers.xlsx --snmp-community public
    python manage.py import_printers_xlsx printers.xlsx --organization "ООО Рога и копыта"
    python manage.py import_printers_xlsx printers.xlsx --create-missing-models

ВАЖНО: Принтеры импортируются ТОЛЬКО в inventory (Опрос устройств).
       Они НЕ добавляются в contracts (Устройства по договору).
       Проверяется уникальность серийного номера в обеих таблицах.

Поиск модели оборудования:
    1. Точное совпадение (без учёта регистра)
    2. Совпадение без пробелов (например "WorkCentre3550" = "WorkCentre 3550")
    3. Название из Excel содержится в названии из БД
    4. Название из БД содержится в названии из Excel
    5. Если модель не найдена и указан --create-missing-models - создаётся новая
"""

import re
from collections import namedtuple
from pathlib import Path

from openpyxl import load_workbook

from django.core.management.base import BaseCommand, CommandError
from django.db import IntegrityError, transaction

from contracts.models import ContractDevice, DeviceModel, Manufacturer
from inventory.models import Organization, Printer

# Заголовки из Excel → внутренние ключи
HEADERS = {
    "№": "row_number",
    "номер": "row_number",
    "n": "row_number",
    "#": "row_number",
    "производитель": "manufacturer",
    "vendor": "manufacturer",
    "бренд": "manufacturer",
    "manufacturer": "manufacturer",
    "модель оборудования": "model",
    "модель": "model",
    "device model": "model",
    "model": "model",
    "серийный номер": "serial_number",
    "серийник": "serial_number",
    "serial number": "serial_number",
    "sn": "serial_number",
    "serial": "serial_number",
    "ip": "ip_address",
    "ip-адрес": "ip_address",
    "ip адрес": "ip_address",
    "ip address": "ip_address",
    "адрес": "ip_address",
}


def norm(val):
    """None -> '', str -> strip + схлопывание пробелов, числа -> str без '.0'."""
    if val is None:
        return ""
    if isinstance(val, (int, float)):
        s = str(int(val)) if float(val).is_integer() else str(val)
    else:
        s = str(val)
    s = s.strip()
    return re.sub(r"\s+", " ", s)


def key_of(header):
    if not header:
        return None
    return HEADERS.get(norm(header).lower())


def ci_get_or_create(model, name_value, name_field="name", **extra_filters):
    """
    Регистронезависимый get_or_create по текстовому полю `name_field`.
    """
    q = {f"{name_field}__iexact": name_value, **extra_filters}
    obj = model.objects.filter(**q).first()
    if obj:
        return obj, False
    try:
        obj = model.objects.create(**{name_field: name_value, **extra_filters})
        return obj, True
    except IntegrityError:
        obj = model.objects.filter(**q).first()
        if obj:
            return obj, False
        raise


def validate_ip(ip_str):
    """Проверка формата IP-адреса."""
    if not ip_str:
        return False
    parts = ip_str.split(".")
    if len(parts) != 4:
        return False
    for part in parts:
        try:
            num = int(part)
            if num < 0 or num > 255:
                return False
        except ValueError:
            return False
    return True


def normalize_model_name(name):
    """Нормализация названия модели для сравнения."""
    if not name:
        return ""
    # Убираем пробелы, дефисы и приводим к нижнему регистру
    return re.sub(r"[\s\-_]+", "", name.lower())


def find_device_model(manufacturer, model_name):
    """
    Поиск модели оборудования с fuzzy matching.

    Стратегии поиска (в порядке приоритета):
    1. Точное совпадение (case-insensitive)
    2. Совпадение без пробелов/дефисов
    3. Название из Excel содержится в названии из БД
    4. Название из БД содержится в названии из Excel

    Returns:
        tuple: (DeviceModel или None, метод_поиска или None)
    """
    # 1. Точное совпадение
    model = DeviceModel.objects.filter(manufacturer=manufacturer, name__iexact=model_name).first()
    if model:
        return model, "exact"

    # 2. Совпадение без пробелов
    normalized_input = normalize_model_name(model_name)
    all_models = DeviceModel.objects.filter(manufacturer=manufacturer)

    for m in all_models:
        if normalize_model_name(m.name) == normalized_input:
            return m, "normalized"

    # 3. Название из Excel содержится в названии из БД
    model = DeviceModel.objects.filter(manufacturer=manufacturer, name__icontains=model_name).first()
    if model:
        return model, "contains_in_db"

    # 4. Название из БД содержится в названии из Excel
    for m in all_models:
        if m.name.lower() in model_name.lower():
            return m, "db_in_input"

    return None, None


BadRow = namedtuple("BadRow", "row reason preview")


class Command(BaseCommand):
    help = "Импорт принтеров из Excel в модуль 'Опрос устройств' (inventory)"

    def add_arguments(self, parser):
        parser.add_argument("xlsx_path", type=str, help="Путь к .xlsx файлу")
        parser.add_argument("--sheet", type=str, default=None, help="Имя листа (по умолчанию первый)")
        parser.add_argument("--dry-run", action="store_true", help="Не сохранять, только показать результат")
        parser.add_argument(
            "--snmp-community",
            type=str,
            default="public",
            help="SNMP community string для создаваемых принтеров (по умолчанию: public)",
        )
        parser.add_argument(
            "--organization",
            type=str,
            default=None,
            help="Название организации для привязки принтеров",
        )
        parser.add_argument(
            "--skip-contract-check",
            action="store_true",
            help="Пропустить проверку дубликатов в ContractDevice (не рекомендуется)",
        )
        parser.add_argument(
            "--create-missing-models",
            action="store_true",
            help="Создавать модели оборудования, если не найдены в справочнике",
        )
        parser.add_argument(
            "--verbose-matching",
            action="store_true",
            help="Показывать подробную информацию о сопоставлении моделей",
        )

    @transaction.atomic
    def handle(self, *args, **opts):
        # --- Открываем Excel ---
        path = Path(opts["xlsx_path"])
        if not path.exists():
            raise CommandError(f"Файл не найден: {path}")

        wb = load_workbook(filename=str(path), read_only=True, data_only=True)
        ws = wb[opts["sheet"]] if opts["sheet"] else wb.worksheets[0]

        # --- Получаем организацию (если указана) ---
        organization = None
        if opts["organization"]:
            organization = Organization.objects.filter(name__iexact=opts["organization"]).first()
            if not organization:
                raise CommandError(f"Организация не найдена: {opts['organization']}")
            self.stdout.write(f"Организация: {organization.name}")

        # --- Заголовки ---
        try:
            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        except StopIteration:
            raise CommandError("Пустой лист: нет строк с данными")

        columns = [key_of(h) for h in header_row]

        # Обязательные колонки
        have = {c for c in columns if c}
        required = {"manufacturer", "model", "serial_number", "ip_address"}
        missing = required - have

        if missing:
            missing_names = ", ".join(sorted(missing))
            raise CommandError(f"В файле отсутствуют обязательные колонки: {missing_names}")

        # --- Счётчики ---
        manufacturers_created = 0
        models_created = 0
        models_found_exact = 0
        models_found_fuzzy = 0
        models_not_found = 0
        printers_created = 0
        printers_skipped_existing = 0
        printers_skipped_contract = 0
        printers_skipped_ip_duplicate = 0
        failed = 0
        blank_rows = 0
        total_rows = 0
        bad_rows = []
        fuzzy_matches = []  # Для отчёта о fuzzy matching

        def row_preview(d):
            return (
                f"{(d.get('manufacturer') or '—')} {(d.get('model') or '—')} | "
                f"SN: {(d.get('serial_number') or '—')} | "
                f"IP: {d.get('ip_address') or '—'}"
            )

        # --- Основной проход по строкам ---
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            total_rows += 1
            data = {}
            for col_key, cell in zip(columns, row):
                if not col_key:
                    continue
                data[col_key] = norm(cell)

            # Полностью пустая строка
            if not any(str(v).strip() if v is not None else "" for v in data.values()):
                blank_rows += 1
                continue

            # Валидация обязательных полей
            missing_fields = []
            field_names = {
                "manufacturer": "Производитель",
                "model": "Модель",
                "serial_number": "Серийный номер",
                "ip_address": "IP-адрес",
            }
            for field in ["manufacturer", "model", "serial_number", "ip_address"]:
                if not data.get(field):
                    missing_fields.append(field_names[field])

            if missing_fields:
                bad_rows.append(
                    BadRow(
                        idx,
                        f"MISSING_VALUES: Отсутствуют обязательные поля: {', '.join(missing_fields)}",
                        f"  Данные из Excel: {row_preview(data)}",
                    )
                )
                failed += 1
                continue

            # Валидация IP
            ip_address = data["ip_address"]
            if not validate_ip(ip_address):
                bad_rows.append(
                    BadRow(
                        idx,
                        f"INVALID_IP: Некорректный формат IP-адреса '{ip_address}'",
                        f"  Данные из Excel: {row_preview(data)}\n"
                        f"  Ожидается формат: XXX.XXX.XXX.XXX (например, 192.168.1.100)",
                    )
                )
                failed += 1
                continue

            serial_number = data["serial_number"]

            try:
                # --- Проверка дубликатов по серийному номеру ---

                # 1. Проверка в Printer (inventory)
                existing_printer = Printer.objects.filter(serial_number__iexact=serial_number).first()
                if existing_printer:
                    org_info = (
                        f" [{existing_printer.organization.name}]"
                        if existing_printer.organization
                        else " [без организации]"
                    )
                    bad_rows.append(
                        BadRow(
                            idx,
                            f"DUPLICATE_IN_INVENTORY: SN '{serial_number}' уже есть в Inventory{org_info}",
                            f"  Пытаемся добавить: {row_preview(data)}\n"
                            f"  Уже есть в БД: {existing_printer.model} | SN: {existing_printer.serial_number} | "
                            f"IP: {existing_printer.ip_address}{org_info}",
                        )
                    )
                    printers_skipped_existing += 1
                    continue

                # 2. Проверка в ContractDevice (contracts) - если не отключено
                if not opts.get("skip_contract_check"):
                    existing_contract = ContractDevice.objects.filter(serial_number__iexact=serial_number).first()
                    if existing_contract:
                        org_info = (
                            f" [{existing_contract.organization.name}]"
                            if existing_contract.organization
                            else " [без организации]"
                        )
                        model_info = (
                            f"{existing_contract.device_model.manufacturer.name} {existing_contract.device_model.name}"
                            if existing_contract.device_model
                            else "Модель не указана"
                        )
                        bad_rows.append(
                            BadRow(
                                idx,
                                f"EXISTS_IN_CONTRACT: SN '{serial_number}' уже есть в 'Устройства по договору'{org_info}",
                                f"  Пытаемся добавить: {row_preview(data)}\n"
                                f"  Уже есть в БД: {model_info} | SN: {existing_contract.serial_number} | "
                                f"Инв.№: {existing_contract.inventory_number or '—'}{org_info}",
                            )
                        )
                        printers_skipped_contract += 1
                        continue

                # 3. Проверка уникальности IP
                existing_ip = Printer.objects.filter(ip_address=ip_address).first()
                if existing_ip:
                    org_info = (
                        f" [{existing_ip.organization.name}]" if existing_ip.organization else " [без организации]"
                    )
                    bad_rows.append(
                        BadRow(
                            idx,
                            f"DUPLICATE_IP: IP '{ip_address}' уже используется",
                            f"  Пытаемся добавить: {row_preview(data)}\n"
                            f"  Конфликтующее устройство: {existing_ip.model} | SN: {existing_ip.serial_number} | "
                            f"IP: {existing_ip.ip_address}{org_info}",
                        )
                    )
                    printers_skipped_ip_duplicate += 1
                    continue

                # --- Создание/получение Manufacturer ---
                mfr_name = data["manufacturer"]
                mfr, mfr_created = ci_get_or_create(Manufacturer, mfr_name)
                if mfr_created:
                    manufacturers_created += 1

                # --- Поиск DeviceModel с fuzzy matching ---
                model_name = data["model"]
                device_model, match_method = find_device_model(mfr, model_name)

                if device_model:
                    if match_method == "exact":
                        models_found_exact += 1
                    else:
                        models_found_fuzzy += 1
                        fuzzy_matches.append(
                            {"row": idx, "input": model_name, "found": device_model.name, "method": match_method}
                        )
                        if opts.get("verbose_matching"):
                            self.stdout.write(
                                f"  [строка {idx}] Fuzzy match: '{model_name}' → '{device_model.name}' ({match_method})"
                            )
                else:
                    models_not_found += 1
                    if opts.get("create_missing_models"):
                        if not opts.get("dry_run"):
                            device_model = DeviceModel.objects.create(
                                manufacturer=mfr,
                                name=model_name,
                                device_type="printer",
                                has_network_port=True,  # Раз есть IP, значит есть сетевой порт
                            )
                        models_created += 1
                    else:
                        bad_rows.append(
                            BadRow(
                                idx,
                                f"MODEL_NOT_FOUND: Модель '{model_name}' не найдена в справочнике",
                                f"  Данные из Excel: {row_preview(data)}\n"
                                f"  Производитель: {mfr_name}\n"
                                f"  💡 Используйте --create-missing-models для автоматического создания модели",
                            )
                        )
                        failed += 1
                        continue

                # --- Создание Printer ---
                if not opts.get("dry_run"):
                    Printer.objects.create(
                        ip_address=ip_address,
                        serial_number=serial_number,
                        model=f"{mfr_name} {model_name}",  # Текстовое поле для совместимости
                        device_model=device_model,
                        snmp_community=opts["snmp_community"],
                        organization=organization,
                    )
                printers_created += 1

            except Exception as e:
                bad_rows.append(
                    BadRow(
                        idx,
                        f"UNEXPECTED: Непредвиденная ошибка - {e.__class__.__name__}",
                        f"  Данные из Excel: {row_preview(data)}\n" f"  Ошибка: {str(e)}",
                    )
                )
                failed += 1
                continue

        # --- Отчёт по fuzzy matching ---
        if fuzzy_matches and not opts.get("verbose_matching"):
            self.stdout.write(self.style.WARNING("\n🔍 FUZZY MATCHING (нечёткое сопоставление моделей):"))
            for fm in fuzzy_matches:
                self.stdout.write(f"  [строка {fm['row']}] '{fm['input']}' → '{fm['found']}' ({fm['method']})")

        # --- Отчёт по проблемным строкам ---
        if bad_rows:
            self.stdout.write(self.style.WARNING("\n❌ Строки, которые НЕ были импортированы:"))
            if organization:
                self.stdout.write(f"   (Импорт в организацию: {organization.name})\n")

            # Группируем ошибки по типам для более понятного вывода
            error_types = {"DUPLICATE": [], "VALIDATION": [], "NOT_FOUND": [], "UNEXPECTED": []}

            for br in bad_rows:
                if any(x in br.reason for x in ["DUPLICATE_IN_INVENTORY", "EXISTS_IN_CONTRACT", "DUPLICATE_IP"]):
                    error_types["DUPLICATE"].append(br)
                elif any(x in br.reason for x in ["MISSING_VALUES", "INVALID_IP"]):
                    error_types["VALIDATION"].append(br)
                elif "MODEL_NOT_FOUND" in br.reason:
                    error_types["NOT_FOUND"].append(br)
                else:
                    error_types["UNEXPECTED"].append(br)

            # Выводим по категориям
            if error_types["DUPLICATE"]:
                self.stdout.write(self.style.WARNING("\n  📌 Дубликаты (уже есть в системе):"))
                for br in error_types["DUPLICATE"]:
                    self.stdout.write(self.style.WARNING(f"\n    [строка {br.row}] {br.reason}"))
                    for line in br.preview.split("\n"):
                        if line.strip():
                            self.stdout.write(f"    {line}")

            if error_types["VALIDATION"]:
                self.stdout.write(self.style.ERROR("\n  ⚠️  Ошибки валидации данных:"))
                for br in error_types["VALIDATION"]:
                    self.stdout.write(self.style.ERROR(f"\n    [строка {br.row}] {br.reason}"))
                    for line in br.preview.split("\n"):
                        if line.strip():
                            self.stdout.write(f"    {line}")

            if error_types["NOT_FOUND"]:
                self.stdout.write(self.style.NOTICE("\n  🔍 Модели не найдены в справочнике:"))
                for br in error_types["NOT_FOUND"]:
                    self.stdout.write(self.style.NOTICE(f"\n    [строка {br.row}] {br.reason}"))
                    for line in br.preview.split("\n"):
                        if line.strip():
                            self.stdout.write(f"    {line}")

            if error_types["UNEXPECTED"]:
                self.stdout.write(self.style.ERROR("\n  💥 Непредвиденные ошибки:"))
                for br in error_types["UNEXPECTED"]:
                    self.stdout.write(self.style.ERROR(f"\n    [строка {br.row}] {br.reason}"))
                    for line in br.preview.split("\n"):
                        if line.strip():
                            self.stdout.write(f"    {line}")

        # --- Итоги ---
        if opts.get("dry_run"):
            transaction.set_rollback(True)
            mode = " (DRY-RUN - изменения не сохранены)"
        else:
            mode = ""

        self.stdout.write("\n" + "=" * 70)
        self.stdout.write(self.style.SUCCESS(f"✅ Импорт завершён{mode}"))
        self.stdout.write("=" * 70)

        self.stdout.write("\n📊 СТАТИСТИКА:")
        self.stdout.write(f"   Обработано строк: {total_rows}")
        self.stdout.write(f"   Пустых строк: {blank_rows}")

        if (
            failed > 0
            or printers_skipped_existing > 0
            or printers_skipped_contract > 0
            or printers_skipped_ip_duplicate > 0
        ):
            self.stdout.write(
                f"\n   ❌ НЕ импортировано: {failed + printers_skipped_existing + printers_skipped_contract + printers_skipped_ip_duplicate}"
            )
            if printers_skipped_existing > 0:
                self.stdout.write(f"      • Дубликаты в Inventory: {printers_skipped_existing}")
            if printers_skipped_contract > 0:
                self.stdout.write(f"      • Есть в Договорах: {printers_skipped_contract}")
            if printers_skipped_ip_duplicate > 0:
                self.stdout.write(f"      • Дубликаты IP: {printers_skipped_ip_duplicate}")
            if failed > 0:
                self.stdout.write(f"      • Прочие ошибки: {failed}")

        self.stdout.write("\n🏭 ПРОИЗВОДИТЕЛИ:")
        self.stdout.write(f"   Создано новых: {manufacturers_created}")

        self.stdout.write("\n📱 МОДЕЛИ ОБОРУДОВАНИЯ:")
        self.stdout.write(f"   Найдено (точное совпадение): {models_found_exact}")
        self.stdout.write(f"   Найдено (нечёткое совпадение): {models_found_fuzzy}")
        self.stdout.write(f"   Не найдено: {models_not_found}")
        self.stdout.write(f"   Создано новых: {models_created}")

        self.stdout.write("\n🖨️ ПРИНТЕРЫ:")
        self.stdout.write(f"   Создано новых: {printers_created}")
        self.stdout.write(f"   Пропущено (уже в Inventory): {printers_skipped_existing}")
        self.stdout.write(f"   Пропущено (есть в Договорах): {printers_skipped_contract}")
        self.stdout.write(f"   Пропущено (дубликат IP): {printers_skipped_ip_duplicate}")

        self.stdout.write("=" * 70)

        if opts.get("dry_run"):
            self.stdout.write(
                self.style.WARNING("\n⚠️  Это был тестовый запуск. Для реального импорта запустите без --dry-run")
            )

        if models_not_found > 0 and not opts.get("create_missing_models"):
            self.stdout.write(
                self.style.WARNING(
                    f"\n💡 Совет: {models_not_found} моделей не найдено. "
                    f"Используйте --create-missing-models для автоматического создания."
                )
            )
