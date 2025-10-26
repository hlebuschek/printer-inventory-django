from django.contrib import messages
from contracts.utils import generate_email_for_device
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.utils.timezone import localtime
from django.contrib.auth.decorators import login_required, permission_required
from django.core.paginator import Paginator
from django.views.decorators.http import require_POST
from django.db.models.functions import TruncDate
from django.utils import timezone
from django.conf import settings

import openpyxl
from openpyxl.utils import get_column_letter

from .models import Printer, InventoryTask, PageCounter, Organization
from contracts.models import DeviceModel
from .forms import PrinterForm
from .services import (
    run_discovery_for_ip,
    extract_serial_from_xml,
    get_printer_inventory_status,
    run_inventory_for_printer,
    inventory_daemon,
    get_glpi_info,
)

# Попытка подтянуть Celery задачи
try:
    from .tasks import run_inventory_task_priority, inventory_daemon_task

    CELERY_AVAILABLE = True
except Exception:
    CELERY_AVAILABLE = False

import json
import logging
from datetime import timedelta

logger = logging.getLogger(__name__)

# ──────────────────────────────────────────────────────────────────────────────
# Fallback пул фоновых задач (если нет Celery)
# ──────────────────────────────────────────────────────────────────────────────
if not CELERY_AVAILABLE:
    from concurrent.futures import ThreadPoolExecutor
    import threading

    EXECUTOR = ThreadPoolExecutor(max_workers=5)
    _RUNNING: set[int] = set()
    _RUNNING_LOCK = threading.Lock()


    def _queue_inventory(pk: int) -> bool:
        """Запускает инвентаризацию в пуле, если ещё не идёт для этого принтера."""
        with _RUNNING_LOCK:
            if pk in _RUNNING:
                return False
            _RUNNING.add(pk)

        def _job():
            try:
                run_inventory_for_printer(pk)
            except Exception as e:
                logger.error(f"Error in inventory job for printer {pk}: {e}")
            finally:
                with _RUNNING_LOCK:
                    _RUNNING.discard(pk)

        EXECUTOR.submit(_job)
        return True
else:
    def _queue_inventory(pk: int) -> bool:
        return True


# ──────────────────────────────────────────────────────────────────────────────
# LIST
# ──────────────────────────────────────────────────────────────────────────────

@login_required
@permission_required("inventory.access_inventory_app", raise_exception=True)
@permission_required("inventory.view_printer", raise_exception=True)
def printer_list(request):
    q_ip = request.GET.get('q_ip', '').strip()
    q_model = request.GET.get('q_model', '').strip()
    q_serial = request.GET.get('q_serial', '').strip()
    q_org = request.GET.get('q_org', '').strip()
    q_rule = request.GET.get('q_rule', '').strip()
    per_page = request.GET.get('per_page', '100').strip()

    try:
        per_page = int(per_page)
        if per_page not in [10, 25, 50, 100, 250, 500, 1000, 2000, 5000]:
            per_page = 100
    except ValueError:
        per_page = 100

    # Базовый запрос с оптимизацией - ДОБАВИЛИ device_model
    qs = Printer.objects.select_related(
        'organization',
        'device_model',  # НОВОЕ
        'device_model__manufacturer'  # НОВОЕ - для избежания N+1
    ).all()

    # Фильтры
    if q_ip:
        qs = qs.filter(ip_address__icontains=q_ip)
    if q_model:
        # ОБНОВЛЕНО: ищем и в старом текстовом поле, и в новом справочнике
        from django.db.models import Q
        qs = qs.filter(
            Q(model__icontains=q_model) |  # старое текстовое поле
            Q(device_model__name__icontains=q_model) |  # название модели
            Q(device_model__manufacturer__name__icontains=q_model)  # производитель
        )
    if q_serial:
        qs = qs.filter(serial_number__icontains=q_serial)

    if q_org == 'none':
        qs = qs.filter(organization__isnull=True)
    elif q_org:
        try:
            qs = qs.filter(organization_id=int(q_org))
        except ValueError:
            qs = qs.filter(organization__name__icontains=q_org)

    if q_rule in ('SN_MAC', 'MAC_ONLY', 'SN_ONLY'):
        qs = qs.filter(last_match_rule=q_rule)
    elif q_rule == 'NONE':
        qs = qs.filter(last_match_rule__isnull=True)

    qs = qs.order_by('ip_address')

    # Пагинация
    paginator = Paginator(qs, per_page)
    page_obj = paginator.get_page(request.GET.get('page'))

    # === ОПТИМИЗАЦИЯ: Убираем N+1 проблему ===
    
    # Получаем ID всех принтеров на текущей странице
    printer_ids = [p.id for p in page_obj]
    
    # Один запрос для получения последних SUCCESS задач для всех принтеров
    # Используем DISTINCT ON (PostgreSQL) для получения только последней задачи
    from django.db.models import OuterRef, Subquery
    
    latest_tasks_subquery = (
        InventoryTask.objects
        .filter(printer=OuterRef('pk'), status='SUCCESS')
        .order_by('-task_timestamp')
        .values('id')[:1]
    )
    
    tasks_dict = {}
    latest_task_ids = (
        InventoryTask.objects
        .filter(id__in=Subquery(latest_tasks_subquery), printer_id__in=printer_ids)
        .select_related('printer')
        .in_bulk()
    )
    
    for task in latest_task_ids.values():
        tasks_dict[task.printer_id] = task
    
    # Один запрос для всех счётчиков
    task_ids = list(latest_task_ids.keys())
    counters = PageCounter.objects.filter(task_id__in=task_ids).select_related('task')
    counters_dict = {c.task_id: c for c in counters}

    # Формируем данные для шаблона
    data = []
    for p in page_obj:
        task = tasks_dict.get(p.id)
        counter = counters_dict.get(task.id) if task else None

        if task:
            last_date_iso = int(task.task_timestamp.timestamp() * 1000)
            last_date = localtime(task.task_timestamp).strftime('%d.%m.%Y %H:%M')
        else:
            last_date_iso = ''
            last_date = '—'

        data.append({
            'printer': p,
            'bw_a4': getattr(counter, 'bw_a4', None),
            'color_a4': getattr(counter, 'color_a4', None),
            'bw_a3': getattr(counter, 'bw_a3', None),
            'color_a3': getattr(counter, 'color_a3', None),
            'total': getattr(counter, 'total_pages', None),
            'drum_black': getattr(counter, 'drum_black', ''),
            'drum_cyan': getattr(counter, 'drum_cyan', ''),
            'drum_magenta': getattr(counter, 'drum_magenta', ''),
            'drum_yellow': getattr(counter, 'drum_yellow', ''),
            'toner_black': getattr(counter, 'toner_black', ''),
            'toner_cyan': getattr(counter, 'toner_cyan', ''),
            'toner_magenta': getattr(counter, 'toner_magenta', ''),
            'toner_yellow': getattr(counter, 'toner_yellow', ''),
            'fuser_kit': getattr(counter, 'fuser_kit', ''),
            'transfer_kit': getattr(counter, 'transfer_kit', ''),
            'waste_toner': getattr(counter, 'waste_toner', ''),
            'last_date': last_date,
            'last_date_iso': last_date_iso,
        })

    per_page_options = [10, 25, 50, 100, 250, 500, 1000, 2000, 5000]

    return render(request, 'inventory/index.html', {
        'data': data,
        'page_obj': page_obj,
        'q_ip': q_ip,
        'q_model': q_model,
        'q_serial': q_serial,
        'q_org': q_org,
        'q_rule': q_rule,
        'per_page': per_page,
        'per_page_options': per_page_options,
        'organizations': Organization.objects.filter(active=True).order_by('name'),
    })

# ──────────────────────────────────────────────────────────────────────────────
# EXPORTS
# ──────────────────────────────────────────────────────────────────────────────

@login_required
@permission_required("inventory.access_inventory_app", raise_exception=True)
@permission_required("inventory.export_printers", raise_exception=True)
def export_excel(request):
    """Экспорт в Excel БЕЗ кэширования."""

    q_ip = request.GET.get("q_ip", "").strip()
    q_model = request.GET.get("q_model", "").strip()
    q_serial = request.GET.get("q_serial", "").strip()
    q_org = request.GET.get("q_org", "").strip()
    q_rule = request.GET.get("q_rule", "").strip()

    # ОБНОВЛЕНО: добавили select_related для device_model
    qs = Printer.objects.select_related(
        "organization",
        "device_model",
        "device_model__manufacturer"
    ).all()

    # Фильтры - ОБНОВЛЕНО
    if q_ip:
        qs = qs.filter(ip_address__icontains=q_ip)
    if q_model:
        from django.db.models import Q
        qs = qs.filter(
            Q(model__icontains=q_model) |
            Q(device_model__name__icontains=q_model) |
            Q(device_model__manufacturer__name__icontains=q_model)
        )
    if q_serial:
        qs = qs.filter(serial_number__icontains=q_serial)

    if q_org == "none":
        qs = qs.filter(organization__isnull=True)
    elif q_org:
        try:
            qs = qs.filter(organization_id=int(q_org))
        except ValueError:
            qs = qs.filter(organization__name__icontains=q_org)

    if q_rule in ("SN_MAC", "MAC_ONLY", "SN_ONLY"):
        qs = qs.filter(last_match_rule=q_rule)
    elif q_rule == "NONE":
        qs = qs.filter(last_match_rule__isnull=True)

    qs = qs.order_by("ip_address")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Printers"

    headers = [
        "Организация", "IP-адрес", "Серийный №", "MAC-адрес", "Модель",
        "ЧБ A4", "Цвет A4", "ЧБ A3", "Цвет A3", "Всего",
        "Тонер K", "Тонер C", "Тонер M", "Тонер Y",
        "Барабан K", "Барабан C", "Барабан M", "Барабан Y",
        "Fuser Kit", "Transfer Kit", "Waste Toner",
        "Правило", "Дата последнего опроса",
        "Статус последнего опроса", "Последняя ошибка",
    ]

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = openpyxl.styles.Font(bold=True)

    col_widths = [len(h) for h in headers]
    date_col_idx = headers.index("Дата последнего опроса") + 1

    rule_label = {
        "SN_MAC": "Серийник+MAC",
        "MAC_ONLY": "Только MAC",
        "SN_ONLY": "Только серийник",
    }

    row_idx = 2
    for p in qs:
        inv_status = get_printer_inventory_status(p.id)
        counters = inv_status.get("counters", {})

        dt_value = None
        if inv_status.get("timestamp"):
            try:
                dt = timezone.datetime.fromisoformat(inv_status["timestamp"].replace("Z", "+00:00"))
                dt_value = localtime(dt).replace(tzinfo=None)
            except Exception:
                pass

        # Последняя задача для статуса/ошибки
        last_task = InventoryTask.objects.filter(printer=p).order_by("-task_timestamp").first()
        if last_task:
            try:
                status_map = dict(InventoryTask.STATUS_CHOICES)
                last_status = status_map.get(last_task.status, last_task.status or "—")
            except Exception:
                last_status = last_task.status or "—"
        else:
            last_status = "—"

        last_error = ""
        if last_task and last_task.status in ["FAILED", "VALIDATION_ERROR", "HISTORICAL_INCONSISTENCY"]:
            last_error = (last_task.error_message or "")[:100]

        values = [
            p.organization.name if p.organization_id else "—",
            p.ip_address,
            p.serial_number,
            p.mac_address or "",
            p.model,
            counters.get("bw_a4", ""),
            counters.get("color_a4", ""),
            counters.get("bw_a3", ""),
            counters.get("color_a3", ""),
            counters.get("total_pages", ""),
            counters.get("toner_black", ""),
            counters.get("toner_cyan", ""),
            counters.get("toner_magenta", ""),
            counters.get("toner_yellow", ""),
            counters.get("drum_black", ""),
            counters.get("drum_cyan", ""),
            counters.get("drum_magenta", ""),
            counters.get("drum_yellow", ""),
            counters.get("fuser_kit", ""),
            counters.get("transfer_kit", ""),
            counters.get("waste_toner", ""),
            rule_label.get(p.last_match_rule, "—"),
            dt_value,
            last_status,
            last_error,
        ]

        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)

            if col_idx == date_col_idx and dt_value:
                cell.number_format = "dd.mm.yyyy hh:mm"

            disp = val if val is not None else ""
            if hasattr(val, "strftime"):
                disp = val.strftime("%d.%m.%Y %H:%M")
            col_widths[col_idx - 1] = max(col_widths[col_idx - 1], len(str(disp)))

        row_idx += 1

    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = min(max(w + 2, 10), 50)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="printers.xlsx"'
    wb.save(response)

    return response


@login_required
@permission_required("inventory.access_inventory_app", raise_exception=True)
@permission_required("inventory.export_amb_report", raise_exception=True)
def export_amb(request):
    """Экспорт AMB отчёта БЕЗ кэширования."""

    if request.method != "POST" or "template" not in request.FILES:
        return render(request, "inventory/export_amb.html")

    wb = openpyxl.load_workbook(request.FILES["template"])
    ws = wb.active

    # Поиск строки заголовков
    header_row = next(
        (r for r in range(1, 11) if any(
            str(c.value).strip().lower() == "серийный номер оборудования" for c in ws[r]
        )),
        None,
    )
    if not header_row:
        return HttpResponse("Строка заголовков не найдена.", status=400)

    headers = {
        str(ws.cell(row=header_row, column=col).value).strip().lower(): col
        for col in range(1, ws.max_column + 1)
        if ws.cell(row=header_row, column=col).value
    }

    lookup = {
        "serial": lambda k: "серийный номер оборудования" in k,
        "a4_bw": lambda k: "ч/б" in k and "конец периода" in k and "а4" in k,
        "a4_color": lambda k: "цветные" in k and "конец периода" in k and "а4" in k,
        "a3_bw": lambda k: "ч/б" in k and "конец периода" in k and "а3" in k,
        "a3_color": lambda k: "цветные" in k and "конец периода" in k and "а3" in k,
    }

    cols = {}
    for internal, cond in lookup.items():
        for key, idx in headers.items():
            if cond(key):
                cols[internal] = idx
                break
        if internal not in cols:
            return HttpResponse(f"Колонка для '{internal}' не найдена.", status=400)

    # Серийники
    serial_cells = []
    for row in range(header_row + 1, ws.max_row + 1):
        raw = ws.cell(row=row, column=cols["serial"]).value
        serial = str(raw).strip() if raw else ""
        if serial:
            serial_cells.append((row, serial))

    serials = {s for _, s in serial_cells}
    if not serials:
        return HttpResponse("В файле нет серийных номеров.", status=400)

    # Данные инвентаризации напрямую из БД
    counters_by_serial = {}
    for serial in serials:
        try:
            printer = Printer.objects.get(serial_number=serial)
            inv_status = get_printer_inventory_status(printer.id)
            counters = inv_status.get("counters", {})
            timestamp = inv_status.get("timestamp")
            if counters and timestamp:
                counters_by_serial[serial] = {"counters": counters, "timestamp": timestamp}
        except Printer.DoesNotExist:
            continue

    # Запись значений
    date_col = ws.max_column + 1
    ws.cell(row=header_row, column=date_col, value="Дата опроса")

    for row_idx, serial in serial_cells:
        data = counters_by_serial.get(serial)
        if not data:
            continue

        counters = data["counters"]
        ws.cell(row=row_idx, column=cols["a4_bw"], value=counters.get("bw_a4") or 0)
        ws.cell(row=row_idx, column=cols["a4_color"], value=counters.get("color_a4") or 0)
        ws.cell(row=row_idx, column=cols["a3_bw"], value=counters.get("bw_a3") or 0)
        ws.cell(row=row_idx, column=cols["a3_color"], value=counters.get("color_a3") or 0)

        try:
            dt = timezone.datetime.fromisoformat(data["timestamp"].replace("Z", "+00:00"))
            dt_local = localtime(dt).replace(tzinfo=None)
            c = ws.cell(row=row_idx, column=date_col, value=dt_local)
            c.number_format = "dd.mm.yyyy hh:mm"
        except Exception:
            pass

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="amb_report.xlsx"'
    wb.save(response)
    wb.close()
    return response


# ──────────────────────────────────────────────────────────────────────────────
# CRUD
# ──────────────────────────────────────────────────────────────────────────────

@login_required
@permission_required("inventory.access_inventory_app", raise_exception=True)
@permission_required("inventory.add_printer", raise_exception=True)
def add_printer(request):
    """Добавление принтера БЕЗ кэширования."""
    form = PrinterForm(request.POST or None)
    if form.is_valid():
        printer = form.save()
        messages.success(request, "Принтер добавлен")
        return redirect("inventory:printer_list")
    return render(request, "inventory/add_printer.html", {"form": form})


@login_required
@permission_required("inventory.access_inventory_app", raise_exception=True)
@permission_required("inventory.change_printer", raise_exception=True)
def edit_printer(request, pk):
    """Редактирование принтера БЕЗ кэширования."""
    printer = get_object_or_404(Printer, pk=pk)
    form = PrinterForm(request.POST or None, instance=printer)

    if request.method == "POST" and form.is_valid():
        printer = form.save()

        if request.headers.get("X-Requested-With") == "XMLHttpRequest":
            return JsonResponse({
                "success": True,
                "printer": {
                    "id": printer.id,
                    "ip_address": printer.ip_address,
                    "serial_number": printer.serial_number,
                    "mac_address": printer.mac_address,
                    "model": printer.model,
                    "snmp_community": printer.snmp_community,
                    "organization": printer.organization.name if printer.organization_id else None,
                    "organization_id": printer.organization_id,
                },
            })

        messages.success(request, "Принтер обновлён")
        return redirect("inventory:printer_list")

    if request.headers.get("X-Requested-With") == "XMLHttpRequest":
        return JsonResponse({"success": False, "error": form.errors.as_json()}, status=400)

    return render(request, "inventory/edit_printer.html", {"form": form})


@login_required
@permission_required("inventory.access_inventory_app", raise_exception=True)
@permission_required("inventory.delete_printer", raise_exception=True)
def delete_printer(request, pk):
    """Удаление принтера."""
    printer = get_object_or_404(Printer, pk=pk)

    if request.method == "POST":
        printer.delete()
        messages.success(request, "Принтер удалён")
        return JsonResponse({"success": True})

    # GET - показываем список с модалкой
    return printer_list(request)


# ──────────────────────────────────────────────────────────────────────────────
# HISTORY
# ──────────────────────────────────────────────────────────────────────────────

@login_required
@permission_required("inventory.access_inventory_app", raise_exception=True)
@permission_required("inventory.view_printer", raise_exception=True)
def history_view(request, pk):
    """История принтера БЕЗ кэширования."""
    printer = get_object_or_404(Printer, pk=pk)

    if request.headers.get("X-Requested-With") == "XMLHttpRequest":
        # AJAX запрос - возвращаем JSON
        daily_tasks = (
            InventoryTask.objects.filter(
                printer=printer,
                status__in=['SUCCESS', 'HISTORICAL_INCONSISTENCY']
            )
            .annotate(day=TruncDate("task_timestamp"))
            .order_by("-day", "-task_timestamp", "-pk")
            .distinct("day")
        )

        daily_list = list(daily_tasks)
        counters = PageCounter.objects.filter(task__in=daily_list)
        counter_by_task_id = {c.task_id: c for c in counters}

        data = []
        for t in daily_list:
            c = counter_by_task_id.get(t.id)

            # Для исторических несоответствий показываем предыдущие данные
            if t.status == 'HISTORICAL_INCONSISTENCY':
                last_valid_task = InventoryTask.objects.filter(
                    printer=printer,
                    status='SUCCESS',
                    task_timestamp__lt=t.task_timestamp
                ).order_by('-task_timestamp').first()

                if last_valid_task:
                    c = PageCounter.objects.filter(task=last_valid_task).first()

            data.append({
                "task_timestamp": localtime(t.task_timestamp).strftime("%Y-%m-%dT%H:%M:%S"),
                "match_rule": t.match_rule,
                "status": t.status,
                "status_display": t.get_status_display(),
                "bw_a4": c.bw_a4 if c else None,
                "color_a4": c.color_a4 if c else None,
                "bw_a3": c.bw_a3 if c else None,
                "color_a3": c.color_a3 if c else None,
                "total_pages": c.total_pages if c else None,
                "drum_black": c.drum_black if c else None,
                "drum_cyan": c.drum_cyan if c else None,
                "drum_magenta": c.drum_magenta if c else None,
                "drum_yellow": c.drum_yellow if c else None,
                "toner_black": c.toner_black if c else None,
                "toner_cyan": c.toner_cyan if c else None,
                "toner_magenta": c.toner_magenta if c else None,
                "toner_yellow": c.toner_yellow if c else None,
                "fuser_kit": c.fuser_kit if c else None,
                "transfer_kit": c.transfer_kit if c else None,
                "waste_toner": c.waste_toner if c else None,
                "is_historical_issue": t.status == 'HISTORICAL_INCONSISTENCY',
                "error_message": t.error_message if t.status == 'HISTORICAL_INCONSISTENCY' else None,
            })

        return JsonResponse(data, safe=False)

    # HTML версия - постранично
    tasks = InventoryTask.objects.filter(
        printer=printer,
        status="SUCCESS"
    ).order_by("-task_timestamp")

    paginator = Paginator(tasks, 50)
    page_obj = paginator.get_page(request.GET.get("page"))
    rows = [(t, PageCounter.objects.filter(task=t).first()) for t in page_obj]

    return render(
        request,
        "inventory/history.html",
        {"printer": printer, "rows": rows, "page_obj": page_obj}
    )


# ──────────────────────────────────────────────────────────────────────────────
# RUNNERS
# ──────────────────────────────────────────────────────────────────────────────

@login_required
@permission_required("inventory.access_inventory_app", raise_exception=True)
@permission_required("inventory.run_inventory", raise_exception=True)
@require_POST
def run_inventory(request, pk):
    """Запуск инвентаризации одного принтера."""
    pk = int(pk)

    if CELERY_AVAILABLE:
        try:
            from inventory.tasks import run_inventory_task_priority

            task = run_inventory_task_priority.apply_async(
                args=[pk, request.user.id],
                priority=9,
            )
            logger.info(f"Queued PRIORITY Celery task {task.id} for printer {pk}")
            return JsonResponse({
                "success": True,
                "celery": True,
                "task_id": task.id,
                "queued": True
            })
        except Exception as e:
            logger.error(f"Error queuing Celery task for printer {pk}: {e}")
            return JsonResponse({
                "success": False,
                "celery": True,
                "error": str(e)
            }, status=500)
    else:
        queued = _queue_inventory(pk)
        return JsonResponse({
            "success": True,
            "celery": False,
            "queued": queued
        })


@login_required
@permission_required("inventory.access_inventory_app", raise_exception=True)
@permission_required("inventory.run_inventory", raise_exception=True)
@require_POST
def run_inventory_all(request):
    """Запуск инвентаризации всех принтеров."""

    if CELERY_AVAILABLE:
        try:
            task = inventory_daemon_task.apply_async(priority=5)
            logger.info(f"User {request.user.id} queued daemon task {task.id}")
            return JsonResponse({
                "success": True,
                "celery": True,
                "task_id": task.id,
                "queued": True
            })
        except Exception as e:
            logger.error(f"Error queuing daemon task: {e}")
            return JsonResponse({
                "success": False,
                "celery": True,
                "error": str(e)
            }, status=500)
    else:
        from concurrent.futures import ThreadPoolExecutor
        executor = ThreadPoolExecutor(max_workers=1)
        executor.submit(inventory_daemon)
        return JsonResponse({
            "success": True,
            "celery": False,
            "queued": True
        })


# ──────────────────────────────────────────────────────────────────────────────
# APIs
# ──────────────────────────────────────────────────────────────────────────────
# inventory/views.py
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.utils.timezone import localtime
from django.contrib.auth.decorators import login_required, permission_required
from django.core.paginator import Paginator
from django.views.decorators.http import require_POST
from django.db.models.functions import TruncDate
from django.utils import timezone
from django.conf import settings

import openpyxl
from openpyxl.utils import get_column_letter

from .models import Printer, InventoryTask, PageCounter, Organization
from .forms import PrinterForm
from .services import (
    run_discovery_for_ip,
    extract_serial_from_xml,
    get_printer_inventory_status,
    run_inventory_for_printer,
    inventory_daemon,
    get_glpi_info,
)

# Попытка подтянуть Celery задачи
try:
    from .tasks import run_inventory_task_priority, inventory_daemon_task

    CELERY_AVAILABLE = True
except Exception:
    CELERY_AVAILABLE = False

import json
import logging
from datetime import timedelta

logger = logging.getLogger(__name__)

# ──────────────────────────────────────────────────────────────────────────────
# Fallback пул фоновых задач (если нет Celery)
# ──────────────────────────────────────────────────────────────────────────────
if not CELERY_AVAILABLE:
    from concurrent.futures import ThreadPoolExecutor
    import threading

    EXECUTOR = ThreadPoolExecutor(max_workers=5)
    _RUNNING: set[int] = set()
    _RUNNING_LOCK = threading.Lock()


    def _queue_inventory(pk: int) -> bool:
        """Запускает инвентаризацию в пуле, если ещё не идёт для этого принтера."""
        with _RUNNING_LOCK:
            if pk in _RUNNING:
                return False
            _RUNNING.add(pk)

        def _job():
            try:
                run_inventory_for_printer(pk)
            except Exception as e:
                logger.error(f"Error in inventory job for printer {pk}: {e}")
            finally:
                with _RUNNING_LOCK:
                    _RUNNING.discard(pk)

        EXECUTOR.submit(_job)
        return True
else:
    def _queue_inventory(pk: int) -> bool:
        return True


# ──────────────────────────────────────────────────────────────────────────────
# LIST
# ──────────────────────────────────────────────────────────────────────────────

@login_required
@permission_required("inventory.access_inventory_app", raise_exception=True)
@permission_required("inventory.view_printer", raise_exception=True)
def printer_list(request):
    """Список принтеров БЕЗ кэширования всей страницы."""

    q_ip = request.GET.get("q_ip", "").strip()
    q_model = request.GET.get("q_model", "").strip()
    q_serial = request.GET.get("q_serial", "").strip()
    q_org = request.GET.get("q_org", "").strip()
    q_rule = request.GET.get("q_rule", "").strip()
    per_page = request.GET.get("per_page", "100").strip()

    try:
        per_page = int(per_page)
        if per_page not in [10, 25, 50, 100, 250, 500, 1000, 2000, 5000]:
            per_page = 100
    except ValueError:
        per_page = 100

    # Оптимизация: используем select_related для организации
    qs = Printer.objects.select_related("organization").all()

    if q_ip:
        qs = qs.filter(ip_address__icontains=q_ip)
    if q_model:
        qs = qs.filter(model__icontains=q_model)
    if q_serial:
        qs = qs.filter(serial_number__icontains=q_serial)

    if q_org == "none":
        qs = qs.filter(organization__isnull=True)
    elif q_org:
        try:
            qs = qs.filter(organization_id=int(q_org))
        except ValueError:
            qs = qs.filter(organization__name__icontains=q_org)

    if q_rule in ("SN_MAC", "MAC_ONLY", "SN_ONLY"):
        qs = qs.filter(last_match_rule=q_rule)
    elif q_rule == "NONE":
        qs = qs.filter(last_match_rule__isnull=True)

    qs = qs.order_by("ip_address")
    paginator = Paginator(qs, per_page)
    page_obj = paginator.get_page(request.GET.get("page"))

    # Получаем данные для текущей страницы
    data = []
    for p in page_obj:
        # Читаем статус напрямую из БД
        inv_status = get_printer_inventory_status(p.id)
        counters = inv_status.get("counters", {})

        last_date_iso = ""
        last_date = "—"
        if inv_status.get("timestamp"):
            try:
                dt = timezone.datetime.fromisoformat(inv_status["timestamp"].replace("Z", "+00:00"))
                last_date_iso = int(dt.timestamp() * 1000)
                last_date = localtime(dt).strftime("%d.%m.%Y %H:%M")
            except Exception:
                pass

        data.append({
            "printer": p,
            "bw_a4": counters.get("bw_a4"),
            "color_a4": counters.get("color_a4"),
            "bw_a3": counters.get("bw_a3"),
            "color_a3": counters.get("color_a3"),
            "total": counters.get("total_pages"),
            "drum_black": counters.get("drum_black", ""),
            "drum_cyan": counters.get("drum_cyan", ""),
            "drum_magenta": counters.get("drum_magenta", ""),
            "drum_yellow": counters.get("drum_yellow", ""),
            "toner_black": counters.get("toner_black", ""),
            "toner_cyan": counters.get("toner_cyan", ""),
            "toner_magenta": counters.get("toner_magenta", ""),
            "toner_yellow": counters.get("toner_yellow", ""),
            "fuser_kit": counters.get("fuser_kit", ""),
            "transfer_kit": counters.get("transfer_kit", ""),
            "waste_toner": counters.get("waste_toner", ""),
            "last_date": last_date,
            "last_date_iso": last_date_iso,
            "is_fresh": False,  # Без кэша нет понятия "свежести"
        })

    # Статистика недавних ошибок
    recent_cutoff = timezone.now() - timedelta(hours=24)
    recent_failed = (
        InventoryTask.objects
        .filter(
            task_timestamp__gte=recent_cutoff,
            status__in=["FAILED", "VALIDATION_ERROR", "HISTORICAL_INCONSISTENCY"]
        )
        .values("printer")
        .distinct()
        .count()
    )

    per_page_options = [10, 25, 50, 100, 250, 500, 1000, 2000, 5000]
    organizations = Organization.objects.filter(active=True).order_by("name")

    context = {
        "data": data,
        "page_obj": page_obj,
        "q_ip": q_ip,
        "q_model": q_model,
        "q_serial": q_serial,
        "q_org": q_org,
        "q_rule": q_rule,
        "per_page": per_page,
        "per_page_options": per_page_options,
        "organizations": organizations,
        "celery_available": CELERY_AVAILABLE,
        "recent_failed": recent_failed,
        "recent_cutoff": recent_cutoff,
    }

    return render(request, "inventory/index.html", context)


# ──────────────────────────────────────────────────────────────────────────────
# EXPORTS
# ──────────────────────────────────────────────────────────────────────────────

@login_required
@permission_required("inventory.access_inventory_app", raise_exception=True)
@permission_required("inventory.export_printers", raise_exception=True)
def export_excel(request):
    """Экспорт в Excel БЕЗ кэширования."""

    q_ip = request.GET.get("q_ip", "").strip()
    q_model = request.GET.get("q_model", "").strip()
    q_serial = request.GET.get("q_serial", "").strip()
    q_org = request.GET.get("q_org", "").strip()
    q_rule = request.GET.get("q_rule", "").strip()

    qs = Printer.objects.select_related("organization").all()

    if q_ip:
        qs = qs.filter(ip_address__icontains=q_ip)
    if q_model:
        qs = qs.filter(model__icontains=q_model)
    if q_serial:
        qs = qs.filter(serial_number__icontains=q_serial)

    if q_org == "none":
        qs = qs.filter(organization__isnull=True)
    elif q_org:
        try:
            qs = qs.filter(organization_id=int(q_org))
        except ValueError:
            qs = qs.filter(organization__name__icontains=q_org)

    if q_rule in ("SN_MAC", "MAC_ONLY", "SN_ONLY"):
        qs = qs.filter(last_match_rule=q_rule)
    elif q_rule == "NONE":
        qs = qs.filter(last_match_rule__isnull=True)

    qs = qs.order_by("ip_address")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Printers"

    headers = [
        "Организация", "IP-адрес", "Серийный №", "MAC-адрес", "Модель",
        "ЧБ A4", "Цвет A4", "ЧБ A3", "Цвет A3", "Всего",
        "Тонер K", "Тонер C", "Тонер M", "Тонер Y",
        "Барабан K", "Барабан C", "Барабан M", "Барабан Y",
        "Fuser Kit", "Transfer Kit", "Waste Toner",
        "Правило", "Дата последнего опроса",
        "Статус последнего опроса", "Последняя ошибка",
    ]

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = openpyxl.styles.Font(bold=True)

    col_widths = [len(h) for h in headers]
    date_col_idx = headers.index("Дата последнего опроса") + 1

    rule_label = {
        "SN_MAC": "Серийник+MAC",
        "MAC_ONLY": "Только MAC",
        "SN_ONLY": "Только серийник",
    }

    row_idx = 2
    for p in qs:
        inv_status = get_printer_inventory_status(p.id)
        counters = inv_status.get("counters", {})

        dt_value = None
        if inv_status.get("timestamp"):
            try:
                dt = timezone.datetime.fromisoformat(inv_status["timestamp"].replace("Z", "+00:00"))
                dt_value = localtime(dt).replace(tzinfo=None)
            except Exception:
                pass

        # Последняя задача для статуса/ошибки
        last_task = InventoryTask.objects.filter(printer=p).order_by("-task_timestamp").first()
        if last_task:
            try:
                status_map = dict(InventoryTask.STATUS_CHOICES)
                last_status = status_map.get(last_task.status, last_task.status or "—")
            except Exception:
                last_status = last_task.status or "—"
        else:
            last_status = "—"

        last_error = ""
        if last_task and last_task.status in ["FAILED", "VALIDATION_ERROR", "HISTORICAL_INCONSISTENCY"]:
            last_error = (last_task.error_message or "")[:100]

        values = [
            p.organization.name if p.organization_id else "—",
            p.ip_address,
            p.serial_number,
            p.mac_address or "",
            p.model,
            counters.get("bw_a4", ""),
            counters.get("color_a4", ""),
            counters.get("bw_a3", ""),
            counters.get("color_a3", ""),
            counters.get("total_pages", ""),
            counters.get("toner_black", ""),
            counters.get("toner_cyan", ""),
            counters.get("toner_magenta", ""),
            counters.get("toner_yellow", ""),
            counters.get("drum_black", ""),
            counters.get("drum_cyan", ""),
            counters.get("drum_magenta", ""),
            counters.get("drum_yellow", ""),
            counters.get("fuser_kit", ""),
            counters.get("transfer_kit", ""),
            counters.get("waste_toner", ""),
            rule_label.get(p.last_match_rule, "—"),
            dt_value,
            last_status,
            last_error,
        ]

        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)

            if col_idx == date_col_idx and dt_value:
                cell.number_format = "dd.mm.yyyy hh:mm"

            disp = val if val is not None else ""
            if hasattr(val, "strftime"):
                disp = val.strftime("%d.%m.%Y %H:%M")
            col_widths[col_idx - 1] = max(col_widths[col_idx - 1], len(str(disp)))

        row_idx += 1

    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = min(max(w + 2, 10), 50)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="printers.xlsx"'
    wb.save(response)

    return response


@login_required
@permission_required("inventory.access_inventory_app", raise_exception=True)
@permission_required("inventory.export_amb_report", raise_exception=True)
def export_amb(request):
    """Экспорт AMB отчёта БЕЗ кэширования."""

    if request.method != "POST" or "template" not in request.FILES:
        return render(request, "inventory/export_amb.html")

    wb = openpyxl.load_workbook(request.FILES["template"])
    ws = wb.active

    # Поиск строки заголовков
    header_row = next(
        (r for r in range(1, 11) if any(
            str(c.value).strip().lower() == "серийный номер оборудования" for c in ws[r]
        )),
        None,
    )
    if not header_row:
        return HttpResponse("Строка заголовков не найдена.", status=400)

    headers = {
        str(ws.cell(row=header_row, column=col).value).strip().lower(): col
        for col in range(1, ws.max_column + 1)
        if ws.cell(row=header_row, column=col).value
    }

    lookup = {
        "serial": lambda k: "серийный номер оборудования" in k,
        "a4_bw": lambda k: "ч/б" in k and "конец периода" in k and "а4" in k,
        "a4_color": lambda k: "цветные" in k and "конец периода" in k and "а4" in k,
        "a3_bw": lambda k: "ч/б" in k and "конец периода" in k and "а3" in k,
        "a3_color": lambda k: "цветные" in k and "конец периода" in k and "а3" in k,
    }

    cols = {}
    for internal, cond in lookup.items():
        for key, idx in headers.items():
            if cond(key):
                cols[internal] = idx
                break
        if internal not in cols:
            return HttpResponse(f"Колонка для '{internal}' не найдена.", status=400)

    # Серийники
    serial_cells = []
    for row in range(header_row + 1, ws.max_row + 1):
        raw = ws.cell(row=row, column=cols["serial"]).value
        serial = str(raw).strip() if raw else ""
        if serial:
            serial_cells.append((row, serial))

    serials = {s for _, s in serial_cells}
    if not serials:
        return HttpResponse("В файле нет серийных номеров.", status=400)

    # Данные инвентаризации напрямую из БД
    counters_by_serial = {}
    for serial in serials:
        try:
            printer = Printer.objects.get(serial_number=serial)
            inv_status = get_printer_inventory_status(printer.id)
            counters = inv_status.get("counters", {})
            timestamp = inv_status.get("timestamp")
            if counters and timestamp:
                counters_by_serial[serial] = {"counters": counters, "timestamp": timestamp}
        except Printer.DoesNotExist:
            continue

    # Запись значений
    date_col = ws.max_column + 1
    ws.cell(row=header_row, column=date_col, value="Дата опроса")

    for row_idx, serial in serial_cells:
        data = counters_by_serial.get(serial)
        if not data:
            continue

        counters = data["counters"]
        ws.cell(row=row_idx, column=cols["a4_bw"], value=counters.get("bw_a4") or 0)
        ws.cell(row=row_idx, column=cols["a4_color"], value=counters.get("color_a4") or 0)
        ws.cell(row=row_idx, column=cols["a3_bw"], value=counters.get("bw_a3") or 0)
        ws.cell(row=row_idx, column=cols["a3_color"], value=counters.get("color_a3") or 0)

        try:
            dt = timezone.datetime.fromisoformat(data["timestamp"].replace("Z", "+00:00"))
            dt_local = localtime(dt).replace(tzinfo=None)
            c = ws.cell(row=row_idx, column=date_col, value=dt_local)
            c.number_format = "dd.mm.yyyy hh:mm"
        except Exception:
            pass

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="amb_report.xlsx"'
    wb.save(response)
    wb.close()
    return response


# ──────────────────────────────────────────────────────────────────────────────
# CRUD
# ──────────────────────────────────────────────────────────────────────────────

@login_required
@permission_required("inventory.access_inventory_app", raise_exception=True)
@permission_required("inventory.add_printer", raise_exception=True)
def add_printer(request):
    """Добавление принтера БЕЗ кэширования."""
    form = PrinterForm(request.POST or None)
    if form.is_valid():
        printer = form.save()
        messages.success(request, "Принтер добавлен")
        return redirect("inventory:printer_list")
    return render(request, "inventory/add_printer.html", {"form": form})


@login_required
@permission_required("inventory.access_inventory_app", raise_exception=True)
@permission_required("inventory.change_printer", raise_exception=True)
def edit_printer(request, pk):
    """Редактирование принтера БЕЗ кэширования."""
    printer = get_object_or_404(Printer, pk=pk)
    form = PrinterForm(request.POST or None, instance=printer)

    if request.method == "POST" and form.is_valid():
        printer = form.save()

        if request.headers.get("X-Requested-With") == "XMLHttpRequest":
            return JsonResponse({
                "success": True,
                "printer": {
                    "id": printer.id,
                    "ip_address": printer.ip_address,
                    "serial_number": printer.serial_number,
                    "mac_address": printer.mac_address,
                    "model": printer.model,
                    "snmp_community": printer.snmp_community,
                    "organization": printer.organization.name if printer.organization_id else None,
                    "organization_id": printer.organization_id,
                },
            })

        messages.success(request, "Принтер обновлён")
        return redirect("inventory:printer_list")

    if request.headers.get("X-Requested-With") == "XMLHttpRequest":
        return JsonResponse({"success": False, "error": form.errors.as_json()}, status=400)

    return render(request, "inventory/edit_printer.html", {"form": form})


@login_required
@permission_required("inventory.access_inventory_app", raise_exception=True)
@permission_required("inventory.delete_printer", raise_exception=True)
def delete_printer(request, pk):
    """Удаление принтера."""
    printer = get_object_or_404(Printer, pk=pk)

    if request.method == "POST":
        printer.delete()
        messages.success(request, "Принтер удалён")
        return JsonResponse({"success": True})

    # GET - показываем список с модалкой
    return printer_list(request)


# ──────────────────────────────────────────────────────────────────────────────
# HISTORY
# ──────────────────────────────────────────────────────────────────────────────

@login_required
@permission_required("inventory.access_inventory_app", raise_exception=True)
@permission_required("inventory.view_printer", raise_exception=True)
def history_view(request, pk):
    """История принтера БЕЗ кэширования."""
    printer = get_object_or_404(Printer, pk=pk)

    if request.headers.get("X-Requested-With") == "XMLHttpRequest":
        # AJAX запрос - возвращаем JSON
        daily_tasks = (
            InventoryTask.objects.filter(
                printer=printer,
                status__in=['SUCCESS', 'HISTORICAL_INCONSISTENCY']
            )
            .annotate(day=TruncDate("task_timestamp"))
            .order_by("-day", "-task_timestamp", "-pk")
            .distinct("day")
        )

        daily_list = list(daily_tasks)
        counters = PageCounter.objects.filter(task__in=daily_list)
        counter_by_task_id = {c.task_id: c for c in counters}

        data = []
        for t in daily_list:
            c = counter_by_task_id.get(t.id)

            # Для исторических несоответствий показываем предыдущие данные
            if t.status == 'HISTORICAL_INCONSISTENCY':
                last_valid_task = InventoryTask.objects.filter(
                    printer=printer,
                    status='SUCCESS',
                    task_timestamp__lt=t.task_timestamp
                ).order_by('-task_timestamp').first()

                if last_valid_task:
                    c = PageCounter.objects.filter(task=last_valid_task).first()

            data.append({
                "task_timestamp": localtime(t.task_timestamp).strftime("%Y-%m-%dT%H:%M:%S"),
                "match_rule": t.match_rule,
                "status": t.status,
                "status_display": t.get_status_display(),
                "bw_a4": c.bw_a4 if c else None,
                "color_a4": c.color_a4 if c else None,
                "bw_a3": c.bw_a3 if c else None,
                "color_a3": c.color_a3 if c else None,
                "total_pages": c.total_pages if c else None,
                "drum_black": c.drum_black if c else None,
                "drum_cyan": c.drum_cyan if c else None,
                "drum_magenta": c.drum_magenta if c else None,
                "drum_yellow": c.drum_yellow if c else None,
                "toner_black": c.toner_black if c else None,
                "toner_cyan": c.toner_cyan if c else None,
                "toner_magenta": c.toner_magenta if c else None,
                "toner_yellow": c.toner_yellow if c else None,
                "fuser_kit": c.fuser_kit if c else None,
                "transfer_kit": c.transfer_kit if c else None,
                "waste_toner": c.waste_toner if c else None,
                "is_historical_issue": t.status == 'HISTORICAL_INCONSISTENCY',
                "error_message": t.error_message if t.status == 'HISTORICAL_INCONSISTENCY' else None,
            })

        return JsonResponse(data, safe=False)

    # HTML версия - постранично
    tasks = InventoryTask.objects.filter(
        printer=printer,
        status="SUCCESS"
    ).order_by("-task_timestamp")

    paginator = Paginator(tasks, 50)
    page_obj = paginator.get_page(request.GET.get("page"))
    rows = [(t, PageCounter.objects.filter(task=t).first()) for t in page_obj]

    return render(
        request,
        "inventory/history.html",
        {"printer": printer, "rows": rows, "page_obj": page_obj}
    )


# ──────────────────────────────────────────────────────────────────────────────
# RUNNERS
# ──────────────────────────────────────────────────────────────────────────────

@login_required
@permission_required("inventory.access_inventory_app", raise_exception=True)
@permission_required("inventory.run_inventory", raise_exception=True)
@require_POST
def run_inventory(request, pk):
    """Запуск инвентаризации одного принтера."""
    pk = int(pk)

    if CELERY_AVAILABLE:
        try:
            from inventory.tasks import run_inventory_task_priority

            task = run_inventory_task_priority.apply_async(
                args=[pk, request.user.id],
                priority=9,
            )
            logger.info(f"Queued PRIORITY Celery task {task.id} for printer {pk}")
            return JsonResponse({
                "success": True,
                "celery": True,
                "task_id": task.id,
                "queued": True
            })
        except Exception as e:
            logger.error(f"Error queuing Celery task for printer {pk}: {e}")
            return JsonResponse({
                "success": False,
                "celery": True,
                "error": str(e)
            }, status=500)
    else:
        queued = _queue_inventory(pk)
        return JsonResponse({
            "success": True,
            "celery": False,
            "queued": queued
        })


@login_required
@permission_required("inventory.access_inventory_app", raise_exception=True)
@permission_required("inventory.run_inventory", raise_exception=True)
@require_POST
def run_inventory_all(request):
    """Запуск инвентаризации всех принтеров."""

    if CELERY_AVAILABLE:
        try:
            task = inventory_daemon_task.apply_async(priority=5)
            logger.info(f"User {request.user.id} queued daemon task {task.id}")
            return JsonResponse({
                "success": True,
                "celery": True,
                "task_id": task.id,
                "queued": True
            })
        except Exception as e:
            logger.error(f"Error queuing daemon task: {e}")
            return JsonResponse({
                "success": False,
                "celery": True,
                "error": str(e)
            }, status=500)
    else:
        from concurrent.futures import ThreadPoolExecutor
        executor = ThreadPoolExecutor(max_workers=1)
        executor.submit(inventory_daemon)
        return JsonResponse({
            "success": True,
            "celery": False,
            "queued": True
        })


# ──────────────────────────────────────────────────────────────────────────────
# APIs
# ──────────────────────────────────────────────────────────────────────────────

@login_required
@permission_required("inventory.access_inventory_app", raise_exception=True)
@permission_required("inventory.view_printer", raise_exception=True)
def api_printers(request):
    """API списка принтеров БЕЗ кэширования."""

    output = []
    for p in Printer.objects.select_related("organization").all():
        inv_status = get_printer_inventory_status(p.id)
        counters = inv_status.get("counters", {})

        ts_ms = ""
        if inv_status.get("timestamp"):
            try:
                dt = timezone.datetime.fromisoformat(
                    inv_status["timestamp"].replace("Z", "+00:00")
                )
                ts_ms = int(dt.timestamp() * 1000)
            except Exception:
                pass

        output.append({
            "id": p.id,
            "ip_address": p.ip_address,
            "serial_number": p.serial_number,
            "mac_address": p.mac_address or "-",
            "model": p.model,
            "snmp_community": p.snmp_community or "public",
            "organization_id": p.organization_id,
            "organization": p.organization.name if p.organization_id else None,
            "last_match_rule": p.last_match_rule,
            "last_match_rule_label": p.get_last_match_rule_display() if p.last_match_rule else None,
            "bw_a4": counters.get("bw_a4", "-"),
            "color_a4": counters.get("color_a4", "-"),
            "bw_a3": counters.get("bw_a3", "-"),
            "color_a3": counters.get("color_a3", "-"),
            "total_pages": counters.get("total_pages", "-"),
            "drum_black": counters.get("drum_black", "-"),
            "drum_cyan": counters.get("drum_cyan", "-"),
            "drum_magenta": counters.get("drum_magenta", "-"),
            "drum_yellow": counters.get("drum_yellow", "-"),
            "toner_black": counters.get("toner_black", "-"),
            "toner_cyan": counters.get("toner_cyan", "-"),
            "toner_magenta": counters.get("toner_magenta", "-"),
            "toner_yellow": counters.get("toner_yellow", "-"),
            "fuser_kit": counters.get("fuser_kit", "-"),
            "transfer_kit": counters.get("transfer_kit", "-"),
            "waste_toner": counters.get("waste_toner", "-"),
            "last_date_iso": ts_ms,
        })

    return JsonResponse(output, safe=False)


@login_required
@permission_required("inventory.access_inventory_app", raise_exception=True)
@permission_required("inventory.view_printer", raise_exception=True)
def api_printer(request, pk):
    """API информации о принтере БЕЗ кэширования."""
    pk = int(pk)
    # ОБНОВЛЕНО: добавили select_related для device_model
    printer = get_object_or_404(
        Printer.objects.select_related('organization', 'device_model', 'device_model__manufacturer'),
        pk=pk
    )

    inv_status = get_printer_inventory_status(pk)
    counters = inv_status.get("counters", {})

    ts_ms = ""
    if inv_status.get("timestamp"):
        try:
            dt = timezone.datetime.fromisoformat(
                inv_status["timestamp"].replace("Z", "+00:00")
            )
            ts_ms = int(dt.timestamp() * 1000)
        except Exception:
            pass

    data = {
        "id": printer.id,
        "ip_address": printer.ip_address,
        "serial_number": printer.serial_number,
        "mac_address": printer.mac_address or "-",
        "model": printer.model_display,  # ИЗМЕНЕНО: используем свойство model_display
        "model_text": printer.model,  # ДОБАВЛЕНО: старое текстовое поле
        "device_model_id": printer.device_model_id,  # ДОБАВЛЕНО
        "device_model_name": str(printer.device_model) if printer.device_model else None,  # ДОБАВЛЕНО
        "snmp_community": printer.snmp_community or "public",
        "organization_id": printer.organization_id,
        "organization": printer.organization.name if printer.organization_id else None,
        "last_match_rule": printer.last_match_rule,
        "last_match_rule_label": printer.get_last_match_rule_display() if printer.last_match_rule else None,
        "bw_a4": counters.get("bw_a4", "-"),
        "color_a4": counters.get("color_a4", "-"),
        "bw_a3": counters.get("bw_a3", "-"),
        "color_a3": counters.get("color_a3", "-"),
        "total_pages": counters.get("total_pages", "-"),
        "drum_black": counters.get("drum_black", "-"),
        "drum_cyan": counters.get("drum_cyan", "-"),
        "drum_magenta": counters.get("drum_magenta", "-"),
        "drum_yellow": counters.get("drum_yellow", "-"),
        "toner_black": counters.get("toner_black", "-"),
        "toner_cyan": counters.get("toner_cyan", "-"),
        "toner_magenta": counters.get("toner_magenta", "-"),
        "toner_yellow": counters.get("toner_yellow", "-"),
        "fuser_kit": counters.get("fuser_kit", "-"),
        "transfer_kit": counters.get("transfer_kit", "-"),
        "waste_toner": counters.get("waste_toner", "-"),
        "last_date_iso": ts_ms,
    }
    return JsonResponse(data)


@login_required
@permission_required("inventory.access_inventory_app", raise_exception=True)
@permission_required("inventory.run_inventory", raise_exception=True)
@require_POST
def api_probe_serial(request):
    """API для получения серийника через SNMP."""
    payload = json.loads(request.body.decode("utf-8"))
    ip = (payload.get("ip") or "").strip()
    community = (payload.get("community") or "public").strip() or "public"

    if not ip:
        return JsonResponse({"ok": False, "error": "ip не передан"}, status=400)

    ok, xml_or_err = run_discovery_for_ip(ip, community)
    if not ok:
        return JsonResponse({"ok": False, "error": xml_or_err}, status=502)

    serial = extract_serial_from_xml(xml_or_err)
    if not serial:
        return JsonResponse({"ok": False, "error": "Серийник не найден в XML"}, status=404)

    return JsonResponse({"ok": True, "serial": serial})


# ──────────────────────────────────────────────────────────────────────────────
# УПРОЩЁННЫЕ СИСТЕМНЫЕ API (без сложного мониторинга)
# ──────────────────────────────────────────────────────────────────────────────

@login_required
@permission_required("inventory.access_inventory_app", raise_exception=True)
def api_system_status(request):
    """Упрощённый API статуса системы."""

    # Базовая статистика Celery (если доступен)
    celery_status = {"available": CELERY_AVAILABLE}
    if CELERY_AVAILABLE:
        try:
            from celery import current_app
            inspect = current_app.control.inspect()
            stats = inspect.stats()

            if stats:
                celery_status.update({
                    "workers_count": len(stats),
                    "broker_url": current_app.conf.broker_url,
                })
        except Exception as e:
            celery_status["error"] = str(e)

    # Простая статистика принтеров из БД
    total_printers = Printer.objects.count()
    recent_cutoff = timezone.now() - timedelta(hours=24)

    recent_success = InventoryTask.objects.filter(
        task_timestamp__gte=recent_cutoff,
        status='SUCCESS'
    ).values('printer').distinct().count()

    printer_stats = {
        "total_printers": total_printers,
        "recent_success": recent_success,
        "success_rate": round(
            (recent_success / total_printers * 100) if total_printers > 0 else 0,
            1
        ),
        "updated_at": timezone.now().isoformat(),
    }

    return JsonResponse({
        "timestamp": timezone.now().isoformat(),
        "celery": celery_status,
        "printers": printer_stats,
        "glpi": get_glpi_info(),
    })


@login_required
@permission_required("inventory.access_inventory_app", raise_exception=True)
def api_status_statistics(request):
    """API статистики по статусам задач."""

    days = int(request.GET.get('days', 7))
    start_date = timezone.now() - timedelta(days=days)

    # Общая статистика
    total_tasks = InventoryTask.objects.filter(
        task_timestamp__gte=start_date
    ).count()

    # Статистика по статусам
    from django.db.models import Count
    status_stats = InventoryTask.objects.filter(
        task_timestamp__gte=start_date
    ).values('status').annotate(
        count=Count('id')
    ).order_by('-count')

    # Преобразуем в читаемый формат
    status_data = []
    for stat in status_stats:
        status_display = dict(InventoryTask.STATUS_CHOICES).get(
            stat['status'],
            stat['status']
        )
        percentage = (stat['count'] / total_tasks * 100) if total_tasks > 0 else 0

        status_data.append({
            'status': stat['status'],
            'status_display': status_display,
            'count': stat['count'],
            'percentage': round(percentage, 1)
        })

    # Топ принтеров с проблемами
    problematic_printers = InventoryTask.objects.filter(
        task_timestamp__gte=start_date,
        status__in=['FAILED', 'VALIDATION_ERROR', 'HISTORICAL_INCONSISTENCY']
    ).values(
        'printer__ip_address',
        'printer__model',
        'status'
    ).annotate(
        error_count=Count('id')
    ).order_by('-error_count')[:10]

    return JsonResponse({
        'period_days': days,
        'total_tasks': total_tasks,
        'status_statistics': status_data,
        'problematic_printers': list(problematic_printers),
        'timestamp': timezone.now().isoformat(),
    })

@login_required
@permission_required("inventory.access_inventory_app", raise_exception=True)
@permission_required("inventory.view_printer", raise_exception=True)
def generate_email_from_inventory(request, pk: int):
    """
    Генерирует .eml файл для принтера из инвентаря.
    Ищет соответствующее устройство в договорах по серийному номеру.
    """
    try:
        printer = Printer.objects.get(pk=pk)
    except Printer.DoesNotExist:
        messages.error(request, 'Принтер не найден')
        return redirect('inventory:printer_list')

    if not printer.serial_number:
        messages.error(
            request,
            f'У принтера {printer.ip_address} отсутствует серийный номер'
        )
        return redirect('inventory:printer_list')

    try:
        return generate_email_for_device(
            serial_number=printer.serial_number,
            user_email=request.user.email or 'user@example.com'
        )
    except Http404:
        messages.error(
            request,
            f'Устройство с серийным номером {printer.serial_number} не найдено в договорах. '
            f'Добавьте его сначала в раздел "Устройства в договоре".'
        )
        return redirect('inventory:printer_list')


@login_required
@permission_required("inventory.access_inventory_app", raise_exception=True)
def api_models_by_manufacturer(request):
    """API для получения моделей по производителю"""
    manufacturer_id = request.GET.get('manufacturer_id')

    if not manufacturer_id:
        return JsonResponse({'models': []})

    try:
        models = DeviceModel.objects.filter(
            manufacturer_id=manufacturer_id,
            device_type='printer'
        ).select_related('manufacturer').order_by('name').values('id', 'name', 'manufacturer__name')

        return JsonResponse({
            'models': [
                {
                    'id': m['id'],
                    'name': m['name'],
                    'full_name': f"{m['manufacturer__name']} {m['name']}"
                }
                for m in models
            ]
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@login_required
@permission_required("inventory.access_inventory_app", raise_exception=True)
def api_all_printer_models(request):
    """API для получения всех моделей принтеров"""
    models = DeviceModel.objects.filter(
        device_type='printer'
    ).select_related('manufacturer').order_by('manufacturer__name', 'name').values(
        'id', 'name', 'manufacturer__name', 'manufacturer__id'
    )

    return JsonResponse({
        'models': [
            {
                'id': m['id'],
                'name': m['name'],
                'manufacturer': m['manufacturer__name'],
                'manufacturer_id': m['manufacturer__id']
            }
            for m in models
        ]
    })

