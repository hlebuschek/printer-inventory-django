# inventory/views/export_views.py
"""
Export views for generating Excel reports and AMB templates.
Handles data export for external systems and reporting.
"""

import logging

from django.contrib.auth.decorators import login_required, permission_required
from django.shortcuts import render
from django.http import HttpResponse
from django.utils.timezone import localtime
from django.utils import timezone

import openpyxl
from openpyxl.utils import get_column_letter

from ..models import Printer, InventoryTask
from ..services import get_printer_inventory_status

logger = logging.getLogger(__name__)


# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
# EXCEL EXPORT
# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

@login_required
@permission_required("inventory.access_inventory_app", raise_exception=True)
@permission_required("inventory.export_printers", raise_exception=True)
def export_excel(request):
    """
    Ğ­ĞºÑĞ¿Ğ¾Ñ€Ñ‚ ÑĞ¿Ğ¸ÑĞºĞ° Ğ¿Ñ€Ğ¸Ğ½Ñ‚ĞµÑ€Ğ¾Ğ² Ğ² Excel Ñ Ñ„Ğ¸Ğ»ÑŒÑ‚Ñ€Ğ°Ñ†Ğ¸ĞµĞ¹.
    Ğ‘Ğ•Ğ— ĞºÑÑˆĞ¸Ñ€Ğ¾Ğ²Ğ°Ğ½Ğ¸Ñ - Ğ´Ğ°Ğ½Ğ½Ñ‹Ğµ Ñ‡Ğ¸Ñ‚Ğ°ÑÑ‚ÑÑ Ğ½Ğ°Ğ¿Ñ€ÑĞ¼ÑƒÑ Ğ¸Ğ· Ğ‘Ğ”.
    """
    from openpyxl.utils import get_column_letter
    import openpyxl
    from django.db.models import Q
    from django.http import HttpResponse
    from django.utils.timezone import localtime
    from ..services import get_printer_inventory_status

    # ĞŸĞ°Ñ€Ğ°Ğ¼ĞµÑ‚Ñ€Ñ‹ Ñ„Ğ¸Ğ»ÑŒÑ‚Ñ€Ğ°Ñ†Ğ¸Ğ¸
    q_ip = request.GET.get("q_ip", "").strip()
    q_serial = request.GET.get("q_serial", "").strip()
    q_org = request.GET.get("q_org", "").strip()
    q_rule = request.GET.get("q_rule", "").strip()

    # ğŸ”¹ ĞĞ¾Ğ²Ñ‹Ğµ Ğ¿Ğ°Ñ€Ğ°Ğ¼ĞµÑ‚Ñ€Ñ‹ Ñ„Ğ¸Ğ»ÑŒÑ‚Ñ€Ğ°Ñ†Ğ¸Ğ¸
    q_manufacturer = request.GET.get("q_manufacturer", "").strip()
    q_device_model = request.GET.get("q_device_model", "").strip()
    q_model_text = request.GET.get("q_model_text", "").strip()

    # Ğ‘Ğ°Ğ·Ğ¾Ğ²Ñ‹Ğ¹ queryset
    qs = Printer.objects.select_related(
        "organization",
        "device_model",
        "device_model__manufacturer"
    ).all()

    # ĞŸÑ€Ğ¸Ğ¼ĞµĞ½ÑĞµĞ¼ Ñ„Ğ¸Ğ»ÑŒÑ‚Ñ€Ñ‹
    if q_ip:
        qs = qs.filter(ip_address__icontains=q_ip)

    # ğŸ”¹ ĞĞ¾Ğ²Ğ°Ñ Ñ„Ğ¸Ğ»ÑŒÑ‚Ñ€Ğ°Ñ†Ğ¸Ñ Ğ¿Ğ¾ Ğ¼Ğ¾Ğ´ĞµĞ»Ğ¸/Ğ¿Ñ€Ğ¾Ğ¸Ğ·Ğ²Ğ¾Ğ´Ğ¸Ñ‚ĞµĞ»Ñ
    if q_device_model:
        qs = qs.filter(device_model_id=q_device_model)
    elif q_manufacturer:
        qs = qs.filter(device_model__manufacturer_id=q_manufacturer)
    elif q_model_text:
        qs = qs.filter(
            Q(model__icontains=q_model_text) |
            Q(device_model__name__icontains=q_model_text) |
            Q(device_model__manufacturer__name__icontains=q_model_text)
        )

    if q_serial:
        qs = qs.filter(serial_number__icontains=q_serial)

    if q_org == "none":
        qs = qs.filter(organization__isnull=True)
    elif q_org:
        try:
            qs = qs.filter(organization_id=int(q_org))
        except ValueError:
            qs = qs.filter(organization__name__icontains=q_org)

    if q_rule in ("SN_MAC", "MAC_ONLY", "SN_ONLY"):
        qs = qs.filter(last_match_rule=q_rule)
    elif q_rule == "NONE":
        qs = qs.filter(last_match_rule__isnull=True)

    qs = qs.order_by("ip_address")

    # â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    # Ğ¡Ğ¾Ğ·Ğ´Ğ°Ñ‘Ğ¼ Excel
    # â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Printers"

    # Ğ—Ğ°Ğ³Ğ¾Ğ»Ğ¾Ğ²ĞºĞ¸ â€” Ğ´Ğ¾Ğ±Ğ°Ğ²Ğ»ĞµĞ½Ğ° ĞºĞ¾Ğ»Ğ¾Ğ½ĞºĞ° "ĞŸÑ€Ğ¾Ğ¸Ğ·Ğ²Ğ¾Ğ´Ğ¸Ñ‚ĞµĞ»ÑŒ"
    headers = [
        "ĞÑ€Ğ³Ğ°Ğ½Ğ¸Ğ·Ğ°Ñ†Ğ¸Ñ", "IP-Ğ°Ğ´Ñ€ĞµÑ", "Ğ¡ĞµÑ€Ğ¸Ğ¹Ğ½Ñ‹Ğ¹ â„–", "MAC-Ğ°Ğ´Ñ€ĞµÑ",
        "ĞŸÑ€Ğ¾Ğ¸Ğ·Ğ²Ğ¾Ğ´Ğ¸Ñ‚ĞµĞ»ÑŒ", "ĞœĞ¾Ğ´ĞµĞ»ÑŒ",
        "Ğ§Ğ‘ A4", "Ğ¦Ğ²ĞµÑ‚ A4", "Ğ§Ğ‘ A3", "Ğ¦Ğ²ĞµÑ‚ A3", "Ğ’ÑĞµĞ³Ğ¾",
        "Ğ¢Ğ¾Ğ½ĞµÑ€ K", "Ğ¢Ğ¾Ğ½ĞµÑ€ C", "Ğ¢Ğ¾Ğ½ĞµÑ€ M", "Ğ¢Ğ¾Ğ½ĞµÑ€ Y",
        "Ğ‘Ğ°Ñ€Ğ°Ğ±Ğ°Ğ½ K", "Ğ‘Ğ°Ñ€Ğ°Ğ±Ğ°Ğ½ C", "Ğ‘Ğ°Ñ€Ğ°Ğ±Ğ°Ğ½ M", "Ğ‘Ğ°Ñ€Ğ°Ğ±Ğ°Ğ½ Y",
        "Fuser Kit", "Transfer Kit", "Waste Toner",
        "ĞŸÑ€Ğ°Ğ²Ğ¸Ğ»Ğ¾", "Ğ”Ğ°Ñ‚Ğ° Ğ¿Ğ¾ÑĞ»ĞµĞ´Ğ½ĞµĞ³Ğ¾ Ğ¾Ğ¿Ñ€Ğ¾ÑĞ°",
        "Ğ¡Ñ‚Ğ°Ñ‚ÑƒÑ Ğ¿Ğ¾ÑĞ»ĞµĞ´Ğ½ĞµĞ³Ğ¾ Ğ¾Ğ¿Ñ€Ğ¾ÑĞ°", "ĞŸĞ¾ÑĞ»ĞµĞ´Ğ½ÑÑ Ğ¾ÑˆĞ¸Ğ±ĞºĞ°",
    ]

    # Ğ—Ğ°Ğ³Ğ¾Ğ»Ğ¾Ğ²ĞºĞ¸ â€” Ğ¶Ğ¸Ñ€Ğ½Ñ‹Ğ¼
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = openpyxl.styles.Font(bold=True)

    col_widths = [len(h) for h in headers]
    date_col_idx = headers.index("Ğ”Ğ°Ñ‚Ğ° Ğ¿Ğ¾ÑĞ»ĞµĞ´Ğ½ĞµĞ³Ğ¾ Ğ¾Ğ¿Ñ€Ğ¾ÑĞ°") + 1

    rule_label = {
        "SN_MAC": "Ğ¡ĞµÑ€Ğ¸Ğ¹Ğ½Ğ¸Ğº+MAC",
        "MAC_ONLY": "Ğ¢Ğ¾Ğ»ÑŒĞºĞ¾ MAC",
        "SN_ONLY": "Ğ¢Ğ¾Ğ»ÑŒĞºĞ¾ ÑĞµÑ€Ğ¸Ğ¹Ğ½Ğ¸Ğº",
    }

    # â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    # Ğ—Ğ°Ğ¿Ğ¾Ğ»Ğ½ÑĞµĞ¼ Ğ´Ğ°Ğ½Ğ½Ñ‹Ğµ
    # â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    row_idx = 2
    for p in qs:
        inv_status = get_printer_inventory_status(p.id)
        counters = inv_status.get("counters", {})

        # Ğ”Ğ°Ñ‚Ğ° Ğ¿Ğ¾ÑĞ»ĞµĞ´Ğ½ĞµĞ³Ğ¾ Ğ¾Ğ¿Ñ€Ğ¾ÑĞ°
        dt_value = None
        if inv_status.get("timestamp"):
            try:
                dt = timezone.datetime.fromisoformat(inv_status["timestamp"].replace("Z", "+00:00"))
                dt_value = localtime(dt).replace(tzinfo=None)
            except Exception:
                pass

        # ĞŸĞ¾ÑĞ»ĞµĞ´Ğ½ÑÑ Ğ·Ğ°Ğ´Ğ°Ñ‡Ğ°
        last_task = InventoryTask.objects.filter(printer=p).order_by("-task_timestamp").first()
        if last_task:
            try:
                status_map = dict(InventoryTask.STATUS_CHOICES)
                last_status = status_map.get(last_task.status, last_task.status or "â€”")
            except Exception:
                last_status = last_task.status or "â€”"
        else:
            last_status = "â€”"

        last_error = ""
        if last_task and last_task.status in ["FAILED", "VALIDATION_ERROR", "HISTORICAL_INCONSISTENCY"]:
            last_error = (last_task.error_message or "")[:100]

        values = [
            p.organization.name if p.organization_id else "â€”",
            p.ip_address,
            p.serial_number,
            p.mac_address or "",
            getattr(p.device_model.manufacturer, "name", "â€”") if p.device_model_id else "â€”",  # ğŸ”¹ ĞŸÑ€Ğ¾Ğ¸Ğ·Ğ²Ğ¾Ğ´Ğ¸Ñ‚ĞµĞ»ÑŒ
            p.model_display,  # ğŸ”¹ ĞœĞ¾Ğ´ĞµĞ»ÑŒ
            counters.get("bw_a4", ""),
            counters.get("color_a4", ""),
            counters.get("bw_a3", ""),
            counters.get("color_a3", ""),
            counters.get("total_pages", ""),
            counters.get("toner_black", ""),
            counters.get("toner_cyan", ""),
            counters.get("toner_magenta", ""),
            counters.get("toner_yellow", ""),
            counters.get("drum_black", ""),
            counters.get("drum_cyan", ""),
            counters.get("drum_magenta", ""),
            counters.get("drum_yellow", ""),
            counters.get("fuser_kit", ""),
            counters.get("transfer_kit", ""),
            counters.get("waste_toner", ""),
            rule_label.get(p.last_match_rule, "â€”"),
            dt_value,
            last_status,
            last_error,
        ]

        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            if col_idx == date_col_idx and dt_value:
                cell.number_format = "dd.mm.yyyy hh:mm"

            disp = val if val is not None else ""
            if hasattr(val, "strftime"):
                disp = val.strftime("%d.%m.%Y %H:%M")
            col_widths[col_idx - 1] = max(col_widths[col_idx - 1], len(str(disp)))

        row_idx += 1

    # â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    # ĞĞ²Ñ‚Ğ¾Ğ¿Ğ¾Ğ´Ğ±Ğ¾Ñ€ ÑˆĞ¸Ñ€Ğ¸Ğ½Ñ‹ ĞºĞ¾Ğ»Ğ¾Ğ½Ğ¾Ğº
    # â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = min(max(w + 2, 10), 50)

    # â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    # ĞÑ‚Ğ²ĞµÑ‚ Ğ¿Ğ¾Ğ»ÑŒĞ·Ğ¾Ğ²Ğ°Ñ‚ĞµĞ»Ñ
    # â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="printers.xlsx"'
    wb.save(response)
    return response



# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
# AMB REPORT EXPORT
# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

@login_required
@permission_required("inventory.access_inventory_app", raise_exception=True)
@permission_required("inventory.export_amb_report", raise_exception=True)
def export_amb(request):
    """
    Ğ­ĞºÑĞ¿Ğ¾Ñ€Ñ‚ AMB Ğ¾Ñ‚Ñ‡Ñ‘Ñ‚Ğ° Ğ½Ğ° Ğ¾ÑĞ½Ğ¾Ğ²Ğµ Ğ·Ğ°Ğ³Ñ€ÑƒĞ¶ĞµĞ½Ğ½Ğ¾Ğ³Ğ¾ ÑˆĞ°Ğ±Ğ»Ğ¾Ğ½Ğ°.
    Ğ‘Ğ•Ğ— ĞºÑÑˆĞ¸Ñ€Ğ¾Ğ²Ğ°Ğ½Ğ¸Ñ - Ğ´Ğ°Ğ½Ğ½Ñ‹Ğµ Ñ‡Ğ¸Ñ‚Ğ°ÑÑ‚ÑÑ Ğ½Ğ°Ğ¿Ñ€ÑĞ¼ÑƒÑ Ğ¸Ğ· Ğ‘Ğ”.
    """
    if request.method != "POST" or "template" not in request.FILES:
        return render(request, "inventory/amb_export_vue.html")

    # Ğ—Ğ°Ğ³Ñ€ÑƒĞ¶Ğ°ĞµĞ¼ ÑˆĞ°Ğ±Ğ»Ğ¾Ğ½
    wb = openpyxl.load_workbook(request.FILES["template"])
    ws = wb.active

    # ĞŸĞ¾Ğ¸ÑĞº ÑÑ‚Ñ€Ğ¾ĞºĞ¸ Ğ·Ğ°Ğ³Ğ¾Ğ»Ğ¾Ğ²ĞºĞ¾Ğ²
    header_row = next(
        (r for r in range(1, 11) if any(
            str(c.value).strip().lower() == "ÑĞµÑ€Ğ¸Ğ¹Ğ½Ñ‹Ğ¹ Ğ½Ğ¾Ğ¼ĞµÑ€ Ğ¾Ğ±Ğ¾Ñ€ÑƒĞ´Ğ¾Ğ²Ğ°Ğ½Ğ¸Ñ" for c in ws[r]
        )),
        None,
    )
    if not header_row:
        return HttpResponse("Ğ¡Ñ‚Ñ€Ğ¾ĞºĞ° Ğ·Ğ°Ğ³Ğ¾Ğ»Ğ¾Ğ²ĞºĞ¾Ğ² Ğ½Ğµ Ğ½Ğ°Ğ¹Ğ´ĞµĞ½Ğ°.", status=400)

    # ĞŸĞ°Ñ€ÑĞ¸Ğ¼ Ğ·Ğ°Ğ³Ğ¾Ğ»Ğ¾Ğ²ĞºĞ¸
    headers = {
        str(ws.cell(row=header_row, column=col).value).strip().lower(): col
        for col in range(1, ws.max_column + 1)
        if ws.cell(row=header_row, column=col).value
    }

    # ĞĞ°Ñ…Ğ¾Ğ´Ğ¸Ğ¼ Ğ½ÑƒĞ¶Ğ½Ñ‹Ğµ ĞºĞ¾Ğ»Ğ¾Ğ½ĞºĞ¸
    lookup = {
        "serial": lambda k: "ÑĞµÑ€Ğ¸Ğ¹Ğ½Ñ‹Ğ¹ Ğ½Ğ¾Ğ¼ĞµÑ€ Ğ¾Ğ±Ğ¾Ñ€ÑƒĞ´Ğ¾Ğ²Ğ°Ğ½Ğ¸Ñ" in k,
        "a4_bw": lambda k: "Ñ‡/Ğ±" in k and "ĞºĞ¾Ğ½ĞµÑ† Ğ¿ĞµÑ€Ğ¸Ğ¾Ğ´Ğ°" in k and "Ğ°4" in k,
        "a4_color": lambda k: "Ñ†Ğ²ĞµÑ‚Ğ½Ñ‹Ğµ" in k and "ĞºĞ¾Ğ½ĞµÑ† Ğ¿ĞµÑ€Ğ¸Ğ¾Ğ´Ğ°" in k and "Ğ°4" in k,
        "a3_bw": lambda k: "Ñ‡/Ğ±" in k and "ĞºĞ¾Ğ½ĞµÑ† Ğ¿ĞµÑ€Ğ¸Ğ¾Ğ´Ğ°" in k and "Ğ°3" in k,
        "a3_color": lambda k: "Ñ†Ğ²ĞµÑ‚Ğ½Ñ‹Ğµ" in k and "ĞºĞ¾Ğ½ĞµÑ† Ğ¿ĞµÑ€Ğ¸Ğ¾Ğ´Ğ°" in k and "Ğ°3" in k,
    }

    cols = {}
    for internal, cond in lookup.items():
        for key, idx in headers.items():
            if cond(key):
                cols[internal] = idx
                break
        if internal not in cols:
            return HttpResponse(f"ĞšĞ¾Ğ»Ğ¾Ğ½ĞºĞ° Ğ´Ğ»Ñ '{internal}' Ğ½Ğµ Ğ½Ğ°Ğ¹Ğ´ĞµĞ½Ğ°.", status=400)

    # Ğ¡Ğ¾Ğ±Ğ¸Ñ€Ğ°ĞµĞ¼ ÑĞµÑ€Ğ¸Ğ¹Ğ½Ñ‹Ğµ Ğ½Ğ¾Ğ¼ĞµÑ€Ğ° Ğ¸Ğ· Ñ„Ğ°Ğ¹Ğ»Ğ°
    serial_cells = []
    for row in range(header_row + 1, ws.max_row + 1):
        raw = ws.cell(row=row, column=cols["serial"]).value
        serial = str(raw).strip() if raw else ""
        if serial:
            serial_cells.append((row, serial))

    serials = {s for _, s in serial_cells}
    if not serials:
        return HttpResponse("Ğ’ Ñ„Ğ°Ğ¹Ğ»Ğµ Ğ½ĞµÑ‚ ÑĞµÑ€Ğ¸Ğ¹Ğ½Ñ‹Ñ… Ğ½Ğ¾Ğ¼ĞµÑ€Ğ¾Ğ².", status=400)

    # ĞŸĞ¾Ğ»ÑƒÑ‡Ğ°ĞµĞ¼ Ğ´Ğ°Ğ½Ğ½Ñ‹Ğµ Ğ¸Ğ½Ğ²ĞµĞ½Ñ‚Ğ°Ñ€Ğ¸Ğ·Ğ°Ñ†Ğ¸Ğ¸ Ğ½Ğ°Ğ¿Ñ€ÑĞ¼ÑƒÑ Ğ¸Ğ· Ğ‘Ğ”
    counters_by_serial = {}
    for serial in serials:
        try:
            printer = Printer.objects.get(serial_number=serial)
            inv_status = get_printer_inventory_status(printer.id)
            counters = inv_status.get("counters", {})
            timestamp = inv_status.get("timestamp")
            if counters and timestamp:
                counters_by_serial[serial] = {"counters": counters, "timestamp": timestamp}
        except Printer.DoesNotExist:
            continue

    # Ğ”Ğ¾Ğ±Ğ°Ğ²Ğ»ÑĞµĞ¼ ĞºĞ¾Ğ»Ğ¾Ğ½ĞºÑƒ Ğ´Ğ»Ñ Ğ´Ğ°Ñ‚Ñ‹ Ğ¾Ğ¿Ñ€Ğ¾ÑĞ°
    date_col = ws.max_column + 1
    ws.cell(row=header_row, column=date_col, value="Ğ”Ğ°Ñ‚Ğ° Ğ¾Ğ¿Ñ€Ğ¾ÑĞ°")

    # Ğ—Ğ°Ğ¿Ğ¾Ğ»Ğ½ÑĞµĞ¼ Ğ´Ğ°Ğ½Ğ½Ñ‹Ğµ
    for row_idx, serial in serial_cells:
        data = counters_by_serial.get(serial)
        if not data:
            continue

        counters = data["counters"]
        ws.cell(row=row_idx, column=cols["a4_bw"], value=counters.get("bw_a4") or 0)
        ws.cell(row=row_idx, column=cols["a4_color"], value=counters.get("color_a4") or 0)
        ws.cell(row=row_idx, column=cols["a3_bw"], value=counters.get("bw_a3") or 0)
        ws.cell(row=row_idx, column=cols["a3_color"], value=counters.get("color_a3") or 0)

        # Ğ”Ğ°Ñ‚Ğ° Ğ¾Ğ¿Ñ€Ğ¾ÑĞ°
        try:
            dt = timezone.datetime.fromisoformat(data["timestamp"].replace("Z", "+00:00"))
            dt_local = localtime(dt).replace(tzinfo=None)
            c = ws.cell(row=row_idx, column=date_col, value=dt_local)
            c.number_format = "dd.mm.yyyy hh:mm"
        except Exception:
            pass

    # Ğ’Ğ¾Ğ·Ğ²Ñ€Ğ°Ñ‰Ğ°ĞµĞ¼ Ñ„Ğ°Ğ¹Ğ»
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="amb_report.xlsx"'
    wb.save(response)
    wb.close()

    return response