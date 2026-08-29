import json
from datetime import datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from django.contrib.auth.decorators import login_required, permission_required
from django.core.exceptions import PermissionDenied
from django.db import IntegrityError, transaction
from django.db.models import Count, Q
from django.http import FileResponse, Http404, HttpResponse, HttpResponseBadRequest, JsonResponse
from django.utils.timezone import now
from django.views.decorators.http import require_POST

from access.services.change_log_service import ChangeLogService

from .models import AcceptanceDocument, ContractDevice, ContractStatus, ServiceProvider
from .utils import SupportEmailNotConfigured, generate_email_for_device

# Поля, доступные с правом manage_device_acceptance (без полного change_contractdevice)
ACCEPTANCE_EDIT_FIELDS = ("status_id", "service_start_month", "initial_counter")


def _check_acceptance_edit_perm(user) -> bool:
    """True — полный редактор, False — только поля приёмки. Иначе PermissionDenied."""
    if user.has_perm("contracts.change_contractdevice"):
        return True
    if user.has_perm("contracts.manage_device_acceptance"):
        return False
    raise PermissionDenied


# ── API: частичное обновление (инлайн-редактор) ──────────────────────────────
@login_required
@permission_required("contracts.access_contracts_app", raise_exception=True)
@require_POST
def contractdevice_update_api(request, pk: int):
    full_edit = _check_acceptance_edit_perm(request.user)
    try:
        obj = ContractDevice.objects.select_related("organization", "city", "model__manufacturer", "status").get(pk=pk)
    except ContractDevice.DoesNotExist:
        raise Http404("Device not found")

    # Сохраняем старые значения для логирования
    old_data = ChangeLogService.get_model_data(obj)

    try:
        payload = json.loads(request.body.decode("utf-8"))
    except Exception:
        return JsonResponse({"ok": False, "error": "Некорректный JSON"}, status=400)

    # разрешённые поля (включая FK)
    allowed = (
        (
            "address",
            "room_number",
            "serial_number",
            "comment",
            "status_id",
            "service_provider_id",
            "organization_id",
            "city_id",
            "model_id",
            "service_start_month",
            "initial_counter",
        )
        if full_edit
        else ACCEPTANCE_EDIT_FIELDS
    )
    data = {k: v for k, v in payload.items() if k in allowed}

    # нормализация текстовых полей
    for key in ("address", "room_number", "serial_number", "comment"):
        if key in data and data[key] is not None:
            data[key] = str(data[key]).strip()

    # обработка даты принятия на обслуживание
    if "service_start_month" in data and data["service_start_month"]:
        try:
            # Принимаем YYYY-MM (от <input type="month">) и YYYY-MM-DD (legacy)
            parts = str(data["service_start_month"]).strip().split("-")
            if len(parts) < 2:
                raise ValueError("expected YYYY-MM")
            year, month = int(parts[0]), int(parts[1])
            obj.service_start_month = datetime(year, month, 1).date()
        except (ValueError, TypeError, AttributeError):
            return JsonResponse({"ok": False, "error": "Некорректный формат месяца обслуживания"}, status=400)
    elif "service_start_month" in data:
        obj.service_start_month = None

    # счётчик при приёмке
    if "initial_counter" in data:
        if data["initial_counter"] in (None, ""):
            obj.initial_counter = None
        else:
            try:
                counter = int(data["initial_counter"])
                if counter < 0:
                    raise ValueError
                obj.initial_counter = counter
            except (TypeError, ValueError):
                return JsonResponse({"ok": False, "error": "Некорректный счётчик при приёмке"}, status=400)

    # FK: организация/город/модель
    if "organization_id" in data and data["organization_id"]:
        try:
            obj.organization_id = int(data["organization_id"])
        except (TypeError, ValueError):
            return JsonResponse({"ok": False, "error": "Некорректная организация"}, status=400)

    if "city_id" in data and data["city_id"]:
        try:
            obj.city_id = int(data["city_id"])
        except (TypeError, ValueError):
            return JsonResponse({"ok": False, "error": "Некорректный город"}, status=400)

    if "model_id" in data and data["model_id"]:
        try:
            obj.model_id = int(data["model_id"])
        except (TypeError, ValueError):
            return JsonResponse({"ok": False, "error": "Некорректная модель"}, status=400)

    # статус
    if "status_id" in data:
        try:
            st = ContractStatus.objects.get(pk=int(data["status_id"]))
            obj.status = st
        except (TypeError, ValueError, ContractStatus.DoesNotExist):
            return JsonResponse({"ok": False, "error": "Статус не найден"}, status=400)

    # подрядчик
    if "service_provider_id" in data:
        try:
            obj.service_provider = ServiceProvider.objects.get(pk=int(data["service_provider_id"]))
        except (TypeError, ValueError, ServiceProvider.DoesNotExist):
            return JsonResponse({"ok": False, "error": "Подрядчик не найден"}, status=400)

    # простые поля
    for k in ("address", "room_number", "serial_number", "comment"):
        if k in data:
            setattr(obj, k, data[k])

    # сохранение с контролем уникальности серийника в рамках организации
    try:
        with transaction.atomic():
            obj.save()
            # Логируем изменения после успешного сохранения
            ChangeLogService.log_update(instance=obj, user=request.user, request=request, old_data=old_data)
    except IntegrityError:
        return JsonResponse(
            {"ok": False, "error": "Нарушение уникальности (серийный номер уже используется в этой организации)."},
            status=400,
        )

    # свежие related-объекты в ответ
    obj.refresh_from_db()
    st = obj.status
    return JsonResponse(
        {
            "ok": True,
            "device": {
                "id": obj.id,
                "address": obj.address,
                "room_number": obj.room_number,
                "serial_number": obj.serial_number,
                "comment": obj.comment,
                "service_start_month": obj.service_start_month.strftime("%Y-%m") if obj.service_start_month else "",
                "service_start_month_display": obj.service_start_month_display,
                "initial_counter": obj.initial_counter,
                "status": {
                    "id": st.id if st else None,
                    "name": st.name if st else "",
                    "color": st.color if st else "#6c757d",
                    "is_active": st.is_active if st else True,
                },
                "organization": {"id": obj.organization_id, "name": str(obj.organization)},
                "city": {"id": obj.city_id, "name": str(obj.city)},
                "manufacturer": {"id": obj.model.manufacturer_id, "name": str(obj.model.manufacturer)},
                "model": {"id": obj.model_id, "name": obj.model.name},
            },
        }
    )


# ── API: удаление ─────────────────────────────────────────────────────────────
@login_required
@permission_required("contracts.access_contracts_app", raise_exception=True)
@permission_required("contracts.delete_contractdevice", raise_exception=True)
@require_POST
def contractdevice_delete_api(request, pk: int):
    try:
        obj = ContractDevice.objects.get(pk=pk)
    except ContractDevice.DoesNotExist:
        raise Http404("Device not found")

    # Логируем удаление ДО фактического удаления
    ChangeLogService.log_delete(instance=obj, user=request.user, request=request)

    obj.delete()
    return JsonResponse({"ok": True})


@login_required
@permission_required("contracts.access_contracts_app", raise_exception=True)
@permission_required("contracts.add_contractdevice", raise_exception=True)
@require_POST
def contractdevice_create_api(request):
    """
    Создать устройство договора.
    Ожидает JSON:
    {
      organization_id, city_id, model_id, status_id, service_provider_id,
      address, room_number, serial_number, comment, service_start_month
    }
    """
    try:
        payload = json.loads(request.body.decode("utf-8"))
    except Exception:
        return JsonResponse({"ok": False, "error": "Некорректный JSON"}, status=400)

    required = ("organization_id", "city_id", "model_id", "status_id", "service_provider_id")
    for k in required:
        if not payload.get(k):
            return JsonResponse({"ok": False, "error": f"Не заполнено поле: {k}"}, status=400)

    # Подрядчик обязателен: от него зависит, куда уходят заявки по устройству
    try:
        provider = ServiceProvider.objects.get(pk=int(payload["service_provider_id"]))
    except (TypeError, ValueError, ServiceProvider.DoesNotExist):
        return JsonResponse({"ok": False, "error": "Подрядчик не найден"}, status=400)

    # нормализуем строки
    for key in ("address", "room_number", "serial_number", "comment"):
        if key in payload and payload[key] is not None:
            payload[key] = str(payload[key]).strip()

    # обработка даты принятия на обслуживание
    service_start_month = None
    if payload.get("service_start_month"):
        try:
            # Принимаем YYYY-MM (от <input type="month">) и YYYY-MM-DD (legacy)
            parts = str(payload["service_start_month"]).strip().split("-")
            if len(parts) < 2:
                raise ValueError("expected YYYY-MM")
            year, month = int(parts[0]), int(parts[1])
            service_start_month = datetime(year, month, 1).date()
        except (ValueError, TypeError, AttributeError):
            return JsonResponse({"ok": False, "error": "Некорректный формат месяца обслуживания"}, status=400)

    initial_counter = None
    if payload.get("initial_counter") not in (None, ""):
        try:
            initial_counter = int(payload["initial_counter"])
            if initial_counter < 0:
                raise ValueError
        except (TypeError, ValueError):
            return JsonResponse({"ok": False, "error": "Некорректный счётчик при приёмке"}, status=400)

    try:
        with transaction.atomic():
            obj = ContractDevice.objects.create(
                organization_id=payload["organization_id"],
                city_id=payload["city_id"],
                model_id=payload["model_id"],
                status_id=payload["status_id"],
                service_provider=provider,
                address=payload.get("address") or "",
                room_number=payload.get("room_number") or "",
                serial_number=payload.get("serial_number") or "",
                comment=payload.get("comment") or "",
                service_start_month=service_start_month,
                initial_counter=initial_counter,
            )
            # Логируем создание
            ChangeLogService.log_create(instance=obj, user=request.user, request=request)
    except IntegrityError:
        return JsonResponse(
            {"ok": False, "error": "Нарушение уникальности (серийный номер в организации)."}, status=400
        )

    # перечитаем для ответа с названиями
    obj = ContractDevice.objects.select_related("organization", "city", "model__manufacturer", "status", "printer").get(
        pk=obj.pk
    )

    st = obj.status
    return JsonResponse(
        {
            "ok": True,
            "device": {
                "id": obj.id,
                "address": obj.address,
                "room_number": obj.room_number,
                "serial_number": obj.serial_number,
                "comment": obj.comment,
                "service_start_month": obj.service_start_month.strftime("%Y-%m") if obj.service_start_month else "",
                "service_start_month_display": obj.service_start_month_display,
                "initial_counter": obj.initial_counter,
                "organization": {"id": obj.organization_id, "name": str(obj.organization)},
                "city": {"id": obj.city_id, "name": str(obj.city)},
                "manufacturer": {"id": obj.model.manufacturer_id, "name": str(obj.model.manufacturer)},
                "model": {"id": obj.model_id, "name": obj.model.name},
                "status": {
                    "id": st.id if st else None,
                    "name": st.name if st else "",
                    "color": st.color if st else "#6c757d",
                    "is_active": st.is_active if st else True,
                },
                "has_printer": bool(obj.printer_id),
                "printer_id": obj.printer_id,
            },
        }
    )


@login_required
@permission_required("contracts.access_contracts_app", raise_exception=True)
@permission_required("contracts.export_contracts", raise_exception=True)
def contractdevice_export_excel(request):
    # 1) собрать queryset — те же фильтры/поиск/сортировка
    qs = ContractDevice.objects.select_related(
        "organization", "city", "model__manufacturer", "status", "printer", "service_provider"
    ).annotate(_acceptance_docs_count=Count("acceptance_documents"))

    g = request.GET

    _filters = {
        "org": ("organization__name__icontains", g.get("org")),
        "city": ("city__name__icontains", g.get("city")),
        "address": ("address__icontains", g.get("address")),
        "room": ("room_number__icontains", g.get("room")),
        "mfr": ("model__manufacturer__name__icontains", g.get("mfr")),
        "model": ("model__name__icontains", g.get("model")),
        "serial": ("serial_number__icontains", g.get("serial")),
        "status": ("status__name__icontains", g.get("status")),
        "provider": ("service_provider__name__icontains", g.get("provider")),
        "service_month": ("service_start_month__icontains", g.get("service_month")),
        "comment": ("comment__icontains", g.get("comment")),
    }
    for key, (lookup, val) in _filters.items():
        if val:
            if key == "service_month":
                # Специальная обработка для фильтра месяца
                filter_val = val.strip()
                if "." in filter_val:
                    try:
                        month, year = filter_val.split(".")
                        month, year = int(month), int(year)
                        qs = qs.filter(service_start_month__year=year, service_start_month__month=month)
                        continue
                    except (ValueError, TypeError):
                        pass
                qs = qs.extra(where=["to_char(service_start_month, 'MM.YYYY') ILIKE %s"], params=[f"%{filter_val}%"])
            else:
                qs = qs.filter(**{lookup: val})

    q = g.get("q")
    if q:
        qs = qs.filter(
            Q(serial_number__icontains=q)
            | Q(address__icontains=q)
            | Q(room_number__icontains=q)
            | Q(comment__icontains=q)
            | Q(model__name__icontains=q)
            | Q(model__manufacturer__name__icontains=q)
            | Q(organization__name__icontains=q)
            | Q(city__name__icontains=q)
            | Q(status__name__icontains=q)
        )

    allowed = {
        "org": "organization__name",
        "city": "city__name",
        "address": "address",
        "room": "room_number",
        "mfr": "model__manufacturer__name",
        "model": "model__name",
        "serial": "serial_number",
        "status": "status__name",
        "provider": "service_provider__name",
        "service_month": "service_start_month",
        "comment": "comment",
    }
    sort = g.get("sort")
    if sort:
        desc = sort.startswith("-")
        key = sort[1:] if desc else sort
        if key in allowed:
            field = allowed[key]
            qs = qs.order_by(("-" if desc else "") + field)
    else:
        qs = qs.order_by("organization__name", "city__name", "address", "room_number")

    # 2) helpers для цвета статуса
    def xl_color(hex_color: str) -> str:
        if not hex_color:
            return "FF6C757D"
        h = hex_color.lstrip("#")
        if len(h) == 3:
            h = "".join(c * 2 for c in h)
        return ("FF" + h.upper())[:8]  # ARGB

    def contrast_font(hex_color: str) -> str:
        try:
            h = hex_color.lstrip("#")
            if len(h) == 3:
                h = "".join(c * 2 for c in h)
            r, g, b = int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16)
            y = (r * 299 + g * 587 + b * 114) / 1000
            return "FF000000" if y > 140 else "FFFFFFFF"  # черный / белый
        except Exception:
            return "FF000000"

    # 2.5) предзагрузка заявок Okdesk по серийным номерам
    # serial -> {all: [issue_id, ...], active: [issue_id, ...], overdue: [issue_id, ...]}
    okdesk_by_serial = {}
    try:
        from integrations.models import OkdeskIssue

        for issue in OkdeskIssue.objects.only(
            "issue_id", "serial_numbers", "status_name", "is_overdue", "author_name", "created_at"
        ):
            serials = [s.strip() for s in issue.serial_numbers.split(",") if s.strip()]
            is_active = issue.status_name != "Закрыта"
            for sn in serials:
                entry = okdesk_by_serial.setdefault(sn, {"all": [], "active": [], "overdue": [], "author": ""})
                entry["all"].append(str(issue.issue_id))
                if is_active:
                    entry["active"].append(str(issue.issue_id))
                    # Автор последней активной заявки
                    if issue.author_name:
                        entry["author"] = issue.author_name
                if issue.is_overdue:
                    entry["overdue"].append(str(issue.issue_id))
    except ImportError:
        pass

    # 3) сформировать книгу
    wb = Workbook()
    ws = wb.active
    ws.title = "Устройства"

    headers = [
        "№",
        "Организация",
        "Город",
        "Адрес",
        "№ кабинета",
        "Производитель",
        "Модель",
        "Серийный номер",
        "Месяц обслуживания",
        "Статус",
        "Подрядчик",
        "Комментарий",
        "Счётчик при приёмке",
        "Документы приёмки",
        "Автор заявки",
        "Заявки Okdesk",
        "Незакрытые заявки",
        "Просроченные заявки",
    ]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True)
        cell.fill = PatternFill("solid", fgColor="FFE9ECEF")  # светло-серый заголовок

    # 4) строки
    row = 2
    for i, d in enumerate(qs.iterator(), start=1):
        ws.cell(row=row, column=1, value=i)
        ws.cell(row=row, column=2, value=str(d.organization))
        ws.cell(row=row, column=3, value=str(d.city))
        ws.cell(row=row, column=4, value=d.address or "").alignment = Alignment(wrap_text=True)
        ws.cell(row=row, column=5, value=d.room_number or "")

        ws.cell(row=row, column=6, value=str(d.model.manufacturer))
        ws.cell(row=row, column=7, value=d.model.name)
        ws.cell(row=row, column=8, value=d.serial_number or "")

        # Месяц обслуживания
        service_month_value = d.service_start_month_display if d.service_start_month else ""
        ws.cell(row=row, column=9, value=service_month_value)

        # Статус
        st_name = d.status.name if d.status else ""
        st_cell = ws.cell(row=row, column=10, value=st_name)
        st_cell.alignment = Alignment(wrap_text=True)
        if d.status and d.status.color:
            st_cell.fill = PatternFill("solid", fgColor=xl_color(d.status.color))
            st_cell.font = Font(color=contrast_font(d.status.color))

        ws.cell(row=row, column=11, value=d.service_provider.name if d.service_provider_id else "")
        ws.cell(row=row, column=12, value=d.comment or "").alignment = Alignment(wrap_text=True)

        # Приёмка
        ws.cell(row=row, column=13, value=d.initial_counter)
        docs_count = d._acceptance_docs_count
        ws.cell(row=row, column=14, value=f"Да ({docs_count})" if docs_count else "Нет")

        # Автор заявки и заявки Okdesk
        sn = d.serial_number or ""
        issues = okdesk_by_serial.get(sn, {})
        ws.cell(row=row, column=15, value=issues.get("author", ""))
        ws.cell(row=row, column=16, value=", ".join(issues.get("all", [])))
        ws.cell(row=row, column=17, value=", ".join(issues.get("active", [])))

        overdue_val = ", ".join(issues.get("overdue", []))
        overdue_cell = ws.cell(row=row, column=18, value=overdue_val)
        if overdue_val:
            overdue_cell.font = Font(color="FFDC3545")  # красный для просроченных

        row += 1

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    # 5) автоширина колонок (ограничим разумно)
    for col_cells in ws.columns:
        letter = col_cells[0].column_letter
        max_len = 0
        for c in col_cells:
            v = c.value
            if v is None:
                continue
            s = str(v)
            if "\n" in s:
                s = max(s.split("\n"), key=len)
            max_len = max(max_len, len(s))
        ws.column_dimensions[letter].width = min(60, max(8, max_len + 2))

    # 6) ответ
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    filename = f"contract_devices_{now().strftime('%Y-%m-%d_%H-%M')}.xlsx"
    return HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@login_required
@permission_required("contracts.access_contracts_app", raise_exception=True)
@permission_required("contracts.view_contractdevice", raise_exception=True)
def contractdevice_lookup_by_serial_api(request):
    serial = (request.GET.get("serial") or "").strip()
    if not serial:
        return JsonResponse({"ok": False, "error": "serial не передан"}, status=400)

    try:
        dev = ContractDevice.objects.select_related(
            "organization", "city", "model__manufacturer", "service_provider"
        ).get(serial_number__iexact=serial)
    except ContractDevice.DoesNotExist:
        return JsonResponse({"ok": True, "found": False})

    return JsonResponse(
        {
            "ok": True,
            "found": True,
            "device": {
                "id": dev.id,
                "serial_number": dev.serial_number,
                "organization": {"id": dev.organization_id, "name": str(dev.organization) if dev.organization else ""},
                "city": {"id": dev.city_id, "name": str(dev.city) if dev.city else ""},
                "model": {"id": dev.model_id, "name": dev.model.name if dev.model_id else ""},
                "manufacturer": {
                    "id": dev.model.manufacturer_id if dev.model_id else None,
                    "name": str(dev.model.manufacturer) if dev.model_id else "",
                },
                "service_start_month": dev.service_start_month_display,
                "service_provider": dev.service_provider.name if dev.service_provider_id else "",
                "okdesk_enabled": dev.okdesk_enabled,
            },
        }
    )


@login_required
@permission_required("contracts.access_contracts_app", raise_exception=True)
@permission_required("contracts.view_contractdevice", raise_exception=True)
def generate_email_msg(request, pk: int):
    """
    Генерирует .eml файл (email) с заявкой на картридж для устройства.
    """
    try:
        return generate_email_for_device(device_id=pk, user_email=request.user.email or "sd@abi.com.ru")
    except SupportEmailNotConfigured as e:
        return HttpResponseBadRequest(str(e), content_type="text/plain; charset=utf-8")


# ── API: документы приёмки (PDF) ──────────────────────────────────────────────
ACCEPTANCE_PDF_MAX_SIZE = 20 * 1024 * 1024  # 20 МБ


def _serialize_acceptance_doc(doc):
    return {
        "id": doc.id,
        "name": doc.original_name,
        "uploaded_at": doc.uploaded_at.isoformat(),
        "uploaded_by": doc.uploaded_by.username if doc.uploaded_by_id else "",
    }


def _log_acceptance_doc_event(device, request, text):
    """Запись в историю изменений устройства о загрузке/удалении документа приёмки."""
    from django.contrib.contenttypes.models import ContentType

    from access.models import EntityChangeLog

    EntityChangeLog.objects.create(
        content_type=ContentType.objects.get_for_model(ContractDevice),
        object_id=device.pk,
        action="update",
        user=request.user,
        changes={"acceptance_documents": {"old": None, "new": text, "label": "Документы приёмки"}},
        object_repr=str(device)[:500],
        ip_address=ChangeLogService.get_ip_from_request(request),
        user_agent=ChangeLogService.get_user_agent(request),
    )


@login_required
@permission_required("contracts.access_contracts_app", raise_exception=True)
@permission_required("contracts.view_contractdevice", raise_exception=True)
def acceptance_doc_download(request, doc_id: int):
    """Отдаёт PDF через авторизованный view (media напрямую не публикуется)."""
    try:
        doc = AcceptanceDocument.objects.get(pk=doc_id)
    except AcceptanceDocument.DoesNotExist:
        raise Http404("Document not found")

    return FileResponse(
        doc.file.open("rb"), content_type="application/pdf", filename=doc.original_name or "document.pdf"
    )


@login_required
@permission_required("contracts.access_contracts_app", raise_exception=True)
@require_POST
def acceptance_docs_upload(request, pk: int):
    _check_acceptance_edit_perm(request.user)
    try:
        device = ContractDevice.objects.get(pk=pk)
    except ContractDevice.DoesNotExist:
        raise Http404("Device not found")

    files = request.FILES.getlist("files")
    if not files:
        return JsonResponse({"ok": False, "error": "Файлы не переданы"}, status=400)

    for file in files:
        if not file.name.lower().endswith(".pdf"):
            return JsonResponse({"ok": False, "error": f"«{file.name}»: допускается только PDF"}, status=400)
        if file.size > ACCEPTANCE_PDF_MAX_SIZE:
            return JsonResponse({"ok": False, "error": f"«{file.name}»: файл больше 20 МБ"}, status=400)
        # сигнатура PDF, чтобы не принять переименованный файл
        head = file.read(5)
        file.seek(0)
        if head != b"%PDF-":
            return JsonResponse({"ok": False, "error": f"«{file.name}»: файл не является PDF"}, status=400)

    docs = []
    with transaction.atomic():
        for file in files:
            docs.append(AcceptanceDocument.objects.create(device=device, file=file, uploaded_by=request.user))

    for doc in docs:
        _log_acceptance_doc_event(device, request, f"загружен файл «{doc.original_name}»")

    return JsonResponse({"ok": True, "documents": [_serialize_acceptance_doc(d) for d in docs]})


@login_required
@permission_required("contracts.access_contracts_app", raise_exception=True)
@require_POST
def acceptance_doc_delete(request, doc_id: int):
    _check_acceptance_edit_perm(request.user)
    try:
        doc = AcceptanceDocument.objects.select_related("device").get(pk=doc_id)
    except AcceptanceDocument.DoesNotExist:
        raise Http404("Document not found")

    device, name = doc.device, doc.original_name
    doc.file.delete(save=False)
    doc.delete()
    _log_acceptance_doc_event(device, request, f"удалён файл «{name}»")

    return JsonResponse({"ok": True})


# ── API: история изменений ────────────────────────────────────────────────────
@login_required
@permission_required("contracts.access_contracts_app", raise_exception=True)
@permission_required("access.view_entity_changes", raise_exception=True)
def contractdevice_change_history(request, pk: int):
    """
    Получение истории изменений устройства
    """
    try:
        device = ContractDevice.objects.get(pk=pk)
    except ContractDevice.DoesNotExist:
        return JsonResponse({"error": "Device not found"}, status=404)

    # Получаем историю изменений
    history = ChangeLogService.get_history(instance=device, limit=100)

    # Форматируем данные для фронтенда
    result = []
    for log in history:
        result.append(
            {
                "id": log.id,
                "action": log.action,
                "action_display": dict(log.ACTION_CHOICES).get(log.action, log.action),
                "user": log.user.username if log.user else "Система",
                "timestamp": log.timestamp.isoformat(),
                "changes": log.get_changes_display(),
                "ip_address": log.ip_address,
            }
        )

    return JsonResponse({"history": result}, safe=False)
