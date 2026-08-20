"""Выгрузка журнала заявок в Excel.

Один экспорт вместо шести окдесковых: те отличались только набором статусов,
а статус здесь — обычный фильтр. Выгружается ровно то, что отобрано в журнале,
включая заявки почтовым подрядчикам, которых в Okdesk нет вовсе.

Разбивка — лист на пару «подрядчик + статус». Разбивка по статусам осталась от
окдесковой выгрузки, а подрядчик добавлен к ней: заявки уходят нескольким
исполнителям одновременно, и разговор с каждым идёт по своим листам.

Статус на листах — подрядчика («В работе», «Требует решения»), а не наш: по нему
видно, что у подрядчика висит. Наши четыре статуса — про акт и K1/K2, они
остаются отдельной колонкой и используются как имя листа там, где своего
статуса у подрядчика нет (почтовые подрядчики, заявки не из Okdesk).
"""

import re
from io import BytesIO

from django.db.models import Case, IntegerField, Value, When
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .models import ServiceRequest
from .services_okdesk_import import provider_statuses
from .services_working_hours import calculator_for_request

HEADERS = (
    "Номер",
    "У подрядчика",
    "Подрядчик",
    "Организация",
    "Город",
    "Адрес",
    "Кабинет",
    "Модель",
    "Серийный номер",
    "Инициатор",
    "Контакты",
    "Описание",
    "Зарегистрирована",
    "Норматив, раб. ч",
    "Срок",
    "Срочность",
    "Восстановлена",
    "Закрыта",
    "Простой, раб. ч",
    "Статус у подрядчика",
    "Статус",
    "Просрочена",
    "Акт",
)

_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill(start_color="495057", end_color="495057", fill_type="solid")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
_TITLE_CHARS = re.compile(r"[\[\]:*?/\\]")
_TITLE_LIMIT = 31
_TITLE_MIN_STATUS = 8
_NO_PROVIDER = "Без подрядчика"
# Статусы на листах идут как в журнале (открытые сверху), а не по алфавиту кода
_STATUS_ORDER = Case(
    *[When(status=code, then=Value(rank)) for rank, (code, _) in enumerate(ServiceRequest.STATUSES)],
    output_field=IntegerField(),
)


def export_requests_excel(queryset):
    """Возвращает (содержимое xlsx, имя файла) по готовому queryset журнала."""
    workbook = Workbook()
    workbook.remove(workbook.active)

    # Связи берём здесь, а не полагаемся на вызывающего: выгрузка читает больше полей,
    # чем журнал на экране. Простой вдобавок считается по производственному календарю,
    # поэтому калькуляторы кэшируются на всю выгрузку.
    calculators = {}
    statuses = provider_statuses(queryset.exclude(external_number="").values_list("external_number", flat=True))
    rows = (
        queryset.select_related(
            "device__organization",
            "device__city__work_schedule",
            "device__model__manufacturer",
            "device__work_schedule",
            "service_provider",
            "initiator",
            "schedule_snapshot",
        )
        # Подрядчик и статус определяют лист, поэтому сортируем сначала по ним; внутри
        # листа остаётся порядок журнала. Так строки приходят группами и пишутся потоком.
        .annotate(status_rank=_STATUS_ORDER)
        .order_by("service_provider__name", "status_rank", "-registered_at")
        .iterator(chunk_size=500)
    )
    # У одного нашего статуса статусов подрядчика несколько, поэтому листы держим
    # в словаре: строки приезжают группами по нашему статусу, а внутри перемешаны
    sheets = {}
    for service_request in rows:
        provider_status = statuses.get(service_request.external_number, "")
        key = (service_request.service_provider_id, provider_status or service_request.status)
        sheet = sheets.get(key)
        if sheet is None:
            sheet = sheets[key] = _new_sheet(workbook, _sheet_title(service_request, provider_status))
        sheet.append(_row(service_request, calculators, provider_status))

    if not workbook.sheetnames:
        _new_sheet(workbook, "Заявки")
    for sheet in workbook.worksheets:
        _autosize(sheet)
        sheet.freeze_panes = "A2"

    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue(), f"service_requests_{timezone.localdate().isoformat()}.xlsx"


def _sheet_title(service_request, provider_status):
    provider = service_request.service_provider.name if service_request.service_provider_id else _NO_PROVIDER
    status = provider_status or service_request.get_status_display()
    # Лист длиннее 31 символа Excel не примет. Имя подрядчика ограничиваем так, чтобы
    # начало статуса осталось видно, а хвост длинных статусов вроде «Отправлено в
    # сторонний сервис» уже обрезаем. Одноимённые листы openpyxl переименовывает сам.
    provider = _TITLE_CHARS.sub("-", provider)[: _TITLE_LIMIT - _TITLE_MIN_STATUS - 3].strip()
    return f"{provider} — {status}"[:_TITLE_LIMIT]


def _new_sheet(workbook, title):
    sheet = workbook.create_sheet(title)
    sheet.append(list(HEADERS))
    for cell in sheet[1]:
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
    return sheet


def _row(service_request, calculators, provider_status):
    device = service_request.device
    return [
        service_request.number,
        service_request.external_number,
        service_request.service_provider.name if service_request.service_provider_id else "",
        device.organization.name if device.organization_id else "",
        device.city.name if device.city_id else "",
        device.address,
        device.room_number,
        str(device.model) if device.model_id else "",
        device.serial_number,
        _initiator(service_request),
        service_request.initiator_contacts,
        service_request.description,
        _moment(service_request.registered_at),
        service_request.sla_hours,
        _moment(service_request.deadline_at),
        service_request.urgency_label,
        _moment(service_request.restored_at),
        _moment(service_request.closed_at),
        # Незакрытая заявка копит простой прямо сейчас — считаем на момент выгрузки
        _downtime(service_request, calculators),
        provider_status,
        service_request.get_status_display(),
        "Да" if service_request.is_overdue else "",
        service_request.act_number,
    ]


def _downtime(service_request, calculators):
    if not service_request.stops_printing:
        return ""
    calculator = calculator_for_request(service_request, calculators)
    return round(service_request.downtime_hours(calculator=calculator), 2) if calculator else ""


def _initiator(service_request):
    if not service_request.initiator_id:
        return ""
    return service_request.initiator.get_full_name() or service_request.initiator.username


def _moment(value):
    return timezone.localtime(value).strftime("%d.%m.%Y %H:%M") if value else ""


def _autosize(sheet, max_width=60):
    for cells in sheet.columns:
        longest = max((len(str(cell.value)) for cell in cells if cell.value is not None), default=10)
        sheet.column_dimensions[get_column_letter(cells[0].column)].width = min(longest + 2, max_width)
