"""
Массовый импорт устройств по договору из Excel.

Две фазы:
  1. parse_workbook + analyze_file — разбор и классификация, ни одной записи в ContractDevice
  2. apply_session — применение решений пользователя

Общий код для веб-интерфейса и management-команды contracts_import_xlsx.

Правила валидации: серийник обязателен; организация, производитель и модель должны
существовать в справочнике; город создаётся только по явному согласию пользователя;
статус берётся из сессии, а не из файла.
"""

import re
from dataclasses import dataclass, field
from datetime import date, datetime

from openpyxl import load_workbook

from django.contrib.contenttypes.models import ContentType
from django.db import IntegrityError, transaction
from django.db.models.functions import Lower
from django.utils import timezone

from access.models import EntityChangeLog
from inventory.models import Organization, Printer

from .models import City, ContractDevice, DeviceModel, ImportFile, ImportRow, ImportSession, Manufacturer

# Заголовки из Excel → внутренние ключи
HEADERS = {
    "№": "rownum",
    "номер": "rownum",
    "организация": "organization",
    "организация, наименование": "organization",
    "город": "city",
    "адрес": "address",
    "№ кабинета": "room",
    "кабинет": "room",
    "номер кабинета": "room",
    "производитель": "manufacturer",
    "vendor": "manufacturer",
    "бренд": "manufacturer",
    "модель оборудования": "model",
    "модель": "model",
    "серийный номер": "serial",
    "sn": "serial",
    "s/n": "serial",
    "serial": "serial",
    "месяц обслуживания": "service_month",
    "месяц принятия на обслуживание": "service_month",
    "дата принятия": "service_month",
    "начало обслуживания": "service_month",
    "комментарий": "comment",
}

REQUIRED_COLUMNS = {"organization", "city", "address", "manufacturer", "model", "serial"}

ANALYZE_BATCH = 500
LOOKUP_CHUNK = 1000
APPLY_CHUNK = 200

# Классы, которые пользователь обязан разобрать вручную перед применением
CONFLICT_CLASSES = (ImportRow.MOVED, ImportRow.DUP_IN_FILE)


class ImportFileError(Exception):
    """Файл невозможно разобрать: нет листа, нет обязательных колонок."""


@dataclass
class ParsedRow:
    row_number: int
    values: dict = field(default_factory=dict)


def norm(value):
    """None -> '', str -> strip + схлопывание пробелов, числа -> str без '.0'."""
    if value is None:
        return ""
    if isinstance(value, (int, float)):
        s = str(int(value)) if float(value).is_integer() else str(value)
    else:
        s = str(value)
    return re.sub(r"\s+", " ", s.strip())


def key_of(header):
    if not header:
        return None
    return HEADERS.get(norm(header).lower())


def parse_service_month(value):
    """
    Парсит месяц обслуживания из MM.YYYY, MM/YYYY, YYYY-MM, YYYY-MM-DD, date/datetime.
    Возвращает date с первым числом месяца или None.
    """
    if not value:
        return None

    if isinstance(value, datetime):
        return value.date().replace(day=1)
    if isinstance(value, date):
        return value.replace(day=1)

    str_val = norm(value)
    if not str_val:
        return None

    try:
        if "." in str_val and len(str_val.split(".")) == 2:
            month_str, year_str = str_val.split(".")
            return date(int(year_str), int(month_str), 1)
        if "/" in str_val and len(str_val.split("/")) == 2:
            month_str, year_str = str_val.split("/")
            return date(int(year_str), int(month_str), 1)
        if "-" in str_val and len(str_val.split("-")) == 2:
            year_str, month_str = str_val.split("-")
            return date(int(year_str), int(month_str), 1)
        if "-" in str_val and len(str_val.split("-")) == 3:
            year_str, month_str, _ = str_val.split("-")
            return date(int(year_str), int(month_str), 1)
        return datetime.fromisoformat(str_val.replace("Z", "+00:00")).date().replace(day=1)
    except (ValueError, TypeError, AttributeError):
        return None


def parse_workbook(file_obj, sheet=None):
    """
    Разбирает Excel в список ParsedRow. Шапка обязана быть в первой строке —
    преамбулу и подвал с подписями пользователь убирает сам.
    """
    workbook = load_workbook(filename=file_obj, read_only=True, data_only=True)
    try:
        if sheet:
            if sheet not in workbook.sheetnames:
                raise ImportFileError(f"В книге нет листа «{sheet}». Доступны: {', '.join(workbook.sheetnames)}")
            worksheet = workbook[sheet]
        else:
            worksheet = workbook.worksheets[0]

        rows = worksheet.iter_rows(values_only=True)
        try:
            header_row = next(rows)
        except StopIteration:
            raise ImportFileError("Лист пустой — нет ни одной строки")

        columns = [key_of(h) for h in header_row]
        missing = REQUIRED_COLUMNS - {c for c in columns if c}
        if missing:
            titles = {
                "organization": "Организация",
                "city": "Город",
                "address": "Адрес",
                "manufacturer": "Производитель",
                "model": "Модель оборудования",
                "serial": "Серийный номер",
            }
            raise ImportFileError(
                "В первой строке файла не найдены обязательные колонки: "
                + ", ".join(sorted(titles[m] for m in missing))
                + ". Убедитесь, что шапка таблицы стоит первой строкой, без преамбулы над ней."
            )

        parsed = []
        for offset, row in enumerate(rows, start=2):
            values = {}
            for col_key, cell in zip(columns, row):
                if not col_key or col_key == "rownum":
                    continue
                values[col_key] = cell if col_key == "service_month" else norm(cell)

            if not any(norm(v) for v in values.values()):
                continue

            parsed.append(ParsedRow(row_number=offset, values=values))

        return parsed, worksheet.title
    finally:
        workbook.close()


class _Reference:
    """Справочники, загруженные один раз на файл: поиск без запросов на строку."""

    def __init__(self):
        self.organizations = {norm(o.name).lower(): o.id for o in Organization.objects.only("id", "name")}
        self.cities = {norm(c.name).lower(): c.id for c in City.objects.only("id", "name")}
        self.manufacturers = {norm(m.name).lower(): m.id for m in Manufacturer.objects.only("id", "name")}
        self.models = {}
        for model_id, manufacturer_id, name in DeviceModel.objects.values_list("id", "manufacturer_id", "name"):
            self.models[(manufacturer_id, norm(name).lower())] = model_id

        self.max_lengths = {
            "address": ContractDevice._meta.get_field("address").max_length,
            "room": ContractDevice._meta.get_field("room_number").max_length,
            "serial": ContractDevice._meta.get_field("serial_number").max_length,
        }


def _chunked(items, size):
    for start in range(0, len(items), size):
        yield items[start : start + size]


def _lookup_devices_by_serial(serials):
    """{sn_lower: [(device_id, organization_id), ...]} — один SN может быть в разных организациях."""
    found = {}
    for chunk in _chunked(serials, LOOKUP_CHUNK):
        qs = (
            ContractDevice.objects.annotate(sn=Lower("serial_number"))
            .filter(sn__in=chunk)
            .values_list("id", "organization_id", "sn")
        )
        for device_id, organization_id, sn in qs:
            found.setdefault(sn, []).append((device_id, organization_id))
    return found


def _lookup_printers_by_serial(serials):
    found = {}
    for chunk in _chunked(serials, LOOKUP_CHUNK):
        qs = Printer.objects.annotate(sn=Lower("serial_number")).filter(sn__in=chunk).values_list("id", "sn")
        for printer_id, sn in qs:
            found.setdefault(sn, printer_id)
    return found


def _resolve_row(parsed_row, refs, devices_by_sn, printers_by_sn):
    """Возвращает (resolved, errors, warnings, matched_device_id)."""
    values = parsed_row.values
    errors = []
    warnings = []
    resolved = {}

    serial = values.get("serial", "")
    sn_lower = serial.lower()
    if not serial:
        errors.append({"code": "NO_SERIAL", "message": "Не указан серийный номер"})
    elif len(serial) > refs.max_lengths["serial"]:
        errors.append({"code": "TOO_LONG", "message": f"Серийный номер длиннее {refs.max_lengths['serial']} символов"})

    organization_name = values.get("organization", "")
    organization_id = refs.organizations.get(organization_name.lower()) if organization_name else None
    if not organization_name:
        errors.append({"code": "NO_ORGANIZATION", "message": "Не указана организация"})
    elif organization_id is None:
        errors.append(
            {
                "code": "UNKNOWN_ORGANIZATION",
                "message": f"Организации «{organization_name}» нет в справочнике",
                "value": organization_name,
            }
        )
    resolved["organization_id"] = organization_id

    manufacturer_name = values.get("manufacturer", "")
    manufacturer_id = refs.manufacturers.get(manufacturer_name.lower()) if manufacturer_name else None
    if not manufacturer_name:
        errors.append({"code": "NO_MANUFACTURER", "message": "Не указан производитель"})
    elif manufacturer_id is None:
        errors.append(
            {
                "code": "UNKNOWN_MANUFACTURER",
                "message": f"Производителя «{manufacturer_name}» нет в справочнике",
                "value": manufacturer_name,
            }
        )
    resolved["manufacturer_id"] = manufacturer_id

    model_name = values.get("model", "")
    model_id = None
    if not model_name:
        errors.append({"code": "NO_MODEL", "message": "Не указана модель оборудования"})
    elif manufacturer_id is not None:
        model_id = refs.models.get((manufacturer_id, model_name.lower()))
        if model_id is None:
            errors.append(
                {
                    "code": "UNKNOWN_MODEL",
                    "message": f"Модели «{model_name}» нет в справочнике производителя «{manufacturer_name}»",
                    "value": model_name,
                }
            )
    resolved["model_id"] = model_id

    city_name = values.get("city", "")
    if not city_name:
        errors.append({"code": "NO_CITY", "message": "Не указан город"})
        resolved["city_id"] = None
    else:
        city_id = refs.cities.get(city_name.lower())
        resolved["city_id"] = city_id
        resolved["city_name"] = city_name
        if city_id is None:
            warnings.append(
                {"code": "NEW_CITY", "message": f"Города «{city_name}» нет в справочнике", "value": city_name}
            )

    address = values.get("address", "")
    if not address:
        errors.append({"code": "NO_ADDRESS", "message": "Не указан адрес"})
    elif len(address) > refs.max_lengths["address"]:
        errors.append({"code": "TOO_LONG", "message": f"Адрес длиннее {refs.max_lengths['address']} символов"})

    room = values.get("room", "")
    if len(room) > refs.max_lengths["room"]:
        errors.append({"code": "TOO_LONG", "message": f"№ кабинета длиннее {refs.max_lengths['room']} символов"})

    service_month = parse_service_month(values.get("service_month"))
    resolved["service_start_month"] = service_month.isoformat() if service_month else None
    if values.get("service_month") and service_month is None:
        warnings.append({"code": "BAD_SERVICE_MONTH", "message": "Не удалось разобрать месяц обслуживания"})

    printer_id = printers_by_sn.get(sn_lower) if sn_lower else None
    resolved["printer_id"] = printer_id
    if printer_id:
        warnings.append({"code": "WILL_LINK_PRINTER", "message": "Серийник найден среди опрашиваемых принтеров"})

    matched_device_id = None
    resolved["matched_same_org"] = False
    if sn_lower and organization_id is not None:
        candidates = devices_by_sn.get(sn_lower, [])
        same_org = [d for d, org in candidates if org == organization_id]
        if same_org:
            matched_device_id = same_org[0]
            resolved["matched_same_org"] = True
        elif candidates:
            matched_device_id = candidates[0][0]

    return resolved, errors, warnings, matched_device_id


def _base_classification(import_row):
    """Класс строки без учёта дублей внутри пачки — выводится из уже сохранённых данных."""
    if import_row.errors:
        return ImportRow.ERROR
    if import_row.matched_device_id is None:
        return ImportRow.NEW
    if import_row.resolved.get("matched_same_org"):
        return ImportRow.MATCH
    return ImportRow.MOVED


def analyze_file(session, file_obj, original_name, sheet=None):
    """
    Разбирает файл, создаёт ImportFile и ImportRow, пересчитывает дубли по всей сессии.
    Повторная загрузка файла с тем же именем заменяет его прежние строки.
    Возвращает созданный ImportFile.
    """
    parsed, sheet_title = parse_workbook(file_obj, sheet=sheet)

    refs = _Reference()
    serials = sorted({row.values.get("serial", "").lower() for row in parsed if row.values.get("serial")})
    devices_by_sn = _lookup_devices_by_serial(serials)
    printers_by_sn = _lookup_printers_by_serial(serials)

    name = original_name[:255]
    with transaction.atomic():
        session.files.filter(original_name=name).delete()
        import_file = ImportFile.objects.create(
            session=session,
            original_name=name,
            sheet_name=sheet_title[:255],
            rows_total=len(parsed),
        )

    pending = []
    for parsed_row in parsed:
        resolved, errors, warnings, matched_device_id = _resolve_row(parsed_row, refs, devices_by_sn, printers_by_sn)
        pending.append(
            ImportRow(
                session=session,
                file=import_file,
                row_number=parsed_row.row_number,
                raw=parsed_row.values,
                sn_lower=parsed_row.values.get("serial", "").lower()[:128],
                resolved=resolved,
                errors=errors,
                warnings=warnings,
                matched_device_id=matched_device_id,
            )
        )

    for chunk in _chunked(pending, ANALYZE_BATCH):
        ImportRow.objects.bulk_create(chunk)

    recompute_session(session)
    return import_file


def recompute_session(session):
    """
    Пересчитывает классификацию по всей сессии: дубли серийников ищутся сквозь все файлы,
    поэтому после добавления каждого файла класс строк из предыдущих может измениться.
    """
    rows = list(session.rows.all())

    counts = {}
    for row in rows:
        if row.errors or not row.sn_lower:
            continue
        counts[row.sn_lower] = counts.get(row.sn_lower, 0) + 1

    changed = []
    for row in rows:
        classification = _base_classification(row)
        if classification != ImportRow.ERROR and counts.get(row.sn_lower, 0) > 1:
            classification = ImportRow.DUP_IN_FILE
        if row.classification != classification:
            row.classification = classification
            changed.append(row)

    for chunk in _chunked(changed, ANALYZE_BATCH):
        ImportRow.objects.bulk_update(chunk, ["classification"])


def session_summary(session):
    """Сводка по сессии для превью."""
    counts = {key: 0 for key, _ in ImportRow.CLASSIFICATION_CHOICES}
    unknown_organizations = set()
    unknown_models = set()
    new_cities = set()
    pending_conflicts = 0

    for row in session.rows.all().only("classification", "decision", "errors", "warnings"):
        counts[row.classification] = counts.get(row.classification, 0) + 1
        if row.classification in CONFLICT_CLASSES and row.decision == ImportRow.PENDING:
            pending_conflicts += 1
        for err in row.errors:
            if err.get("code") == "UNKNOWN_ORGANIZATION":
                unknown_organizations.add(err.get("value", ""))
            elif err.get("code") == "UNKNOWN_MODEL":
                unknown_models.add(err.get("value", ""))
        for warn in row.warnings:
            if warn.get("code") == "NEW_CITY":
                new_cities.add(warn.get("value", ""))

    return {
        "counts": counts,
        "total": sum(counts.values()),
        "pending_conflicts": pending_conflicts,
        "unknown_organizations": sorted(unknown_organizations),
        "unknown_models": sorted(unknown_models),
        "new_cities": sorted(new_cities),
    }


# ─── Применение ───────────────────────────────────────────────────────────────


def rows_to_apply(session):
    """
    Строки, которые уйдут в БД: без ошибок, не пропущенные пользователем.
    Конфликтные классы требуют явного «применить».
    """
    return (
        session.rows.exclude(classification=ImportRow.ERROR)
        .exclude(decision=ImportRow.SKIP)
        .exclude(classification__in=CONFLICT_CLASSES, decision=ImportRow.PENDING)
    )


def _ensure_cities(rows):
    """Создаёт недостающие города и возвращает {lower(name): id} по всем городам."""
    existing = {norm(c.name).lower(): c.id for c in City.objects.only("id", "name")}

    wanted = {}
    for row in rows:
        name = row.resolved.get("city_name") or ""
        if name and name.lower() not in existing:
            wanted[name.lower()] = name

    for lowered, name in wanted.items():
        city, _ = City.objects.get_or_create(name=name)
        existing[lowered] = city.id

    return existing


def _busy_printers(rows):
    """{printer_id: contract_device_id} — принтеры, уже занятые связью 1:1."""
    printer_ids = {row.resolved.get("printer_id") for row in rows if row.resolved.get("printer_id")}
    if not printer_ids:
        return {}
    return dict(ContractDevice.objects.filter(printer_id__in=printer_ids).values_list("printer_id", "id"))


def _device_fields(row, session, cities, printer_id):
    raw = row.raw
    service_month = row.resolved.get("service_start_month")
    fields = {
        "organization_id": row.resolved["organization_id"],
        "city_id": cities[(row.resolved.get("city_name") or "").lower()],
        "address": raw.get("address", ""),
        "room_number": raw.get("room", ""),
        "model_id": row.resolved["model_id"],
        "serial_number": raw.get("serial", ""),
        "status": session.target_status,
        "service_start_month": date.fromisoformat(service_month) if service_month else None,
        "printer_id": printer_id,
    }
    # Подрядчик не задан — не трогаем уже проставленный у существующего устройства
    if session.service_provider_id:
        fields["service_provider_id"] = session.service_provider_id
    if session.contract_id:
        fields["contract_id"] = session.contract_id
    return fields


def _apply_row(row, session, cities, busy_printers, content_type, user, logs):
    """Создаёт или обновляет одно устройство. Комментарий не трогаем — он ведётся руками."""
    printer_id = row.resolved.get("printer_id")
    if printer_id and busy_printers.get(printer_id) not in (None, row.matched_device_id):
        printer_id = None

    fields = _device_fields(row, session, cities, printer_id)
    device = row.matched_device if row.matched_device_id else None

    if device is None:
        device = ContractDevice.objects.create(**fields)
        logs.append(
            EntityChangeLog(
                content_type=content_type,
                object_id=device.id,
                action="create",
                user=user,
                object_repr=str(device)[:500],
            )
        )
    else:
        changes = {}
        for name, value in fields.items():
            old = getattr(device, name)
            if old != value:
                changes[name] = {"old": str(old) if old is not None else None, "new": str(value) if value else None}
                setattr(device, name, value)
        if changes:
            device.save(update_fields=[*fields.keys(), "updated_at"])
            logs.append(
                EntityChangeLog(
                    content_type=content_type,
                    object_id=device.id,
                    action="update",
                    user=user,
                    changes=changes,
                    object_repr=str(device)[:500],
                )
            )

    if printer_id:
        busy_printers[printer_id] = device.id

    row.applied_device = device
    row.apply_error = ""
    return device


def apply_session(session, user=None, create_cities=False):
    """
    Применяет решения по сессии. Чанками, чтобы одна битая строка не откатывала всю загрузку
    и чтобы не держать блокировки на 3000 строк.
    """
    rows = list(rows_to_apply(session).select_related("matched_device"))

    missing_cities = [r for r in rows if r.resolved.get("city_id") is None]
    if missing_cities and not create_cities:
        raise ImportFileError(
            "В файлах есть города, которых нет в справочнике. "
            "Подтвердите их создание или приведите названия к справочнику."
        )

    cities = _ensure_cities(rows) if rows else {}
    busy_printers = _busy_printers(rows)
    content_type = ContentType.objects.get_for_model(ContractDevice)

    created = updated = failed = 0
    errors = []

    for chunk in _chunked(rows, APPLY_CHUNK):
        printers_before = dict(busy_printers)
        try:
            with transaction.atomic():
                logs = []
                chunk_created = chunk_updated = 0
                for row in chunk:
                    is_new = row.matched_device_id is None
                    _apply_row(row, session, cities, busy_printers, content_type, user, logs)
                    chunk_created, chunk_updated = (
                        (chunk_created + 1, chunk_updated) if is_new else (chunk_created, chunk_updated + 1)
                    )
                EntityChangeLog.objects.bulk_create(logs)
                ImportRow.objects.bulk_update(chunk, ["applied_device", "apply_error"])
            created += chunk_created
            updated += chunk_updated
        except IntegrityError:
            # Чанк откатился целиком — повторяем построчно, чтобы потерять только виноватые строки
            busy_printers = printers_before
            chunk_created, chunk_updated, chunk_failed, chunk_errors = _apply_chunk_row_by_row(
                chunk, session, cities, busy_printers, content_type, user
            )
            created += chunk_created
            updated += chunk_updated
            failed += chunk_failed
            errors.extend(chunk_errors)

    session.stats = {"created": created, "updated": updated, "failed": failed, "total": len(rows)}
    session.state = ImportSession.APPLIED
    session.applied_at = timezone.now()
    session.save(update_fields=["stats", "state", "applied_at"])

    return {**session.stats, "errors": errors}


def _apply_chunk_row_by_row(chunk, session, cities, busy_printers, content_type, user):
    created = updated = failed = 0
    errors = []
    for row in chunk:
        is_new = row.matched_device_id is None
        try:
            with transaction.atomic():
                logs = []
                _apply_row(row, session, cities, busy_printers, content_type, user, logs)
                EntityChangeLog.objects.bulk_create(logs)
                row.save(update_fields=["applied_device", "apply_error"])
            created, updated = (created + 1, updated) if is_new else (created, updated + 1)
        except (IntegrityError, ValueError) as exc:
            failed += 1
            row.applied_device = None
            row.apply_error = f"{exc.__class__.__name__}: {exc}"
            row.save(update_fields=["applied_device", "apply_error"])
            errors.append({"row_id": row.id, "row_number": row.row_number, "error": row.apply_error})
    return created, updated, failed, errors


def find_missing_devices(session):
    """
    Устройства организаций, затронутых сессией, которые не попали ни в один файл.
    Импорт их не трогает — решение об удалении принимает пользователь.
    """
    organization_ids = {
        row.resolved.get("organization_id")
        for row in session.rows.exclude(classification=ImportRow.ERROR).only("resolved")
        if row.resolved.get("organization_id")
    }
    if not organization_ids:
        return ContractDevice.objects.none()

    applied = session.rows.filter(applied_device__isnull=False).values("applied_device")
    return (
        ContractDevice.objects.filter(organization_id__in=organization_ids)
        .exclude(id__in=applied)
        .select_related("organization", "city", "model", "model__manufacturer", "status")
        .order_by("organization__name", "city__name", "address")
    )
