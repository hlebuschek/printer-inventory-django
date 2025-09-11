"""
Импорт устройств по договору из Excel в БД.

Ожидаемая структура листа (строка заголовков):
№ | Организация | Город | Адрес | № кабинета | Производитель | Модель оборудования | Серийный номер | Месяц обслуживания | Статус

Пример запуска:
    python manage.py import_contracts_xlsx contracts.xlsx
    python manage.py import_contracts_xlsx contracts.xlsx --sheet "Лист1" --truncate
    python manage.py import_contracts_xlsx contracts.xlsx --no-merge-empty-sn
    python manage.py import_contracts_xlsx contracts.xlsx --dry-run
"""

import re
from pathlib import Path
from collections import namedtuple
from datetime import datetime, date

from django.core.management.base import BaseCommand, CommandError
from openpyxl import load_workbook
from django.db import transaction, IntegrityError

from inventory.models import Organization, Printer
from contracts.models import City, Manufacturer, DeviceModel, ContractDevice, ContractStatus


# Заголовки из Excel → внутренние ключи
HEADERS = {
    "№": "rownum",
    "номер": "rownum",

    "организация": "organization",
    "организация, наименование": "organization",

    "город": "city",

    "адрес": "address",

    "№ кабинета": "room",
    "кабинет": "room",
    "номер кабинета": "room",

    "производитель": "manufacturer",
    "vendor": "manufacturer",
    "бренд": "manufacturer",

    "модель оборудования": "model",
    "модель": "model",

    "серийный номер": "serial",
    "sn": "serial",
    "s/n": "serial",
    "serial": "serial",

    "месяц обслуживания": "service_month",
    "месяц принятия на обслуживание": "service_month",
    "дата принятия": "service_month",
    "начало обслуживания": "service_month",

    "статус": "status",
    "status": "status",

    "комментарий": "comment",
}


def norm(val):
    """None -> '', str -> strip + схлопывание пробелов, числа -> str без '.0'."""
    if val is None:
        return ""
    if isinstance(val, (int, float)):
        s = str(int(val)) if float(val).is_integer() else str(val)
    else:
        s = str(val)
    s = s.strip()
    return re.sub(r"\s+", " ", s)


def key_of(header):
    if not header:
        return None
    return HEADERS.get(norm(header).lower())


def parse_service_month(value):
    """
    Парсит месяц обслуживания из различных форматов:
    - MM.YYYY (например: 01.2024)
    - MM/YYYY (например: 01/2024)
    - YYYY-MM (например: 2024-01)
    - datetime объект
    - строка в формате ISO

    Возвращает datetime.date с первым числом месяца или None
    """
    if not value:
        return None

    # Если это уже datetime или date объект
    if isinstance(value, (datetime, date)):
        if isinstance(value, datetime):
            return value.date().replace(day=1)
        return value.replace(day=1)

    # Преобразуем в строку и очищаем
    str_val = norm(value)
    if not str_val:
        return None

    try:
        # Формат MM.YYYY
        if '.' in str_val and len(str_val.split('.')) == 2:
            month_str, year_str = str_val.split('.')
            month, year = int(month_str), int(year_str)
            return date(year, month, 1)

        # Формат MM/YYYY
        elif '/' in str_val and len(str_val.split('/')) == 2:
            month_str, year_str = str_val.split('/')
            month, year = int(month_str), int(year_str)
            return date(year, month, 1)

        # Формат YYYY-MM
        elif '-' in str_val and len(str_val.split('-')) == 2:
            year_str, month_str = str_val.split('-')
            month, year = int(month_str), int(year_str)
            return date(year, month, 1)

        # Формат YYYY-MM-DD (берем только год и месяц)
        elif '-' in str_val and len(str_val.split('-')) == 3:
            year_str, month_str, _ = str_val.split('-')
            month, year = int(month_str), int(year_str)
            return date(year, month, 1)

        # Попытка парсинга как ISO дата
        else:
            parsed_date = datetime.fromisoformat(str_val.replace('Z', '+00:00'))
            return parsed_date.date().replace(day=1)

    except (ValueError, TypeError, AttributeError):
        # Если не удалось распарсить, возвращаем None
        return None

    return None


def ci_get_or_create(model, name_value, name_field="name", **extra_filters):
    """
    Регистронезависимый get_or_create по текстовому полю `name_field`.
    Корректно работает с UniqueConstraint(Lower(name)).
    """
    q = {f"{name_field}__iexact": name_value, **extra_filters}
    obj = model.objects.filter(**q).first()
    if obj:
        return obj, False
    try:
        obj = model.objects.create(**{name_field: name_value, **extra_filters})
        return obj, True
    except IntegrityError:
        obj = model.objects.filter(**q).first()
        if obj:
            return obj, False
        raise


BadRow = namedtuple("BadRow", "row reason preview")


class Command(BaseCommand):
    help = "Импорт устройств по договору из Excel (см. docstring в файле)"

    def add_arguments(self, parser):
        parser.add_argument("xlsx_path", type=str, help="Путь к .xlsx файлу")
        parser.add_argument("--sheet", type=str, default=None, help="Имя листа (по умолчанию первый)")
        parser.add_argument("--dry-run", action="store_true", help="Не сохранять, только показать, что будет сделано")
        parser.add_argument(
            "--truncate",
            action="store_true",
            help="Если значение длиннее лимита поля модели — обрезать вместо ошибки",
        )
        parser.add_argument(
            "--no-merge-empty-sn",
            action="store_true",
            help="Не сливать строки с пустым SN по адресу/кабинету/модели — всегда создавать новую запись",
        )

    @transaction.atomic
    def handle(self, *args, **opts):
        # --- Открываем Excel ---
        path = Path(opts["xlsx_path"])
        if not path.exists():
            raise CommandError(f"Файл не найден: {path}")

        wb = load_workbook(filename=str(path), read_only=True, data_only=True)
        ws = wb[opts["sheet"]] if opts["sheet"] else wb.worksheets[0]

        # --- Лимиты полей моделей ---
        ROOM_MAX = ContractDevice._meta.get_field("room_number").max_length
        SN_MAX = ContractDevice._meta.get_field("serial_number").max_length
        ADDR_MAX = ContractDevice._meta.get_field("address").max_length
        STAT_MAX = ContractStatus._meta.get_field("name").max_length
        CITY_MAX = City._meta.get_field("name").max_length
        MFR_MAX = Manufacturer._meta.get_field("name").max_length
        DMOD_MAX = DeviceModel._meta.get_field("name").max_length

        def clip(value, maxlen, label):
            """Обрезает по maxlen или кидает CommandError (если нет --truncate)."""
            if value and len(value) > maxlen:
                if opts.get("truncate", False):
                    self.stdout.write(self.style.WARNING(
                        f"[TRUNCATE] {label}: '{value[:50]}...' ({len(value)} симв.) → {maxlen}"
                    ))
                    return value[:maxlen]
                raise CommandError(
                    f"[ДЛИНА] {label}: '{value[:120]}' ({len(value)} симв.) превышает лимит {maxlen}. "
                    f"Запусти с --truncate или поправь данные."
                )
            return value

        # --- Заголовки ---
        try:
            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        except StopIteration:
            raise CommandError("Пустой лист: нет строк с данными")

        columns = [key_of(h) for h in header_row]

        required = {"organization", "city", "address", "manufacturer", "model", "status"}
        have = {c for c in columns if c}
        if not required.issubset(have):
            missing = required - have
            raise CommandError(
                "В файле отсутствуют обязательные колонки: " + ", ".join(sorted(missing))
            )

        # --- Счётчики и статистика ---
        created = 0
        updated_rows = 0
        updated_ids = set()
        dup_row_updates = 0
        skipped = 0
        failed = 0
        blank_rows = 0
        total_rows = 0
        bad_rows = []

        initial_total = ContractDevice.objects.count()

        def row_preview(d):
            service_month_display = ""
            if d.get('service_month'):
                parsed_month = parse_service_month(d['service_month'])
                if parsed_month:
                    service_month_display = f" • {parsed_month.strftime('%m.%Y')}"

            return (
                f"{(d.get('organization') or '—')} • {(d.get('city') or '—')} • "
                f"{(d.get('address') or '—')} • {(d.get('room') or '—')} • "
                f"{(d.get('manufacturer') or '—')} {(d.get('model') or '—')} • "
                f"SN:{(d.get('serial') or '—')}{service_month_display} • {(d.get('status') or '—')}"
            )

        # --- Основной проход по строкам ---
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            total_rows += 1
            data = {}
            for col_key, cell in zip(columns, row):
                if not col_key:
                    continue
                data[col_key] = norm(cell) if col_key != 'service_month' else cell

            # Полностью пустая строка — пропускаем
            if not any(str(v).strip() if v is not None else "" for v in data.values()):
                blank_rows += 1
                continue

            # Валидация минимального набора
            missing_vals = [k for k in ("organization", "address", "manufacturer", "model") if not data.get(k)]
            if missing_vals:
                skipped += 1
                bad_rows.append(BadRow(idx, f"MISSING_VALUES: {', '.join(missing_vals)}", row_preview(data)))
                failed += 1
                continue

            try:
                # Нормализация с учётом лимитов
                org_name = data["organization"]
                city_name = clip(data.get("city") or "—", CITY_MAX, "Город")
                mfr_name = clip(data["manufacturer"], MFR_MAX, "Производитель")
                model_nm = clip(data["model"], DMOD_MAX, "Модель оборудования")
                status_nm = clip(data.get("status") or "—", STAT_MAX, "Статус")
                serial = clip(data.get("serial") or "", SN_MAX, "Серийный номер")
                room = clip(data.get("room") or "", ROOM_MAX, "№ кабинета")
                address = clip(data["address"], ADDR_MAX, "Адрес")
                comment = data.get("comment") or ""

                # Парсинг месяца обслуживания
                service_start_month = None
                if data.get("service_month"):
                    service_start_month = parse_service_month(data["service_month"])
                    if service_start_month is None and data["service_month"]:
                        self.stdout.write(self.style.WARNING(
                            f"[row {idx}] Не удалось распарсить месяц обслуживания: '{data['service_month']}'"
                        ))

                # Справочники (CI)
                org, _ = ci_get_or_create(Organization, org_name)
                city, _ = ci_get_or_create(City, city_name)
                mfr, _ = ci_get_or_create(Manufacturer, mfr_name)
                dev_model, _ = DeviceModel.objects.get_or_create(manufacturer=mfr, name=model_nm)
                status, _ = ci_get_or_create(ContractStatus, status_nm)

                # Поиск существующей записи
                obj = None
                if serial:
                    obj = ContractDevice.objects.filter(
                        organization=org, serial_number__iexact=serial
                    ).first()
                if not obj and not opts.get("no_merge_empty_sn", False):
                    obj = ContractDevice.objects.filter(
                        organization=org,
                        address__iexact=address,
                        room_number__iexact=room,
                        model=dev_model,
                    ).first()

                # Автопривязка к Printer по SN
                printer = Printer.objects.filter(serial_number__iexact=serial).first() if serial else None

                fields = dict(
                    organization=org,
                    city=city,
                    address=address,
                    room_number=room,
                    model=dev_model,
                    serial_number=serial,
                    service_start_month=service_start_month,
                    status=status,
                    comment=comment,
                    printer=printer or None,
                )

                if obj:
                    # UPDATE
                    for k, v in fields.items():
                        setattr(obj, k, v)
                    if not opts.get("dry_run", False):
                        try:
                            with transaction.atomic():
                                obj.save()
                        except IntegrityError as e:
                            bad_rows.append(
                                BadRow(idx, f"INTEGRITY_ERROR on save: {e.__class__.__name__}: {e}", row_preview(data))
                            )
                            failed += 1
                            continue
                    updated_rows += 1
                    if obj.id in updated_ids:
                        dup_row_updates += 1
                    else:
                        updated_ids.add(obj.id)
                else:
                    # CREATE
                    if not opts.get("dry_run", False):
                        try:
                            with transaction.atomic():
                                obj = ContractDevice.objects.create(**fields)
                        except IntegrityError as e:
                            bad_rows.append(
                                BadRow(idx, f"INTEGRITY_ERROR on create: {e.__class__.__name__}: {e}", row_preview(data))
                            )
                            failed += 1
                            continue
                    created += 1

            except CommandError as e:
                bad_rows.append(BadRow(idx, f"LENGTH_ERROR: {e}", row_preview(data)))
                failed += 1
                continue
            except Exception as e:
                bad_rows.append(BadRow(idx, f"UNEXPECTED: {e.__class__.__name__}: {e}", row_preview(data)))
                failed += 1
                continue

        # --- Отчёт по проблемным строкам ---
        if bad_rows:
            self.stdout.write(self.style.WARNING("\nСтроки, которые НЕ были импортированы:"))
            for br in bad_rows:
                self.stdout.write(self.style.WARNING(f"[row {br.row}] {br.reason} | {br.preview}"))

        unique_updated = len(updated_ids)

        # --- Итоги и коммит/роллбек ---
        if opts.get("dry_run", False):
            final_total = ContractDevice.objects.count()
            transaction.set_rollback(True)
        else:
            final_total = ContractDevice.objects.count()

        msg = (
            "Импорт завершён: "
            f"создано {created}, "
            f"обновлено (строк) {updated_rows}, "
            f"обновлено (уникально) {unique_updated}, "
            f"повторных обновлений {dup_row_updates}, "
            f"пропущено {skipped}, "
            f"ошибок {failed}, "
            f"пустых строк {blank_rows}. "
            f"(строк обработано: {total_rows}; было в БД: {initial_total}, стало: {final_total})"
        )
        if opts.get("dry_run", False):
            msg += " (dry-run)"
        self.stdout.write(self.style.SUCCESS(msg))