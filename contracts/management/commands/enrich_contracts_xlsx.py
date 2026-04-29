"""
Обогащение Excel со списком устройств: стандартизация производителя/модели
по справочнику contracts (Manufacturer/DeviceModel) + проверка серийников в GLPI
+ расчёт среднего количества страниц/мес.

За один проход читает исходный файл и пишет новый xlsx с добавленными колонками:
    - Производитель (эталон)
    - Модель (эталон)
    - Статус сопоставления
    - GLPI: статус
    - GLPI: кол-во карточек
    - GLPI: ID карточек
    - GLPI: по какому SN найдено
    - Среднее страниц/мес
    - Источник среднего

Пример:
    python manage.py enrich_contracts_xlsx "список финал.xlsx"
    python manage.py enrich_contracts_xlsx "список финал.xlsx" --output out.xlsx
    python manage.py enrich_contracts_xlsx "список финал.xlsx" --skip-glpi --skip-average
    python manage.py enrich_contracts_xlsx "список финал.xlsx" --start-row 500 --limit 1000
"""

import re
from datetime import date, datetime
from difflib import SequenceMatcher, get_close_matches
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import requests
from openpyxl import Workbook, load_workbook

from django.conf import settings
from django.core.management.base import BaseCommand, CommandError
from django.db.models import Avg, Count

from contracts.models import DeviceModel, Manufacturer
from integrations.glpi.client import GLPIAPIError, GLPIClient
from monthly_report.models import MonthlyReport


# ──────────────── Стандартизация: алиасы и нормализация ──────────────────────

MANUFACTURER_ALIASES: Dict[str, List[str]] = {
    "hp": ["hp", "hewlett"],
    "h.p.": ["hp", "hewlett"],
    "hewlett packard": ["hewlett", "hp"],
    "hewlett-packard": ["hewlett", "hp"],
    "hewlett packard company": ["hewlett", "hp"],
    "hewlett-packard company": ["hewlett", "hp"],
    "hewlettpackard": ["hewlett", "hp"],
    "kyocera": ["kyocera"],
    "kyocera corporation": ["kyocera"],
    "kyocera mita": ["kyocera"],
    "kyocera mita corporation": ["kyocera"],
    "xerox": ["xerox"],
    "xerox corporation": ["xerox"],
    "fuji xerox": ["xerox"],
    "canon": ["canon"],
    "canon inc": ["canon"],
    "canon inc.": ["canon"],
    "canon k.k.": ["canon"],
    "ricoh": ["ricoh"],
    "ricoh company ltd": ["ricoh"],
    "brother": ["brother"],
    "brother industries": ["brother"],
    "brother industries ltd": ["brother"],
    "samsung": ["samsung"],
    "samsung electronics": ["samsung"],
    "samsung electronics co ltd": ["samsung"],
    "epson": ["epson"],
    "seiko epson": ["epson"],
    "seiko epson corp": ["epson"],
    "seiko epson corporation": ["epson"],
    "konica minolta": ["konica"],
    "konica minolta inc": ["konica"],
    "lexmark": ["lexmark"],
    "lexmark international": ["lexmark"],
    "oki": ["oki"],
    "oki data": ["oki"],
    "oki electric": ["oki"],
    "pantum": ["pantum"],
    "sharp": ["sharp"],
    "sharp corporation": ["sharp"],
    "toshiba": ["toshiba"],
    "toshiba tec": ["toshiba"],
    "dell": ["dell"],
    "avision": ["avision"],
    "sindoh": ["sindoh"],
    "катюша": ["катюша"],
    "katusha": ["катюша"],
    "katusha it": ["катюша"],
}

CORP_NOISE = re.compile(
    r"\b("
    r"corporation|corp\.?|company|co\.?|inc\.?|ltd\.?|limited|gmbh|sa|s\.a\.|kk|k\.k\.|ag|llc|plc|s\.l\."
    r")\b",
    flags=re.IGNORECASE,
)
MODEL_NOISE_PATTERNS = [
    re.compile(r"\bv\s*\d+(\s*\.\s*\d+)*\b", flags=re.IGNORECASE),
    re.compile(r"-\s*printer\s*$", flags=re.IGNORECASE),
    re.compile(r"\bseries\b", flags=re.IGNORECASE),
]


def norm_space(s) -> str:
    if s is None:
        return ""
    if isinstance(s, (int, float)):
        s = str(int(s)) if float(s).is_integer() else str(s)
    s = str(s).strip()
    return re.sub(r"\s+", " ", s)


def normalize_mfr_string(raw: str) -> str:
    s = norm_space(raw).lower()
    s = CORP_NOISE.sub("", s)
    s = re.sub(r"[,\.]", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def normalize_model_string(raw: str) -> str:
    s = norm_space(raw)
    for p in MODEL_NOISE_PATTERNS:
        s = p.sub("", s)
    s = re.sub(r"\s*\.\s*", ".", s)
    return re.sub(r"\s+", " ", s).strip(" -.,")


def _similarity(a: str, b: str) -> float:
    return SequenceMatcher(None, a.lower(), b.lower()).ratio()


class ManufacturerResolver:
    def __init__(self) -> None:
        self.db: List[Manufacturer] = list(Manufacturer.objects.all())
        self._by_lower: Dict[str, Manufacturer] = {m.name.lower(): m for m in self.db}
        self._names_lower: List[str] = list(self._by_lower.keys())

    def resolve(self, raw: str) -> Tuple[Optional[Manufacturer], str]:
        if not raw:
            return None, "not_found"
        raw_norm = normalize_mfr_string(raw)

        if raw_norm and raw_norm in self._by_lower:
            return self._by_lower[raw_norm], "exact"

        alias_keys = MANUFACTURER_ALIASES.get(raw_norm)
        if alias_keys:
            for key in alias_keys:
                for name_lower, mfr in self._by_lower.items():
                    if key in name_lower:
                        return mfr, "alias"
            return None, "alias_no_db"

        for name_lower, mfr in self._by_lower.items():
            if name_lower and (name_lower in raw_norm or raw_norm in name_lower):
                return mfr, "partial"

        if raw_norm and self._names_lower:
            close = get_close_matches(raw_norm, self._names_lower, n=1, cutoff=0.75)
            if close:
                return self._by_lower[close[0]], "fuzzy"

        return None, "not_found"


class ModelResolver:
    def __init__(self) -> None:
        self._cache: Dict[int, List[DeviceModel]] = {}

    def _models_for(self, mfr: Manufacturer) -> List[DeviceModel]:
        if mfr.id not in self._cache:
            self._cache[mfr.id] = list(DeviceModel.objects.filter(manufacturer=mfr))
        return self._cache[mfr.id]

    def resolve(
        self, raw_model: str, mfr: Optional[Manufacturer], model_cutoff: float
    ) -> Tuple[Optional[DeviceModel], str]:
        if not mfr:
            return None, "no_mfr"
        if not raw_model:
            return None, "empty"

        candidates = self._models_for(mfr)
        if not candidates:
            return None, "no_catalog"

        cleaned = normalize_model_string(raw_model)
        cleaned_lower = cleaned.lower()

        mfr_tokens = mfr.name.lower().split()
        stripped = cleaned_lower
        for tok in mfr_tokens:
            stripped = re.sub(rf"^\s*{re.escape(tok)}\s+", "", stripped)

        for dm in candidates:
            if dm.name.lower() in (cleaned_lower, stripped):
                return dm, "exact"

        substring_matches: List[Tuple[DeviceModel, int]] = []
        for dm in candidates:
            dmn = dm.name.lower()
            if dmn and (dmn in cleaned_lower or dmn in stripped):
                substring_matches.append((dm, len(dmn)))
        if substring_matches:
            substring_matches.sort(key=lambda p: p[1], reverse=True)
            return substring_matches[0][0], "partial"

        best_dm: Optional[DeviceModel] = None
        best_ratio = 0.0
        for dm in candidates:
            for src in (cleaned_lower, stripped):
                r = _similarity(src, dm.name)
                if r > best_ratio:
                    best_ratio = r
                    best_dm = dm
        if best_dm and best_ratio >= model_cutoff:
            return best_dm, f"fuzzy({best_ratio:.2f})"
        return None, "not_found"


# ──────────────── GLPI ─────────────────────────────────────────────────────────

SERIAL_MAIN_KEYS = ("серийный номер",)
SERIAL_LABEL_KEYS = ("на бирке",)


def _find_col(header_lower: List[str], keys: Tuple[str, ...], exclude: Tuple[str, ...] = ()) -> Optional[int]:
    for i, h in enumerate(header_lower):
        if any(k in h for k in keys) and not any(x in h for x in exclude):
            return i
    return None


def _extract_ids(items: List[Dict]) -> List[str]:
    ids: List[str] = []
    for it in items or []:
        val = it.get("2") if isinstance(it, dict) else None
        if val is None and isinstance(it, dict):
            val = it.get("id")
        if val is not None:
            ids.append(str(val))
    return ids


# ──────────────── Среднее страниц/мес ──────────────────────────────────────────


def _parse_date(raw) -> Optional[date]:
    if raw is None:
        return None
    s = str(raw).strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            return datetime.strptime(s[: len(fmt) + 2], fmt).date()
        except ValueError:
            continue
    return None


def _months_between(d1: date, d2: date) -> float:
    if d2 < d1:
        d1, d2 = d2, d1
    months = (d2.year - d1.year) * 12 + (d2.month - d1.month) + (d2.day - d1.day) / 30.44
    return max(months, 0.0)


def _avg_from_points(points: List[Tuple[date, int]]) -> Optional[float]:
    """points — [(date, counter)]; считаем (last - first) / месяцев_между."""
    if len(points) < 2:
        return None
    points = sorted(points, key=lambda p: p[0])
    first_date, first_val = points[0]
    last_date, last_val = points[-1]
    span = _months_between(first_date, last_date)
    delta = last_val - first_val
    if span <= 0 or delta <= 0:
        return None
    return delta / span


def _avg_from_monthly_report(serial: str) -> Optional[Tuple[float, int]]:
    """Возвращает (avg_pages_per_month, months_count) или None."""
    if not serial:
        return None
    agg = MonthlyReport.objects.filter(serial_number__iexact=serial).aggregate(
        avg=Avg("total_prints"), cnt=Count("id")
    )
    cnt = agg["cnt"] or 0
    avg = agg["avg"]
    if cnt >= 1 and avg is not None:
        return (float(avg), int(cnt))
    return None


def _avg_from_printer_log(client: GLPIClient, printer_id: int) -> Optional[Tuple[float, int]]:
    """Считает среднее по истории SNMP-счётчика (PrinterLog).

    Если записей мало, дополняет данными из detail: date_creation как начальная
    точка (counter=0) и last_pages_counter + сегодня как конечная.
    """
    try:
        resp = requests.get(
            f"{client.url}/Printer/{printer_id}/PrinterLog/",
            headers=client._get_headers(with_session=True),
            params={"range": "0-9999", "sort": "date", "order": "ASC"},
            timeout=15,
            verify=client.verify_ssl,
        )
    except requests.RequestException:
        return None
    if resp.status_code not in (200, 206):
        return None
    rows = resp.json() or []
    points: List[Tuple[date, int]] = []
    for r in rows:
        if not isinstance(r, dict):
            continue
        d = _parse_date(r.get("date") or r.get("date_creation") or r.get("date_mod"))
        val = r.get("total_pages")
        try:
            ival = int(val) if val is not None else None
        except (ValueError, TypeError):
            ival = None
        if d and ival is not None:
            points.append((d, ival))

    # Дополняем из карточки принтера, если точек мало
    if len(points) < 2:
        try:
            detail_resp = requests.get(
                f"{client.url}/Printer/{printer_id}",
                headers=client._get_headers(with_session=True),
                timeout=15,
                verify=client.verify_ssl,
            )
            if detail_resp.status_code == 200:
                detail = detail_resp.json() or {}
                created = _parse_date(detail.get("date_creation"))
                counter = detail.get("last_pages_counter")
                try:
                    counter = int(counter) if counter is not None else None
                except (ValueError, TypeError):
                    counter = None
                if created and counter and counter > 0:
                    today = date.today()
                    # Добавляем текущий счётчик (на сегодня) если ещё нет
                    if not any(p[0] == today for p in points):
                        points.append((today, counter))
                    # Добавляем стартовую точку (date_creation, 0) если нет более ранней
                    if not any(p[0] <= created for p in points):
                        points.append((created, 0))
        except requests.RequestException:
            pass

    avg = _avg_from_points(points)
    if avg is None:
        return None
    return (avg, len(points))


def _avg_from_change_log(
    client: GLPIClient, printer_id: int, counter_field_id: int
) -> Optional[Tuple[float, int]]:
    """Считает среднее по журналу изменений last_pages_counter (/Printer/{id}/Log/).

    Если Log-записей несколько — считаем (last-first)/months_between по точкам
    (date_mod, new_value). Если всего одна запись, но в ней есть old_value —
    используем date_creation принтера как начальную точку для old_value.
    """
    try:
        resp = requests.get(
            f"{client.url}/Printer/{printer_id}/Log/",
            headers=client._get_headers(with_session=True),
            params={"range": "0-9999"},
            timeout=15,
            verify=client.verify_ssl,
        )
    except requests.RequestException:
        return None
    if resp.status_code not in (200, 206):
        return None
    rows = resp.json() or []

    points: List[Tuple[date, int]] = []
    earliest_old: Optional[Tuple[date, int]] = None
    for r in rows:
        if not isinstance(r, dict):
            continue
        if r.get("id_search_option") != counter_field_id:
            continue
        d = _parse_date(r.get("date_mod"))
        if not d:
            continue
        try:
            new_val = int(str(r.get("new_value")).strip())
        except (ValueError, TypeError, AttributeError):
            new_val = None
        try:
            old_val = int(str(r.get("old_value")).strip())
        except (ValueError, TypeError, AttributeError):
            old_val = None
        if new_val is not None:
            points.append((d, new_val))
        if old_val is not None and (earliest_old is None or d < earliest_old[0]):
            earliest_old = (d, old_val)

    # Мало точек — подкладываем old_value самой ранней записи в качестве
    # стартовой точки, датой берём date_creation принтера.
    if len(points) < 2 and earliest_old:
        try:
            detail_resp = requests.get(
                f"{client.url}/Printer/{printer_id}",
                headers=client._get_headers(with_session=True),
                timeout=15,
                verify=client.verify_ssl,
            )
            if detail_resp.status_code == 200:
                detail = detail_resp.json() or {}
                created = _parse_date(detail.get("date_creation") or detail.get("date_mod"))
                if created and created < earliest_old[0]:
                    points.append((created, earliest_old[1]))
        except requests.RequestException:
            pass

    avg = _avg_from_points(points)
    if avg is None:
        return None
    return (avg, len(points))


# ──────────────── Команда ──────────────────────────────────────────────────────


class Command(BaseCommand):
    help = (
        "Обогащение Excel в два шага: (1) стандартизация производителя/модели по "
        "справочнику Manufacturer/DeviceModel, (2) проверка серийников в GLPI. "
        "Результат пишется в новый .xlsx."
    )

    def add_arguments(self, parser):
        parser.add_argument("xlsx_path", type=str, help="Путь к исходному .xlsx")
        parser.add_argument("--sheet", type=str, default=None, help="Имя листа (по умолчанию первый)")
        parser.add_argument(
            "--output", type=str, default=None, help="Путь к результату (по умолчанию <имя>_enriched.xlsx)"
        )
        parser.add_argument("--mfr-col", type=str, default="производитель", help="Заголовок колонки производителя")
        parser.add_argument("--model-col", type=str, default="модель", help="Заголовок колонки модели")
        parser.add_argument(
            "--model-cutoff", type=float, default=0.7, help="Порог схожести для fuzzy по моделям (0..1)"
        )
        parser.add_argument("--skip-standardize", action="store_true", help="Не делать нормализацию производителя/модели")
        parser.add_argument("--skip-glpi", action="store_true", help="Не обращаться в GLPI")
        parser.add_argument("--skip-average", action="store_true", help="Не считать среднее страниц/мес")
        parser.add_argument("--start-row", type=int, default=2, help="С какой строки начать обрабатывать (минимум 2)")
        parser.add_argument("--limit", type=int, default=0, help="Ограничить число обрабатываемых строк (0 = все)")
        parser.add_argument("--progress-every", type=int, default=50, help="Как часто печатать прогресс")
        parser.add_argument(
            "--counter-field-id",
            type=int,
            default=None,
            help="GLPI id_search_option для last_pages_counter (иначе settings.GLPI_LAST_COUNTER_FIELD_ID, дефолт 12)",
        )

    def handle(self, *args, **opts):
        src = Path(opts["xlsx_path"])
        if not src.exists():
            raise CommandError(f"Файл не найден: {src}")

        if opts["skip_standardize"] and opts["skip_glpi"] and opts["skip_average"]:
            raise CommandError("Все шаги отключены — нечего делать.")

        # Для среднего нужен GLPI-поиск, чтобы узнать ID карточки.
        # Если GLPI пропущен, но среднее просят — считаем только monthly_report.
        average_needs_glpi = not opts["skip_average"] and not opts["skip_glpi"]
        counter_field_id = opts.get("counter_field_id") or getattr(settings, "GLPI_LAST_COUNTER_FIELD_ID", 12)

        out_path = Path(opts["output"]) if opts["output"] else src.with_name(f"{src.stem}_enriched.xlsx")
        start_row = max(opts["start_row"], 2)
        limit = max(opts["limit"], 0)

        wb_in = load_workbook(filename=str(src), read_only=True, data_only=True)
        ws_in = wb_in[opts["sheet"]] if opts["sheet"] else wb_in.worksheets[0]

        try:
            header = next(ws_in.iter_rows(min_row=1, max_row=1, values_only=True))
        except StopIteration:
            raise CommandError("Пустой лист")

        header_list = [norm_space(h) for h in header]
        lower_header = [h.lower() for h in header_list]

        # Колонки для стандартизации
        mfr_idx = model_idx = None
        if not opts["skip_standardize"]:
            mfr_key = opts["mfr_col"].strip().lower()
            model_key = opts["model_col"].strip().lower()
            try:
                mfr_idx = next(i for i, h in enumerate(lower_header) if h == mfr_key or mfr_key in h)
                model_idx = next(i for i, h in enumerate(lower_header) if h == model_key or model_key in h)
            except StopIteration:
                raise CommandError(
                    f"Не нашёл колонки производителя/модели. Заголовки: {header_list}. "
                    f"Используй --mfr-col / --model-col."
                )

        # Колонки серийников — нужны и для GLPI, и для расчёта среднего
        main_idx = label_idx = None
        if not opts["skip_glpi"] or not opts["skip_average"]:
            label_idx = _find_col(lower_header, SERIAL_LABEL_KEYS)
            main_idx = _find_col(lower_header, SERIAL_MAIN_KEYS, exclude=SERIAL_LABEL_KEYS)
            if main_idx is None and label_idx is None:
                raise CommandError(f"Не нашёл колонок с серийниками. Заголовки: {header_list}")
            self.stdout.write(
                f"Серийники: main_idx={main_idx} "
                f"({header_list[main_idx] if main_idx is not None else '—'}), "
                f"label_idx={label_idx} ({header_list[label_idx] if label_idx is not None else '—'})"
            )

        # Выходные колонки
        extra_cols: List[str] = []
        if not opts["skip_standardize"]:
            extra_cols += ["Производитель (эталон)", "Модель (эталон)", "Статус сопоставления"]
        if not opts["skip_glpi"]:
            extra_cols += ["GLPI: статус", "GLPI: кол-во карточек", "GLPI: ID карточек", "GLPI: по какому SN найдено"]
        if not opts["skip_average"]:
            extra_cols += ["Среднее страниц/мес", "Источник среднего"]

        # Читаем все строки в память (для 5-10к строк это OK, даёт чистые этапы)
        self.stdout.write("Чтение файла…")
        all_rows: List[List] = [list(r) for r in ws_in.iter_rows(min_row=2, values_only=True)]
        total = len(all_rows)
        # индексы строк для обработки (0-based внутри all_rows)
        to_process: List[int] = []
        for i in range(total):
            excel_row_num = i + 2  # 1-based + header
            if excel_row_num < start_row:
                continue
            if limit and len(to_process) >= limit:
                break
            to_process.append(i)
        proc_total = len(to_process)
        self.stdout.write(f"Всего строк в файле: {total}, будет обработано: {proc_total}")

        # Кэши
        mfr_cache: Dict[str, Tuple[Optional[Manufacturer], str]] = {}
        model_cache: Dict[Tuple[Optional[int], str], Tuple[Optional[DeviceModel], str]] = {}
        glpi_cache: Dict[str, Tuple[str, int, str]] = {}
        avg_mr_cache: Dict[str, Optional[Tuple[float, int]]] = {}
        avg_glpi_cache: Dict[int, Optional[Tuple[float, int, str]]] = {}
        # Хранилище результатов по индексу строки (только для обработанных)
        std_extras: Dict[int, List] = {}
        glpi_extras: Dict[int, List] = {}
        avg_extras: Dict[int, List] = {}

        stats = {
            "mfr_ok": 0,
            "model_ok": 0,
            "glpi_checked": 0,
            "glpi_cache_hits": 0,
            "found_single": 0,
            "found_multiple": 0,
            "not_found": 0,
            "error": 0,
            "empty_sn": 0,
            "avg_from_mr": 0,
            "avg_from_printer_log": 0,
            "avg_from_change_log": 0,
            "avg_missing": 0,
        }

        step_num = 0
        total_steps = (
            (0 if opts["skip_standardize"] else 1)
            + (0 if opts["skip_glpi"] else 1)
            + (0 if opts["skip_average"] else 1)
        )

        def progress(stage_num: int, stage_total: int, done: int, extra: str = ""):
            pct = (done * 100 // stage_total) if stage_total else 100
            msg = f"  [{done}/{stage_total}] ({pct}%)"
            if extra:
                msg += f"  {extra}"
            self.stdout.write(msg)

        # ── Шаг 1: стандартизация ──────────────────────────────────────────────

        if not opts["skip_standardize"]:
            step_num += 1
            self.stdout.write("")
            self.stdout.write(self.style.MIGRATE_HEADING(f"▶ Шаг {step_num}/{total_steps}: стандартизация производителя/модели"))
            mfr_resolver = ManufacturerResolver()
            model_resolver = ModelResolver()

            for done, idx in enumerate(to_process, start=1):
                row = all_rows[idx]
                raw_mfr = norm_space(row[mfr_idx]) if mfr_idx < len(row) else ""
                raw_model = norm_space(row[model_idx]) if model_idx < len(row) else ""

                if raw_mfr in mfr_cache:
                    mfr_obj, mfr_status = mfr_cache[raw_mfr]
                else:
                    mfr_obj, mfr_status = mfr_resolver.resolve(raw_mfr)
                    mfr_cache[raw_mfr] = (mfr_obj, mfr_status)

                key = (mfr_obj.id if mfr_obj else None, raw_model)
                if key in model_cache:
                    model_obj, model_status = model_cache[key]
                else:
                    model_obj, model_status = model_resolver.resolve(raw_model, mfr_obj, opts["model_cutoff"])
                    model_cache[key] = (model_obj, model_status)

                if mfr_obj:
                    stats["mfr_ok"] += 1
                if model_obj:
                    stats["model_ok"] += 1

                std_extras[idx] = [
                    mfr_obj.name if mfr_obj else "",
                    model_obj.name if model_obj else "",
                    f"mfr:{mfr_status} / model:{model_status}",
                ]

                if opts["progress_every"] and (done % opts["progress_every"] == 0 or done == proc_total):
                    progress(1, proc_total, done, f"mfr_ok={stats['mfr_ok']} model_ok={stats['model_ok']}")

            self.stdout.write(
                self.style.SUCCESS(
                    f"  ✔ Шаг 1 завершён: производитель {stats['mfr_ok']}/{proc_total}, "
                    f"модель {stats['model_ok']}/{proc_total}"
                )
            )

        # ── Шаг 2: GLPI поиск ──────────────────────────────────────────────────
        glpi_client: Optional[GLPIClient] = None
        glpi_needed = (not opts["skip_glpi"]) or average_needs_glpi

        def _save_partial():
            self._write_output(
                out_path, ws_in.title, header_list, extra_cols, all_rows,
                std_extras, glpi_extras, avg_extras,
            )

        try:
            if glpi_needed:
                glpi_client = GLPIClient()
                glpi_client.init_session()

            if not opts["skip_glpi"]:
                step_num += 1
                self.stdout.write("")
                self.stdout.write(
                    self.style.MIGRATE_HEADING(f"▶ Шаг {step_num}/{total_steps}: проверка серийников в GLPI")
                )

                def probe_glpi(serial: str) -> Tuple[str, int, str]:
                    key = serial.upper()
                    if key in glpi_cache:
                        stats["glpi_cache_hits"] += 1
                        return glpi_cache[key]
                    assert glpi_client is not None
                    status, items, err = glpi_client.search_printer_by_serial(serial)
                    stats["glpi_checked"] += 1
                    if status == "ERROR":
                        result = (f"ERROR: {err or ''}".strip(), 0, "")
                    else:
                        ids = _extract_ids(items)
                        result = (status, len(ids), ",".join(ids))
                    glpi_cache[key] = result
                    return result

                for done, idx in enumerate(to_process, start=1):
                    row = all_rows[idx]
                    main_sn = norm_space(row[main_idx]) if main_idx is not None and main_idx < len(row) else ""
                    label_sn = norm_space(row[label_idx]) if label_idx is not None and label_idx < len(row) else ""

                    results: List[Tuple[str, Tuple[str, int, str]]] = []
                    seen = set()
                    for origin, sn in (("main", main_sn), ("label", label_sn)):
                        if not sn:
                            continue
                        k = sn.upper()
                        if k in seen:
                            continue
                        seen.add(k)
                        try:
                            results.append((origin, probe_glpi(sn)))
                        except GLPIAPIError as e:
                            results.append((origin, (f"ERROR: {e}", 0, "")))

                    if not results:
                        stats["empty_sn"] += 1
                        glpi_extras[idx] = ["", "", "", ""]
                    else:
                        priority = {"FOUND_MULTIPLE": 3, "FOUND_SINGLE": 2, "NOT_FOUND": 1}
                        results.sort(key=lambda p: priority.get(p[1][0], 0), reverse=True)
                        origin_val, (status_val, count_val, ids_val) = results[0]
                        if status_val == "FOUND_SINGLE":
                            stats["found_single"] += 1
                        elif status_val == "FOUND_MULTIPLE":
                            stats["found_multiple"] += 1
                        elif status_val == "NOT_FOUND":
                            stats["not_found"] += 1
                        else:
                            stats["error"] += 1
                        glpi_extras[idx] = [status_val, count_val, ids_val, origin_val]

                    if opts["progress_every"] and (done % opts["progress_every"] == 0 or done == proc_total):
                        progress(
                            step_num,
                            proc_total,
                            done,
                            f"запросов={stats['glpi_checked']} cache={stats['glpi_cache_hits']} "
                            f"single={stats['found_single']} multi={stats['found_multiple']} "
                            f"not_found={stats['not_found']} err={stats['error']}",
                        )

                self.stdout.write(
                    self.style.SUCCESS(
                        f"  ✔ Шаг {step_num} завершён: FOUND_SINGLE={stats['found_single']}, "
                        f"FOUND_MULTIPLE={stats['found_multiple']}, NOT_FOUND={stats['not_found']}, "
                        f"ERROR={stats['error']}, пустых SN={stats['empty_sn']}"
                    )
                )

            # ── Шаг 3: среднее страниц/мес ─────────────────────────────────────
            if not opts["skip_average"]:
                step_num += 1
                self.stdout.write("")
                self.stdout.write(
                    self.style.MIGRATE_HEADING(
                        f"▶ Шаг {step_num}/{total_steps}: среднее страниц/мес "
                        f"(monthly_report → GLPI PrinterLog → GLPI Log, counter_field_id={counter_field_id})"
                    )
                )

                def compute_avg_for_row(row, idx) -> Tuple[Optional[float], str]:
                    # 1) monthly_report — по любому из серийников
                    sns = []
                    if main_idx is not None and main_idx < len(row):
                        main_sn = norm_space(row[main_idx])
                        if main_sn:
                            sns.append(main_sn)
                    if label_idx is not None and label_idx < len(row):
                        label_sn = norm_space(row[label_idx])
                        if label_sn and label_sn.upper() not in {s.upper() for s in sns}:
                            sns.append(label_sn)

                    for sn in sns:
                        key = sn.upper()
                        if key in avg_mr_cache:
                            mr = avg_mr_cache[key]
                        else:
                            mr = _avg_from_monthly_report(sn)
                            avg_mr_cache[key] = mr
                        if mr:
                            avg_val, months_cnt = mr
                            stats["avg_from_mr"] += 1
                            return avg_val, f"monthly_report ({months_cnt} мес.)"

                    # 2) + 3) GLPI — нужен ID карточки
                    if glpi_client is None:
                        return None, ""
                    glpi_ids_str = glpi_extras.get(idx, ["", "", "", ""])[2] if glpi_extras.get(idx) else ""
                    if not glpi_ids_str:
                        return None, ""
                    first_id = glpi_ids_str.split(",")[0].strip()
                    if not first_id:
                        return None, ""
                    try:
                        pid = int(first_id)
                    except ValueError:
                        return None, ""

                    if pid in avg_glpi_cache:
                        cached = avg_glpi_cache[pid]
                        if cached is None:
                            return None, ""
                        avg_val, points_cnt, source = cached
                        if source == "printer_log":
                            stats["avg_from_printer_log"] += 1
                            return avg_val, f"GLPI PrinterLog ({points_cnt} точек)"
                        stats["avg_from_change_log"] += 1
                        return avg_val, f"GLPI Log ({points_cnt} точек)"

                    # 2) PrinterLog
                    pl = _avg_from_printer_log(glpi_client, pid)
                    if pl:
                        avg_val, points_cnt = pl
                        avg_glpi_cache[pid] = (avg_val, points_cnt, "printer_log")
                        stats["avg_from_printer_log"] += 1
                        return avg_val, f"GLPI PrinterLog ({points_cnt} точек)"

                    # 3) change log
                    cl = _avg_from_change_log(glpi_client, pid, counter_field_id)
                    if cl:
                        avg_val, points_cnt = cl
                        avg_glpi_cache[pid] = (avg_val, points_cnt, "change_log")
                        stats["avg_from_change_log"] += 1
                        return avg_val, f"GLPI Log ({points_cnt} точек)"

                    avg_glpi_cache[pid] = None
                    return None, ""

                for done, idx in enumerate(to_process, start=1):
                    row = all_rows[idx]
                    avg_val, src = compute_avg_for_row(row, idx)
                    if avg_val is None:
                        stats["avg_missing"] += 1
                        avg_extras[idx] = ["", ""]
                    else:
                        avg_extras[idx] = [round(avg_val, 1), src]

                    if opts["progress_every"] and (done % opts["progress_every"] == 0 or done == proc_total):
                        progress(
                            step_num,
                            proc_total,
                            done,
                            f"mr={stats['avg_from_mr']} printer_log={stats['avg_from_printer_log']} "
                            f"change_log={stats['avg_from_change_log']} miss={stats['avg_missing']}",
                        )

                self.stdout.write(
                    self.style.SUCCESS(
                        f"  ✔ Шаг {step_num} завершён: monthly_report={stats['avg_from_mr']}, "
                        f"PrinterLog={stats['avg_from_printer_log']}, "
                        f"Log={stats['avg_from_change_log']}, без данных={stats['avg_missing']}"
                    )
                )

        except GLPIAPIError as e:
            _save_partial()
            raise CommandError(f"Ошибка GLPI API: {e}. Частичный результат сохранён в {out_path}")
        finally:
            if glpi_client is not None:
                try:
                    glpi_client.kill_session()
                except Exception:
                    pass

        # ── Запись результата ──────────────────────────────────────────────────
        self.stdout.write("")
        self.stdout.write("Запись результата…")
        _save_partial()

        self.stdout.write("")
        self.stdout.write(self.style.SUCCESS(f"Готово: {out_path}"))
        self.stdout.write(f"  строк обработано: {proc_total} из {total}")
        if not opts["skip_standardize"]:
            self.stdout.write(f"  производитель сопоставлен: {stats['mfr_ok']}/{proc_total}")
            self.stdout.write(f"  модель сопоставлена: {stats['model_ok']}/{proc_total}")
        if not opts["skip_glpi"]:
            self.stdout.write(
                f"  GLPI запросов: {stats['glpi_checked']} (cache hits: {stats['glpi_cache_hits']})"
            )
            self.stdout.write(f"    FOUND_SINGLE:   {stats['found_single']}")
            self.stdout.write(f"    FOUND_MULTIPLE: {stats['found_multiple']}")
            self.stdout.write(f"    NOT_FOUND:      {stats['not_found']}")
            self.stdout.write(f"    ERROR:          {stats['error']}")
            self.stdout.write(f"    пустые SN:      {stats['empty_sn']}")
        if not opts["skip_average"]:
            self.stdout.write("  Среднее страниц/мес:")
            self.stdout.write(f"    из monthly_report:    {stats['avg_from_mr']}")
            self.stdout.write(f"    из GLPI PrinterLog:   {stats['avg_from_printer_log']}")
            self.stdout.write(f"    из GLPI Log:          {stats['avg_from_change_log']}")
            self.stdout.write(f"    без данных:           {stats['avg_missing']}")

    @staticmethod
    def _write_output(out_path, sheet_title, header_list, extra_cols, all_rows, std_extras, glpi_extras, avg_extras):
        def sanitize(row_values):
            return [v if (v is None or isinstance(v, (int, float, str, bool))) else str(v) for v in row_values]

        wb_out = Workbook(write_only=True)
        ws_out = wb_out.create_sheet(sheet_title)
        ws_out.append(list(header_list) + extra_cols)

        n_std = 3 if any(c == "Производитель (эталон)" for c in extra_cols) else 0
        n_glpi = 4 if any(c.startswith("GLPI:") for c in extra_cols) else 0
        n_avg = 2 if any(c == "Среднее страниц/мес" for c in extra_cols) else 0

        for idx, row in enumerate(all_rows):
            extra: List = []
            if n_std:
                extra += std_extras.get(idx, [""] * n_std)
            if n_glpi:
                extra += glpi_extras.get(idx, [""] * n_glpi)
            if n_avg:
                extra += avg_extras.get(idx, [""] * n_avg)
            ws_out.append(sanitize(list(row) + extra))

        wb_out.save(str(out_path))
