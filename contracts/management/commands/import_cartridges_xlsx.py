# contracts/management/commands/import_cartridges_xlsx.py
"""
Импорт картриджей для моделей принтеров из Excel.

Ожидаемая структура листа (строка заголовков):

Вариант 1 (по ID модели - рекомендуется):
ID | Модель оборудования | Картридж | Артикул | Цвет | Ресурс | Комментарий

Вариант 2 (по производителю и модели):
Производитель | Модель оборудования | Картридж | Артикул | Цвет | Ресурс | Комментарий

При использовании варианта 1:
- ID - обязательное поле, должно совпадать с ID модели в базе данных
- Модель оборудования - опциональное, используется для дополнительной валидации
  (если указано, то проверяется соответствие названия модели ID)

Примеры запуска:
    python manage.py import_cartridges_xlsx cartridges.xlsx
    python manage.py import_cartridges_xlsx cartridges.xlsx --sheet "Лист1"
    python manage.py import_cartridges_xlsx cartridges.xlsx --dry-run
    python manage.py import_cartridges_xlsx cartridges.xlsx --primary
"""

import re
from collections import namedtuple
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

from django.core.management.base import BaseCommand, CommandError
from django.db import IntegrityError, transaction

from contracts.models import Cartridge, DeviceModel, DeviceModelCartridge, Manufacturer

# Заголовки из Excel → внутренние ключи
HEADERS = {
    "id": "model_id",
    "id модели": "model_id",
    "model id": "model_id",
    "device model id": "model_id",
    "производитель": "manufacturer",
    "vendor": "manufacturer",
    "бренд": "manufacturer",
    "модель оборудования": "model",
    "модель": "model",
    "device model": "model",
    "картридж": "cartridge",
    "название картриджа": "cartridge",
    "cartridge name": "cartridge",
    "toner": "cartridge",
    "артикул": "part_number",
    "part number": "part_number",
    "pn": "part_number",
    "партномер": "part_number",
    "цвет": "color",
    "color": "color",
    "ресурс": "capacity",
    "capacity": "capacity",
    "yield": "capacity",
    "комментарий": "comment",
    "примечание": "comment",
    "comment": "comment",
}


def norm(val):
    """None -> '', str -> strip + схлопывание пробелов, числа -> str без '.0'."""
    if val is None:
        return ""
    if isinstance(val, (int, float)):
        s = str(int(val)) if float(val).is_integer() else str(val)
    else:
        s = str(val)
    s = s.strip()
    return re.sub(r"\s+", " ", s)


def key_of(header):
    if not header:
        return None
    return HEADERS.get(norm(header).lower())


def parse_color(color_str):
    """Парсинг цвета картриджа"""
    if not color_str:
        return "black"

    color_str = color_str.lower().strip()

    color_map = {
        "черный": "black",
        "black": "black",
        "k": "black",
        "чёрный": "black",
        "голубой": "cyan",
        "cyan": "cyan",
        "c": "cyan",
        "синий": "cyan",
        "пурпурный": "magenta",
        "magenta": "magenta",
        "m": "magenta",
        "розовый": "magenta",
        "желтый": "yellow",
        "yellow": "yellow",
        "y": "yellow",
        "жёлтый": "yellow",
        "цветной": "color",
        "color": "color",
        "трёхцветный": "color",
        "триколор": "color",
    }

    return color_map.get(color_str, "other")


def ci_get_or_create(model, name_value, name_field="name", **extra_filters):
    """
    Регистронезависимый get_or_create по текстовому полю `name_field`.
    """
    q = {f"{name_field}__iexact": name_value, **extra_filters}
    obj = model.objects.filter(**q).first()
    if obj:
        return obj, False
    try:
        obj = model.objects.create(**{name_field: name_value, **extra_filters})
        return obj, True
    except IntegrityError:
        obj = model.objects.filter(**q).first()
        if obj:
            return obj, False
        raise


BadRow = namedtuple("BadRow", "row reason preview")


class Command(BaseCommand):
    help = "Импорт картриджей для моделей принтеров из Excel"

    def add_arguments(self, parser):
        parser.add_argument("xlsx_path", type=str, help="Путь к .xlsx файлу")
        parser.add_argument("--sheet", type=str, default=None, help="Имя листа (по умолчанию первый)")
        parser.add_argument("--dry-run", action="store_true", help="Не сохранять, только показать результат")
        parser.add_argument(
            "--primary",
            action="store_true",
            help="Помечать все импортируемые картриджи как основные (is_primary=True)",
        )
        parser.add_argument(
            "--skip-existing",
            action="store_true",
            help="Пропускать существующие связи модель-картридж (по умолчанию обновляет)",
        )

    @transaction.atomic
    def handle(self, *args, **opts):
        # --- Открываем Excel ---
        path = Path(opts["xlsx_path"])
        if not path.exists():
            raise CommandError(f"Файл не найден: {path}")

        wb = load_workbook(filename=str(path), read_only=True, data_only=True)
        ws = wb[opts["sheet"]] if opts["sheet"] else wb.worksheets[0]

        # --- Заголовки ---
        try:
            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        except StopIteration:
            raise CommandError("Пустой лист: нет строк с данными")

        columns = [key_of(h) for h in header_row]

        # Обязательные колонки - либо ID модели, либо производитель+модель
        have = {c for c in columns if c}
        has_model_id = "model_id" in have
        has_manufacturer_and_model = {"manufacturer", "model"}.issubset(have)

        if not has_model_id and not has_manufacturer_and_model:
            raise CommandError("В файле должны быть либо 'ID модели', либо 'Производитель' + 'Модель оборудования'")

        if "cartridge" not in have:
            raise CommandError("В файле отсутствует обязательная колонка 'Картридж'")

        # --- Счётчики ---
        cartridges_created = 0
        cartridges_found = 0
        links_created = 0
        links_updated = 0
        links_skipped = 0
        failed = 0
        blank_rows = 0
        total_rows = 0
        bad_rows = []

        def row_preview(d):
            return (
                f"{(d.get('manufacturer') or '—')} {(d.get('model') or '—')} → "
                f"Картридж: {(d.get('cartridge') or '—')} "
                f"({d.get('part_number') or '—'}) "
                f"{d.get('color') or '—'}"
            )

        # --- Основной проход по строкам ---
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            total_rows += 1
            data = {}
            for col_key, cell in zip(columns, row):
                if not col_key:
                    continue
                data[col_key] = norm(cell)

            # Полностью пустая строка
            if not any(str(v).strip() if v is not None else "" for v in data.values()):
                blank_rows += 1
                continue

            # Валидация обязательных полей
            if not data.get("cartridge"):
                bad_rows.append(BadRow(idx, "MISSING_VALUES: cartridge", row_preview(data)))
                failed += 1
                continue

            # Проверяем, что есть либо ID модели, либо производитель+модель
            has_model_id = bool(data.get("model_id"))
            has_manufacturer_and_model = bool(data.get("manufacturer") and data.get("model"))

            if not has_model_id and not has_manufacturer_and_model:
                bad_rows.append(
                    BadRow(
                        idx,
                        "MISSING_VALUES: требуется либо 'ID модели', либо 'Производитель' + 'Модель'",
                        row_preview(data),
                    )
                )
                failed += 1
                continue

            try:
                device_model = None

                # Вариант 1: Поиск по ID модели
                if has_model_id:
                    model_id_str = data["model_id"]
                    try:
                        model_id = int(model_id_str)
                        device_model = DeviceModel.objects.filter(id=model_id).first()

                        if not device_model:
                            bad_rows.append(BadRow(idx, f"MODEL_NOT_FOUND_BY_ID: ID={model_id}", row_preview(data)))
                            failed += 1
                            continue

                        # Дополнительная валидация: проверяем название модели, если оно указано
                        if data.get("model"):
                            if device_model.name.lower() != data["model"].lower():
                                bad_rows.append(
                                    BadRow(
                                        idx,
                                        f"MODEL_NAME_MISMATCH: ID={model_id} ожидается '{device_model.name}', в файле '{data['model']}'",
                                        row_preview(data),
                                    )
                                )
                                failed += 1
                                continue

                    except ValueError:
                        bad_rows.append(
                            BadRow(idx, f"INVALID_MODEL_ID: '{model_id_str}' не является числом", row_preview(data))
                        )
                        failed += 1
                        continue

                # Вариант 2: Поиск по производителю и названию модели
                else:
                    mfr_name = data["manufacturer"]
                    mfr, _ = ci_get_or_create(Manufacturer, mfr_name)

                    model_name = data["model"]
                    device_model = DeviceModel.objects.filter(manufacturer=mfr, name__iexact=model_name).first()

                    if not device_model:
                        bad_rows.append(BadRow(idx, f"MODEL_NOT_FOUND: {mfr_name} {model_name}", row_preview(data)))
                        failed += 1
                        continue

                # Данные картриджа
                cartridge_name = data["cartridge"]
                part_number = data.get("part_number") or ""
                color = parse_color(data.get("color"))
                capacity = data.get("capacity") or ""
                comment = data.get("comment") or ""

                # Создаём или находим картридж
                cartridge = None
                if part_number:
                    # Ищем по названию и артикулу (регистронезависимо)
                    cartridge = Cartridge.objects.filter(
                        name__iexact=cartridge_name, part_number__iexact=part_number
                    ).first()

                if not cartridge:
                    # Ищем только по названию
                    cartridge = Cartridge.objects.filter(name__iexact=cartridge_name).first()

                if cartridge:
                    cartridges_found += 1
                    # Обновляем данные, если они изменились
                    updated = False
                    if part_number and cartridge.part_number != part_number:
                        cartridge.part_number = part_number
                        updated = True
                    if color and cartridge.color != color:
                        cartridge.color = color
                        updated = True
                    if capacity and cartridge.capacity != capacity:
                        cartridge.capacity = capacity
                        updated = True
                    if comment and not cartridge.comment:
                        cartridge.comment = comment
                        updated = True

                    if updated and not opts.get("dry_run"):
                        cartridge.save()
                else:
                    # Создаём новый картридж
                    if not opts.get("dry_run"):
                        cartridge = Cartridge.objects.create(
                            name=cartridge_name,
                            part_number=part_number,
                            color=color,
                            capacity=capacity,
                            comment=comment,
                            is_active=True,
                        )
                    cartridges_created += 1

                # Создаём или обновляем связь модель-картридж
                if cartridge:
                    link = DeviceModelCartridge.objects.filter(device_model=device_model, cartridge=cartridge).first()

                    is_primary = opts.get("primary", False)

                    if link:
                        if opts.get("skip_existing"):
                            links_skipped += 1
                        else:
                            # Обновляем связь
                            if link.is_primary != is_primary or (comment and not link.comment):
                                link.is_primary = is_primary
                                if comment and not link.comment:
                                    link.comment = comment
                                if not opts.get("dry_run"):
                                    link.save()
                                links_updated += 1
                            else:
                                links_skipped += 1
                    else:
                        # Создаём новую связь
                        if not opts.get("dry_run"):
                            DeviceModelCartridge.objects.create(
                                device_model=device_model, cartridge=cartridge, is_primary=is_primary, comment=comment
                            )
                        links_created += 1

            except Exception as e:
                bad_rows.append(BadRow(idx, f"UNEXPECTED: {e.__class__.__name__}: {e}", row_preview(data)))
                failed += 1
                continue

        # --- Отчёт по проблемным строкам ---
        if bad_rows:
            self.stdout.write(self.style.WARNING("\n❌ Строки, которые НЕ были импортированы:"))
            for br in bad_rows:
                self.stdout.write(self.style.WARNING(f"  [строка {br.row}] {br.reason}"))
                self.stdout.write(f"    {br.preview}")

        # --- Итоги ---
        if opts.get("dry_run"):
            transaction.set_rollback(True)
            mode = " (DRY-RUN - изменения не сохранены)"
        else:
            mode = ""

        self.stdout.write("\n" + "=" * 70)
        self.stdout.write(self.style.SUCCESS(f"✅ Импорт завершён{mode}"))
        self.stdout.write("=" * 70)

        self.stdout.write(f"\n📊 СТАТИСТИКА:")
        self.stdout.write(f"   Обработано строк: {total_rows}")
        self.stdout.write(f"   Пустых строк: {blank_rows}")
        self.stdout.write(f"   Ошибок: {failed}")

        self.stdout.write(f"\n🎨 КАРТРИДЖИ:")
        self.stdout.write(f"   Создано новых: {cartridges_created}")
        self.stdout.write(f"   Найдено существующих: {cartridges_found}")

        self.stdout.write(f"\n🔗 СВЯЗИ МОДЕЛЬ-КАРТРИДЖ:")
        self.stdout.write(f"   Создано новых: {links_created}")
        self.stdout.write(f"   Обновлено: {links_updated}")
        self.stdout.write(f"   Пропущено существующих: {links_skipped}")

        if opts.get("primary"):
            self.stdout.write(f"\n⭐ Все картриджи помечены как основные (is_primary=True)")

        self.stdout.write("=" * 70)

        if opts.get("dry_run"):
            self.stdout.write(
                self.style.WARNING("\n⚠️  Это был тестовый запуск. Для реального импорта запустите без --dry-run")
            )
