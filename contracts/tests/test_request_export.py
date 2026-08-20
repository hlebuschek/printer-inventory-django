from io import BytesIO

from django.contrib.auth import get_user_model
from django.contrib.auth.models import Permission
from django.test import override_settings
from django.urls import reverse
from openpyxl import load_workbook

from contracts.models import ServiceProvider, ServiceRequest
from contracts.services_request_export import HEADERS, export_requests_excel
from integrations.models import OkdeskIssue

from .test_service_requests import ServiceRequestBase, dt


# Выгрузка уходит в очередь exports; в тестах гоняем задачу синхронно
@override_settings(CELERY_TASK_ALWAYS_EAGER=True, CELERY_TASK_EAGER_PROPAGATES=True)
class RequestExportTests(ServiceRequestBase):
    """Выгрузка журнала: один экспорт вместо шести окдесковых, статус — обычный фильтр."""

    def setUp(self):
        self.user = self.login("access_contracts_app", "view_servicerequest", "export_service_requests")

    def login(self, *codenames):
        User = get_user_model()
        user = User.objects.create_user(
            username=f"tester{User.objects.count()}", password="pass12345", first_name="Иван", last_name="Петров"
        )
        for codename in codenames:
            user.user_permissions.add(Permission.objects.get(codename=codename))
        self.client.force_login(user, backend="django.contrib.auth.backends.ModelBackend")
        return user

    def _workbook(self, **params):
        started = self.client.get(reverse("contracts:api_requests_export"), params)
        self.assertEqual(started.status_code, 202)

        response = self.client.get(started.json()["download_url"])
        self.assertEqual(response.status_code, 200)
        return load_workbook(BytesIO(response.content))

    def _sheet(self, **params):
        return self._workbook(**params).active

    def test_export_requires_permission(self):
        self.login("access_contracts_app", "view_servicerequest")
        response = self.client.get(reverse("contracts:api_requests_export"))
        self.assertEqual(response.status_code, 403)

    def test_headers_and_row_content(self):
        self.make_request(initiator=self.user, act_number="АКТ-7", service_provider=self.provider)
        sheet = self._sheet()

        self.assertEqual([cell.value for cell in sheet[1]], list(HEADERS))
        row = dict(zip(HEADERS, [cell.value for cell in sheet[2]]))
        self.assertEqual(row["Подрядчик"], "Подрядчик")
        self.assertEqual(row["Серийный номер"], "SN001")
        self.assertEqual(row["Инициатор"], "Иван Петров")
        self.assertEqual(row["Акт"], "АКТ-7")
        self.assertEqual(row["Просрочена"], "Да")

    def test_sheet_per_okdesk_status(self):
        # Разбивка по статусам подрядчика — как в окдесковой выгрузке: наши четыре статуса
        # для подрядчика слишком грубые, «В работе» и «Требует решения» для нас одинаково «Открыта»
        for issue_id, status_name in ((101, "В работе"), (102, "Требует решения"), (103, "Заявка собрана")):
            OkdeskIssue.objects.create(issue_id=issue_id, title=f"Заявка {issue_id}", status_name=status_name)
            self.make_request(
                description=f"Заявка {issue_id}", service_provider=self.provider, external_number=str(issue_id)
            )

        workbook = self._workbook(status="")

        self.assertCountEqual(
            workbook.sheetnames,
            ["Подрядчик — В работе", "Подрядчик — Требует решения", "Подрядчик — Заявка собрана"],
        )
        row = dict(zip(HEADERS, [cell.value for cell in workbook["Подрядчик — В работе"][2]]))
        self.assertEqual(row["Статус у подрядчика"], "В работе")
        self.assertEqual(row["Статус"], "Открыта")

    def test_sheet_per_provider_and_own_status_without_okdesk(self):
        # У почтовых подрядчиков своего статуса нет — лист берёт наш
        other = ServiceProvider.objects.create(name="Второй подрядчик", code="second")
        self.make_request(description="Открытая первого", service_provider=self.provider)
        self.make_request(
            description="Выполненная первого",
            service_provider=self.provider,
            restored_at=dt("2026-03-03", 11),
            closed_at=dt("2026-03-03", 12),
        )
        self.make_request(description="Открытая второго", service_provider=other)
        self.make_request(description="Ничья")

        workbook = self._workbook(status="")

        expected = {
            "Подрядчик — Открыта": ["Открытая первого"],
            "Подрядчик — Выполнена": ["Выполненная первого"],
            "Второй подрядчик — Открыта": ["Открытая второго"],
            "Без подрядчика — Открыта": ["Ничья"],
        }
        self.assertCountEqual(workbook.sheetnames, expected)
        for title, descriptions in expected.items():
            sheet = workbook[title]
            self.assertEqual([cell.value for cell in sheet[1]], list(HEADERS))
            self.assertEqual(
                [row[0] for row in sheet.iter_rows(min_row=2, min_col=12, max_col=12, values_only=True)], descriptions
            )

    def test_statuses_of_one_provider_go_in_journal_order(self):
        self.make_request(description="Открытая", service_provider=self.provider)
        self.make_request(
            description="Выполненная",
            service_provider=self.provider,
            restored_at=dt("2026-03-03", 11),
            closed_at=dt("2026-03-03", 12),
        )
        self.make_request(description="Отклонённая", service_provider=self.provider, status=ServiceRequest.REJECTED)

        workbook = self._workbook(status="")

        self.assertEqual(workbook.sheetnames, ["Подрядчик — Открыта", "Подрядчик — Выполнена", "Подрядчик — Отклонена"])

    def test_sheet_title_fits_excel_limits(self):
        # Excel не принимает лист длиннее 31 символа и служебные символы в имени
        long_name = ServiceProvider.objects.create(name="Очень длинное имя подрядчика", code="long")
        slashed = ServiceProvider.objects.create(name="Сеть/Сервис", code="slash")
        self.make_request(service_provider=long_name)
        self.make_request(service_provider=slashed)
        # Длинный статус подрядчика теснит имя подрядчика, но не убирает его совсем
        OkdeskIssue.objects.create(issue_id=101, title="Заявка", status_name="Отправлено в сторонний сервис")
        self.make_request(service_provider=self.provider, external_number="101")

        workbook = self._workbook(status="")

        self.assertCountEqual(
            workbook.sheetnames,
            ["Очень длинное имя по — Открыта", "Сеть-Сервис — Открыта", "Подрядчик — Отправлено в сторон"],
        )
        self.assertLessEqual(max(len(title) for title in workbook.sheetnames), 31)

    def test_default_filter_keeps_only_unclosed(self):
        self.make_request(description="Открытая")
        self.make_request(description="Закрытая", restored_at=dt("2026-03-03", 11), closed_at=dt("2026-03-03", 12))
        sheet = self._sheet(status="active")

        descriptions = [row[0] for row in sheet.iter_rows(min_row=2, min_col=12, max_col=12, values_only=True)]
        self.assertEqual(descriptions, ["Открытая"])

    def test_status_filter_selects_closed(self):
        self.make_request(description="Открытая")
        self.make_request(description="Закрытая", restored_at=dt("2026-03-03", 11), closed_at=dt("2026-03-03", 12))
        sheet = self._sheet(status=ServiceRequest.CLOSED)

        descriptions = [row[0] for row in sheet.iter_rows(min_row=2, min_col=12, max_col=12, values_only=True)]
        self.assertEqual(descriptions, ["Закрытая"])

    def test_search_narrows_export(self):
        self.make_request(description="Открытая")
        self.make_request(description="Вторая", act_number="АКТ-9")
        sheet = self._sheet(status="", q="АКТ-9")

        descriptions = [row[0] for row in sheet.iter_rows(min_row=2, min_col=12, max_col=12, values_only=True)]
        self.assertEqual(descriptions, ["Вторая"])

    def test_export_does_not_query_per_row(self):
        # Простой считается по производственному календарю, а статус подрядчика лежит
        # в зеркале Okdesk: без общего кэша и то и другое читалось на каждой строке
        for issue_id in range(101, 106):
            OkdeskIssue.objects.create(issue_id=issue_id, title=f"Заявка {issue_id}", status_name="В работе")
            self.make_request(external_number=str(issue_id))

        with self.assertNumQueries(5):
            export_requests_excel(ServiceRequest.objects.all())

    def test_ready_file_is_not_served_to_another_user(self):
        # Выборка зависит от прав заказчика, поэтому чужой task_id не должен отдавать файл
        self.make_request()
        started = self.client.get(reverse("contracts:api_requests_export"))

        self.login("access_contracts_app", "view_servicerequest", "export_service_requests")
        self.assertEqual(self.client.get(started.json()["download_url"]).status_code, 403)

    def test_export_covers_whole_journal_not_only_own_requests(self):
        # Право на выгрузку даёт общую картину, даже если на экране видны только свои заявки
        restricted = self.login("access_contracts_app", "create_service_request", "export_service_requests")
        self.make_request(description="Моя", initiator=restricted)
        self.make_request(description="Чужая", initiator=self.user)

        sheet = self._sheet(status="")

        descriptions = [row[0] for row in sheet.iter_rows(min_row=2, min_col=12, max_col=12, values_only=True)]
        self.assertCountEqual(descriptions, ["Моя", "Чужая"])

    def test_mine_filter_narrows_export_to_own_requests(self):
        restricted = self.login("access_contracts_app", "create_service_request", "export_service_requests")
        self.make_request(description="Моя", initiator=restricted)
        self.make_request(description="Чужая", initiator=self.user)

        sheet = self._sheet(status="", mine="1")

        descriptions = [row[0] for row in sheet.iter_rows(min_row=2, min_col=12, max_col=12, values_only=True)]
        self.assertEqual(descriptions, ["Моя"])
