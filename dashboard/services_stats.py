"""
Чистые расчётные функции статистики по устройствам (без кэша).

Используются:
  - кэширующими обёртками в dashboard.services (виджеты),
  - Celery-задачей build_statistics_export_task (полная XLSX-выгрузка).

Все функции принимают org_id (id Organization) и не делают побочных эффектов,
кроме чтения из БД.
"""

import logging
from datetime import date, timedelta

from django.db.models import Count, F, Max, Sum
from django.utils import timezone

logger = logging.getLogger(__name__)


# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────


def _org_name(org_id):
    """Имя организации по id (MonthlyReport.organization — строковое поле)."""
    if not org_id:
        return None
    from inventory.models import Organization

    try:
        return Organization.objects.get(pk=org_id).name
    except Organization.DoesNotExist:
        return None


def _month_cutoff(months):
    """Первое число месяца N месяцев назад. months<=0 → None (без отсечки)."""
    if not months or months <= 0:
        return None
    today = date.today()
    year, month = today.year, today.month - months
    while month <= 0:
        month += 12
        year -= 1
    return date(year, month, 1)


# ─────────────────────────────────────────────────────────────────────────────
# 1. Молчащие принтеры (нет успешного опроса N дней)
# ─────────────────────────────────────────────────────────────────────────────


def compute_silent_printers(org_id=None, days=7, limit=None):
    """
    Активные принтеры без единого SUCCESS-опроса за последние `days` дней.

    Возвращает:
      {
        total: всего активных,
        silent: молчащих,
        active_ok: опрошенных хотя бы раз за период,
        percentage: доля молчащих, %,
        items: [{printer_id, ip_address, model, serial_number, organization,
                 last_success}],
      }
    """
    from inventory.models import InventoryTask, Printer

    cutoff = timezone.now() - timedelta(days=days)

    printers = Printer.objects.filter(is_active=True)
    if org_id:
        printers = printers.filter(organization_id=org_id)
    total = printers.count()

    recent_ok_ids = (
        InventoryTask.objects.filter(status="SUCCESS", task_timestamp__gte=cutoff, printer__is_active=True)
        .values_list("printer_id", flat=True)
        .distinct()
    )
    if org_id:
        recent_ok_ids = recent_ok_ids.filter(printer__organization_id=org_id)
    recent_ok_ids = set(recent_ok_ids)

    silent_qs = (
        printers.exclude(id__in=recent_ok_ids)
        .select_related("organization", "device_model")
        .order_by("organization__name", "ip_address")
    )

    # Последний успешный опрос (за всё время) для молчащих принтеров.
    last_ok_map = dict(
        InventoryTask.objects.filter(status="SUCCESS", printer__in=silent_qs)
        .values("printer_id")
        .annotate(ts=Max("task_timestamp"))
        .values_list("printer_id", "ts")
    )

    items = []
    rows = silent_qs[:limit] if limit else silent_qs
    for p in rows:
        last_ok = last_ok_map.get(p.id)
        items.append(
            {
                "printer_id": p.id,
                "ip_address": p.ip_address,
                "model": p.device_model.name if p.device_model else p.model,
                "serial_number": p.serial_number,
                "organization": p.organization.name if p.organization else "—",
                "last_success": last_ok.isoformat() if last_ok else None,
            }
        )

    # Молчащие = активные минус опрошенные за период (среди активных).
    silent = total - printers.filter(id__in=recent_ok_ids).count()

    return {
        "total": total,
        "silent": silent,
        "active_ok": total - silent,
        "percentage": round(silent / total * 100) if total else 0,
        "items": items,
    }


# ─────────────────────────────────────────────────────────────────────────────
# 2. Топ по объёму печати (сумма total_prints за период)
# ─────────────────────────────────────────────────────────────────────────────


def compute_top_by_volume(org_id=None, months=0, limit=10):
    """
    Топ устройств по суммарному объёму печати за период (из MonthlyReport).

    months=0 → все месяцы, иначе последние N.
    Возвращает [{serial_number, model, organization, total, months_count}].
    """
    from monthly_report.models import MonthlyReport

    qs = MonthlyReport.objects.all()
    cutoff = _month_cutoff(months)
    if cutoff:
        qs = qs.filter(month__gte=cutoff)
    if org_id:
        name = _org_name(org_id)
        qs = qs.filter(organization=name) if name else qs.none()

    rows = (
        qs.values("serial_number", "equipment_model", "organization")
        .annotate(total=Sum("total_prints"), months_count=Count("month", distinct=True))
        .order_by("-total")[:limit]
    )

    return [
        {
            "serial_number": r["serial_number"],
            "model": r["equipment_model"] or "—",
            "organization": r["organization"] or "—",
            "total": r["total"] or 0,
            "months_count": r["months_count"],
        }
        for r in rows
    ]


# ─────────────────────────────────────────────────────────────────────────────
# 3. Распределение по вендорам (3 источника: опрос / договоры / отчёты)
# ─────────────────────────────────────────────────────────────────────────────

UNKNOWN_MFR = "— не определён —"


def _normalize_serial(s):
    return (s or "").strip().lower()


def _serial_manufacturer_maps(org_id=None):
    """
    Карты «серийник → производитель» из договоров и из опроса.

    Договоры — приоритетный источник; опрос — запасной (для старых отчётов,
    которых уже нет в актуальных договорах).
    """
    from contracts.models import ContractDevice
    from inventory.models import Printer

    cd_qs = ContractDevice.objects.exclude(serial_number="").values_list("serial_number", "model__manufacturer__name")
    if org_id:
        cd_qs = cd_qs.filter(organization_id=org_id)
    contract_map = {_normalize_serial(sn): mfr for sn, mfr in cd_qs if _normalize_serial(sn) and mfr}

    pr_qs = (
        Printer.objects.exclude(serial_number__isnull=True)
        .exclude(serial_number="")
        .values_list("serial_number", "device_model__manufacturer__name")
    )
    if org_id:
        pr_qs = pr_qs.filter(organization_id=org_id)
    printer_map = {_normalize_serial(sn): mfr for sn, mfr in pr_qs if _normalize_serial(sn) and mfr}

    return contract_map, printer_map


def _manufacturer_aliases():
    """[(подстрока_в_lowercase, каноническое_имя)] для разбора текста модели."""
    from contracts.models import Manufacturer

    names = list(Manufacturer.objects.values_list("name", flat=True))
    aliases = []
    for name in names:
        low = name.lower()
        aliases.append((low, name))
        # Спец-синонимы для распространённых вендоров.
        if "hewlett" in low or low == "hp":
            aliases.append(("hp", name))
            aliases.append(("hewlett", name))
        if "kyocera" in low:
            aliases.append(("kyocera", name))
    return aliases


def _resolve_manufacturer(serial, model_str, contract_map, printer_map, aliases):
    """Вендор по серийнику (договоры → опрос) с откатом на разбор строки модели."""
    key = _normalize_serial(serial)
    if key:
        if key in contract_map:
            return contract_map[key]
        if key in printer_map:
            return printer_map[key]
    low = (model_str or "").lower()
    for needle, canonical in aliases:
        if needle and needle in low:
            return canonical
    return UNKNOWN_MFR


def _sorted_mfr_counts(counter):
    return sorted(
        ({"manufacturer": k, "count": v} for k, v in counter.items()),
        key=lambda r: (r["manufacturer"] == UNKNOWN_MFR, -r["count"]),
    )


def compute_manufacturer_distribution(source="polling", org_id=None, month_from=None, month_to=None):
    """
    Распределение устройств по производителям из выбранного источника.

    source:
      "polling"   — активные опрашиваемые принтеры (inventory.Printer);
      "contracts" — все устройства по договорам (contracts.ContractDevice);
      "monthly"   — уникальные серийники из MonthlyReport за диапазон месяцев;
                    вендор определяется матчем серийника по договорам/опросу,
                    иначе разбором строки модели, иначе «— не определён —».

    month_from / month_to — границы (date, первый день месяца) для source="monthly".
    Возвращает [{manufacturer, count}], отсортировано по убыванию count,
    «не определён» — в конце.
    """
    from collections import Counter

    if source == "contracts":
        from contracts.models import ContractDevice

        qs = ContractDevice.objects.all()
        if org_id:
            qs = qs.filter(organization_id=org_id)
        rows = qs.values("model__manufacturer__name").annotate(count=Count("id"))
        counter = Counter()
        for r in rows:
            counter[r["model__manufacturer__name"] or UNKNOWN_MFR] += r["count"]
        return _sorted_mfr_counts(counter)

    if source == "monthly":
        from monthly_report.models import MonthlyReport

        qs = MonthlyReport.objects.all()
        if month_from:
            qs = qs.filter(month__gte=month_from)
        if month_to:
            qs = qs.filter(month__lte=month_to)
        if org_id:
            name = _org_name(org_id)
            qs = qs.filter(organization=name) if name else qs.none()

        # Один представительный equipment_model на серийник.
        per_serial = qs.values("serial_number").annotate(model=Max("equipment_model"))
        contract_map, printer_map = _serial_manufacturer_maps(org_id)
        aliases = _manufacturer_aliases()
        counter = Counter()
        for r in per_serial:
            mfr = _resolve_manufacturer(r["serial_number"], r["model"], contract_map, printer_map, aliases)
            counter[mfr] += 1
        return _sorted_mfr_counts(counter)

    # source == "polling" (по умолчанию)
    from inventory.models import Printer

    qs = Printer.objects.filter(is_active=True)
    if org_id:
        qs = qs.filter(organization_id=org_id)
    rows = qs.values("device_model__manufacturer__name").annotate(count=Count("id"))
    counter = Counter()
    for r in rows:
        counter[r["device_model__manufacturer__name"] or UNKNOWN_MFR] += r["count"]
    return _sorted_mfr_counts(counter)


def get_report_months():
    """Доступные месяцы (date, первый день) из MonthlyReport, по убыванию."""
    from monthly_report.models import MonthlyReport

    return list(MonthlyReport.objects.values_list("month", flat=True).distinct().order_by("-month"))


# ─────────────────────────────────────────────────────────────────────────────
# 4. Средняя месячная нагрузка по устройствам (тяжёлый расчёт)
# ─────────────────────────────────────────────────────────────────────────────


def _device_month_deltas(org_id=None, months=0, with_model=False):
    """
    Дельты (end - start) по A4/A3 ЧБ/цвет, агрегированные по (serial_number, month).

    Одно устройство может иметь несколько строк в одном месяце (отдельно A4 и A3),
    поэтому суммируем дельты по всем строкам устройства за месяц. organization
    (и equipment_model) берём через Max — значение стабильно в пределах устройства.

    Возвращает list dict'ов, отсортированный по (serial_number, month).
    """
    from monthly_report.models import MonthlyReport

    qs = MonthlyReport.objects.all()
    cutoff = _month_cutoff(months)
    if cutoff:
        qs = qs.filter(month__gte=cutoff)
    if org_id:
        name = _org_name(org_id)
        qs = qs.filter(organization=name) if name else qs.none()

    annotations = {
        "organization": Max("organization"),
        "d_a4_bw": Sum(F("a4_bw_end") - F("a4_bw_start")),
        "d_a4_color": Sum(F("a4_color_end") - F("a4_color_start")),
        "d_a3_bw": Sum(F("a3_bw_end") - F("a3_bw_start")),
        "d_a3_color": Sum(F("a3_color_end") - F("a3_color_start")),
    }
    if with_model:
        annotations["equipment_model"] = Max("equipment_model")

    return list(qs.values("serial_number", "month").annotate(**annotations).order_by("serial_number", "month"))


def compute_device_monthly_avg(org_id=None, months=0):
    """
    Средняя месячная нагрузка ПО КАЖДОМУ ПРИНТЕРУ (из MonthlyReport).

    Для каждого серийника считаем дельты (end - start) с разбивкой A4/A3 ЧБ/цвет
    (отрицательные дельты = сброс счётчика → 0), делим на число месяцев с данными.

    Возвращает [{serial_number, model, organization, first_month, last_month,
                 months_count, avg_a4_bw, avg_a4_color, avg_a3_bw, avg_a3_color,
                 total, avg}].
    """
    from itertools import groupby

    rows_list = _device_month_deltas(org_id=org_id, months=months, with_model=True)

    result = []
    for serial, group in groupby(rows_list, key=lambda r: r["serial_number"]):
        md = list(group)
        n = len(md)

        def _sum(field):
            return sum(max(0, m[field] or 0) for m in md)

        a4_bw, a4_color = _sum("d_a4_bw"), _sum("d_a4_color")
        a3_bw, a3_color = _sum("d_a3_bw"), _sum("d_a3_color")
        total = a4_bw + a4_color + a3_bw + a3_color

        result.append(
            {
                "serial_number": serial,
                "model": md[0]["equipment_model"] or "—",
                "organization": md[0]["organization"] or "—",
                "first_month": md[0]["month"].isoformat(),
                "last_month": md[-1]["month"].isoformat(),
                "months_count": n,
                "avg_a4_bw": round(a4_bw / n, 1) if n else 0,
                "avg_a4_color": round(a4_color / n, 1) if n else 0,
                "avg_a3_bw": round(a3_bw / n, 1) if n else 0,
                "avg_a3_color": round(a3_color / n, 1) if n else 0,
                "total": total,
                "avg": round(total / n, 1) if n else 0,
            }
        )

    result.sort(key=lambda r: (r["organization"], r["serial_number"]))
    return result


def compute_org_monthly_avg(org_id=None, months=0):
    """
    Средняя месячная нагрузка ПО ОРГАНИЗАЦИЯМ (из MonthlyReport).

    Суммируем дельты всех принтеров организации (A4/A3 ЧБ/цвет, отрицательные → 0),
    считаем число уникальных устройств и месяцев с данными, делим итог на месяцы.

    Возвращает [{organization, printers, months_count, a4_bw, a4_color, a3_bw,
                 a3_color, total, avg}].
    """
    from collections import defaultdict

    rows_list = _device_month_deltas(org_id=org_id, months=months)

    org_data = defaultdict(
        lambda: {"serials": set(), "months": set(), "a4_bw": 0, "a4_color": 0, "a3_bw": 0, "a3_color": 0}
    )
    for r in rows_list:
        org = r["organization"] or "—"
        d = org_data[org]
        d["serials"].add(r["serial_number"])
        d["months"].add(r["month"])
        d["a4_bw"] += max(0, r["d_a4_bw"] or 0)
        d["a4_color"] += max(0, r["d_a4_color"] or 0)
        d["a3_bw"] += max(0, r["d_a3_bw"] or 0)
        d["a3_color"] += max(0, r["d_a3_color"] or 0)

    result = []
    for org, d in org_data.items():
        mc = len(d["months"])
        total = d["a4_bw"] + d["a4_color"] + d["a3_bw"] + d["a3_color"]
        result.append(
            {
                "organization": org,
                "printers": len(d["serials"]),
                "months_count": mc,
                "a4_bw": d["a4_bw"],
                "a4_color": d["a4_color"],
                "a3_bw": d["a3_bw"],
                "a3_color": d["a3_color"],
                "total": total,
                "avg": round(total / mc, 1) if mc else 0,
            }
        )

    result.sort(key=lambda r: r["organization"])
    return result


# ─────────────────────────────────────────────────────────────────────────────
# 5. Средняя нагрузка по принтерам из СЕТЕВОГО ОПРОСА (PageCounter)
# ─────────────────────────────────────────────────────────────────────────────


def compute_printer_polling_avg(org_id=None, include_inactive=False):
    """
    Средняя месячная нагрузка по принтерам на основе данных СЕТЕВОГО ОПРОСА.

    Берём первый и последний SUCCESS-снапшот PageCounter каждого принтера,
    вычитаем дельты (A4/A3 ЧБ/цвет, отрицательные = сброс → 0), делим на число
    месяцев между опросами (включая граничные).

    Возвращает [{organization, ip_address, serial_number, model, first_date,
                 last_date, months_count, avg_a4_bw, avg_a4_color, avg_a3_bw,
                 avg_a3_color, avg}].
    """
    from itertools import groupby

    from inventory.models import PageCounter, Printer

    printers = Printer.objects.select_related("organization", "device_model", "device_model__manufacturer")
    if not include_inactive:
        printers = printers.filter(is_active=True)
    if org_id:
        printers = printers.filter(organization_id=org_id)
    printers = list(printers.order_by("organization__name", "ip_address"))

    if not printers:
        return []

    printer_ids = [p.id for p in printers]
    counters = (
        PageCounter.objects.filter(task__printer_id__in=printer_ids, task__status="SUCCESS")
        .order_by("task__printer_id", "recorded_at")
        .values("task__printer_id", "recorded_at", "bw_a4", "color_a4", "bw_a3", "color_a3")
    )
    counters_by_printer = {pid: list(grp) for pid, grp in groupby(counters, key=lambda c: c["task__printer_id"])}

    result = []
    for p in printers:
        org = p.organization.name if p.organization_id else "—"
        rows = counters_by_printer.get(p.id, [])
        if not rows:
            result.append(
                {
                    "organization": org,
                    "ip_address": p.ip_address,
                    "serial_number": p.serial_number,
                    "model": p.model_display,
                    "first_date": None,
                    "last_date": None,
                    "months_count": 0,
                    "avg_a4_bw": 0,
                    "avg_a4_color": 0,
                    "avg_a3_bw": 0,
                    "avg_a3_color": 0,
                    "avg": 0,
                }
            )
            continue

        first, last = rows[0], rows[-1]
        f_dt, l_dt = first["recorded_at"], last["recorded_at"]
        f_date = f_dt.date() if hasattr(f_dt, "date") else f_dt
        l_date = l_dt.date() if hasattr(l_dt, "date") else l_dt
        n = (l_date.year - f_date.year) * 12 + (l_date.month - f_date.month) + 1

        def _delta(field):
            a, b = first.get(field), last.get(field)
            if a is None or b is None:
                return 0
            return max(0, b - a)

        a4_bw, a4_color = _delta("bw_a4"), _delta("color_a4")
        a3_bw, a3_color = _delta("bw_a3"), _delta("color_a3")
        total = a4_bw + a4_color + a3_bw + a3_color

        result.append(
            {
                "organization": org,
                "ip_address": p.ip_address,
                "serial_number": p.serial_number,
                "model": p.model_display,
                "first_date": f_date.isoformat(),
                "last_date": l_date.isoformat(),
                "months_count": n,
                "avg_a4_bw": round(a4_bw / n, 1) if n else 0,
                "avg_a4_color": round(a4_color / n, 1) if n else 0,
                "avg_a3_bw": round(a3_bw / n, 1) if n else 0,
                "avg_a3_color": round(a3_color / n, 1) if n else 0,
                "avg": round(total / n, 1) if n else 0,
            }
        )

    return result


# ─────────────────────────────────────────────────────────────────────────────
# 6. Сборка полного XLSX-отчёта (с колбэком прогресса)
# ─────────────────────────────────────────────────────────────────────────────


def _noop_progress(percent, message):  # pragma: no cover
    pass


def _style_header(ws, cols):
    from openpyxl.styles import Alignment, Font, PatternFill

    fill = PatternFill(fill_type="solid", fgColor="1F7A4A")
    font = Font(bold=True, color="FFFFFF")
    for col in range(1, cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")


def build_statistics_workbook(org_id=None, days=7, months=0, month_from=None, month_to=None, progress=None):
    """
    Собирает многолистовой XLSX-отчёт по статистике устройств.

    month_from / month_to (date, первый день месяца) — диапазон для листа
    «По вендорам (отчёты)». Если не заданы — берётся весь доступный период.

    progress(percent:int, message:str) — колбэк для отображения прогресса.
    Возвращает (content: bytes, filename: str).
    """
    import io
    from datetime import datetime

    import openpyxl

    progress = progress or _noop_progress

    wb = openpyxl.Workbook()

    # ── Лист 1: Молчащие принтеры ────────────────────────────────────────────
    progress(5, "Считаю молчащие принтеры…")
    silent = compute_silent_printers(org_id=org_id, days=days)
    ws = wb.active
    ws.title = "Молчащие"
    ws.append(["Организация", "IP-адрес", "Модель", "Серийный номер", "Последний успешный опрос"])
    _style_header(ws, 5)
    for it in silent["items"]:
        last_ok = ""
        if it["last_success"]:
            last_ok = datetime.fromisoformat(it["last_success"]).strftime("%d.%m.%Y %H:%M")
        ws.append([it["organization"], it["ip_address"], it["model"], it["serial_number"], last_ok])
    for col, w in zip("ABCDE", (28, 16, 32, 20, 22)):
        ws.column_dimensions[col].width = w

    # ── Лист 2: Топ по объёму ────────────────────────────────────────────────
    progress(30, "Считаю топ по объёму печати…")
    top = compute_top_by_volume(org_id=org_id, months=months, limit=50)
    ws = wb.create_sheet("Топ по объёму")
    ws.append(["#", "Организация", "Модель", "Серийный номер", "Отпечатков", "Месяцев"])
    _style_header(ws, 6)
    for idx, r in enumerate(top, start=1):
        ws.append([idx, r["organization"], r["model"], r["serial_number"], r["total"], r["months_count"]])
    for col, w in zip("ABCDEF", (5, 28, 32, 20, 14, 10)):
        ws.column_dimensions[col].width = w

    # ── Листы 3a-3c: По вендорам (3 источника) ───────────────────────────────
    progress(45, "Считаю распределение по вендорам…")
    vendor_sheets = [
        ("По вендорам (опрос)", "polling", "Устройств", {}),
        ("По вендорам (договоры)", "contracts", "Устройств", {}),
        ("По вендорам (отчёты)", "monthly", "Устройств", {"month_from": month_from, "month_to": month_to}),
    ]
    for title, source, count_label, extra in vendor_sheets:
        dist = compute_manufacturer_distribution(source=source, org_id=org_id, **extra)
        ws = wb.create_sheet(title)
        ws.append(["Производитель", count_label])
        _style_header(ws, 2)
        for r in dist:
            ws.append([r["manufacturer"], r["count"]])
        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 14

    # ── Лист 4: Средняя нагрузка по организациям ─────────────────────────────
    progress(60, "Считаю среднюю нагрузку по организациям…")
    org_rows = compute_org_monthly_avg(org_id=org_id, months=months)
    ws = wb.create_sheet("По организациям")
    ws.append(
        [
            "Организация",
            "Принтеров",
            "Месяцев",
            "ЧБ A4 (всего)",
            "Цвет A4 (всего)",
            "ЧБ A3 (всего)",
            "Цвет A3 (всего)",
            "Итого за период",
            "Среднее / мес",
        ]
    )
    _style_header(ws, 9)
    for r in org_rows:
        ws.append(
            [
                r["organization"],
                r["printers"],
                r["months_count"],
                r["a4_bw"],
                r["a4_color"],
                r["a3_bw"],
                r["a3_color"],
                r["total"],
                r["avg"],
            ]
        )
    for col, w in zip("ABCDEFGHI", (28, 11, 9, 14, 15, 14, 15, 16, 14)):
        ws.column_dimensions[col].width = w

    # ── Лист 5: Средняя нагрузка по принтерам (по отчётам) ────────────────────
    progress(60, "Считаю среднюю нагрузку по принтерам (отчёты)…")
    avg_rows = compute_device_monthly_avg(org_id=org_id, months=months)
    ws = wb.create_sheet("По принтерам (отчёты)")
    ws.append(
        [
            "Организация",
            "Серийный номер",
            "Модель",
            "Первый месяц",
            "Последний месяц",
            "Месяцев",
            "Средн. ЧБ A4/мес",
            "Средн. Цвет A4/мес",
            "Средн. ЧБ A3/мес",
            "Средн. Цвет A3/мес",
            "Итого средн./мес",
        ]
    )
    _style_header(ws, 11)
    n = len(avg_rows) or 1
    for i, r in enumerate(avg_rows, start=1):
        ws.append(
            [
                r["organization"],
                r["serial_number"],
                r["model"],
                r["first_month"],
                r["last_month"],
                r["months_count"],
                r["avg_a4_bw"],
                r["avg_a4_color"],
                r["avg_a3_bw"],
                r["avg_a3_color"],
                r["avg"],
            ]
        )
        if i % 200 == 0:
            progress(60 + int(15 * i / n), f"Отчёты: {i}/{len(avg_rows)}…")
    for col, w in zip("ABCDEFGHIJK", (28, 20, 32, 13, 14, 9, 16, 17, 16, 17, 16)):
        ws.column_dimensions[col].width = w

    # ── Лист 6: Средняя нагрузка по принтерам (по сетевому опросу) ─────────────
    progress(78, "Считаю среднюю нагрузку по принтерам (сеть)…")
    poll_rows = compute_printer_polling_avg(org_id=org_id)
    ws = wb.create_sheet("По принтерам (сеть)")
    ws.append(
        [
            "Организация",
            "IP-адрес",
            "Серийный номер",
            "Модель",
            "Первый опрос",
            "Последний опрос",
            "Месяцев",
            "Средн. ЧБ A4/мес",
            "Средн. Цвет A4/мес",
            "Средн. ЧБ A3/мес",
            "Средн. Цвет A3/мес",
            "Итого средн./мес",
        ]
    )
    _style_header(ws, 12)
    n2 = len(poll_rows) or 1
    for i, r in enumerate(poll_rows, start=1):
        ws.append(
            [
                r["organization"],
                r["ip_address"],
                r["serial_number"],
                r["model"],
                r["first_date"],
                r["last_date"],
                r["months_count"],
                r["avg_a4_bw"],
                r["avg_a4_color"],
                r["avg_a3_bw"],
                r["avg_a3_color"],
                r["avg"],
            ]
        )
        if i % 200 == 0:
            progress(78 + int(18 * i / n2), f"Сеть: {i}/{len(poll_rows)}…")
    for col, w in zip("ABCDEFGHIJKL", (28, 16, 20, 32, 13, 14, 9, 16, 17, 16, 17, 16)):
        ws.column_dimensions[col].width = w

    progress(98, "Формирую файл…")
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    suffix = "all" if not org_id else f"org{org_id}"
    filename = f'device_stats_{suffix}_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
    progress(100, "Готово")
    return buf.getvalue(), filename
