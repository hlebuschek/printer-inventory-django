# dashboard/views/api_views.py
"""API endpoints для дашборда."""

import io
import logging
from datetime import datetime

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

from django.contrib.auth.decorators import login_required, permission_required
from django.http import HttpResponse, JsonResponse
from django.views.decorators.http import require_GET, require_POST

from dashboard import services
from dashboard.api_docs_decorators import (
    api_glpi_cross_check_schema,
    api_low_consumables_schema,
    api_org_devices_schema,
    api_org_summary_schema,
    api_organizations_schema,
    api_printer_status_schema,
    api_poll_stats_schema,
    api_print_trend_schema,
    api_problem_printers_schema,
    api_recent_activity_schema,
)

logger = logging.getLogger(__name__)


def _ok(data):
    return JsonResponse({"ok": True, "data": data})


def _err(msg, status=400):
    return JsonResponse({"ok": False, "error": msg}, status=status)


def _parse_int(value, default=None):
    try:
        return int(value) if value else default
    except (ValueError, TypeError):
        return default


@login_required
@api_printer_status_schema
@permission_required("dashboard.access_dashboard_app", raise_exception=False)
@require_GET
def api_printer_status(request):
    org_id = _parse_int(request.GET.get("org"))
    try:
        data = services.get_printer_status(org_id=org_id)
        return _ok(data)
    except Exception as e:
        logger.exception("api_printer_status error")
        return _err(str(e), status=500)


@login_required
@api_poll_stats_schema
@permission_required("dashboard.access_dashboard_app", raise_exception=False)
@require_GET
def api_poll_stats(request):
    org_id = _parse_int(request.GET.get("org"))
    period_days = _parse_int(request.GET.get("period"), default=7)
    if period_days not in (7, 30, 90):
        period_days = 7
    try:
        data = services.get_poll_stats(org_id=org_id, period_days=period_days)
        return _ok(data)
    except Exception as e:
        logger.exception("api_poll_stats error")
        return _err(str(e), status=500)


@login_required
@api_low_consumables_schema
@permission_required("dashboard.access_dashboard_app", raise_exception=False)
@require_GET
def api_low_consumables(request):
    org_id = _parse_int(request.GET.get("org"))
    threshold = _parse_int(request.GET.get("threshold"), default=20)
    try:
        data = services.get_low_consumables(org_id=org_id, threshold=threshold)
        return _ok(data)
    except Exception as e:
        logger.exception("api_low_consumables error")
        return _err(str(e), status=500)


@login_required
@api_problem_printers_schema
@permission_required("dashboard.access_dashboard_app", raise_exception=False)
@require_GET
def api_problem_printers(request):
    org_id = _parse_int(request.GET.get("org"))
    period_days = _parse_int(request.GET.get("period"), default=7)
    if period_days not in (7, 30, 90):
        period_days = 7
    try:
        data = services.get_problem_printers(org_id=org_id, period_days=period_days)
        return _ok(data)
    except Exception as e:
        logger.exception("api_problem_printers error")
        return _err(str(e), status=500)


@login_required
@api_print_trend_schema
@permission_required("dashboard.access_dashboard_app", raise_exception=False)
@require_GET
def api_print_trend(request):
    org_id = _parse_int(request.GET.get("org"))
    months = _parse_int(request.GET.get("months"), default=0)
    if months not in (0, 6, 12):
        months = 0
    try:
        data = services.get_print_trend(org_id=org_id, months=months)
        return _ok(data)
    except Exception as e:
        logger.exception("api_print_trend error")
        return _err(str(e), status=500)


@login_required
@api_org_devices_schema
@permission_required("dashboard.access_dashboard_app", raise_exception=False)
@require_GET
def api_org_devices(request):
    org_id = _parse_int(request.GET.get("org"))
    if not org_id:
        return _err("Параметр org обязателен")
    status_filter = request.GET.get("status")  # 'online' | 'offline' | None
    if status_filter not in ("online", "offline", None):
        status_filter = None
    try:
        data = services.get_org_devices(org_id=org_id, status_filter=status_filter)
        return _ok(data)
    except Exception as e:
        logger.exception("api_org_devices error")
        return _err(str(e), status=500)


@login_required
@api_org_summary_schema
@permission_required("dashboard.access_dashboard_app", raise_exception=False)
@require_GET
def api_org_summary(request):
    try:
        data = services.get_org_summary()
        return _ok(data)
    except Exception as e:
        logger.exception("api_org_summary error")
        return _err(str(e), status=500)


@login_required
@api_recent_activity_schema
@permission_required("dashboard.access_dashboard_app", raise_exception=False)
@require_GET
def api_recent_activity(request):
    org_id = _parse_int(request.GET.get("org"))
    limit = _parse_int(request.GET.get("limit"), default=20)
    if not 1 <= (limit or 20) <= 100:
        limit = 20
    try:
        data = services.get_recent_activity(org_id=org_id, limit=limit)
        return _ok(data)
    except Exception as e:
        logger.exception("api_recent_activity error")
        return _err(str(e), status=500)


@login_required
@api_organizations_schema
@permission_required("dashboard.access_dashboard_app", raise_exception=False)
@require_GET
def api_organizations(request):
    try:
        data = services.get_organizations()
        return _ok(data)
    except Exception as e:
        logger.exception("api_organizations error")
        return _err(str(e), status=500)


@login_required
@api_glpi_cross_check_schema
@permission_required("dashboard.access_dashboard_app", raise_exception=False)
@require_GET
def api_glpi_cross_check(request):
    org_id = _parse_int(request.GET.get("org"))
    try:
        data = services.get_glpi_cross_check(org_id=org_id)
        return _ok(data)
    except Exception as e:
        logger.exception("api_glpi_cross_check error")
        return _err(str(e), status=500)


@login_required
@permission_required("dashboard.access_dashboard_app", raise_exception=False)
@require_POST
def api_glpi_cross_check_refresh(request):
    try:
        from integrations.tasks import cross_check_glpi_task

        result = cross_check_glpi_task.delay()
        return _ok({"task_id": result.id, "message": "Кросс-проверка запущена"})
    except Exception as e:
        logger.exception("api_glpi_cross_check_refresh error")
        return _err(str(e), status=500)


# ─────────────────────────────────────────────────────────────────────────────
# Excel exports
# ─────────────────────────────────────────────────────────────────────────────


def _make_xlsx_response(wb: "openpyxl.Workbook", filename: str) -> HttpResponse:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = HttpResponse(
        buf,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def _header_style(ws, row, cols):
    """Применить стиль заголовка к строке row."""
    fill = PatternFill(fill_type="solid", fgColor="1F7A4A")
    font = Font(bold=True, color="FFFFFF")
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")


@login_required
@permission_required("dashboard.access_dashboard_app", raise_exception=False)
@require_GET
def export_print_trend(request):
    org_id = _parse_int(request.GET.get("org"))
    months = _parse_int(request.GET.get("months"), default=0)
    if months not in (0, 6, 12):
        months = 0
    try:
        data = services.get_print_trend(org_id=org_id, months=months)
    except Exception as e:
        logger.exception("export_print_trend error")
        return _err(str(e), status=500)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Тренд печати"

    headers = ["Месяц", "Отпечатков"]
    ws.append(headers)
    _header_style(ws, 1, len(headers))

    total = 0
    for row in data:
        ws.append([row["label"], row["total"]])
        total += row["total"]

    # Итого
    ws.append(["Итого", total])
    bold = Font(bold=True)
    for col in range(1, 3):
        ws.cell(row=ws.max_row, column=col).font = bold

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 18

    period_label = "все" if months == 0 else f"{months}мес"
    filename = f'print_trend_{period_label}_{datetime.now().strftime("%Y%m%d")}.xlsx'
    return _make_xlsx_response(wb, filename)


@login_required
@permission_required("dashboard.access_dashboard_app", raise_exception=False)
@require_GET
def export_poll_stats(request):
    org_id = _parse_int(request.GET.get("org"))
    period_days = _parse_int(request.GET.get("period"), default=7)
    if period_days not in (7, 30, 90):
        period_days = 7
    try:
        data = services.get_poll_stats(org_id=org_id, period_days=period_days)
    except Exception as e:
        logger.exception("export_poll_stats error")
        return _err(str(e), status=500)

    STATUS_LABELS = {
        "SUCCESS": "Успешно",
        "FAILED": "Ошибка",
        "VALIDATION_ERROR": "Ошибка валидации",
        "HISTORICAL_INCONSISTENCY": "Расхождение истории",
    }

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Статистика опросов"

    headers = ["Статус", "Количество"]
    ws.append(headers)
    _header_style(ws, 1, len(headers))

    total = 0
    for row in data:
        ws.append([STATUS_LABELS.get(row["status"], row["status"]), row["count"]])
        total += row["count"]

    ws.append(["Итого", total])
    bold = Font(bold=True)
    for col in range(1, 3):
        ws.cell(row=ws.max_row, column=col).font = bold

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 14

    filename = f'poll_stats_{period_days}d_{datetime.now().strftime("%Y%m%d")}.xlsx'
    return _make_xlsx_response(wb, filename)


@login_required
@permission_required("dashboard.access_dashboard_app", raise_exception=False)
@require_GET
def export_org_devices(request):
    org_id = _parse_int(request.GET.get("org"))
    if not org_id:
        return _err("Параметр org обязателен")
    status_filter = request.GET.get("status")
    if status_filter not in ("online", "offline", None):
        status_filter = None

    try:
        data = services.get_org_devices(org_id=org_id, status_filter=status_filter)
    except Exception as e:
        logger.exception("export_org_devices error")
        return _err(str(e), status=500)

    from inventory.models import Organization

    try:
        org_name = Organization.objects.get(pk=org_id).name
    except Organization.DoesNotExist:
        org_name = f"org_{org_id}"

    STATUS_LABELS = {
        "SUCCESS": "Успешно",
        "FAILED": "Ошибка",
        "VALIDATION_ERROR": "Ошибка валидации",
        "HISTORICAL_INCONSISTENCY": "Расхождение истории",
        "—": "—",
    }

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Устройства"

    headers = ["IP-адрес", "Модель", "Серийный номер", "Статус", "Последний опрос", "Online"]
    ws.append(headers)
    _header_style(ws, 1, len(headers))

    green = PatternFill(fill_type="solid", fgColor="D4EDDA")
    red = PatternFill(fill_type="solid", fgColor="F8D7DA")

    for row in data:
        last_poll = ""
        if row["last_poll"]:
            from datetime import datetime as dt

            last_poll = dt.fromisoformat(row["last_poll"]).strftime("%d.%m.%Y %H:%M")
        ws_row = [
            row["ip_address"],
            row["model"],
            row["serial_number"],
            STATUS_LABELS.get(row["last_status"], row["last_status"]),
            last_poll,
            "Да" if row["is_online"] else "Нет",
        ]
        ws.append(ws_row)
        fill = green if row["is_online"] else red
        for col in range(1, len(headers) + 1):
            ws.cell(row=ws.max_row, column=col).fill = fill

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 8

    suffix = f"_{status_filter}" if status_filter else ""
    safe_name = "".join(c if c.isalnum() else "_" for c in org_name)[:30]
    filename = f'devices_{safe_name}{suffix}_{datetime.now().strftime("%Y%m%d")}.xlsx'
    return _make_xlsx_response(wb, filename)


# ─────────────────────────────────────────────────────────────────────────────
# Статистика устройств — виджеты
# ─────────────────────────────────────────────────────────────────────────────


@login_required
@permission_required("dashboard.access_dashboard_app", raise_exception=False)
@require_GET
def api_silent_printers(request):
    org_id = _parse_int(request.GET.get("org"))
    days = _parse_int(request.GET.get("days"), default=7)
    if days not in (7, 14, 30):
        days = 7
    try:
        data = services.get_silent_printers(org_id=org_id, days=days)
        return _ok(data)
    except Exception as e:
        logger.exception("api_silent_printers error")
        return _err(str(e), status=500)


@login_required
@permission_required("dashboard.access_dashboard_app", raise_exception=False)
@require_GET
def api_top_by_volume(request):
    org_id = _parse_int(request.GET.get("org"))
    months = _parse_int(request.GET.get("months"), default=0)
    if months not in (0, 6, 12):
        months = 0
    limit = _parse_int(request.GET.get("limit"), default=10)
    if not 1 <= (limit or 10) <= 50:
        limit = 10
    try:
        data = services.get_top_by_volume(org_id=org_id, months=months, limit=limit)
        return _ok(data)
    except Exception as e:
        logger.exception("api_top_by_volume error")
        return _err(str(e), status=500)


def _parse_month(value):
    """'YYYY-MM' или 'YYYY-MM-DD' → date (первый день месяца). Иначе None."""
    if not value:
        return None
    from datetime import date

    try:
        parts = value.split("-")
        return date(int(parts[0]), int(parts[1]), 1)
    except (ValueError, IndexError, TypeError):
        return None


@login_required
@permission_required("dashboard.access_dashboard_app", raise_exception=False)
@require_GET
def api_manufacturer_distribution(request):
    org_id = _parse_int(request.GET.get("org"))
    source = request.GET.get("source", "polling")
    if source not in ("polling", "contracts", "monthly"):
        source = "polling"
    month_from = _parse_month(request.GET.get("month_from"))
    month_to = _parse_month(request.GET.get("month_to"))
    try:
        data = services.get_manufacturer_distribution(
            source=source, org_id=org_id, month_from=month_from, month_to=month_to
        )
        return _ok(data)
    except Exception as e:
        logger.exception("api_manufacturer_distribution error")
        return _err(str(e), status=500)


@login_required
@permission_required("dashboard.access_dashboard_app", raise_exception=False)
@require_GET
def api_report_months(request):
    try:
        months = services.get_report_months()
        return _ok([m.strftime("%Y-%m") for m in months])
    except Exception as e:
        logger.exception("api_report_months error")
        return _err(str(e), status=500)


# ─────────────────────────────────────────────────────────────────────────────
# Полная XLSX-выгрузка статистики — Celery + polling прогресса
# ─────────────────────────────────────────────────────────────────────────────


@login_required
@permission_required("dashboard.access_dashboard_app", raise_exception=False)
@require_POST
def start_statistics_export(request):
    org_id = _parse_int(request.GET.get("org"))
    days = _parse_int(request.GET.get("days"), default=7)
    if days not in (7, 14, 30):
        days = 7
    months = _parse_int(request.GET.get("months"), default=0)
    if months not in (0, 6, 12):
        months = 0
    month_from = request.GET.get("month_from") or None
    month_to = request.GET.get("month_to") or None

    from dashboard.tasks import build_statistics_export_task

    try:
        task_id = build_statistics_export_task.delay(
            org_id=org_id, days=days, months=months, month_from=month_from, month_to=month_to
        ).id
    except Exception as e:
        logger.exception("start_statistics_export enqueue failed")
        return _err(f"Не удалось поставить задачу: {e}", status=500)

    return JsonResponse(
        {
            "ok": True,
            "task_id": task_id,
            "status_url": f"/dashboard/api/statistics-export/{task_id}/status/",
            "download_url": f"/dashboard/api/statistics-export/{task_id}/download/",
        },
        status=202,
    )


@login_required
@permission_required("dashboard.access_dashboard_app", raise_exception=False)
@require_GET
def statistics_export_status(request, task_id):
    from django.core.cache import cache

    from dashboard.tasks import progress_key

    state = cache.get(progress_key(task_id))
    if state is None:
        # Прогресс ещё не записан или истёк — проверим, не упала ли задача.
        from celery.result import AsyncResult

        res = AsyncResult(task_id)
        if res.failed():
            return _ok({"percent": 100, "message": "Ошибка задачи", "log": [], "done": True, "error": str(res.result)})
        return _ok({"percent": 0, "message": "Ожидание запуска…", "log": [], "done": False})

    payload = dict(state)
    if payload.get("done") and not payload.get("error"):
        payload["download_url"] = f"/dashboard/api/statistics-export/{task_id}/download/"
    return _ok(payload)


@login_required
@permission_required("dashboard.access_dashboard_app", raise_exception=False)
@require_GET
def statistics_export_download(request, task_id):
    from base64 import b64decode

    from django.core.cache import cache

    from dashboard.tasks import result_key

    payload = cache.get(result_key(task_id))
    if not payload:
        return _err("Файл не готов или истёк срок хранения. Запустите выгрузку заново.", status=404)

    content = b64decode(payload["content_b64"])
    resp = HttpResponse(
        content,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="{payload["filename"]}"'
    cache.delete(result_key(task_id))
    return resp
