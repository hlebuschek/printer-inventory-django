# monthly_report/management/commands/monthly_report_org_avg.py
"""
Management command: средняя нагрузка ПО ОРГАНИЗАЦИЯМ за период.

В отличие от monthly_report_avg (одна строка = один принтер), здесь
суммируем все принтеры организации в одну строку.

Алгоритм:
1. Группируем MonthlyReport по (organization, serial_number, month) — суммируем
   (end - start) по всем строкам устройства в этом месяце (могут быть отдельные
   строки на A4 и A3).
2. Группируем по organization — суммируем дельты по всем устройствам, считаем
   уникальные серийники и месяцы с данными.
3. Отрицательные дельты (сброс счётчика) зануляем.

Использование:
    python manage.py monthly_report_org_avg --month-from 2025-04 --month-to 2026-03
    python manage.py monthly_report_org_avg --org-name "Больница"
    python manage.py monthly_report_org_avg --output /tmp/org_stats.xlsx
    python manage.py monthly_report_org_avg --format csv
"""

import csv
import logging
import os
from collections import defaultdict
from datetime import date

import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from django.core.management.base import BaseCommand, CommandError
from django.db.models import F, Max, Sum

from monthly_report.models import MonthlyReport

logger = logging.getLogger(__name__)

HEADERS = [
    "Организация",
    "Принтеров",
    "Месяцев",
    "ЧБ A4 (всего)",
    "Цвет A4 (всего)",
    "ЧБ A3 (всего)",
    "Цвет A3 (всего)",
    "Итого за период",
    "Среднее / мес",
]


class Command(BaseCommand):
    help = "Экспорт средней нагрузки печати по организациям за период"

    def add_arguments(self, parser):
        parser.add_argument(
            "--org-name",
            metavar="НАЗВАНИЕ",
            help="Фильтр по названию организации (частичное совпадение)",
        )
        parser.add_argument(
            "--month-from",
            metavar="ГГГГ-ММ",
            help="Начало периода включительно, например: 2025-04",
        )
        parser.add_argument(
            "--month-to",
            metavar="ГГГГ-ММ",
            help="Конец периода включительно, например: 2026-03",
        )
        parser.add_argument(
            "--output",
            metavar="ПУТЬ",
            help="Путь к файлу вывода (по умолчанию: ./monthly_report_org_avg_YYYYMMDD.xlsx или .csv)",
        )
        parser.add_argument(
            "--format",
            choices=["excel", "csv"],
            default="excel",
            dest="fmt",
            help="Формат вывода: excel (по умолчанию) или csv",
        )

    def handle(self, *args, **options):
        org_name = options["org_name"]
        month_from_str = options["month_from"]
        month_to_str = options["month_to"]
        output_path = options["output"]
        fmt = options["fmt"]

        month_from = _parse_month(month_from_str, "month-from") if month_from_str else None
        month_to = _parse_month(month_to_str, "month-to") if month_to_str else None

        qs = MonthlyReport.objects.all()

        if org_name:
            qs = qs.filter(organization__icontains=org_name)
            self.stdout.write(f"Фильтр: организация содержит «{org_name}»")

        if month_from:
            qs = qs.filter(month__gte=month_from)
        if month_to:
            qs = qs.filter(month__lte=month_to)

        # Шаг 1: суммируем дельты по (organization, serial_number, month).
        # Текстовое поле organization может незначительно отличаться между строками
        # одного устройства — берём Max, чтобы получить стабильный ключ.
        device_month_qs = (
            qs.values("serial_number", "month")
            .annotate(
                organization=Max("organization"),
                delta_a4_bw=Sum(F("a4_bw_end") - F("a4_bw_start")),
                delta_a4_color=Sum(F("a4_color_end") - F("a4_color_start")),
                delta_a3_bw=Sum(F("a3_bw_end") - F("a3_bw_start")),
                delta_a3_color=Sum(F("a3_color_end") - F("a3_color_start")),
            )
            .order_by("organization", "serial_number", "month")
        )

        device_month_rows = list(device_month_qs)

        if not device_month_rows:
            self.stdout.write(self.style.WARNING("Данные не найдены по заданным критериям."))
            return

        self.stdout.write(f"Записей device×month: {len(device_month_rows)}")

        # Шаг 2: агрегируем по организации.
        org_data = defaultdict(
            lambda: {
                "serials": set(),
                "months": set(),
                "a4_bw": 0,
                "a4_color": 0,
                "a3_bw": 0,
                "a3_color": 0,
            }
        )

        for r in device_month_rows:
            org = r["organization"] or "—"
            d = org_data[org]
            d["serials"].add(r["serial_number"])
            d["months"].add(r["month"])
            d["a4_bw"] += max(0, r.get("delta_a4_bw") or 0)
            d["a4_color"] += max(0, r.get("delta_a4_color") or 0)
            d["a3_bw"] += max(0, r.get("delta_a3_bw") or 0)
            d["a3_color"] += max(0, r.get("delta_a3_color") or 0)

        rows = []
        for org, d in org_data.items():
            num_printers = len(d["serials"])
            num_months = len(d["months"])
            total = d["a4_bw"] + d["a4_color"] + d["a3_bw"] + d["a3_color"]
            avg_per_month = round(total / num_months, 1) if num_months else 0.0
            rows.append(
                {
                    "org": org,
                    "num_printers": num_printers,
                    "num_months": num_months,
                    "a4_bw": d["a4_bw"],
                    "a4_color": d["a4_color"],
                    "a3_bw": d["a3_bw"],
                    "a3_color": d["a3_color"],
                    "total": total,
                    "avg_per_month": avg_per_month,
                }
            )

        rows.sort(key=lambda r: r["org"])

        self.stdout.write(f"Организаций: {len(rows)}")

        today_str = date.today().strftime("%Y%m%d")
        ext = "xlsx" if fmt == "excel" else "csv"

        if not output_path:
            output_path = os.path.join(os.getcwd(), f"monthly_report_org_avg_{today_str}.{ext}")

        if fmt == "excel":
            _write_excel(rows, output_path)
        else:
            _write_csv(rows, output_path)

        self.stdout.write(self.style.SUCCESS(f"Файл сохранён: {output_path}"))


def _parse_month(value: str, param_name: str) -> date:
    """Парсит строку ГГГГ-ММ → первый день месяца."""
    try:
        year, month = value.split("-")
        return date(int(year), int(month), 1)
    except (ValueError, AttributeError):
        raise CommandError(f"Неверный формат --{param_name}: '{value}'. Ожидается ГГГГ-ММ, например: 2025-04")


def _row_to_list(row: dict) -> list:
    return [
        row["org"],
        row["num_printers"],
        row["num_months"],
        row["a4_bw"],
        row["a4_color"],
        row["a3_bw"],
        row["a3_color"],
        row["total"],
        row["avg_per_month"],
    ]


def _write_excel(rows: list, path: str) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "По организациям"

    bold = Font(bold=True)
    col_widths = [len(h) + 2 for h in HEADERS]

    for col_idx, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = bold

    int_cols = {2, 3, 4, 5, 6, 7, 8}  # счётчики и итого
    float_cols = {9}  # среднее/мес

    for row_idx, row in enumerate(rows, 2):
        values = _row_to_list(row)
        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)

            if col_idx in float_cols and value is not None:
                cell.number_format = "0.0"
            elif col_idx in int_cols and value is not None:
                cell.number_format = "#,##0"

            cell_len = len(str(value)) if value is not None else 0
            if cell_len > col_widths[col_idx - 1]:
                col_widths[col_idx - 1] = cell_len

    for col_idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = min(width + 2, 60)

    ws.freeze_panes = "A2"

    wb.save(path)


def _write_csv(rows: list, path: str) -> None:
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f, delimiter=";")
        writer.writerow(HEADERS)
        for row in rows:
            values = _row_to_list(row)
            formatted = []
            for v in values:
                if v is None:
                    formatted.append("")
                elif isinstance(v, float):
                    formatted.append(str(v).replace(".", ","))
                else:
                    formatted.append(v)
            writer.writerow(formatted)
