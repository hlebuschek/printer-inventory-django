# monthly_report/management/commands/monthly_report_avg.py
"""
Management command: средняя нагрузка принтеров по месяцам на основе MonthlyReport.

Особенность данных: одно устройство может иметь несколько строк в рамках одного
месяца (например, отдельно A4 и A3). Команда суммирует их перед расчётом среднего.

Алгоритм:
1. Группируем MonthlyReport по (serial_number, month) — суммируем (end - start)
   по всем строкам устройства в этом месяце.
2. Для каждого устройства вычисляем среднее по всем месяцам с данными.
3. Отрицательные дельты (сброс счётчика) зануляем.

Использование:
    python manage.py monthly_report_avg
    python manage.py monthly_report_avg --org-name "Больница"
    python manage.py monthly_report_avg --org-name "БЭК" --format csv
    python manage.py monthly_report_avg --month-from 2025-01 --month-to 2025-12
    python manage.py monthly_report_avg --output /tmp/stats.xlsx
"""

import csv
import logging
import os
from datetime import date
from itertools import groupby

import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from django.core.management.base import BaseCommand, CommandError
from django.db.models import F, Max, Sum

from monthly_report.models import MonthlyReport

logger = logging.getLogger(__name__)

HEADERS = [
    "Организация",
    "Серийный номер",
    "Модель оборудования",
    "Первый месяц",
    "Последний месяц",
    "Месяцев с данными",
    "Среднее ЧБ А4 / мес",
    "Среднее Цвет А4 / мес",
    "Среднее ЧБ А3 / мес",
    "Среднее Цвет А3 / мес",
    "Итого среднее / мес",
]


class Command(BaseCommand):
    help = "Экспорт средней нагрузки принтеров по месяцам на основе ежемесячных отчётов"

    def add_arguments(self, parser):
        parser.add_argument(
            "--org-name",
            metavar="НАЗВАНИЕ",
            help="Фильтр по названию организации (частичное совпадение)",
        )
        parser.add_argument(
            "--month-from",
            metavar="ГГГГ-ММ",
            help="Начало периода включительно, например: 2025-01",
        )
        parser.add_argument(
            "--month-to",
            metavar="ГГГГ-ММ",
            help="Конец периода включительно, например: 2025-12",
        )
        parser.add_argument(
            "--output",
            metavar="ПУТЬ",
            help="Путь к файлу вывода (по умолчанию: ./monthly_report_avg_YYYYMMDD.xlsx или .csv)",
        )
        parser.add_argument(
            "--format",
            choices=["excel", "csv"],
            default="excel",
            dest="fmt",
            help="Формат вывода: excel (по умолчанию) или csv",
        )

    def handle(self, *args, **options):
        org_name = options["org_name"]
        month_from_str = options["month_from"]
        month_to_str = options["month_to"]
        output_path = options["output"]
        fmt = options["fmt"]

        # ── Парсинг дат периода ───────────────────────────────────────────────
        month_from = _parse_month(month_from_str, "month-from") if month_from_str else None
        month_to = _parse_month(month_to_str, "month-to") if month_to_str else None

        # ── Агрегирующий запрос: дельты по (serial_number, month) ─────────────
        # Каждое устройство может иметь несколько строк в рамках одного месяца
        # (например, одна строка с A4-данными, другая с A3-данными).
        # Суммируем (end - start) по всем строкам за тот же месяц.
        qs = MonthlyReport.objects.all()

        if org_name:
            qs = qs.filter(organization__icontains=org_name)
            self.stdout.write(f"Фильтр: организация содержит «{org_name}»")

        if month_from:
            qs = qs.filter(month__gte=month_from)
        if month_to:
            qs = qs.filter(month__lte=month_to)

        monthly_deltas = (
            qs
            .values("serial_number", "month")
            .annotate(
                # Берём представительные текстовые поля через Max
                # (значение одинаково для всех строк одного устройства в месяце)
                organization=Max("organization"),
                equipment_model=Max("equipment_model"),
                # Суммируем дельты по всем строкам устройства в этом месяце
                delta_a4_bw=Sum(F("a4_bw_end") - F("a4_bw_start")),
                delta_a4_color=Sum(F("a4_color_end") - F("a4_color_start")),
                delta_a3_bw=Sum(F("a3_bw_end") - F("a3_bw_start")),
                delta_a3_color=Sum(F("a3_color_end") - F("a3_color_start")),
            )
            .order_by("organization", "serial_number", "month")
        )

        monthly_deltas = list(monthly_deltas)

        if not monthly_deltas:
            self.stdout.write(self.style.WARNING("Данные не найдены по заданным критериям."))
            return

        self.stdout.write(f"Записей device×month: {len(monthly_deltas)}")

        # ── Группируем по serial_number → считаем среднее ─────────────────────
        rows = []
        for serial_number, group in groupby(monthly_deltas, key=lambda r: r["serial_number"]):
            months_data = list(group)
            row = _build_row(serial_number, months_data)
            rows.append(row)

        # Сортируем по организации → серийнику
        rows.sort(key=lambda r: (r["org"], r["serial"]))

        self.stdout.write(f"Устройств: {len(rows)}")

        # ── Вывод ─────────────────────────────────────────────────────────────
        today_str = date.today().strftime("%Y%m%d")
        ext = "xlsx" if fmt == "excel" else "csv"

        if not output_path:
            output_path = os.path.join(os.getcwd(), f"monthly_report_avg_{today_str}.{ext}")

        if fmt == "excel":
            _write_excel(rows, output_path)
        else:
            _write_csv(rows, output_path)

        self.stdout.write(self.style.SUCCESS(f"Файл сохранён: {output_path}"))


# ── Вспомогательные функции ───────────────────────────────────────────────────

def _parse_month(value: str, param_name: str) -> date:
    """Парсит строку ГГГГ-ММ → первый день месяца."""
    try:
        year, month = value.split("-")
        return date(int(year), int(month), 1)
    except (ValueError, AttributeError):
        raise CommandError(
            f"Неверный формат --{param_name}: '{value}'. Ожидается ГГГГ-ММ, например: 2025-01"
        )


def _build_row(serial_number: str, months_data: list) -> dict:
    """
    Считает среднемесячную нагрузку для одного устройства.

    months_data — список dict'ов (уже агрегированных по serial_number+month),
    отсортированных по month.
    """
    org = months_data[0]["organization"] or "—"
    model = months_data[0]["equipment_model"] or "—"

    first_month: date = months_data[0]["month"]
    last_month: date = months_data[-1]["month"]
    num_months = len(months_data)

    def safe_sum(field: str) -> float:
        """Сумма дельт по всем месяцам (отрицательные дельты → 0)."""
        total = 0.0
        for m in months_data:
            val = m.get(field) or 0
            total += max(0, val)
        return total

    total_a4_bw = safe_sum("delta_a4_bw")
    total_a4_color = safe_sum("delta_a4_color")
    total_a3_bw = safe_sum("delta_a3_bw")
    total_a3_color = safe_sum("delta_a3_color")

    avg_a4_bw = round(total_a4_bw / num_months, 1)
    avg_a4_color = round(total_a4_color / num_months, 1)
    avg_a3_bw = round(total_a3_bw / num_months, 1)
    avg_a3_color = round(total_a3_color / num_months, 1)
    avg_total = round((total_a4_bw + total_a4_color + total_a3_bw + total_a3_color) / num_months, 1)

    return {
        "org": org,
        "serial": serial_number,
        "model": model,
        "first_month": first_month,
        "last_month": last_month,
        "num_months": num_months,
        "avg_a4_bw": avg_a4_bw,
        "avg_a4_color": avg_a4_color,
        "avg_a3_bw": avg_a3_bw,
        "avg_a3_color": avg_a3_color,
        "avg_total": avg_total,
    }


def _row_to_list(row: dict) -> list:
    return [
        row["org"],
        row["serial"],
        row["model"],
        row["first_month"],
        row["last_month"],
        row["num_months"],
        row["avg_a4_bw"],
        row["avg_a4_color"],
        row["avg_a3_bw"],
        row["avg_a3_color"],
        row["avg_total"],
    ]


def _write_excel(rows: list, path: str) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Средняя нагрузка"

    bold = Font(bold=True)
    col_widths = [len(h) + 2 for h in HEADERS]

    for col_idx, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = bold

    date_fmt = "mm.yyyy"
    num_fmt = "0.0"
    int_fmt = "0"

    date_cols = {4, 5}      # Первый месяц, Последний месяц
    int_cols = {6}           # Месяцев с данными
    float_cols = {7, 8, 9, 10, 11}  # Средние значения

    for row_idx, row in enumerate(rows, 2):
        values = _row_to_list(row)
        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)

            if col_idx in date_cols and value is not None:
                cell.number_format = date_fmt
            elif col_idx in float_cols and value is not None:
                cell.number_format = num_fmt
            elif col_idx in int_cols and value is not None:
                cell.number_format = int_fmt

            cell_len = len(str(value)) if value is not None else 0
            if cell_len > col_widths[col_idx - 1]:
                col_widths[col_idx - 1] = cell_len

    for col_idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = min(width + 2, 60)

    wb.save(path)


def _write_csv(rows: list, path: str) -> None:
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f, delimiter=";")
        writer.writerow(HEADERS)
        for row in rows:
            values = _row_to_list(row)
            formatted = []
            for v in values:
                if isinstance(v, date):
                    formatted.append(v.strftime("%m.%Y"))
                elif v is None:
                    formatted.append("")
                elif isinstance(v, float):
                    formatted.append(str(v).replace(".", ","))
                else:
                    formatted.append(v)
            writer.writerow(formatted)
