# monthly_report/management/commands/monthly_report_capabilities.py
"""
Management command: классификация принтеров по типу печати на основе MonthlyReport.

Категории (взаимоисключающие, по «максимальной» способности):
    A3 цветной — есть отпечатки A3 цвет
    A3 ЧБ      — есть отпечатки A3 ЧБ, но нет A3 цвет
    A4 цветной — нет A3, но есть A4 цвет
    A4 ЧБ      — только A4 ЧБ

Сигнал способности: max(end) > 0 по соответствующему счётчику за период.
Если у серийника видны и A4, и A3 счётчики — принтер считается A3-формата.

Использование:
    python manage.py monthly_report_capabilities
    python manage.py monthly_report_capabilities --org-name "Больница"
    python manage.py monthly_report_capabilities --month-from 2025-01 --month-to 2025-12
    python manage.py monthly_report_capabilities --detail --output /tmp/caps.xlsx
    python manage.py monthly_report_capabilities --format csv --output /tmp/caps.csv
"""

import csv
import logging
import os
from collections import Counter, defaultdict
from datetime import date

import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from django.core.management.base import BaseCommand, CommandError
from django.db.models import Max

from monthly_report.models import MonthlyReport

logger = logging.getLogger(__name__)

CAT_A3_COLOR = "A3 цветной"
CAT_A3_BW = "A3 ЧБ"
CAT_A4_COLOR = "A4 цветной"
CAT_A4_BW = "A4 ЧБ"
CAT_NONE = "Без данных"

CATEGORY_ORDER = [CAT_A4_BW, CAT_A4_COLOR, CAT_A3_BW, CAT_A3_COLOR, CAT_NONE]

MATRIX_HEADERS = ["Организация", *CATEGORY_ORDER, "Всего"]
DETAIL_HEADERS = [
    "Организация",
    "Серийный номер",
    "Модель оборудования",
    "Категория",
    "Макс A4 ЧБ",
    "Макс A4 цвет",
    "Макс A3 ЧБ",
    "Макс A3 цвет",
]


class Command(BaseCommand):
    help = "Классификация принтеров по формату/цвету на основе ежемесячных отчётов"

    def add_arguments(self, parser):
        parser.add_argument(
            "--org-name",
            metavar="НАЗВАНИЕ",
            help="Фильтр по названию организации (частичное совпадение)",
        )
        parser.add_argument(
            "--month-from",
            metavar="ГГГГ-ММ",
            help="Начало периода включительно, например: 2025-01",
        )
        parser.add_argument(
            "--month-to",
            metavar="ГГГГ-ММ",
            help="Конец периода включительно, например: 2025-12",
        )
        parser.add_argument(
            "--detail",
            action="store_true",
            help="Добавить детальный лист со списком принтеров и их категорией",
        )
        parser.add_argument(
            "--output",
            metavar="ПУТЬ",
            help="Путь к файлу вывода (по умолчанию: ./monthly_report_capabilities_YYYYMMDD.xlsx или .csv)",
        )
        parser.add_argument(
            "--format",
            choices=["excel", "csv"],
            default="excel",
            dest="fmt",
            help="Формат вывода: excel (по умолчанию) или csv",
        )

    def handle(self, *args, **options):
        org_name = options["org_name"]
        month_from_str = options["month_from"]
        month_to_str = options["month_to"]
        detail = options["detail"]
        output_path = options["output"]
        fmt = options["fmt"]

        month_from = _parse_month(month_from_str, "month-from") if month_from_str else None
        month_to = _parse_month(month_to_str, "month-to") if month_to_str else None

        qs = MonthlyReport.objects.all()

        if org_name:
            qs = qs.filter(organization__icontains=org_name)
            self.stdout.write(f"Фильтр: организация содержит «{org_name}»")
        if month_from:
            qs = qs.filter(month__gte=month_from)
        if month_to:
            qs = qs.filter(month__lte=month_to)

        # Агрегируем по серийнику: максимальные значения end-счётчиков за период.
        # Если хотя бы один месяц содержит ненулевой end — счётчик «есть».
        aggregated = (
            qs.values("serial_number")
            .annotate(
                organization=Max("organization"),
                equipment_model=Max("equipment_model"),
                max_a4_bw=Max("a4_bw_end"),
                max_a4_color=Max("a4_color_end"),
                max_a3_bw=Max("a3_bw_end"),
                max_a3_color=Max("a3_color_end"),
            )
            .order_by("organization", "serial_number")
        )

        rows = []
        matrix: dict[str, Counter] = defaultdict(Counter)
        for entry in aggregated:
            serial = entry["serial_number"]
            if not serial:
                continue
            org = entry["organization"] or "—"
            category = _classify(entry)
            matrix[org][category] += 1
            rows.append(
                {
                    "org": org,
                    "serial": serial,
                    "model": entry["equipment_model"] or "—",
                    "category": category,
                    "max_a4_bw": entry["max_a4_bw"] or 0,
                    "max_a4_color": entry["max_a4_color"] or 0,
                    "max_a3_bw": entry["max_a3_bw"] or 0,
                    "max_a3_color": entry["max_a3_color"] or 0,
                }
            )

        if not rows:
            self.stdout.write(self.style.WARNING("Данные не найдены по заданным критериям."))
            return

        matrix_rows = _build_matrix_rows(matrix)

        total_printers = sum(sum(c.values()) for c in matrix.values())
        self.stdout.write(f"Организаций: {len(matrix)}, всего принтеров: {total_printers}")
        for cat in CATEGORY_ORDER:
            cnt = sum(c.get(cat, 0) for c in matrix.values())
            if cnt:
                self.stdout.write(f"  {cat}: {cnt}")

        today_str = date.today().strftime("%Y%m%d")
        ext = "xlsx" if fmt == "excel" else "csv"
        if not output_path:
            output_path = os.path.join(os.getcwd(), f"monthly_report_capabilities_{today_str}.{ext}")

        if fmt == "excel":
            _write_excel(matrix_rows, rows if detail else None, output_path)
        else:
            _write_csv(matrix_rows, rows if detail else None, output_path)

        self.stdout.write(self.style.SUCCESS(f"Файл сохранён: {output_path}"))


def _parse_month(value: str, param_name: str) -> date:
    try:
        year, month = value.split("-")
        return date(int(year), int(month), 1)
    except (ValueError, AttributeError):
        raise CommandError(f"Неверный формат --{param_name}: '{value}'. Ожидается ГГГГ-ММ, например: 2025-01")


def _classify(entry: dict) -> str:
    a4_bw = entry["max_a4_bw"] or 0
    a4_color = entry["max_a4_color"] or 0
    a3_bw = entry["max_a3_bw"] or 0
    a3_color = entry["max_a3_color"] or 0

    if a3_color > 0:
        return CAT_A3_COLOR
    if a3_bw > 0:
        return CAT_A3_BW
    if a4_color > 0:
        return CAT_A4_COLOR
    if a4_bw > 0:
        return CAT_A4_BW
    return CAT_NONE


def _build_matrix_rows(matrix: dict) -> list:
    """
    Строит строки матрицы: [Организация, A4 ЧБ, A4 цв, A3 ЧБ, A3 цв, Без данных, Всего].
    Последняя строка — итог по всем организациям.
    Колонка категории включается в шапку, только если есть хоть один принтер этой категории.
    """
    sorted_orgs = sorted(matrix.keys(), key=lambda s: s.lower())

    result = []
    for org in sorted_orgs:
        counts = matrix[org]
        row = [org]
        total = 0
        for cat in CATEGORY_ORDER:
            cnt = counts.get(cat, 0)
            row.append(cnt)
            total += cnt
        row.append(total)
        result.append(row)

    # Итоговая строка
    totals_row = ["Итого"]
    grand_total = 0
    for cat in CATEGORY_ORDER:
        cnt = sum(c.get(cat, 0) for c in matrix.values())
        totals_row.append(cnt)
        grand_total += cnt
    totals_row.append(grand_total)
    result.append(totals_row)
    return result


def _write_excel(matrix_rows: list, detail_rows: list | None, path: str) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Матрица"

    bold = Font(bold=True)
    col_widths = [len(h) + 2 for h in MATRIX_HEADERS]

    for col_idx, header in enumerate(MATRIX_HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = bold

    last_row_idx = len(matrix_rows) + 1
    for row_idx, row in enumerate(matrix_rows, 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == last_row_idx or col_idx == len(MATRIX_HEADERS):
                cell.font = bold
            if col_idx > 1:
                cell.number_format = "0"
            cell_len = len(str(value)) if value is not None else 0
            if cell_len > col_widths[col_idx - 1]:
                col_widths[col_idx - 1] = cell_len

    for col_idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = min(width + 2, 60)
    ws.freeze_panes = "B2"

    if detail_rows:
        ws_detail = wb.create_sheet("Принтеры")
        col_widths = [len(h) + 2 for h in DETAIL_HEADERS]

        for col_idx, header in enumerate(DETAIL_HEADERS, 1):
            cell = ws_detail.cell(row=1, column=col_idx, value=header)
            cell.font = bold

        for row_idx, row in enumerate(detail_rows, 2):
            values = [
                row["org"],
                row["serial"],
                row["model"],
                row["category"],
                row["max_a4_bw"],
                row["max_a4_color"],
                row["max_a3_bw"],
                row["max_a3_color"],
            ]
            for col_idx, value in enumerate(values, 1):
                cell = ws_detail.cell(row=row_idx, column=col_idx, value=value)
                if col_idx >= 5:
                    cell.number_format = "0"
                cell_len = len(str(value)) if value is not None else 0
                if cell_len > col_widths[col_idx - 1]:
                    col_widths[col_idx - 1] = cell_len

        for col_idx, width in enumerate(col_widths, 1):
            ws_detail.column_dimensions[get_column_letter(col_idx)].width = min(width + 2, 60)

    wb.save(path)


def _write_csv(matrix_rows: list, detail_rows: list | None, path: str) -> None:
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f, delimiter=";")
        writer.writerow(MATRIX_HEADERS)
        for row in matrix_rows:
            writer.writerow(row)

        if detail_rows:
            writer.writerow([])
            writer.writerow(DETAIL_HEADERS)
            for row in detail_rows:
                writer.writerow(
                    [
                        row["org"],
                        row["serial"],
                        row["model"],
                        row["category"],
                        row["max_a4_bw"],
                        row["max_a4_color"],
                        row["max_a3_bw"],
                        row["max_a3_color"],
                    ]
                )
