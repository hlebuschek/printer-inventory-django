import io
from decimal import Decimal

from openpyxl import load_workbook

from contracts.models import Contract
from monthly_report.services.cost_metrics import month_cost_rows
from monthly_report.services.excel_export import export_month_to_excel
from monthly_report.tests_sla_metrics import MARCH, MetricsBase, dt


class CostBase(MetricsBase):
    @classmethod
    def setUpTestData(cls):
        super().setUpTestData()
        cls.contract = Contract.objects.create(
            number="Д-1",
            provider=cls.provider,
            price_a4_bw=Decimal("0.90"),
            price_a4_color=Decimal("5.00"),
            price_a3_bw=Decimal("1.80"),
            price_a3_color=Decimal("10.00"),
        )
        cls.device.contract = cls.contract
        cls.device.save(update_fields=["contract"])

    def make_priced_row(self, **kwargs):
        kwargs.setdefault("a4_bw_start", 100)
        kwargs.setdefault("a4_bw_end", 200)  # 100 × 0.90 = 90
        kwargs.setdefault("a4_color_start", 10)
        kwargs.setdefault("a4_color_end", 20)  # 10 × 5.00 = 50
        kwargs.setdefault("a3_bw_start", 5)
        kwargs.setdefault("a3_bw_end", 15)  # 10 × 1.80 = 18
        kwargs.setdefault("a3_color_start", 0)
        kwargs.setdefault("a3_color_end", 5)  # 5 × 10.00 = 50
        return self.make_row(**kwargs)

    def cost_rows(self):
        rows, totals = month_cost_rows(MARCH)
        return rows, totals


class BaseCostTests(CostBase):
    def test_base_cost_sums_page_types_by_contract_prices(self):
        self.make_priced_row()
        rows, totals = self.cost_rows()

        self.assertEqual(rows[0]["pages"], 125)
        self.assertEqual(rows[0]["base"], Decimal("208.00"))
        self.assertEqual(totals["base"], Decimal("208.00"))

    def test_row_without_contract_gets_note_instead_of_price(self):
        self.make_row(serial="НЕТ-В-ДОГОВОРЕ")
        rows, totals = self.cost_rows()

        self.assertIsNone(rows[0]["base"])
        self.assertEqual(rows[0]["note"], "устройство не найдено в договорах")
        self.assertEqual(totals["base"], Decimal("0"))

    def test_duplicate_rows_split_a4_and_a3(self):
        # Как в recompute_group: первая строка дубля — только A4, вторая — только A3
        first = self.make_priced_row(inventory_number="INV-1")
        second = self.make_priced_row(inventory_number="INV-2")
        rows, _ = self.cost_rows()
        by_id = {row["report"].id: row for row in rows}

        self.assertEqual(by_id[first.id]["base"], Decimal("140.00"))  # 90 + 50
        self.assertEqual(by_id[second.id]["base"], Decimal("68.00"))  # 18 + 50


class ActualCostTests(CostBase):
    def test_meeting_targets_gives_full_price(self):
        # Заявок нет: K1 = K2 = 100 — целевые показатели (98/85) соблюдены, F = B
        self.make_priced_row()
        rows, _ = self.cost_rows()

        self.assertEqual(rows[0]["actual"], Decimal("208.00"))
        self.assertEqual(rows[0]["note"], "целевые показатели соблюдены")

    def test_below_target_multiplies_by_k1_k2(self):
        # Две заявки без остановки печати (K1 = 100), одна просрочена: K2 = 50 < 85
        self.make_request(stops_printing=False, restored_at=dt("2026-03-05", 10), closed_at=dt("2026-03-05", 11))
        self.make_request(
            registered_at=dt("2026-03-10", 9),
            stops_printing=False,
            restored_at=dt("2026-03-10", 12),
            closed_at=dt("2026-03-10", 13),
        )
        self.make_priced_row()
        rows, _ = self.cost_rows()

        self.assertEqual(rows[0]["actual"], Decimal("104.00"))
        self.assertEqual(rows[0]["note"], "ниже целевых показателей")

    def test_vat_is_added_by_contract_rate(self):
        self.make_priced_row()
        rows, totals = self.cost_rows()

        self.assertEqual(rows[0]["vat"], Decimal("41.60"))  # 20% от 208
        self.assertEqual(rows[0]["with_vat"], Decimal("249.60"))
        self.assertEqual(totals["with_vat"], Decimal("249.60"))


class CostExportTests(CostBase):
    def test_cost_sheet_contains_act_columns_and_total(self):
        self.make_priced_row()
        sheet = load_workbook(io.BytesIO(export_month_to_excel(MARCH).content))["Стоимость (акт)"]

        self.assertEqual(sheet.cell(row=2, column=4).value, "Д-1")
        self.assertEqual(sheet.cell(row=2, column=5).value, 125)
        self.assertEqual(sheet.cell(row=2, column=6).value, 208.0)
        self.assertEqual(sheet.cell(row=2, column=9).value, 208.0)
        self.assertEqual(sheet.cell(row=2, column=11).value, 249.6)
        self.assertEqual(sheet.cell(row=3, column=1).value, "Итого")
        self.assertEqual(sheet.cell(row=3, column=11).value, 249.6)
