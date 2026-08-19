from __future__ import annotations

import io
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from django.http import HttpResponse

from django.utils import timezone

from .cost_metrics import month_cost_rows
from .sla_metrics import month_journal_rows, month_metrics_applied

COST_HEADERS = [
    ("№ п/п", 8),
    ("Модель и наименование оборудования", 35),
    ("Серийный номер оборудования", 20),
    ("Договор", 16),
    ("Количество отпечатков за отчётный период, шт.", 14),
    ("Базовая стоимость Услуги (B), руб. без НДС", 16),
    ("K1, %", 10),
    ("K2, %", 10),
    ("Фактическая стоимость (F = B×K1×K2), руб. без НДС", 16),
    ("НДС, руб.", 12),
    ("Фактическая стоимость с НДС, руб.", 16),
    ("Примечание", 32),
]

JOURNAL_HEADERS = [
    ("Номер заявки", 15),
    ("Серийный номер", 20),
    ("Объект", 35),
    ("Описание", 40),
    ("Зарегистрирована", 18),
    ("Нормативный срок", 18),
    ("Восстановлена", 18),
    ("Дата акта", 18),
    ("Номер акта", 15),
    ("Простой в месяце, раб. ч", 14),
    ("Учтена в W", 12),
    ("Просрочена", 12),
]


def _moment(value):
    """Дата в местном времени строкой: Excel не хранит часовой пояс и молча его теряет."""
    return timezone.localtime(value).strftime("%d.%m.%Y %H:%M") if value else ""


def _write_journal_sheet(wb, month_dt, header_font, header_fill, header_alignment, thin_border):
    """Лист с заявками, из которых сложились D, L и W: без него цифры нечем подтвердить."""
    ws = wb.create_sheet("Заявки")

    for col_idx, (title, width) in enumerate(JOURNAL_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 40

    for row_idx, row in enumerate(month_journal_rows(month_dt), start=2):
        values = [
            row["number"],
            row["serial_number"],
            row["address"],
            row["description"],
            _moment(row["registered_at"]),
            _moment(row["deadline_at"]),
            _moment(row["restored_at"]),
            _moment(row["closed_at"]),
            row["act_number"],
            row["downtime"],
            "да" if row["counts_in_w"] else "нет",
            "да" if row["is_overdue"] else "нет",
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border

    ws.freeze_panes = "A2"


def _write_cost_sheet(wb, month_dt, header_font, header_fill, header_alignment, thin_border):
    """Лист по форме акта сдачи-приёмки (Прил. 3 ТЗ): B, F = B×K1×K2 и НДС по каждому устройству."""
    ws = wb.create_sheet("Стоимость (акт)")

    for col_idx, (title, width) in enumerate(COST_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 40

    rows, totals = month_cost_rows(month_dt)
    for row_idx, row in enumerate(rows, start=2):
        report = row["report"]
        values = [
            report.order_number,
            report.equipment_model,
            report.serial_number,
            row["contract"].number if row["contract"] else "",
            row["pages"],
            row["base"],
            round(report.k1, 2) if report.k1 is not None else None,
            round(report.k2, 2) if report.k2 is not None else None,
            row["actual"],
            row["vat"],
            row["with_vat"],
            row["note"],
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border

    total_row = len(rows) + 2
    total_values = {
        1: "Итого",
        5: totals["pages"],
        6: totals["base"],
        9: totals["actual"],
        10: totals["vat"],
        11: totals["with_vat"],
    }
    for col_idx, value in total_values.items():
        cell = ws.cell(row=total_row, column=col_idx, value=value)
        cell.font = Font(bold=True)
        cell.border = thin_border

    ws.freeze_panes = "A2"


def export_month_to_excel(month_dt: date) -> HttpResponse:
    """
    Экспортирует данные месяца в Excel в том же формате, что и загрузка.
    """
    # Показатели качества пересчитываются перед выгрузкой: файл идёт на сверку с отчётом
    # подрядчика, и устаревшие K1/K2 в нём дороже лишних секунд экспорта. Результат
    # только в памяти: сохранение чисел месяца — отдельная операция под своим правом.
    reports, _, _ = month_metrics_applied(month_dt)
    reports.sort(key=lambda r: (r.order_number, r.organization, r.city, r.equipment_model, r.serial_number))

    # Создаем книгу Excel
    wb = Workbook()
    ws = wb.active
    ws.title = f"{month_dt.strftime('%Y-%m')}"

    # Определяем заголовки (точно как в загрузке)
    headers = [
        "№ п/п",
        "Организация",
        "Филиал",
        "Город",
        "Адрес",
        "Модель и наименование оборудования",
        "Серийный номер оборудования",
        "Инв номер",
        "A4 ч/б начало",
        "A4 ч/б конец",
        "A4 цвет начало",
        "A4 цвет конец",
        "A3 ч/б начало",
        "A3 ч/б конец",
        "A3 цвет начало",
        "A3 цвет конец",
        "Итого отпечатков шт.",
        "Нормативное время доступности (A)",
        "Фактическое время недоступности (D)",
        "K1 = ((A - D)/A)*100%",
        "Количество не просроченных запросов (L)",
        "Общее количество запросов (W)",
        "K2 = (L/W)*100%",
    ]

    # Стили для заголовков
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin")
    )

    # Записываем заголовки
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Устанавливаем ширину колонок
    column_widths = {
        1: 8,  # № п/п
        2: 25,  # Организация
        3: 20,  # Филиал
        4: 15,  # Город
        5: 30,  # Адрес
        6: 35,  # Модель
        7: 20,  # Серийный номер
        8: 15,  # Инв номер
        9: 12,  # A4 ч/б начало
        10: 12,  # A4 ч/б конец
        11: 12,  # A4 цвет начало
        12: 12,  # A4 цвет конец
        13: 12,  # A3 ч/б начало
        14: 12,  # A3 ч/б конец
        15: 12,  # A3 цвет начало
        16: 12,  # A3 цвет конец
        17: 15,  # Итого отпечатков
        18: 12,  # A
        19: 12,  # D
        20: 12,  # K1
        21: 12,  # L
        22: 12,  # W
        23: 12,  # K2
    }

    for col_idx, width in column_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Высота строки заголовка
    ws.row_dimensions[1].height = 40

    # Записываем данные
    for row_idx, report in enumerate(reports, start=2):
        row_data = [
            report.order_number,
            report.organization,
            report.branch,
            report.city,
            report.address,
            report.equipment_model,
            report.serial_number,
            report.inventory_number,
            report.a4_bw_start,
            report.a4_bw_end,
            report.a4_color_start,
            report.a4_color_end,
            report.a3_bw_start,
            report.a3_bw_end,
            report.a3_color_start,
            report.a3_color_end,
            report.total_prints,
            report.normative_availability,
            report.actual_downtime,
            round(report.k1, 2) if report.k1 is not None else None,
            report.non_overdue_requests,
            report.total_requests,
            round(report.k2, 2) if report.k2 is not None else None,
        ]

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border

            # Выравнивание
            if col_idx == 1:  # № п/п
                cell.alignment = Alignment(horizontal="center")
            elif col_idx >= 9:  # Числовые поля
                cell.alignment = Alignment(horizontal="right")
            else:
                cell.alignment = Alignment(horizontal="left")

    # Закрепляем первую строку
    ws.freeze_panes = "A2"

    _write_cost_sheet(wb, month_dt, header_font, header_fill, header_alignment, thin_border)
    _write_journal_sheet(wb, month_dt, header_font, header_fill, header_alignment, thin_border)

    # Сохраняем в байты
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    # Создаем HTTP ответ
    filename = f"monthly_report_{month_dt.strftime('%Y-%m')}.xlsx"
    response = HttpResponse(
        output.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    return response
