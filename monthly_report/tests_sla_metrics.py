import io
from datetime import date

from openpyxl import load_workbook

from contracts.models import ServiceRequest
from contracts.tests.test_service_requests import ServiceRequestBase, dt
from monthly_report.models import MonthlyReport
from monthly_report.services.excel_export import export_month_to_excel
from monthly_report.services.sla_metrics import recalculate_month_metrics

MARCH = date(2026, 3, 1)
APRIL = date(2026, 4, 1)


class MetricsBase(ServiceRequestBase):
    def make_row(self, month=MARCH, serial=None, **kwargs):
        kwargs.setdefault("organization", "Тест")
        kwargs.setdefault("branch", "")
        kwargs.setdefault("city", "Иркутск")
        kwargs.setdefault("address", "ул. Ленина, 1")
        kwargs.setdefault("equipment_model", "HP M404")
        kwargs.setdefault("inventory_number", "")
        return MonthlyReport.objects.create(
            month=month, serial_number=serial if serial is not None else self.device.serial_number, **kwargs
        )

    def row_after_recalc(self, row, month=MARCH):
        recalculate_month_metrics(month)
        row.refresh_from_db()
        return row


class AvailabilityTests(MetricsBase):
    def test_month_without_requests_gives_full_availability(self):
        row = self.row_after_recalc(self.make_row())

        # Март 2026: 22 рабочих дня по 9 часов
        self.assertEqual(row.normative_availability, 198.0)
        self.assertEqual(row.actual_downtime, 0.0)
        self.assertEqual(row.k1, 100.0)

    def test_downtime_is_clipped_to_reported_month(self):
        # Простой с вечера 31 марта до утра 1 апреля делится между отчётными периодами
        self.make_request(registered_at=dt("2026-03-31", 15), restored_at=dt("2026-04-01", 10))
        march = self.row_after_recalc(self.make_row())
        april = self.row_after_recalc(self.make_row(month=APRIL), month=APRIL)

        self.assertEqual(march.actual_downtime, 2.0)
        self.assertEqual(april.actual_downtime, 2.0)

    def test_request_without_print_stop_does_not_touch_k1(self):
        self.make_request(stops_printing=False, restored_at=dt("2026-03-03", 10))
        row = self.row_after_recalc(self.make_row())

        self.assertEqual(row.actual_downtime, 0.0)
        self.assertEqual(row.k1, 100.0)

    def test_rejected_request_is_not_counted(self):
        self.make_request(status=ServiceRequest.REJECTED, restored_at=dt("2026-03-03", 10))
        row = self.row_after_recalc(self.make_row())

        self.assertEqual(row.actual_downtime, 0.0)
        self.assertEqual(row.total_requests, 0)

    def test_open_request_counts_downtime_until_month_end(self):
        # Заявка так и не закрыта: простой идёт до конца отчётного периода
        self.make_request(registered_at=dt("2026-03-30", 15))
        row = self.row_after_recalc(self.make_row())

        # 2 часа 30 марта + два полных рабочих дня 31 марта и... март заканчивается 31-м
        self.assertEqual(row.actual_downtime, 11.0)


class IncidentTests(MetricsBase):
    def test_request_belongs_to_month_of_its_deadline(self):
        # Подана 31 марта в 15:00, норматив 8 рабочих часов → срок истекает 1 апреля
        self.make_request(
            registered_at=dt("2026-03-31", 15), restored_at=dt("2026-04-01", 10), closed_at=dt("2026-04-01", 11)
        )
        march = self.row_after_recalc(self.make_row())
        april = self.row_after_recalc(self.make_row(month=APRIL), month=APRIL)

        self.assertEqual(march.total_requests, 0)
        self.assertEqual(april.total_requests, 1)
        self.assertEqual(april.non_overdue_requests, 1)
        self.assertEqual(april.k2, 100.0)

    def test_overdue_request_lowers_k2(self):
        self.make_request(restored_at=dt("2026-03-05", 10), closed_at=dt("2026-03-05", 11))
        row = self.row_after_recalc(self.make_row())

        self.assertEqual(row.total_requests, 1)
        self.assertEqual(row.non_overdue_requests, 0)
        self.assertEqual(row.k2, 0.0)

    def test_request_outside_sla_is_not_counted_in_w(self):
        self.make_request(counts_in_sla=False, restored_at=dt("2026-03-03", 10))
        row = self.row_after_recalc(self.make_row())

        self.assertEqual(row.total_requests, 0)
        # Заявок с нормативным сроком не было — K2 не должен обнулять оплату
        self.assertEqual(row.k2, 100.0)

    def test_reopen_does_not_count_as_second_request(self):
        origin = self.make_request(restored_at=dt("2026-03-02", 15), closed_at=dt("2026-03-02", 16))
        self.make_request(registered_at=dt("2026-03-04", 10), reopened_from=origin)
        row = self.row_after_recalc(self.make_row())

        self.assertEqual(row.total_requests, 1)


class RowMatchingTests(MetricsBase):
    def test_row_without_contract_device_keeps_imported_numbers(self):
        row = self.make_row(serial="НЕТ-В-ДОГОВОРЕ", normative_availability=100.0, actual_downtime=5.0)
        row = self.row_after_recalc(row)

        self.assertEqual(row.normative_availability, 100.0)
        self.assertEqual(row.actual_downtime, 5.0)

    def test_serial_matching_ignores_case_and_spaces(self):
        row = self.row_after_recalc(self.make_row(serial=f"  {self.device.serial_number.lower()} "))

        self.assertEqual(row.normative_availability, 198.0)

    def test_duplicate_rows_get_the_same_metrics(self):
        self.make_request(restored_at=dt("2026-03-02", 16))
        first = self.make_row(inventory_number="INV-1")
        second = self.make_row(inventory_number="INV-2")
        recalculate_month_metrics(MARCH)
        first.refresh_from_db()
        second.refresh_from_db()

        # K1/K2 — характеристики единицы оборудования, дубли строк делят их, а не суммируют
        self.assertEqual(first.actual_downtime, 2.0)
        self.assertEqual(second.actual_downtime, 2.0)
        self.assertEqual(first.k1, second.k1)


class StatsTests(MetricsBase):
    def test_stats_report_unmatched_rows(self):
        self.make_row()
        self.make_row(serial="НЕТ-В-ДОГОВОРЕ")

        stats = recalculate_month_metrics(MARCH)

        self.assertEqual(stats["updated_rows"], 1)
        self.assertEqual(stats["devices_matched"], 1)
        self.assertEqual(stats["rows_without_device"], 1)


class ExportTests(MetricsBase):
    def sheets(self):
        return load_workbook(io.BytesIO(export_month_to_excel(MARCH).content))

    def test_export_refreshes_stale_metrics(self):
        self.make_request(restored_at=dt("2026-03-02", 16))
        self.make_row(normative_availability=1.0, actual_downtime=999.0, k1=0.0)

        row = self.sheets()["2026-03"][2]

        self.assertEqual(row[17].value, 198.0)
        self.assertEqual(row[18].value, 2.0)
        self.assertEqual(row[19].value, 98.99)

    def test_export_does_not_persist_metrics(self):
        # Выгрузка доступна по праву на чтение, а перезапись чисел месяца — по
        # sync_from_inventory, поэтому экспорт обязан оставить БД нетронутой
        self.make_request(restored_at=dt("2026-03-02", 16))
        row = self.make_row(normative_availability=1.0, actual_downtime=999.0, k1=0.0)

        self.sheets()
        row.refresh_from_db()

        self.assertEqual(row.normative_availability, 1.0)
        self.assertEqual(row.actual_downtime, 999.0)
        self.assertEqual(row.k1, 0.0)

    def test_journal_sheet_shows_requests_behind_the_numbers(self):
        service_request = self.make_request(restored_at=dt("2026-03-02", 16), closed_at=dt("2026-03-02", 17))
        self.make_row()

        journal = self.sheets()["Заявки"]

        self.assertEqual(journal.cell(row=2, column=1).value, service_request.number)
        self.assertEqual(journal.cell(row=2, column=10).value, 2.0)
        self.assertEqual(journal.cell(row=2, column=11).value, "да")
        self.assertEqual(journal.cell(row=2, column=12).value, "нет")

    def test_journal_sheet_is_empty_without_requests(self):
        self.make_row()

        self.assertEqual(self.sheets()["Заявки"].max_row, 1)
