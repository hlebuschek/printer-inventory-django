"""
Диагностика обогащения серийниками для одной заявки Okdesk.

Запускать на проде из корня проекта:
    cd /var/www/printer-inventory
    source .venv/bin/activate
    python diag_okdesk_enrich.py             # по умолчанию issue_id=14105
    python diag_okdesk_enrich.py 12345       # другой issue_id
"""

import io
import logging
import os
import sys

# --- Django setup ---
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "printer_inventory.settings")
import django  # noqa: E402

django.setup()

# Подробный лог, чтобы все debug-сообщения из okdesk_enrichment были видны
logging.basicConfig(
    level=logging.DEBUG,
    stream=sys.stdout,
    format="%(levelname)s %(name)s: %(message)s",
)

import requests  # noqa: E402
from django.conf import settings  # noqa: E402

from integrations.okdesk_enrichment import (  # noqa: E402
    build_reference_serials,
    enrich_issue,
    find_serials_in_text,
    parse_html_table,
)


def section(title):
    print("\n" + "=" * 70)
    print(title)
    print("=" * 70)


def main():
    issue_id = int(sys.argv[1]) if len(sys.argv) > 1 else 14105

    api_token = getattr(settings, "OKDESK_API_TOKEN", None)
    api_url = getattr(settings, "OKDESK_API_URL", "https://abikom.okdesk.ru/api/v1")
    verify = getattr(settings, "OKDESK_VERIFY_SSL", True)

    print(f"ISSUE_ID = {issue_id}")
    print(f"OKDESK_API_URL = {api_url}")
    print(f"OKDESK_VERIFY_SSL = {verify}")
    print(f"OKDESK_API_TOKEN set = {bool(api_token)}")

    if not api_token:
        print("FAIL: OKDESK_API_TOKEN не настроен")
        return

    # 1. openpyxl
    section("1) openpyxl установлен?")
    try:
        import openpyxl

        print(f"OK openpyxl {openpyxl.__version__}")
    except Exception as e:
        print(f"FAIL openpyxl: {e!r}")
        openpyxl = None

    # 2. GET /issues/<id>/
    section(f"2) GET /issues/{issue_id}/")
    try:
        r = requests.get(
            f"{api_url}/issues/{issue_id}/",
            params={"api_token": api_token},
            verify=verify,
            timeout=30,
        )
        print(f"status: {r.status_code}")
        r.raise_for_status()
        detail = r.json()
    except Exception as e:
        print(f"FAIL detail request: {e!r}")
        return

    title = detail.get("title") or ""
    desc = detail.get("description") or ""
    equipments = detail.get("equipments") or []
    attachments = detail.get("attachments") or []

    print(f"title: {title}")
    print(f"equipments len: {len(equipments)}")
    if equipments:
        for eq in equipments:
            print(f"  - eq id={eq.get('id')} sn={eq.get('serial_number')!r}")
    print(f"description bytes: {len(desc)}")
    print(f"description preview: {desc[:400]!r}")
    print(f"attachments: {[(a.get('id'), a.get('attachment_file_name')) for a in attachments]}")

    # 3. Справочник серийников
    section("3) Справочник серийников из ContractDevice")
    ref, lookup = build_reference_serials()
    print(f"reference_serials count: {len(ref)}")
    print(f"reference_lookup count:  {len(lookup)}")
    # Проверим, есть ли там R4Q9425700 (известный для 14105)
    test_sn = "R4Q9425700"
    print(f"'{test_sn}' in reference_serials: {test_sn in ref}")
    print(f"'{test_sn}' in reference_lookup: {test_sn in lookup}")

    # 4. Что находит каждый шаг enrich по отдельности
    section("4) Что находит каждый шаг (изолированно)")
    print(f"title  -> {find_serials_in_text(title, ref)}")
    print(f"table  -> {parse_html_table(desc, lookup)}")
    print(f"text   -> {find_serials_in_text(title + ' ' + desc, ref)}")

    # 5. Excel-вложения с подробным логом
    section("5) Excel-вложения (расширенный лог)")
    if not attachments:
        print("Нет вложений вообще.")
    elif openpyxl is None:
        print("Пропуск — openpyxl не установлен.")
    else:
        for att in attachments:
            name = (att.get("attachment_file_name") or "").lower()
            if not name.endswith((".xlsx", ".xls")):
                print(f"skip non-excel: {name}")
                continue

            print(f"\n--- {name} (id={att.get('id')}) ---")
            try:
                meta = requests.get(
                    f"{api_url}/issues/{issue_id}/attachments/{att['id']}",
                    params={"api_token": api_token},
                    verify=verify,
                    timeout=15,
                )
                print(f"meta status: {meta.status_code}")
                meta.raise_for_status()
                meta_json = meta.json()
                url = meta_json.get("attachment_url")
                if not url:
                    print(f"FAIL: пустой attachment_url, ответ: {meta_json}")
                    continue

                host = url.split("/")[2] if "://" in url else url
                print(f"attachment_url host: {host}")

                file_resp = requests.get(url, verify=verify, timeout=60)
                print(f"file status: {file_resp.status_code} | bytes: {len(file_resp.content)}")
                file_resp.raise_for_status()

                wb = openpyxl.load_workbook(io.BytesIO(file_resp.content), read_only=True)
                ws = wb.active
                print(f"sheet: {ws.title} | rows~{ws.max_row} | cols~{ws.max_column}")

                # Первая строка — то, что смотрит реальный код
                row1 = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
                print(f"row1 (где парсер ищет заголовок): {row1}")

                # Расширенный поиск заголовка по первым 10 строкам
                found_header = False
                for ridx, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), 1):
                    for j, cell in enumerate(row):
                        if cell and "серийн" in str(cell).lower():
                            print(f"FOUND header «серийн…» at row={ridx} col={j}: {cell!r}")
                            found_header = True
                if not found_header:
                    print("⚠️  Заголовок «серийн…» не найден в первых 10 строках")

                wb.close()
            except Exception as e:
                print(f"FAIL excel: {e!r}")

    # 6. Финальный результат enrich_issue (как реально работает прод)
    section("6) Полный enrich_issue (как в Celery)")
    res = enrich_issue(
        issue_id=issue_id,
        title=title,
        equipments=equipments,
        reference_serials=ref,
        reference_lookup=lookup,
        api_token=api_token,
        api_url=api_url,
    )
    print(f"RESULT: serial_numbers={res[0]!r}, source={res[1]!r}")


if __name__ == "__main__":
    main()
